"""
IPEDS Completions Explorer
Streamlit app — academic years 2013-14 through 2022-23
(IPEDS file C{YYYY}_A reports awards conferred July YYYY-1 through June YYYY,
i.e. AY (YYYY-1)-YYYY. The DB stores the file's YYYY in the `year` column,
so DB year=2023 means AY 2022-23. Re-run setup_ipeds.py to load AY 2023-24.)
"""

import sqlite3
import urllib.request
from contextlib import contextmanager
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests as _requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from plotly.subplots import make_subplots
import streamlit as st

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing

    _HAS_STATSMODELS = True
except ImportError:
    _HAS_STATSMODELS = False

import rankings as _rankings  # Rankings/grading engine — see rankings.py

# ── Config ────────────────────────────────────────────────────────────────────
# Feature flag: hide the Job Posting Trends section from the UI. The Coresignal
# backend (run_coresignal_trend, _resolve_coresignal_titles, etc. ~line 1290+)
# stays available and can be exercised programmatically — only the rendering
# block in main() is gated on this flag. Flip to True to re-expose the section.
SHOW_JOB_POSTINGS_UI = False

# Feature flag: hide the Distance Education section from the UI. The DEP query
# helpers (run_dep_query, run_dep_by_state_query) and the export pipeline
# entry stay loaded — only the rendering block is gated. Flip to True to
# re-expose the section.
SHOW_DISTANCE_EDUCATION_UI = False

# GitHub Release URL for the database (used on Streamlit Community Cloud)
_GITHUB_DB_URL = (
    "https://github.com/bradycolby-wq/ipeds-completions/releases/"
    "download/v1.6/ipeds.db"
)


def _get_db_path() -> Path:
    """Return path to ipeds.db, downloading from GitHub Release if needed."""
    local = Path(__file__).parent / "ipeds.db"
    if local.exists():
        return local  # local development

    # Cloud deployment: download to a writable cache location
    cache_dir = Path.home() / ".cache" / "ipeds"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = cache_dir / "ipeds.db"
    version_file = cache_dir / "db_version.txt"

    # Re-download when the release URL changes (new version uploaded)
    current_version = _GITHUB_DB_URL
    cached_version = version_file.read_text().strip() if version_file.exists() else ""

    if not cached.exists() or cached_version != current_version:
        with st.spinner("Downloading database (~600 MB) — this takes ~60 seconds on first launch..."):
            urllib.request.urlretrieve(_GITHUB_DB_URL, cached)
            version_file.write_text(current_version)

    return cached


DB_PATH = _get_db_path()

st.set_page_config(
    page_title="IPEDS Completions Explorer",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Authentication gate (Google OAuth via st.login) ─────────────────────────
ALLOWED_EMAILS = {
    "brady.colby@validatedinsights.com",
}

# Detect whether the [auth] section actually loaded into secrets. If it
# didn't, Streamlit's st.user lacks is_logged_in and crashes the whole app
# with an unhelpful AttributeError. Show a clear config error instead.
if "auth" not in st.secrets:
    st.error(
        "Auth is not configured: the `[auth]` section is missing from "
        "secrets.toml. Check that `start.sh` is the service start command "
        "and that REDIRECT_URI, COOKIE_SECRET, GOOGLE_CLIENT_ID, and "
        "GOOGLE_CLIENT_SECRET environment variables are set."
    )
    st.stop()

if not st.user.is_logged_in:
    # Split-screen login: solid VI-Orange brand panel on the left (no
    # gradients per VI brand standards), clean white sign-in form on the
    # right. Montserrat throughout. The left panel is position:fixed so it
    # escapes Streamlit's centered block-container; the block-container is
    # then shifted into the right half to hold the actual form + button.
    st.html(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');

        /* Hide chrome on the login screen */
        section[data-testid="stSidebar"] { display: none !important; }
        div[data-testid="collapsedControl"] { display: none !important; }
        [data-testid="stHeader"] { background: transparent !important; }

        [data-testid="stAppViewContainer"], [data-testid="stApp"] { background: #ffffff !important; }

        /* ── Left brand panel ───────────────────────────────────────── */
        .vi-auth-left {
            position: fixed; top: 0; left: 0;
            width: 50vw; height: 100vh;
            background-color: #F26822;      /* VI Orange — solid, no gradient */
            /* Faint swoosh / linear pattern (brand element). The SVG is
               percent-encoded (no raw angle brackets) so Streamlit's st.html
               sanitizer keeps the surrounding stylesheet intact. */
            background-image: url("data:image/svg+xml;utf8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 600 600' preserveAspectRatio='xMaxYMin slice'%3E%3Cg fill='none' stroke='%23ffffff' stroke-width='2' stroke-opacity='0.13'%3E%3Cpath d='M-80 680 A 760 760 0 0 1 680 -80'/%3E%3Cpath d='M-80 560 A 640 640 0 0 1 560 -80'/%3E%3Cpath d='M-80 440 A 520 520 0 0 1 440 -80'/%3E%3Cpath d='M-80 320 A 400 400 0 0 1 320 -80'/%3E%3Cpath d='M-80 200 A 280 280 0 0 1 200 -80'/%3E%3C/g%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: top right;
            background-size: 120% 120%;
            overflow: hidden; z-index: 0;
            display: flex; flex-direction: column; justify-content: center;
            padding: 4rem 4.75rem;
            box-sizing: border-box;
        }
        .vi-auth-mark {
            font-family: 'Montserrat', sans-serif;
            font-weight: 800; font-size: 1.05rem; letter-spacing: 0.18em;
            color: #ffffff; text-transform: uppercase;
            opacity: 0.92; margin: 0 0 1.75rem 0; position: relative; z-index: 1;
        }
        .vi-auth-headline {
            font-family: 'Montserrat', sans-serif;
            font-weight: 800; font-size: 3.4rem; line-height: 1.08;
            color: #ffffff; margin: 0; letter-spacing: -0.01em;
            position: relative; z-index: 1;
        }
        .vi-auth-tag {
            font-family: 'Montserrat', sans-serif;
            font-weight: 400; font-size: 1.15rem; line-height: 1.55;
            color: rgba(255, 255, 255, 0.92);
            margin: 1.5rem 0 0 0; max-width: 26rem;
            position: relative; z-index: 1;
        }
        .vi-auth-foot {
            position: absolute; left: 4.75rem; bottom: 2.25rem;
            font-family: 'Montserrat', sans-serif; font-weight: 400;
            font-size: 0.85rem; color: rgba(255, 255, 255, 0.7);
            z-index: 1;
        }

        /* ── Right form: push Streamlit's block-container into right half ─ */
        .block-container {
            margin-left: 50vw !important;
            max-width: 50vw !important;
            min-height: 100vh;
            display: flex; flex-direction: column; justify-content: center;
            padding: 2rem 5.5rem !important;
        }
        [data-testid="stImageContainer"] img,
        [data-testid="stImage"] img { width: 150px !important; max-width: 150px !important; height: auto !important; }

        .vi-form-eyebrow {
            font-family: 'Montserrat', sans-serif; font-weight: 400;
            color: #999999; font-size: 1rem; margin: 1.75rem 0 0.35rem 0;
        }
        .vi-form-title {
            font-family: 'Montserrat', sans-serif; font-weight: 700;
            font-size: 2.25rem; color: #333333; line-height: 1.1;
            letter-spacing: -0.01em; margin: 0 0 2rem 0;
        }

        /* Google-branded sign-in button: white, gray border, G icon left. */
        div[data-testid="stButton"] button {
            background: #ffffff !important;
            color: #3c4043 !important;
            border: 1px solid #dadce0 !important;
            border-radius: 10px !important;
            font-family: 'Roboto', 'Montserrat', Arial, sans-serif !important;
            font-weight: 500 !important;
            font-size: 15px !important;
            letter-spacing: 0.15px !important;
            height: 52px !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            gap: 12px !important;
            transition: box-shadow 0.15s ease, background-color 0.15s ease !important;
        }
        div[data-testid="stButton"] button::before {
            content: '';
            display: inline-block;
            width: 20px;
            height: 20px;
            background-image: url("data:image/svg+xml;utf8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 48 48'%3E%3Cpath fill='%23FFC107' d='M43.611 20.083H42V20H24v8h11.303c-1.649 4.657-6.08 8-11.303 8-6.627 0-12-5.373-12-12s5.373-12 12-12c3.059 0 5.842 1.154 7.961 3.039l5.657-5.657C34.046 6.053 29.268 4 24 4 12.955 4 4 12.955 4 24s8.955 20 20 20 20-8.955 20-20c0-1.341-.138-2.65-.389-3.917z'/%3E%3Cpath fill='%23FF3D00' d='M6.306 14.691l6.571 4.819C14.655 15.108 18.961 12 24 12c3.059 0 5.842 1.154 7.961 3.039l5.657-5.657C34.046 6.053 29.268 4 24 4 16.318 4 9.656 8.337 6.306 14.691z'/%3E%3Cpath fill='%234CAF50' d='M24 44c5.166 0 9.86-1.977 13.409-5.192l-6.19-5.238A11.91 11.91 0 0 1 24 36c-5.202 0-9.619-3.317-11.283-7.946l-6.522 5.025C9.505 39.556 16.227 44 24 44z'/%3E%3Cpath fill='%231976D2' d='M43.611 20.083H42V20H24v8h11.303a12.04 12.04 0 0 1-4.087 5.571l.003-.002 6.19 5.238C36.971 39.205 44 34 44 24c0-1.341-.138-2.65-.389-3.917z'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-size: contain;
            flex-shrink: 0;
        }
        div[data-testid="stButton"] button:hover {
            background-color: #f8f9fa !important;
            box-shadow: 0 1px 3px rgba(60,64,67,0.2),
                        0 1px 2px rgba(60,64,67,0.1) !important;
            color: #202124 !important;
        }
        div[data-testid="stButton"] button:focus {
            outline: none !important;
            box-shadow: 0 0 0 3px rgba(242, 104, 34, 0.25) !important;
        }

        /* ── Responsive: stack on narrow viewports ───────────────────── */
        @media (max-width: 820px) {
            .vi-auth-left { display: none !important; }
            .block-container {
                margin-left: 0 !important;
                max-width: 100% !important;
                padding: 2rem 1.75rem !important;
            }
        }
        </style>
        """
    )

    # Brand panel rendered in its own st.html call. Streamlit's HTML
    # sanitizer drops a <style> block when it shares a call with other
    # markup, so the stylesheet above must be passed alone.
    st.html(
        """
        <div class="vi-auth-left">
            <p class="vi-auth-mark">Validated Insights</p>
            <h1 class="vi-auth-headline">VI&nbsp;Data<br>Explorer</h1>
            <p class="vi-auth-tag">IPEDS completions, program rankings, and labor-market
               intelligence &mdash; all in one place.</p>
            <p class="vi-auth-foot">&copy; Validated Insights, Inc. All rights reserved.</p>
        </div>
        """
    )

    # Right-half form content (logo renders correctly on the white side)
    st.image("vi-logo.png")
    st.markdown(
        '<p class="vi-form-eyebrow">Please sign in to continue</p>'
        '<h2 class="vi-form-title">Welcome back</h2>',
        unsafe_allow_html=True,
    )
    if st.button("Sign in with Google", use_container_width=True):
        st.login()
    st.stop()

# Allowlist gate — Google sign-in succeeded but email isn't approved
_user_email = (st.user.email or "").lower()
if _user_email not in ALLOWED_EMAILS:
    st.error(
        f"This app is not authorized for {st.user.email}. "
        "Contact Brady for access."
    )
    if st.button("Sign out"):
        st.logout()
    st.stop()

# ── Reference data ────────────────────────────────────────────────────────────

AWARD_LEVELS = {
    1:  "Less than 1-year certificate (pre-2020)",
    2:  "1–2 year certificate",
    3:  "Associate's degree",
    4:  "2–4 year certificate",
    5:  "Bachelor's degree",
    6:  "Post-baccalaureate certificate",
    7:  "Master's degree",
    8:  "Post-master's certificate",
    17: "Doctorate – Research/Scholarship",
    18: "Doctorate – Professional Practice",
    19: "Doctorate – Other",
    20: "Certificate – under 300 clock hours (2020+)",
    21: "Certificate – 300 to 899 clock hours (2020+)",
}

# VI brand chart sequence — exact order specified in vi-branding guidelines.
# Trailing entries (#E87537, #333333) are reserves for the rare >6-series chart.
CHART_COLORS = [
    "#F26822",  # VI Orange
    "#666666",  # VI Gray 2
    "#0F86C1",  # VI Blue
    "#FAA94D",  # Yellow Orange
    "#999999",  # VI Gray 3
    "#6FB6DA",  # Medium Blue
    "#E87537",  # Medium Orange — reserve
    "#333333",  # VI Gray 1 — reserve
]

# State abbreviation -> FIPS code (for BLS OES state area queries)
STABBR_TO_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08",
    "CT": "09", "DE": "10", "DC": "11", "FL": "12", "GA": "13", "HI": "15",
    "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21",
    "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27",
    "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39",
    "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46",
    "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53",
    "WV": "54", "WI": "55", "WY": "56",
}

# Inverse — used to map OES area_code (FIPS) back to state abbreviations.
FIPS_TO_STABBR = {fips: abbr for abbr, fips in STABBR_TO_FIPS.items()}

# Territories excluded from the platform
EXCLUDED_TERRITORIES = {"PR", "VI", "GU", "AS", "MP", "MH", "FM", "PW"}

EMPLOYMENT_COLORS = [
    "#0f86c1", "#e87537", "#6fb6da", "#f26822", "#faa94d",
    "#333333", "#8B5CF6", "#10B981", "#EF4444", "#F59E0B",
    "#6366F1", "#EC4899", "#14B8A6", "#F97316", "#8B5CF6",
]

# VI brand palette derived from the primary orange (#f26822).
# Used for sequential choropleths so all maps share visual language.
VI_BRAND_ORANGE = "#f26822"
VI_BRAND_BLUE = "#0f86c1"
VI_INK = "#1F2937"
VI_MUTED = "#6B7280"
VI_HAIRLINE = "#E5E7EB"

VI_CHOROPLETH_SCALE = [
    [0.00, "#FFF5EE"],
    [0.20, "#FCD7B4"],
    [0.45, "#F9A66B"],
    [0.70, "#F26822"],
    [1.00, "#A03D0A"],
]


def vi_choropleth(
    locations,
    values,
    *,
    title: str,
    colorbar_title: str = "",
    hover_format: str = "{:,.0f}",
    hover_label: str = "Value",
    height: int = 360,
):
    """Return a VI-branded US-states choropleth Plotly figure.

    locations: iterable of 2-letter state abbreviations.
    values:    iterable of numeric values, same length as locations.
    """
    hover_template = (
        f"<b>%{{location}}</b><br>{hover_label}: %{{customdata}}"
        "<extra></extra>"
    )
    if hover_format.endswith("%}"):
        # Already a percent format string like "{:.1f}%"
        custom = [hover_format.format(v) if v is not None else "—" for v in values]
    else:
        custom = [
            hover_format.format(v) if (v is not None and pd.notna(v)) else "—"
            for v in values
        ]

    fig = go.Figure(go.Choropleth(
        locations=list(locations),
        z=list(values),
        customdata=custom,
        locationmode="USA-states",
        colorscale=VI_CHOROPLETH_SCALE,
        marker=dict(line=dict(color="white", width=0.6)),
        colorbar=dict(
            title=dict(text=colorbar_title, font=dict(size=11, color=VI_MUTED)),
            thickness=10,
            len=0.6,
            outlinewidth=0,
            tickfont=dict(size=10, color=VI_MUTED),
        ),
        hovertemplate=hover_template,
    ))
    _has_title = bool(title)
    fig.update_layout(
        title=(
            dict(
                text=f"<b>{title}</b>",
                font=dict(size=14, color=VI_INK),
                x=0, xanchor="left",
            )
            if _has_title else None
        ),
        geo=dict(
            scope="usa",
            bgcolor="white",
            lakecolor="white",
            showlakes=True,
            landcolor="#FAFAFA",
            subunitcolor="white",
            projection_type="albers usa",
        ),
        height=height,
        margin=dict(t=46 if _has_title else 8, b=8, l=8, r=8),
        paper_bgcolor="white",
        font=dict(family="Montserrat, Arial, sans-serif", size=12, color=VI_INK),
    )
    return fig


def vi_ranking_bar(
    labels,
    values,
    *,
    title: str,
    value_label: str = "Value",
    value_format: str = "{:,.0f}",
    height: int = 360,
    truncate_label_at: int = 28,
):
    """Horizontal bar chart for a state/metro ranking next to a choropleth.

    Sorted descending by `values` (passed in any order — sorted internally).
    Top of the chart = largest value. Bars use the VI primary orange.
    """
    df = pd.DataFrame({"label": list(labels), "value": list(values)})
    df = df.dropna(subset=["value"])
    if df.empty:
        return None
    # Sort descending; keep the top entries. Reverse so Plotly renders the
    # largest at the TOP of a horizontal bar chart.
    df = df.sort_values("value", ascending=True)

    def _truncate(s: str) -> str:
        s = str(s)
        if len(s) <= truncate_label_at + 2:
            return s
        return s[:truncate_label_at] + "…"

    df["display_label"] = df["label"].apply(_truncate)
    df["text"] = df["value"].apply(
        lambda v: value_format.format(v) if pd.notna(v) else ""
    )

    fig = go.Figure(go.Bar(
        x=df["value"],
        y=df["display_label"],
        orientation="h",
        marker=dict(
            color=VI_BRAND_ORANGE,
            line=dict(color="white", width=0.5),
        ),
        text=df["text"],
        textposition="outside",
        textfont=dict(
            size=10, family="Montserrat, Arial, sans-serif", color=VI_INK,
        ),
        cliponaxis=False,
        hovertemplate=(
            f"<b>%{{customdata}}</b><br>{value_label}: %{{text}}<extra></extra>"
        ),
        customdata=df["label"],
    ))
    _has_title = bool(title)
    fig.update_layout(
        title=(
            dict(
                text=f"<b>{title}</b>",
                font=dict(size=14, color=VI_INK),
                x=0, xanchor="left",
            )
            if _has_title else None
        ),
        xaxis=dict(
            title="", showgrid=True, gridcolor="#F3F4F6",
            tickformat=",", showline=False, zeroline=False,
        ),
        yaxis=dict(title="", showgrid=False, automargin=True),
        height=height,
        margin=dict(t=46 if _has_title else 8, b=20, l=4, r=40),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(family="Montserrat, Arial, sans-serif", size=11, color=VI_INK),
        bargap=0.18,
    )
    return fig


# ── NCES projection constants ────────────────────────────────────────────────
# Maps IPEDS award level codes to NCES projection categories
NCES_CATEGORY_MAP = {
    3: "associates",
    5: "bachelors",
    7: "masters",
    17: "doctors",
    18: "doctors",
    19: "doctors",
}

# NCES Projections of Education Statistics to 2032, Table 318.10
# Projected total degrees conferred nationally, by category and academic year.
# Year key follows NCES convention: year = START of academic year
#   (e.g. 2024 = AY 2024-25).
# Note: this differs from our IPEDS DB convention, where year = END of AY
#   (e.g. DB year 2025 = AY 2024-25). When looking these up against DB years
#   we bridge with `nces_year = db_year - 1` (see `_nces_growth_index`).
NCES_PROJECTIONS = {
    "associates": {2024: 1029185, 2025: 1047212, 2026: 1067132, 2027: 1085468, 2028: 1100217},
    "bachelors":  {2024: 2167569, 2025: 2217039, 2026: 2270050, 2027: 2319984, 2028: 2363718},
    "masters":    {2024: 864457,  2025: 886365,  2026: 907435,  2027: 925313,  2028: 943396},
    "doctors":    {2024: 203053,  2025: 205173,  2026: 207292,  2027: 210434,  2028: 215090},
}


# ── DB helpers ────────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA query_only = ON")
    return conn


def ensure_cbsa_index():
    """Add CBSA index if missing. Silently skip if DB is read-only."""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_inst_cbsa ON institutions(cbsa)")
        conn.commit()
        conn.close()
    except sqlite3.OperationalError:
        pass  # read-only filesystem; index should already exist


def ensure_award_levels():
    """Ensure award_levels table has 2020+ codes 20 and 21 so the completions
    view populates award_level_name for them. Silently skip if DB is read-only."""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.executemany(
            "INSERT OR REPLACE INTO award_levels VALUES (?, ?)",
            [(k, v) for k, v in AWARD_LEVELS.items() if k in (20, 21)],
        )
        conn.commit()
        conn.close()
    except sqlite3.OperationalError:
        pass


@st.cache_data(show_spinner=False, ttl=3600)
def get_data_windows() -> dict:
    """Return the actual year/date span of each data source in the DB.

    Used by footnotes so the displayed coverage stays in sync with whatever
    data has been loaded, instead of hardcoded year ranges that drift after
    the next IPEDS / OES / Scorecard refresh.

    Keys: completions, dep, oes, scorecard_status, trends.
    Each value is a (min, max) tuple of ints (years) or strings (dates),
    or None if the table doesn't exist / is empty.
    """
    conn = get_conn()
    out = {
        "completions": None, "dep": None, "oes": None,
        "scorecard_status": None, "trends": None,
    }
    try:
        out["completions"] = conn.execute(
            "SELECT MIN(year), MAX(year) FROM completions"
        ).fetchone()
    except Exception:
        pass
    try:
        out["dep"] = conn.execute(
            "SELECT MIN(year), MAX(year) FROM completions_dep"
        ).fetchone()
    except Exception:
        pass
    try:
        out["oes"] = conn.execute(
            "SELECT MIN(year), MAX(year) FROM oes_employment"
        ).fetchone()
    except Exception:
        pass
    try:
        # College Scorecard FoS file is a "most-recent cohorts" snapshot
        # without a per-row year, so we just confirm presence.
        cnt = conn.execute(
            "SELECT COUNT(*) FROM college_scorecard"
        ).fetchone()[0]
        out["scorecard_status"] = "present" if cnt else None
    except Exception:
        pass
    try:
        out["trends"] = conn.execute(
            "SELECT MIN(date), MAX(date) FROM google_trends_time"
        ).fetchone()
    except Exception:
        pass
    conn.close()
    return out


def _ay_label(year_end: int) -> str:
    """Render an IPEDS DB year (= end of academic year) as 'YYYY-YY'.

    DB convention: year=2024 ⇒ AY 2023-24.
    """
    return f"{year_end - 1}–{str(year_end)[-2:]}"


@st.cache_data(show_spinner=False)
def load_states():
    conn = get_conn()
    rows = conn.execute(
        "SELECT DISTINCT stabbr FROM institutions "
        "WHERE stabbr IS NOT NULL AND stabbr != '' ORDER BY stabbr"
    ).fetchall()
    conn.close()
    return [r[0] for r in rows if r[0] not in EXCLUDED_TERRITORIES]


@st.cache_data(show_spinner=False)
def load_cbsas():
    """Return BLS OES metro areas that have IPEDS institutions, sorted by name.
    Excludes metros that are solely in excluded territories (PR, VI, etc.)."""
    territory_placeholders = ",".join(f"'{t}'" for t in EXCLUDED_TERRITORIES)
    conn = get_conn()
    rows = conn.execute(f"""
        SELECT b.cbsa, b.area_name
        FROM bls_oes_areas b
        INNER JOIN (
            SELECT DISTINCT cbsa FROM institutions
            WHERE cbsa IS NOT NULL AND CAST(cbsa AS INTEGER) > 0
              AND stabbr NOT IN ({territory_placeholders})
        ) i ON i.cbsa = b.cbsa
        WHERE b.area_type = 'metro'
        ORDER BY b.area_name
    """).fetchall()
    conn.close()
    return [(r[0], r[1]) for r in rows]


@st.cache_data(show_spinner=False)
def load_cip_options():
    """Return sorted list of (cipcode, display_label) for all codes with data.

    Retired CIP 2010 codes (those remapped to a new CIP 2020 code) are
    hidden — their historical data is pulled in via expand_cip_patterns()
    when the user picks the current CIP 2020 code.
    """
    conn = get_conn()
    rows = conn.execute("""
        SELECT c.cipcode, COALESCE(t.ciptitle, c.cipcode) AS title
        FROM (SELECT DISTINCT cipcode FROM completions) c
        LEFT JOIN cip_taxonomy t ON c.cipcode = t.cipcode
        WHERE c.cipcode NOT IN (
            SELECT old_cipcode FROM cip_crosswalk
            WHERE old_cipcode != '__CHECKED__'
        )
        ORDER BY c.cipcode
    """).fetchall()
    conn.close()
    return [(r[0], f"{r[0]} \u2013 {r[1]}") for r in rows]


@st.cache_data(show_spinner=False)
def load_cip_crosswalk() -> dict[str, list[str]]:
    """Return mapping: new_cipcode -> [old_cipcode, ...] from the crosswalk table."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT new_cipcode, old_cipcode FROM cip_crosswalk "
            "WHERE old_cipcode != '__CHECKED__'"
        ).fetchall()
    except Exception:
        rows = []
    conn.close()
    result: dict[str, list[str]] = {}
    for new, old in rows:
        result.setdefault(new, []).append(old)
    return result


@st.cache_data(show_spinner=False)
def load_cip_crosswalk_reverse() -> dict[str, list[str]]:
    """Return mapping: old_cipcode -> [new_cipcode, ...] from the crosswalk table."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT new_cipcode, old_cipcode FROM cip_crosswalk "
            "WHERE old_cipcode != '__CHECKED__'"
        ).fetchall()
    except Exception:
        rows = []
    conn.close()
    result: dict[str, list[str]] = {}
    for new, old in rows:
        result.setdefault(old, []).append(new)
    return result


def expand_cip_patterns(cip_patterns: tuple) -> tuple:
    """Add predecessor and successor CIP codes for any selected exact codes.

    Expansion is bidirectional: a CIP 2020 code gets its CIP 2010 predecessors
    added, and a retired CIP 2010 code gets its CIP 2020 successor added, so
    the query spans the full 2014–2023 window regardless of which taxonomy
    the selected code belongs to.
    """
    if not cip_patterns:
        return cip_patterns
    forward = load_cip_crosswalk()
    reverse = load_cip_crosswalk_reverse()
    expanded = list(cip_patterns)
    for code in cip_patterns:
        if "%" in code:  # only exact codes have crosswalk entries
            continue
        for old in forward.get(code, []):
            if old not in expanded:
                expanded.append(old)
        for new in reverse.get(code, []):
            if new not in expanded:
                expanded.append(new)
    return tuple(expanded)


@st.cache_data(show_spinner=False, ttl=600)
def run_national_totals(awlevels: tuple):
    """Return {year: total_completions} nationally for the given award levels."""
    conn = get_conn()
    ph = ",".join("?" * len(awlevels))
    df = pd.read_sql_query(
        f"SELECT year, SUM(ctotalt) AS completions "
        f"FROM completions "
        f"WHERE majornum = 1 AND ctotalt > 0 AND awlevel IN ({ph}) "
        f"GROUP BY year ORDER BY year",
        conn,
        params=list(awlevels),
    )
    conn.close()
    return dict(zip(df["year"], df["completions"]))


def _nces_growth_index(selected_awlevels, proj_years):
    """Return {year: growth_index} based on NCES projections.

    The growth index is relative to an estimated base year (the year before the
    first projection year), computed by back-extrapolating from NCES using the
    average annual growth rate over the projection window.
    """
    cats = {NCES_CATEGORY_MAP[al] for al in selected_awlevels if al in NCES_CATEGORY_MAP}
    if not cats:
        return None

    # Combined NCES totals by year.
    # Bridge between conventions: our `proj_years` use the DB convention
    # (year = end of AY); NCES_PROJECTIONS keys use NCES convention
    # (year = start of AY). For DB year Y (= AY (Y-1)-Y) the matching NCES
    # key is Y-1.
    nces = {}
    for y in proj_years:
        nces[y] = sum(NCES_PROJECTIONS.get(c, {}).get(y - 1, 0) for c in cats)

    first_y, last_y = proj_years[0], proj_years[-1]
    if nces.get(first_y, 0) <= 0:
        return None

    # Average annual growth over the NCES projection period
    n = last_y - first_y
    if n > 0 and nces[last_y] > 0:
        cagr = (nces[last_y] / nces[first_y]) ** (1 / n) - 1
    else:
        cagr = 0

    # Back-extrapolate one year to estimate NCES equivalent for our base year
    nces_base = nces[first_y] / (1 + cagr) if cagr else nces[first_y]

    return {y: nces.get(y, nces[last_y]) / nces_base for y in proj_years}


def compute_projection(sel_dict, national_dict, selected_awlevels, n_forward=5):
    """NCES-constrained top-down projection.

    1.  Compute selection's historical *share* of the national total for the
        chosen award levels.
    2.  Project shares forward using recent-weighted linear trend (last 5 years,
        recency-weighted) to preserve current momentum.
    3.  Project national totals forward using NCES growth indices (or Holt
        fallback for levels without NCES coverage).
    4.  Result = projected_share × projected_national.

    Returns list[(year, projected_completions)] or None on failure.
    """
    if not _HAS_STATSMODELS:
        return None

    years = sorted(set(sel_dict) & set(national_dict))
    if len(years) < 3:
        return None

    last_year = years[-1]
    proj_years = list(range(last_year + 1, last_year + n_forward + 1))

    # ── Historical shares ────────────────────────────────────────────────────
    shares = np.array([
        sel_dict[y] / national_dict[y] if national_dict[y] > 0 else 0
        for y in years
    ])

    # ── Project shares (recent-weighted linear trend) ──────────────────────
    # Use the most recent 5 years of share data with recency weighting
    # so the projection preserves current momentum (both growth and decline)
    # rather than dampening it like Holt exponential smoothing does.
    recent_n = min(5, len(shares))
    recent_shares = shares[-recent_n:]
    if recent_shares.std() < 1e-10:
        proj_shares = np.full(n_forward, shares[-1])
    else:
        x = np.arange(recent_n)
        weights = np.linspace(0.5, 1.0, recent_n)  # 2× weight on most recent
        slope = np.polyfit(x, recent_shares, 1, w=weights)[0]
        proj_shares = shares[-1] + slope * np.arange(1, n_forward + 1)
    proj_shares = np.clip(proj_shares, 0, 1)

    # ── Project national totals ──────────────────────────────────────────────
    last_national = national_dict[last_year]
    growth = _nces_growth_index(selected_awlevels, proj_years)

    if growth:
        proj_nationals = np.array([last_national * growth[y] for y in proj_years])
    else:
        # No NCES coverage → Holt on national totals
        nat_vals = np.array([national_dict[y] for y in years])
        try:
            fit = ExponentialSmoothing(
                nat_vals, trend="add", initialization_method="estimated",
            ).fit(optimized=True, use_brute=True)
            proj_nationals = fit.forecast(n_forward)
        except Exception:
            slope = np.polyfit(np.arange(len(nat_vals)), nat_vals, 1)[0]
            proj_nationals = nat_vals[-1] + slope * np.arange(1, n_forward + 1)
        proj_nationals = np.maximum(proj_nationals, 0)

    # ── Final constrained projection ─────────────────────────────────────────
    result = proj_shares * proj_nationals
    return list(zip(proj_years, np.maximum(result, 0).astype(int)))


def compute_emp_cagr_projection(sel_dict: dict, emp_cagr: float | None, n_forward: int = 5):
    """Project completions using employment CAGR blended with recent momentum.

    Blends the BLS employment-projection CAGR with the selection's own recent
    3-year completions CAGR so the projection incorporates actual field-level
    momentum.  Near-term years lean toward completions momentum (60 %);
    later years decay toward the employment rate (70 %).
    Returns list[(year, projected_completions)] or None.
    """
    if emp_cagr is None:
        return None
    years = sorted(sel_dict.keys())
    if not years:
        return None
    last_year = years[-1]
    last_val = sel_dict[last_year]
    if last_val <= 0:
        return None

    # Recent 3-year completions CAGR
    idx_3 = max(0, len(years) - 4)
    val_3ago = sel_dict[years[idx_3]]
    n_yrs = last_year - years[idx_3]
    if val_3ago > 0 and n_yrs > 0:
        recent_cagr = (last_val / val_3ago) ** (1 / n_yrs) - 1
    else:
        recent_cagr = emp_cagr

    proj_years = list(range(last_year + 1, last_year + n_forward + 1))
    result = []
    for i, y in enumerate(proj_years):
        # Decay from 60 % recent / 40 % emp → 30 % recent / 70 % emp
        w = max(0.30, 0.60 - 0.075 * i)
        rate = w * recent_cagr + (1 - w) * emp_cagr
        projected = last_val * (1 + rate) ** (i + 1)
        result.append((y, max(int(round(projected)), 0)))
    return result


def compute_blended_projection(nces_proj, emp_proj, weight_nces: float = 0.5):
    """Blend NCES-constrained and employment-CAGR projections (50/50 average).

    Both inputs are list[(year, value)]. Returns list[(year, blended_value)] or None.
    """
    if not nces_proj or not emp_proj:
        return None
    nces_d = dict(nces_proj)
    emp_d = dict(emp_proj)
    common = sorted(set(nces_d) & set(emp_d))
    if not common:
        return None
    w = weight_nces
    return [
        (y, max(int(round(w * nces_d[y] + (1 - w) * emp_d[y])), 0))
        for y in common
    ]


def apply_capacity_adjustment(projection, program_counts: dict):
    """Adjust projection values by the trend in number of programs offered.

    Computes the 3-year CAGR of program counts and applies it as a
    multiplicative capacity factor: if programs are growing, projections
    are nudged up; if programs are shrinking, projections are nudged down.

    Returns (adjusted_projection, capacity_cagr) or (projection, None) if
    insufficient data.
    """
    if not projection or not program_counts:
        return projection, None

    years = sorted(program_counts.keys())
    if len(years) < 4:
        return projection, None

    last_year = years[-1]
    last_count = program_counts[last_year]
    yr_3ago = last_year - 3
    count_3ago = program_counts.get(yr_3ago)

    if not count_3ago or count_3ago <= 0 or last_count <= 0:
        return projection, None

    cap_cagr = (last_count / count_3ago) ** (1 / 3) - 1

    # Only adjust if the change is meaningful (>0.5%/yr abs)
    if abs(cap_cagr) < 0.005:
        return projection, cap_cagr

    adjusted = []
    for y, val in projection:
        n = y - last_year
        factor = (1 + cap_cagr) ** n
        adjusted.append((y, max(int(round(val * factor)), 0)))
    return adjusted, cap_cagr


def compute_unified_projection(
    sel_dict: dict,
    emp_cagr: float | None,
    program_counts: dict | None,
    n_forward: int = 5,
):
    """Single unified projection blending trend regression, employment growth,
    and program capacity.

    Components (weighted):
      1. Recent-weighted linear regression of completions (base trend)
      2. Employment CAGR from BLS projections (demand signal)
      3. Program count CAGR (capacity signal)

    Returns (list[(year, value)], dict with component info) or (None, {}).
    """
    years = sorted(sel_dict.keys())
    if len(years) < 3:
        return None, {}

    last_year = years[-1]
    last_val = sel_dict[last_year]
    if last_val <= 0:
        return None, {}

    proj_years = list(range(last_year + 1, last_year + n_forward + 1))

    # ── Component 1: Recent-weighted linear regression CAGR ───────────────
    recent_n = min(5, len(years))
    recent_years = years[-recent_n:]
    recent_vals = np.array([sel_dict[y] for y in recent_years], dtype=float)
    x = np.arange(recent_n)
    weights = np.linspace(0.5, 1.0, recent_n)
    slope = np.polyfit(x, recent_vals, 1, w=weights)[0]
    # Convert slope to an approximate CAGR
    mid_val = recent_vals.mean()
    trend_cagr = slope / mid_val if mid_val > 0 else 0.0

    # ── Component 2: Employment CAGR ──────────────────────────────────────
    has_emp = emp_cagr is not None

    # ── Component 3: Program capacity CAGR ────────────────────────────────
    cap_cagr = None
    if program_counts and len(program_counts) >= 4:
        pc_years = sorted(program_counts.keys())
        pc_last = program_counts[pc_years[-1]]
        yr3 = pc_years[-1] - 3
        pc_3ago = program_counts.get(yr3)
        if pc_3ago and pc_3ago > 0 and pc_last > 0:
            cap_cagr = (pc_last / pc_3ago) ** (1 / 3) - 1

    # ── Blend into a single growth rate ───────────────────────────────────
    # Assign weights based on what's available:
    #   - trend always gets weight
    #   - employment gets weight if available
    #   - capacity gets weight if available and meaningful
    components = {"trend": trend_cagr}
    w_trend = 0.50
    w_emp = 0.0
    w_cap = 0.0

    if has_emp and cap_cagr is not None and abs(cap_cagr) >= 0.005:
        # All three available
        w_trend = 0.40
        w_emp = 0.35
        w_cap = 0.25
        components["employment"] = emp_cagr
        components["capacity"] = cap_cagr
    elif has_emp:
        # Trend + employment
        w_trend = 0.55
        w_emp = 0.45
        components["employment"] = emp_cagr
    elif cap_cagr is not None and abs(cap_cagr) >= 0.005:
        # Trend + capacity
        w_trend = 0.65
        w_cap = 0.35
        components["capacity"] = cap_cagr

    blended_rate = w_trend * trend_cagr + w_emp * (emp_cagr or 0) + w_cap * (cap_cagr or 0)
    components["blended_rate"] = blended_rate
    components["weights"] = {
        k: v for k, v in [("trend", w_trend), ("employment", w_emp), ("capacity", w_cap)] if v > 0
    }

    # ── Project forward ───────────────────────────────────────────────────
    result = []
    for i, y in enumerate(proj_years):
        projected = last_val * (1 + blended_rate) ** (i + 1)
        result.append((y, max(int(round(projected)), 0)))

    return result, components


@st.cache_data(show_spinner=False, ttl=600)
def run_institution_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Return year × institution completions using the same filters as run_query."""
    cip_patterns = expand_cip_patterns(cip_patterns)

    conn = get_conn()
    params = []
    where = [
        "majornum = 1",
        "ctotalt IS NOT NULL",
        "ctotalt > 0",
    ]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"awlevel IN ({placeholders})")
        params.extend(awlevels)

    if geo_key == "state" and geo_values:
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"stabbr IN ({placeholders})")
        params.extend(geo_values)
    elif geo_key == "metro" and geo_values:
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"cbsa IN ({placeholders})")
        params.extend(geo_values)

    where_sql = "WHERE " + " AND ".join(where)

    sql = f"""
        SELECT
            year,
            unitid,
            MAX(instnm)       AS instnm,
            MAX(city)         AS city,
            MAX(stabbr)       AS stabbr,
            MAX(control_name) AS control_name,
            SUM(ctotalt)      AS completions
        FROM completions_view
        {where_sql}
        GROUP BY unitid, year
        ORDER BY unitid, year
    """

    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


@st.cache_data(show_spinner=False, ttl=600)
def run_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
    split_by_level: bool,
):
    # Transparently include predecessor CIP 2010 codes for historical continuity
    cip_patterns = expand_cip_patterns(cip_patterns)

    conn = get_conn()
    params = []
    where = [
        "majornum = 1",
        "ctotalt IS NOT NULL",
        "ctotalt > 0",
    ]

    # CIP filter
    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    # Award level filter
    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"awlevel IN ({placeholders})")
        params.extend(awlevels)

    # Geography filter
    if geo_key == "state" and geo_values:
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"stabbr IN ({placeholders})")
        params.extend(geo_values)
    elif geo_key == "metro" and geo_values:
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"cbsa IN ({placeholders})")
        params.extend(geo_values)

    where_sql = "WHERE " + " AND ".join(where)

    if split_by_level:
        select   = "year, awlevel, award_level_name, SUM(ctotalt) AS completions"
        group_by = "year, awlevel, award_level_name"
        order_by = "year, awlevel"
    else:
        select   = "year, SUM(ctotalt) AS completions"
        group_by = "year"
        order_by = "year"

    sql = f"""
        SELECT {select}
        FROM completions_view
        {where_sql}
        GROUP BY {group_by}
        ORDER BY {order_by}
    """

    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


@st.cache_data(show_spinner=False, ttl=600)
def run_program_count_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Count distinct programs (institutions) reporting ≥1 completion per year.

    A 'program' = unique unitid offering the selected CIP/award-level with
    at least one completion. Returns dict {year: program_count}.
    """
    cip_patterns = expand_cip_patterns(cip_patterns)

    conn = get_conn()
    params = []
    where = [
        "majornum = 1",
        "ctotalt IS NOT NULL",
        "ctotalt > 0",
    ]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"awlevel IN ({placeholders})")
        params.extend(awlevels)

    if geo_key == "state" and geo_values:
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"stabbr IN ({placeholders})")
        params.extend(geo_values)
    elif geo_key == "metro" and geo_values:
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"cbsa IN ({placeholders})")
        params.extend(geo_values)

    where_sql = "WHERE " + " AND ".join(where)

    sql = f"""
        SELECT year, COUNT(DISTINCT unitid) AS program_count
        FROM completions_view
        {where_sql}
        GROUP BY year
        ORDER BY year
    """

    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return dict(zip(df["year"], df["program_count"]))


@st.cache_data(show_spinner=False, ttl=600)
def run_distance_ed_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """DE program counts and DE completions per year.

    Joins completions with distance_ed_status (from IPEDS IC survey) on
    (year, unitid) to identify institutions that offer distance education
    PROGRAMS (distpgs = 1).

    Using distnced = 1 was too restrictive (only exclusively-online
    institutions, ~2% of completions).  Using distnced IN (1, 2) captured
    nearly every institution (~100%).  The distpgs flag is the middle
    ground: institutions that offer at least some fully-online programs
    (~57% of completions).

    Only includes programs with at least 1 completion (ctotalt > 0).

    Returns dict with 'de_program_counts' and 'de_completions' keyed by year,
    or None if distance_ed_status table doesn't exist / no data.
    """
    cip_patterns = expand_cip_patterns(cip_patterns)

    conn = get_conn()

    # Gracefully handle missing table
    try:
        conn.execute("SELECT 1 FROM distance_ed_status LIMIT 1")
    except Exception:
        conn.close()
        return None

    params = []
    where = [
        "c.majornum = 1",
        "c.ctotalt IS NOT NULL",
        "c.ctotalt > 0",
        "d.distpgs = 1",
    ]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("c.cipcode LIKE ?" if "%" in p else "c.cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"c.awlevel IN ({placeholders})")
        params.extend(awlevels)

    # Geography requires joining institutions
    geo_join = ""
    if geo_key == "state" and geo_values:
        geo_join = "INNER JOIN institutions i ON c.unitid = i.unitid AND c.year = i.year"
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"i.stabbr IN ({placeholders})")
        params.extend(geo_values)
    elif geo_key == "metro" and geo_values:
        geo_join = "INNER JOIN institutions i ON c.unitid = i.unitid AND c.year = i.year"
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"i.cbsa IN ({placeholders})")
        params.extend(geo_values)

    where_sql = "WHERE " + " AND ".join(where)

    sql = f"""
        SELECT
            c.year,
            COUNT(DISTINCT c.unitid) AS de_program_count,
            SUM(c.ctotalt)           AS de_completions
        FROM completions c
        INNER JOIN distance_ed_status d
            ON c.year = d.year
            AND c.unitid = d.unitid
        {geo_join}
        {where_sql}
        GROUP BY c.year
        ORDER BY c.year
    """

    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()

    if df.empty:
        return None

    return {
        "de_program_counts": dict(zip(df["year"], df["de_program_count"])),
        "de_completions": dict(zip(df["year"], df["de_completions"])),
    }


@st.cache_data(show_spinner=False, ttl=600)
def run_dep_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Distance education program counts from completions_dep table.

    Returns DataFrame with columns:
      year, programs, programs_de, programs_de_some, programs_de_any
    aggregated across matching CIPs, award levels, and geography.
    Or None if completions_dep table doesn't exist / no data.

    Uses 6-digit CIP codes from completions_dep (filters out 2-digit
    summary rows to avoid double-counting). Joins institutions table
    for geographic filtering.
    """
    cip_patterns = expand_cip_patterns(cip_patterns)

    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM completions_dep LIMIT 1")
    except Exception:
        conn.close()
        return None

    params = []
    where = [
        "d.programs > 0",
        "LENGTH(d.cipcode) >= 5",  # exclude 2-digit summary rows
    ]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append(
                "d.cipcode LIKE ?" if "%" in p else "d.cipcode = ?"
            )
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"d.awlevel IN ({placeholders})")
        params.extend(awlevels)

    geo_join = ""
    if geo_key == "state" and geo_values:
        geo_join = (
            "INNER JOIN institutions i "
            "ON d.unitid = i.unitid AND d.year = i.year"
        )
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"i.stabbr IN ({placeholders})")
        params.extend(geo_values)
    elif geo_key == "metro" and geo_values:
        geo_join = (
            "INNER JOIN institutions i "
            "ON d.unitid = i.unitid AND d.year = i.year"
        )
        placeholders = ",".join("?" * len(geo_values))
        where.append(f"i.cbsa IN ({placeholders})")
        params.extend(geo_values)

    where_sql = "WHERE " + " AND ".join(where)

    sql = f"""
        SELECT
            d.year,
            SUM(d.programs)         AS programs,
            SUM(d.programs_de)      AS programs_de,
            SUM(d.programs_de_some) AS programs_de_some,
            SUM(d.programs_de_any)  AS programs_de_any
        FROM completions_dep d
        {geo_join}
        {where_sql}
        GROUP BY d.year
        ORDER BY d.year
    """

    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()

    if df.empty:
        return None

    # Compute percentage columns
    df["pct_de_any"] = (
        100.0 * df["programs_de_any"] / df["programs"]
    ).round(1).where(df["programs"] > 0)

    return df


@st.cache_data(show_spinner=False)
def run_completions_by_state_query(
    cip_patterns: tuple,
    awlevels: tuple,
    year: int,
):
    """Total completions by state for a single year, across all 50 + DC.

    Returns DataFrame with columns: stabbr, completions.
    """
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()

    params = [year]
    where = ["c.year = ?"]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("c.cipcode LIKE ?" if "%" in p else "c.cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"c.awlevel IN ({placeholders})")
        params.extend(awlevels)

    excluded = ",".join(f"'{s}'" for s in EXCLUDED_TERRITORIES)
    where.append(f"i.stabbr NOT IN ({excluded})")
    where.append("i.stabbr IS NOT NULL AND i.stabbr != ''")

    sql = f"""
        SELECT i.stabbr AS stabbr, SUM(c.ctotalt) AS completions
        FROM completions c
        INNER JOIN institutions i
          ON c.unitid = i.unitid AND c.year = i.year
        WHERE {' AND '.join(where)}
        GROUP BY i.stabbr
        ORDER BY completions DESC
    """
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


@st.cache_data(show_spinner=False)
def run_completions_by_metro_query(
    cip_patterns: tuple,
    awlevels: tuple,
    year: int,
    top_n: int = 25,
):
    """Top metros by completions for a single year.

    Returns DataFrame: cbsa, cbsa_name, completions (sorted desc).
    Excludes institutions with no CBSA assignment (rural).
    """
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()

    params = [year]
    where = ["c.year = ?"]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("c.cipcode LIKE ?" if "%" in p else "c.cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"c.awlevel IN ({placeholders})")
        params.extend(awlevels)

    excluded = ",".join(f"'{s}'" for s in EXCLUDED_TERRITORIES)
    where.append(f"i.stabbr NOT IN ({excluded})")
    where.append("i.cbsa IS NOT NULL")
    where.append("CAST(i.cbsa AS INTEGER) > 0")

    params.append(top_n)
    sql = f"""
        SELECT
            i.cbsa AS cbsa,
            COALESCE(n.cbsanm, i.cbsa) AS cbsa_name,
            SUM(c.ctotalt) AS completions
        FROM completions c
        INNER JOIN institutions i
          ON c.unitid = i.unitid AND c.year = i.year
        LEFT JOIN cbsa_names n ON i.cbsa = n.cbsa
        WHERE {' AND '.join(where)}
        GROUP BY i.cbsa
        HAVING completions > 0
        ORDER BY completions DESC
        LIMIT ?
    """
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


@st.cache_data(show_spinner=False)
def run_employment_by_metro_query(soc_codes: tuple, year: int, top_n: int = 25):
    """Top metros by total employment for the given SOC codes in a single year.

    Returns DataFrame: cbsa, cbsa_name, tot_emp (sorted desc).
    OES area_type=4 = metro; BLS area_code is 00 + 5-digit CBSA. Aggregates
    across all matched occupations.
    """
    if not soc_codes:
        return pd.DataFrame()
    conn = get_conn()
    soc_ph = ",".join("?" * len(soc_codes))
    sql = f"""
        SELECT
            area_code,
            MAX(area_title) AS area_title,
            SUM(tot_emp) AS tot_emp
        FROM oes_employment
        WHERE year = ?
          AND area_type = 4
          AND occ_code IN ({soc_ph})
          AND tot_emp IS NOT NULL
        GROUP BY area_code
        HAVING tot_emp > 0
        ORDER BY tot_emp DESC
        LIMIT ?
    """
    df = pd.read_sql_query(sql, conn, params=[year] + list(soc_codes) + [top_n])
    conn.close()
    if df.empty:
        return df
    # BLS area_code is 7-digit "00" + 5-digit CBSA — strip the leading "00".
    df["cbsa"] = df["area_code"].astype(str).str.zfill(7).str[2:]
    df["cbsa_name"] = df["area_title"]
    return df[["cbsa", "cbsa_name", "tot_emp"]].reset_index(drop=True)


# ── Growth queries (post-COVID CAGR per geography) ──────────────────────────

def _cagr(start, end, years):
    """Compound annual growth rate as a decimal (e.g. 0.05 = 5%/yr).

    Returns NaN if either bookend is non-positive or the span is zero.
    """
    if start is None or end is None or years <= 0:
        return float("nan")
    if start <= 0 or end <= 0:
        return float("nan")
    return (end / start) ** (1.0 / years) - 1.0


@st.cache_data(show_spinner=False)
def run_completions_state_cagr(
    cip_patterns: tuple, awlevels: tuple, base_year: int, end_year: int,
):
    """CAGR by state from `base_year` to `end_year`. Decimal cagr.

    Used by the Completions Growth (post-COVID CAGR) view. Returns
    DataFrame: stabbr, base, end, cagr.
    """
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()
    params = [base_year, end_year]
    where = ["c.year IN (?, ?)"]

    if cip_patterns:
        cc = []
        for p in cip_patterns:
            cc.append("c.cipcode LIKE ?" if "%" in p else "c.cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cc)})")

    if awlevels:
        ph = ",".join("?" * len(awlevels))
        where.append(f"c.awlevel IN ({ph})")
        params.extend(awlevels)

    excluded = ",".join(f"'{s}'" for s in EXCLUDED_TERRITORIES)
    where.append(f"i.stabbr NOT IN ({excluded})")
    where.append("i.stabbr IS NOT NULL AND i.stabbr != ''")

    sql = f"""
        SELECT i.stabbr AS stabbr, c.year AS year,
               SUM(c.ctotalt) AS completions
        FROM completions c
        INNER JOIN institutions i
          ON c.unitid = i.unitid AND c.year = i.year
        WHERE {' AND '.join(where)}
        GROUP BY i.stabbr, c.year
    """
    raw = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    if raw.empty:
        return raw

    pivot = raw.pivot(index="stabbr", columns="year", values="completions")
    span = end_year - base_year
    pivot["base"] = pivot.get(base_year)
    pivot["end"] = pivot.get(end_year)
    pivot["cagr"] = pivot.apply(
        lambda r: _cagr(r["base"], r["end"], span), axis=1,
    )
    return (
        pivot.reset_index()[["stabbr", "base", "end", "cagr"]]
        .sort_values("cagr", ascending=False, na_position="last")
        .reset_index(drop=True)
    )


@st.cache_data(show_spinner=False)
def run_completions_metro_cagr(
    cip_patterns: tuple, awlevels: tuple,
    base_year: int, end_year: int, top_n: int = 25,
):
    """CAGR by CBSA from base_year → end_year. Top-N by end-year volume."""
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()
    params = [base_year, end_year]
    where = ["c.year IN (?, ?)"]

    if cip_patterns:
        cc = []
        for p in cip_patterns:
            cc.append("c.cipcode LIKE ?" if "%" in p else "c.cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cc)})")

    if awlevels:
        ph = ",".join("?" * len(awlevels))
        where.append(f"c.awlevel IN ({ph})")
        params.extend(awlevels)

    excluded = ",".join(f"'{s}'" for s in EXCLUDED_TERRITORIES)
    where.append(f"i.stabbr NOT IN ({excluded})")
    where.append("i.cbsa IS NOT NULL AND CAST(i.cbsa AS INTEGER) > 0")

    sql = f"""
        SELECT i.cbsa AS cbsa,
               COALESCE(n.cbsanm, i.cbsa) AS cbsa_name,
               c.year AS year,
               SUM(c.ctotalt) AS completions
        FROM completions c
        INNER JOIN institutions i
          ON c.unitid = i.unitid AND c.year = i.year
        LEFT JOIN cbsa_names n ON i.cbsa = n.cbsa
        WHERE {' AND '.join(where)}
        GROUP BY i.cbsa, c.year
    """
    raw = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    if raw.empty:
        return raw

    meta = raw.drop_duplicates("cbsa")[["cbsa", "cbsa_name"]]
    pivot = raw.pivot(index="cbsa", columns="year", values="completions")
    span = end_year - base_year
    pivot["base"] = pivot.get(base_year)
    pivot["end"] = pivot.get(end_year)
    pivot["cagr"] = pivot.apply(
        lambda r: _cagr(r["base"], r["end"], span), axis=1,
    )
    out = pivot.reset_index().merge(meta, on="cbsa", how="left")
    # Filter: require at least 50 completions in end-year for noise control.
    out = out[out["end"].fillna(0) >= 50]
    out = out.dropna(subset=["cagr"])
    out = out.sort_values("cagr", ascending=False).head(top_n).reset_index(drop=True)
    return out[["cbsa", "cbsa_name", "base", "end", "cagr"]]


@st.cache_data(show_spinner=False)
def run_employment_state_cagr(
    soc_codes: tuple, base_year: int, end_year: int,
):
    """OES total-employment CAGR by state from base_year → end_year."""
    if not soc_codes:
        return pd.DataFrame()
    conn = get_conn()
    soc_ph = ",".join("?" * len(soc_codes))
    sql = f"""
        SELECT area_code, year, SUM(tot_emp) AS tot_emp
        FROM oes_employment
        WHERE year IN (?, ?)
          AND area_type = 2
          AND occ_code IN ({soc_ph})
          AND tot_emp IS NOT NULL
        GROUP BY area_code, year
    """
    raw = pd.read_sql_query(
        sql, conn, params=[base_year, end_year] + list(soc_codes),
    )
    conn.close()
    if raw.empty:
        return raw

    raw["stabbr"] = raw["area_code"].map(FIPS_TO_STABBR)
    raw = raw.dropna(subset=["stabbr"])
    raw = raw[~raw["stabbr"].isin(EXCLUDED_TERRITORIES)]

    pivot = raw.pivot(index="stabbr", columns="year", values="tot_emp")
    span = end_year - base_year
    pivot["base"] = pivot.get(base_year)
    pivot["end"] = pivot.get(end_year)
    pivot["cagr"] = pivot.apply(
        lambda r: _cagr(r["base"], r["end"], span), axis=1,
    )
    return (
        pivot.reset_index()[["stabbr", "base", "end", "cagr"]]
        .sort_values("cagr", ascending=False, na_position="last")
        .reset_index(drop=True)
    )


@st.cache_data(show_spinner=False)
def run_employment_metro_cagr(
    soc_codes: tuple, base_year: int, end_year: int, top_n: int = 25,
):
    """OES total-employment CAGR by metro (CBSA) from base_year → end_year."""
    if not soc_codes:
        return pd.DataFrame()
    conn = get_conn()
    soc_ph = ",".join("?" * len(soc_codes))
    sql = f"""
        SELECT area_code, MAX(area_title) AS area_title, year,
               SUM(tot_emp) AS tot_emp
        FROM oes_employment
        WHERE year IN (?, ?)
          AND area_type = 4
          AND occ_code IN ({soc_ph})
          AND tot_emp IS NOT NULL
        GROUP BY area_code, year
    """
    raw = pd.read_sql_query(
        sql, conn, params=[base_year, end_year] + list(soc_codes),
    )
    conn.close()
    if raw.empty:
        return raw

    raw["cbsa"] = raw["area_code"].astype(str).str.zfill(7).str[2:]
    meta = raw.drop_duplicates("cbsa")[["cbsa", "area_title"]].rename(
        columns={"area_title": "cbsa_name"}
    )
    pivot = raw.pivot(index="cbsa", columns="year", values="tot_emp")
    span = end_year - base_year
    pivot["base"] = pivot.get(base_year)
    pivot["end"] = pivot.get(end_year)
    pivot["cagr"] = pivot.apply(
        lambda r: _cagr(r["base"], r["end"], span), axis=1,
    )
    out = pivot.reset_index().merge(meta, on="cbsa", how="left")
    # Noise filter: drop metros with under 100 employed in end-year.
    out = out[out["end"].fillna(0) >= 100]
    out = out.dropna(subset=["cagr"])
    out = out.sort_values("cagr", ascending=False).head(top_n).reset_index(drop=True)
    return out[["cbsa", "cbsa_name", "base", "end", "cagr"]]


# ── Projection queries (BLS Employment Projections, geo-level) ──────────────

@st.cache_data(show_spinner=False)
def run_employment_projection_state(soc_codes: tuple):
    """Weighted-average projected CAGR by state across the given SOC codes.

    Aggregates `employment_projections` rows where geo_level='state',
    summing base_emp and proj_emp across the matched occupations, then
    deriving an aggregate CAGR.

    Returns DataFrame: stabbr, base_total, proj_total, cagr (decimal).
    """
    if not soc_codes:
        return pd.DataFrame()
    conn = get_conn()
    soc_ph = ",".join("?" * len(soc_codes))
    sql = f"""
        SELECT geo_code, geo_name,
               MAX(base_year) AS base_year, MAX(proj_year) AS proj_year,
               SUM(base_emp)  AS base_total,
               SUM(proj_emp)  AS proj_total
        FROM employment_projections
        WHERE geo_level = 'state'
          AND occ_code IN ({soc_ph})
        GROUP BY geo_code, geo_name
    """
    df = pd.read_sql_query(sql, conn, params=list(soc_codes))
    conn.close()
    if df.empty:
        return df

    df["stabbr"] = df["geo_code"].map(FIPS_TO_STABBR)
    df = df.dropna(subset=["stabbr"])
    df = df[~df["stabbr"].isin(EXCLUDED_TERRITORIES)]
    df["span"] = (df["proj_year"] - df["base_year"]).clip(lower=1)
    df["cagr"] = df.apply(
        lambda r: _cagr(r["base_total"], r["proj_total"], r["span"]),
        axis=1,
    )
    return (
        df[["stabbr", "base_total", "proj_total", "cagr"]]
        .sort_values("cagr", ascending=False, na_position="last")
        .reset_index(drop=True)
    )


@st.cache_data(show_spinner=False)
def run_employment_projection_metro(soc_codes: tuple, top_n: int = 25):
    """Weighted-average projected CAGR by metro (CBSA)."""
    if not soc_codes:
        return pd.DataFrame()
    conn = get_conn()
    soc_ph = ",".join("?" * len(soc_codes))
    sql = f"""
        SELECT geo_code, MAX(geo_name) AS geo_name,
               MAX(base_year) AS base_year, MAX(proj_year) AS proj_year,
               SUM(base_emp)  AS base_total,
               SUM(proj_emp)  AS proj_total
        FROM employment_projections
        WHERE geo_level = 'metro'
          AND occ_code IN ({soc_ph})
        GROUP BY geo_code
    """
    df = pd.read_sql_query(sql, conn, params=list(soc_codes))
    conn.close()
    if df.empty:
        return df

    df["span"] = (df["proj_year"] - df["base_year"]).clip(lower=1)
    df["cagr"] = df.apply(
        lambda r: _cagr(r["base_total"], r["proj_total"], r["span"]),
        axis=1,
    )
    df = df.rename(columns={"geo_code": "cbsa", "geo_name": "cbsa_name"})
    df = df.dropna(subset=["cagr"])
    # Noise filter: require ≥100 base employed.
    df = df[df["base_total"].fillna(0) >= 100]
    return (
        df.sort_values("cagr", ascending=False)
          .head(top_n)
          .reset_index(drop=True)[
              ["cbsa", "cbsa_name", "base_total", "proj_total", "cagr"]
          ]
    )


# ── Search Traffic state-level history (rolling 12-mo change) ───────────────

@st.cache_data(show_spinner=False)
def search_traffic_state_time_coverage(cip_patterns: tuple) -> int:
    """How many (CIP, state) pairs have time-series data loaded?

    Cheap pre-flight check used by the UI to decide whether the Growth
    toggle has anything to render. Returns 0 if the table doesn't exist.
    """
    if not cip_patterns:
        return 0
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM google_trends_state_time LIMIT 1")
    except Exception:
        conn.close()
        return 0
    cip_clauses, params = [], []
    for p in cip_patterns:
        cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
        params.append(p)
    where = " OR ".join(cip_clauses) if cip_clauses else "1=1"
    n = conn.execute(
        f"SELECT COUNT(DISTINCT cipcode || '|' || state_abbr) "
        f"FROM google_trends_state_time WHERE ({where})",
        params,
    ).fetchone()[0]
    conn.close()
    return int(n or 0)


@st.cache_data(show_spinner=False)
def run_search_traffic_state_growth(
    cip_patterns: tuple, window_months: int = 12,
):
    """Rolling 12-month change in search interest per state.

    Compares the average interest over the most recent `window_months`
    versus the prior `window_months` for each state, summed across the
    selected CIP codes. Returns DataFrame: stabbr, recent, prior, pct_change.
    Negative when interest is declining. Returns empty DataFrame if no
    state-level time-series data is loaded.
    """
    if not cip_patterns:
        return pd.DataFrame()
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM google_trends_state_time LIMIT 1")
    except Exception:
        conn.close()
        return pd.DataFrame()

    cip_clauses, params = [], []
    for p in cip_patterns:
        cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
        params.append(p)
    cip_where = " OR ".join(cip_clauses) if cip_clauses else "1=1"

    # Pull the full time series for the selected CIPs, then partition
    # in pandas. This is cleaner than writing two SQL CTEs and the row
    # count is small (≤ 51 states × ≤ 65 months × ≤ N CIPs).
    sql = f"""
        SELECT cipcode, state_abbr, date,
               COALESCE(interest, 0) AS interest
        FROM google_trends_state_time
        WHERE ({cip_where})
    """
    raw = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    if raw.empty:
        return raw

    raw["date"] = pd.to_datetime(raw["date"])
    # Sum across the selected CIPs so multi-CIP selections aggregate
    # cleanly; for a single CIP this is a no-op.
    by_state_date = (
        raw.groupby(["state_abbr", "date"], as_index=False)["interest"].sum()
    )

    # Determine the cutoff between "recent" and "prior" windows. Use
    # the latest date observed across the dataset so partial-month
    # bookends don't silently shift one state's window.
    latest = by_state_date["date"].max()
    cutoff_recent_start = latest - pd.DateOffset(months=window_months - 1)
    cutoff_prior_start  = latest - pd.DateOffset(months=2 * window_months - 1)

    by_state_date["bucket"] = "older"
    by_state_date.loc[
        by_state_date["date"] >= cutoff_recent_start, "bucket"
    ] = "recent"
    by_state_date.loc[
        (by_state_date["date"] >= cutoff_prior_start)
        & (by_state_date["date"] < cutoff_recent_start), "bucket"
    ] = "prior"

    pivot = (
        by_state_date[by_state_date["bucket"].isin(["recent", "prior"])]
        .groupby(["state_abbr", "bucket"], as_index=False)["interest"]
        .mean()
        .pivot(index="state_abbr", columns="bucket", values="interest")
        .reset_index()
    )
    if "recent" not in pivot.columns or "prior" not in pivot.columns:
        return pd.DataFrame()

    pivot["pct_change"] = pivot.apply(
        lambda r: (r["recent"] - r["prior"]) / r["prior"]
        if pd.notna(r["prior"]) and r["prior"] > 0 else float("nan"),
        axis=1,
    )
    pivot = pivot.rename(columns={"state_abbr": "stabbr"})
    pivot = pivot[~pivot["stabbr"].isin(EXCLUDED_TERRITORIES)]
    return (
        pivot.dropna(subset=["pct_change"])
             .sort_values("pct_change", ascending=False)
             .reset_index(drop=True)
        [["stabbr", "prior", "recent", "pct_change"]]
    )


@st.cache_data(show_spinner=False)
def run_dep_by_state_query(
    cip_patterns: tuple,
    awlevels: tuple,
    year: int,
):
    """DE share by state for a single year.

    Returns DataFrame: stabbr, programs, programs_de_any, pct_de_any.
    """
    cip_patterns = expand_cip_patterns(cip_patterns)
    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM completions_dep LIMIT 1")
    except Exception:
        conn.close()
        return pd.DataFrame()

    params = [year]
    where = [
        "d.year = ?",
        "d.programs > 0",
        "LENGTH(d.cipcode) >= 5",
    ]

    if cip_patterns:
        cip_clauses = []
        for p in cip_patterns:
            cip_clauses.append("d.cipcode LIKE ?" if "%" in p else "d.cipcode = ?")
            params.append(p)
        where.append(f"({' OR '.join(cip_clauses)})")

    if awlevels:
        placeholders = ",".join("?" * len(awlevels))
        where.append(f"d.awlevel IN ({placeholders})")
        params.extend(awlevels)

    excluded = ",".join(f"'{s}'" for s in EXCLUDED_TERRITORIES)
    where.append(f"i.stabbr NOT IN ({excluded})")
    where.append("i.stabbr IS NOT NULL AND i.stabbr != ''")

    sql = f"""
        SELECT
            i.stabbr             AS stabbr,
            SUM(d.programs)      AS programs,
            SUM(d.programs_de_any) AS programs_de_any
        FROM completions_dep d
        INNER JOIN institutions i
          ON d.unitid = i.unitid AND d.year = i.year
        WHERE {' AND '.join(where)}
        GROUP BY i.stabbr
    """
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()

    if df.empty:
        return df
    df["pct_de_any"] = (
        100.0 * df["programs_de_any"] / df["programs"]
    ).round(1).where(df["programs"] > 0)
    return df


@st.cache_data(show_spinner=False)
def run_employment_by_state_query(soc_codes: tuple, year: int):
    """Total employment by state for the given SOC codes in a single year.

    Returns DataFrame: stabbr, tot_emp.
    Uses OES state-level rows (area_type=2). Aggregates across multiple
    occupations; states with suppressed values for a particular SOC simply
    don't contribute.
    """
    if not soc_codes:
        return pd.DataFrame()
    conn = get_conn()
    soc_ph = ",".join("?" * len(soc_codes))
    sql = f"""
        SELECT area_code, SUM(tot_emp) AS tot_emp
        FROM oes_employment
        WHERE year = ?
          AND area_type = 2
          AND occ_code IN ({soc_ph})
          AND tot_emp IS NOT NULL
        GROUP BY area_code
    """
    df = pd.read_sql_query(sql, conn, params=[year] + list(soc_codes))
    conn.close()
    if df.empty:
        return df
    df["stabbr"] = df["area_code"].map(FIPS_TO_STABBR)
    df = df.dropna(subset=["stabbr"])
    df = df[~df["stabbr"].isin(EXCLUDED_TERRITORIES)]
    return df[["stabbr", "tot_emp"]].reset_index(drop=True)


_CORESIGNAL_BASE = "https://api.coresignal.com/cdapi/v2/job_base"

# Generic SOC titles to exclude from Coresignal searches (too broad / noisy)
_GENERIC_SOC_TITLES = {
    "Managers, All Other",
    "Teachers and Instructors, All Other, Except Substitute Teachers",
    "First-Line Supervisors of Office and Administrative Support Workers",
    "Postsecondary Teachers",
    "Education Administrators, Postsecondary",
    "Education Administrators, All Other",
}


def _resolve_coresignal_titles(cip_patterns: tuple, awlevels: tuple) -> list:
    """Map CIP codes to simplified occupation titles for Coresignal searches."""
    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM cip_soc_crosswalk LIMIT 1")
    except Exception:
        conn.close()
        return []

    params = []
    cip_clauses = []
    for p in cip_patterns:
        cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
        params.append(p)

    awlevel_filter = ""
    if awlevels:
        undergrad = any(a in (1, 2, 3, 4, 5, 20, 21) for a in awlevels)
        grad = any(a in (6, 7, 8, 17, 18, 19) for a in awlevels)
        if undergrad and not grad:
            awlevel_filter = "AND awlevel_group = 'undergrad'"
        elif grad and not undergrad:
            awlevel_filter = "AND awlevel_group = 'graduate'"

    sql = f"""
        SELECT DISTINCT soc_title
        FROM cip_soc_crosswalk
        WHERE ({' OR '.join(cip_clauses)}) {awlevel_filter}
        ORDER BY soc_title
    """
    soc_titles = [r[0] for r in conn.execute(sql, params).fetchall()]
    conn.close()

    soc_titles = [t for t in soc_titles if t not in _GENERIC_SOC_TITLES]
    simplified = []
    for t in soc_titles:
        t = t.split(",")[0].strip()
        if t.endswith("s") and not t.endswith("ss"):
            t = t[:-1]
        if t and t not in simplified:
            simplified.append(t)
    return simplified[:3]


@st.cache_data(show_spinner=False, ttl=3600)
def run_coresignal_trend(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Query Coresignal for monthly job posting counts over the last 12 months.

    Uses the x-total-pages response header to get exact totals without
    paginating (1 credit per month per title). Also collects a small sample
    of recent postings for a detail table.

    Returns dict with 'trend_df', 'current_active', and 'search_titles',
    or None if the API key is missing or no results found.
    """
    from datetime import datetime, timedelta
    import calendar

    api_key = st.secrets["coresignal"]["api_key"] if "coresignal" in st.secrets else ""
    if not api_key:
        return None

    cip_patterns = expand_cip_patterns(cip_patterns)
    search_titles = _resolve_coresignal_titles(cip_patterns, awlevels)
    if not search_titles:
        return None

    headers = {"accept": "application/json", "apikey": api_key, "Content-Type": "application/json"}
    session = _requests.Session()
    session.headers.update(headers)

    # Resolve geography to Coresignal-compatible location strings.
    # States: use abbreviations directly (e.g., "TX"). Metro filtering not supported.
    cs_locations = []  # empty = national (no location filter)
    if geo_key == "state" and geo_values:
        cs_locations = list(geo_values)

    def _search_total(body: dict) -> int:
        """Run a search/filter call and return estimated total from headers."""
        try:
            resp = session.post(
                f"{_CORESIGNAL_BASE}/search/filter", json=body, timeout=30
            )
            if resp.status_code == 200:
                total_pages = int(resp.headers.get("x-total-pages", 1))
                items_per_page = int(resp.headers.get("x-items-per-page", 1000))
                return total_pages * items_per_page
        except Exception:
            pass
        return 0

    def _query_postings(extra_filters: dict) -> int:
        """Sum posting counts across all search titles and locations."""
        total = 0
        locs = cs_locations or [None]  # None = no location filter
        for title in search_titles:
            for loc in locs:
                body = {"title": title, "country": "United States", **extra_filters}
                if loc is not None:
                    body["location"] = loc
                total += _search_total(body)
        return total

    # Build list of last 12 months
    today = datetime.now()
    months = []
    for i in range(11, -1, -1):
        dt = today.replace(day=1) - timedelta(days=i * 30)
        dt = dt.replace(day=1)
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        end_day = today.day if (dt.year == today.year and dt.month == today.month) else last_day
        months.append({
            "label": dt.strftime("%Y-%m"),
            "gte": f"{dt.year}-{dt.month:02d}-01 00:00:00",
            "lte": f"{dt.year}-{dt.month:02d}-{end_day:02d} 23:59:59",
        })

    # Query posting counts per month
    trend_rows = []
    for m in months:
        month_total = _query_postings({
            "created_at_gte": m["gte"],
            "created_at_lte": m["lte"],
        })
        trend_rows.append({"month": m["label"], "postings": month_total})

    trend_df = pd.DataFrame(trend_rows)

    if trend_df["postings"].sum() == 0:
        return "empty"

    # Get current active postings count
    current_active = _query_postings({"application_active": True})

    return {
        "trend_df": trend_df,
        "current_active": current_active,
        "search_titles": search_titles,
    }


@st.cache_data(show_spinner=False, ttl=600)
def run_employment_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Query OES employment data for occupations related to selected CIP codes.

    Handles SOC version differences:
      - 2015-2018 OES data uses SOC 2010 codes
      - 2019-2024 OES data uses SOC 2018 codes
      - Some 2019-2020 data uses BLS combined codes (e.g. 15-1256)
        that are remapped to SOC 2018 detail codes
      - CIP-SOC crosswalk maps CIP -> SOC 2018
      - soc_2010_to_2018 table bridges the gap for older years

    Uses awlevel_group filtering:
      - undergrad levels (1-5): include "all" + exclude "graduate"-only
      - graduate levels (6-19): include "all" + "graduate"
      - mixed: include all mappings
    """
    conn = get_conn()

    # Determine award-level group filter
    # undergrad: awlevel 1-5, 20, 21; graduate: awlevel 6+
    UNDERGRAD_LEVELS = {1, 2, 3, 4, 5, 20, 21}
    GRADUATE_LEVELS = {6, 7, 8, 17, 18, 19}
    has_undergrad = bool(set(awlevels) & UNDERGRAD_LEVELS)
    has_graduate = bool(set(awlevels) & GRADUATE_LEVELS)

    if has_undergrad and has_graduate:
        awlevel_filter = ""  # include all mappings
    elif has_graduate:
        awlevel_filter = " AND awlevel_group IN ('all', 'graduate')"
    else:
        # Undergrad only: exclude graduate-only mappings
        awlevel_filter = " AND awlevel_group = 'all'"

    # 1. Find SOC 2018 codes mapped to the selected CIP codes
    if cip_patterns:
        cip_clauses = []
        cip_params = []
        for p in cip_patterns:
            cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
            cip_params.append(p)
        cip_where = " OR ".join(cip_clauses)
        soc_rows = conn.execute(
            f"SELECT DISTINCT soc_code FROM cip_soc_crosswalk WHERE ({cip_where}){awlevel_filter}",
            cip_params,
        ).fetchall()
    else:
        soc_rows = conn.execute(
            f"SELECT DISTINCT soc_code FROM cip_soc_crosswalk WHERE 1=1{awlevel_filter}"
        ).fetchall()

    soc_2018_codes = [r[0] for r in soc_rows]
    if not soc_2018_codes:
        conn.close()
        return pd.DataFrame()

    # 2. Also get SOC 2010 equivalents for querying older data
    soc_ph = ",".join("?" * len(soc_2018_codes))
    soc_2010_rows = conn.execute(
        f"SELECT DISTINCT soc_2010, soc_2018 FROM soc_2010_to_2018 "
        f"WHERE soc_2018 IN ({soc_ph})",
        soc_2018_codes,
    ).fetchall()
    soc_2010_codes = [r[0] for r in soc_2010_rows]
    # For many-to-many mappings, prefer the lowest SOC 2018 code
    # (BLS assigns lower codes to primary occupations, e.g. 15-1252
    # Software Developers before 15-1253 Software QA)
    soc_2010_to_2018_map: dict[str, str] = {}
    for s2010, s2018 in sorted(soc_2010_rows, key=lambda x: x[1]):
        if s2010 not in soc_2010_to_2018_map:
            soc_2010_to_2018_map[s2010] = s2018
    # Override: "All Other" codes (XX-XX99) should map to their "All Other"
    # successor, not inflate a specific occupation (e.g. 15-1199 -> 15-1299)
    for s2010, s2018 in soc_2010_rows:
        if s2010[-2:] == "99" and s2018[-2:] == "99" and s2010[:3] == s2018[:3]:
            soc_2010_to_2018_map[s2010] = s2018

    # 2a. Handle BLS combined codes (2019-2020)
    # BLS sometimes publishes combined codes when detail codes can't be
    # separately disclosed. Include them and remap to detail codes.
    _COMBINED_SOC = {
        "15-1245": {"15-1242", "15-1243"},  # DB Admins + Architects
        "15-1256": {"15-1252", "15-1253"},  # Software Devs + QA
        "15-1257": {"15-1254", "15-1255"},  # Web Devs + Designers
        "15-2098": {"15-2051", "15-2099"},  # Data Scientists + Math Sci Other
    }
    combined_remap = {}  # combined_code -> detail_code to remap to
    for combined, details in _COMBINED_SOC.items():
        overlap = details & set(soc_2018_codes)
        if overlap:
            combined_remap[combined] = sorted(overlap)[0]
    all_soc_2018 = list(set(soc_2018_codes) | set(combined_remap.keys()))

    # 3. Build area filter
    area_where = ""
    area_params = []
    if geo_key == "national":
        area_where = "AND area_type = 1"
    elif geo_key == "state" and geo_values:
        fips_codes = [STABBR_TO_FIPS.get(s, "") for s in geo_values]
        fips_codes = [f for f in fips_codes if f]
        if fips_codes:
            ph = ",".join("?" * len(fips_codes))
            area_where = f"AND area_type = 2 AND area_code IN ({ph})"
            area_params = fips_codes
        else:
            conn.close()
            return pd.DataFrame()
    elif geo_key == "metro" and geo_values:
        # Our geo_values are 5-digit CBSAs; BLS uses 00+CBSA (7-digit)
        bls_codes = ["00" + str(c).zfill(5) for c in geo_values]
        ph = ",".join("?" * len(bls_codes))
        area_where = f"AND area_type = 4 AND area_code IN ({ph})"
        area_params = bls_codes

    # 4. Query: UNION of SOC 2018 data (2019+) and SOC 2010 data (2015-2018)
    # For 2019+ data, use SOC 2018 codes (+ any BLS combined codes)
    soc_ph_2018 = ",".join("?" * len(all_soc_2018))
    params_2018 = all_soc_2018 + area_params

    sql_2018 = f"""
        SELECT year, occ_code, occ_title,
               SUM(tot_emp) AS tot_emp,
               CASE WHEN COUNT(CASE WHEN a_mean IS NOT NULL THEN 1 END) > 0
                    THEN CAST(SUM(CASE WHEN a_mean IS NOT NULL THEN tot_emp * a_mean ELSE 0 END)
                         / NULLIF(SUM(CASE WHEN a_mean IS NOT NULL THEN tot_emp ELSE 0 END), 0) AS INTEGER)
                    ELSE NULL END AS a_mean,
               CASE WHEN COUNT(CASE WHEN a_median IS NOT NULL THEN 1 END) > 0
                    THEN CAST(SUM(CASE WHEN a_median IS NOT NULL THEN tot_emp * a_median ELSE 0 END)
                         / NULLIF(SUM(CASE WHEN a_median IS NOT NULL THEN tot_emp ELSE 0 END), 0) AS INTEGER)
                    ELSE NULL END AS a_median
        FROM oes_employment
        WHERE year >= 2019
          AND occ_code IN ({soc_ph_2018})
          {area_where}
        GROUP BY year, occ_code, occ_title
    """

    # For pre-2019 data (2015-2018), use SOC 2010 codes and map to 2018
    dfs = [pd.read_sql_query(sql_2018, conn, params=params_2018)]

    if soc_2010_codes:
        soc_ph_2010 = ",".join("?" * len(soc_2010_codes))
        params_2010 = soc_2010_codes + area_params

        sql_2010 = f"""
            SELECT year, occ_code, occ_title,
                   SUM(tot_emp) AS tot_emp,
                   CASE WHEN COUNT(CASE WHEN a_mean IS NOT NULL THEN 1 END) > 0
                        THEN CAST(SUM(CASE WHEN a_mean IS NOT NULL THEN tot_emp * a_mean ELSE 0 END)
                             / NULLIF(SUM(CASE WHEN a_mean IS NOT NULL THEN tot_emp ELSE 0 END), 0) AS INTEGER)
                        ELSE NULL END AS a_mean,
                   CASE WHEN COUNT(CASE WHEN a_median IS NOT NULL THEN 1 END) > 0
                        THEN CAST(SUM(CASE WHEN a_median IS NOT NULL THEN tot_emp * a_median ELSE 0 END)
                             / NULLIF(SUM(CASE WHEN a_median IS NOT NULL THEN tot_emp ELSE 0 END), 0) AS INTEGER)
                        ELSE NULL END AS a_median
            FROM oes_employment
            WHERE year < 2019
              AND occ_code IN ({soc_ph_2010})
              {area_where}
            GROUP BY year, occ_code, occ_title
        """

        df_2010 = pd.read_sql_query(sql_2010, conn, params=params_2010)
        # Map SOC 2010 codes to SOC 2018 for consistent time series
        if not df_2010.empty:
            df_2010["occ_code"] = df_2010["occ_code"].map(
                lambda x: soc_2010_to_2018_map.get(x, x)
            )
            # Re-aggregate after remapping (multiple 2010 codes may map to one 2018 code)
            df_2010 = df_2010.groupby(["year", "occ_code"]).agg({
                "occ_title": "first",
                "tot_emp": "sum",
                "a_mean": "first",
                "a_median": "first",
            }).reset_index()
            dfs.append(df_2010)

    conn.close()

    if not dfs or all(d.empty for d in dfs):
        return pd.DataFrame()

    result = pd.concat(dfs, ignore_index=True)

    # Remap BLS combined codes to their detail equivalents
    if combined_remap:
        result["occ_code"] = result["occ_code"].map(
            lambda x: combined_remap.get(x, x)
        )
        # Re-aggregate after remapping (combined code may merge with detail code)
        result = result.groupby(["year", "occ_code"]).agg({
            "occ_title": "first",
            "tot_emp": "sum",
            "a_mean": "first",
            "a_median": "first",
        }).reset_index()

    # Update occ_title: use most recent year's title (most accurate post-reclassification)
    title_source = result.sort_values("year", ascending=False).drop_duplicates("occ_code")
    title_map = title_source.set_index("occ_code")["occ_title"].to_dict()
    result["occ_title"] = result["occ_code"].map(lambda x: title_map.get(x, x))

    return result.sort_values(["occ_code", "year"]).reset_index(drop=True)


# ── Automation risk lookup ───────────────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=3600)
def get_automation_risk(soc_codes: tuple) -> pd.DataFrame:
    """Return LMII automation risk scores for the given SOC codes.

    Source: LMI Institute Automation Exposure Index (2019 OES). Each row
    scores an occupation 1-10, where 1 = least exposed to automation and
    10 = most exposed. SOC 2010 codes are bridged through soc_2010_to_2018
    so older OES years still resolve.

    Returns DataFrame: occ_code, risk_score, composite. Empty when the
    table is absent (e.g. loader hasn't run).
    """
    if not soc_codes:
        return pd.DataFrame(columns=["occ_code", "risk_score", "composite"])

    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM occ_automation_risk LIMIT 1")
    except Exception:
        conn.close()
        return pd.DataFrame(columns=["occ_code", "risk_score", "composite"])

    # Direct match on the provided code, OR bridge through SOC 2010→2018
    # so that older OES rows (which use SOC 2010 codes) still find a score.
    sql = f"""
        WITH wanted(code) AS (
            VALUES {",".join(f"(?)" for _ in soc_codes)}
        )
        SELECT w.code AS occ_code,
               COALESCE(r_direct.risk_score, r_bridged.risk_score) AS risk_score,
               COALESCE(r_direct.composite,  r_bridged.composite)  AS composite
        FROM wanted w
        LEFT JOIN occ_automation_risk r_direct
               ON r_direct.occ_code = w.code
        LEFT JOIN soc_2010_to_2018 x
               ON x.soc_2010 = w.code
        LEFT JOIN occ_automation_risk r_bridged
               ON r_bridged.occ_code = x.soc_2018
    """
    df = pd.read_sql_query(sql, conn, params=list(soc_codes))
    conn.close()

    # If SOC 2010 maps to multiple 2018 codes, the JOIN yields duplicates —
    # keep the highest risk match for the same input SOC (conservative).
    df = (
        df.dropna(subset=["risk_score"])
          .sort_values("risk_score", ascending=False)
          .drop_duplicates(subset=["occ_code"], keep="first")
    )
    return df.reset_index(drop=True)


# ── College Scorecard query ──────────────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=600)
def run_scorecard_query(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Query College Scorecard graduate outcomes for selected filters.

    Matches CIP codes at 4-digit level (XX.XX) since Scorecard data is
    reported at that granularity vs IPEDS 6-digit (XX.XXXX).
    """
    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM college_scorecard LIMIT 1")
    except Exception:
        conn.close()
        return pd.DataFrame()

    params: list = []
    where = ["sc.earn_mdn_4yr IS NOT NULL"]

    # ── CIP filter (truncate 6-digit → 4-digit) ──────────────────────────
    if cip_patterns:
        sc_cip_set: set[str] = set()
        like_clauses: list[str] = []
        for p in cip_patterns:
            if "%" in p:
                # Wildcard: "52.%" or "52.01%" → use LIKE
                like_clauses.append("sc.cipcode LIKE ?")
                params.append(p[:5] if len(p) > 5 else p)
            else:
                # Exact 6-digit: "52.0101" → truncate to "52.01"
                sc_cip_set.add(p[:5])

        cip_parts: list[str] = list(like_clauses)
        if sc_cip_set:
            ph = ",".join("?" * len(sc_cip_set))
            cip_parts.append(f"sc.cipcode IN ({ph})")
            params.extend(sorted(sc_cip_set))
        if cip_parts:
            where.append(f"({' OR '.join(cip_parts)})")

    # ── Award level filter (pre-mapped in table) ─────────────────────────
    if awlevels:
        ph = ",".join("?" * len(awlevels))
        where.append(f"sc.awlevel IN ({ph})")
        params.extend(awlevels)

    # ── Geography filter (join institutions for state/metro) ─────────────
    # Always join institutions to exclude territories (PR, VI, GU, etc.)
    # Use each institution's most recent IPEDS year so closed/merged schools
    # still match rather than requiring all to exist in the global max year.
    join_inst = (
        "INNER JOIN ("
        "  SELECT unitid, MAX(year) AS max_year"
        "  FROM institutions GROUP BY unitid"
        ") imax ON sc.unitid = imax.unitid "
        "INNER JOIN institutions i "
        "ON i.unitid = imax.unitid AND i.year = imax.max_year"
    )
    territory_ph = ",".join("?" * len(EXCLUDED_TERRITORIES))
    where.append(f"i.stabbr NOT IN ({territory_ph})")
    params.extend(sorted(EXCLUDED_TERRITORIES))

    if geo_key == "state" and geo_values:
        ph = ",".join("?" * len(geo_values))
        where.append(f"i.stabbr IN ({ph})")
        params.extend(geo_values)
    elif geo_key == "metro" and geo_values:
        ph = ",".join("?" * len(geo_values))
        where.append(f"i.cbsa IN ({ph})")
        params.extend(geo_values)

    where_sql = " AND ".join(where)

    sql = f"""
        SELECT
            sc.unitid,
            sc.instnm,
            i.city || ', ' || i.stabbr AS city,
            CASE i.control
                WHEN 1 THEN 'Public'
                WHEN 2 THEN 'Private'
                WHEN 3 THEN 'For-Profit'
                ELSE 'Unknown'
            END AS control_name,
            sc.earn_mdn_4yr,
            sc.debt_all_stgp_eval_mdn,
            sc.debt_to_earnings,
            sc.distance
        FROM college_scorecard sc
        {join_inst}
        WHERE {where_sql}
        ORDER BY sc.debt_to_earnings ASC
    """

    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


@st.cache_data(show_spinner=False, ttl=3600)
def get_projection_coverage():
    """Load metro projection coverage tracking data."""
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql("SELECT * FROM metro_projection_coverage ORDER BY state_abbr, cbsa_name", conn)
        return df
    except Exception:
        return None
    finally:
        conn.close()


@st.cache_data(show_spinner=False, ttl=600)
def get_employment_projections(
    soc_codes: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Fetch projected growth (CAGR) for given SOC codes and geography.

    Returns DataFrame with columns: occ_code, cagr, base_year, proj_year, source.
    Uses best available geo match: metro > state > national.
    """
    conn = get_conn()

    # Check if projections table exists
    try:
        conn.execute("SELECT 1 FROM employment_projections LIMIT 1")
    except Exception:
        conn.close()
        return pd.DataFrame()

    if not soc_codes:
        conn.close()
        return pd.DataFrame()

    soc_ph = ",".join("?" * len(soc_codes))

    # Build geo filter based on geo_key
    if geo_key == "metro" and geo_values:
        # Try metro first, then fall back to state/national
        cbsa_ph = ",".join("?" * len(geo_values))
        # For metro, we might have multiple CBSAs — average across them
        sql = f"""
            SELECT occ_code, AVG(cagr) as cagr,
                   MIN(base_year) as base_year, MAX(proj_year) as proj_year,
                   'metro' as geo_level
            FROM employment_projections
            WHERE occ_code IN ({soc_ph})
              AND geo_level = 'metro'
              AND geo_code IN ({cbsa_ph})
            GROUP BY occ_code
        """
        params = list(soc_codes) + list(geo_values)
        df = pd.read_sql_query(sql, conn, params=params)

    elif geo_key == "state" and geo_values:
        fips_codes = [STABBR_TO_FIPS.get(s, "") for s in geo_values]
        fips_codes = [f for f in fips_codes if f]
        if fips_codes:
            fips_ph = ",".join("?" * len(fips_codes))
            sql = f"""
                SELECT occ_code, AVG(cagr) as cagr,
                       MIN(base_year) as base_year, MAX(proj_year) as proj_year,
                       'state' as geo_level
                FROM employment_projections
                WHERE occ_code IN ({soc_ph})
                  AND geo_level = 'state'
                  AND geo_code IN ({fips_ph})
                GROUP BY occ_code
            """
            params = list(soc_codes) + fips_codes
            df = pd.read_sql_query(sql, conn, params=params)
        else:
            df = pd.DataFrame()
    else:
        # National
        sql = f"""
            SELECT occ_code, cagr, base_year, proj_year, 'national' as geo_level
            FROM employment_projections
            WHERE occ_code IN ({soc_ph})
              AND geo_level = 'national'
        """
        df = pd.read_sql_query(sql, conn, params=list(soc_codes))

    # If metro/state returned nothing, fall back to national
    if df.empty and geo_key != "national":
        sql = f"""
            SELECT occ_code, cagr, base_year, proj_year, 'national' as geo_level
            FROM employment_projections
            WHERE occ_code IN ({soc_ph})
              AND geo_level = 'national'
        """
        df = pd.read_sql_query(sql, conn, params=list(soc_codes))

    conn.close()
    return df


@st.cache_data(show_spinner=False)
def resolve_soc_codes_for_cips(cip_patterns: tuple, awlevels: tuple) -> tuple:
    """SOC 2018 codes that map to the given CIPs in `cip_soc_crosswalk`.

    Honors the awlevel_group filter (undergrad-only excludes graduate-only
    mappings, etc.) the same way `run_employment_query` does. Returns a
    tuple suitable for caching keys.
    """
    if not cip_patterns:
        return tuple()
    UNDERGRAD = {1, 2, 3, 4, 5, 20, 21}
    GRADUATE = {6, 7, 8, 17, 18, 19}
    has_ug = bool(set(awlevels) & UNDERGRAD)
    has_gr = bool(set(awlevels) & GRADUATE)
    if has_ug and has_gr:
        awf = ""
    elif has_gr:
        awf = " AND awlevel_group IN ('all', 'graduate')"
    else:
        awf = " AND awlevel_group = 'all'"

    conn = get_conn()
    cip_clauses, cip_params = [], []
    for p in cip_patterns:
        cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
        cip_params.append(p)
    cip_where = " OR ".join(cip_clauses)
    rows = conn.execute(
        f"SELECT DISTINCT soc_code FROM cip_soc_crosswalk "
        f"WHERE ({cip_where}){awf}",
        cip_params,
    ).fetchall()
    conn.close()
    return tuple(r[0] for r in rows)


@st.cache_data(show_spinner=False, ttl=600)
def get_emp_proj_cagr(
    cip_patterns: tuple,
    awlevels: tuple,
    geo_key: str,
    geo_values: tuple,
) -> float | None:
    """Return weighted-average employment projection CAGR for SOC codes
    mapped to the given CIP codes/geography, or None if unavailable."""
    if not cip_patterns:
        return None  # "All CIPs" → no meaningful SOC mapping

    conn = get_conn()

    # Award-level group filter (mirrors run_employment_query)
    UNDERGRAD = {1, 2, 3, 4, 5, 20, 21}
    GRADUATE = {6, 7, 8, 17, 18, 19}
    has_ug = bool(set(awlevels) & UNDERGRAD)
    has_gr = bool(set(awlevels) & GRADUATE)
    if has_ug and has_gr:
        awf = ""
    elif has_gr:
        awf = " AND awlevel_group IN ('all', 'graduate')"
    else:
        awf = " AND awlevel_group = 'all'"

    # SOC 2018 codes from CIP-SOC crosswalk
    cip_clauses, cip_params = [], []
    for p in cip_patterns:
        cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
        cip_params.append(p)
    cip_where = " OR ".join(cip_clauses)
    soc_rows = conn.execute(
        f"SELECT DISTINCT soc_code FROM cip_soc_crosswalk WHERE ({cip_where}){awf}",
        cip_params,
    ).fetchall()
    soc_codes = tuple(r[0] for r in soc_rows)
    if not soc_codes:
        conn.close()
        return None

    # Latest-year employment weights
    soc_ph = ",".join("?" * len(soc_codes))
    area_where, area_params = "", []
    if geo_key == "national":
        area_where = "AND area_type = 1"
    elif geo_key == "state" and geo_values:
        fips = [STABBR_TO_FIPS.get(s, "") for s in geo_values]
        fips = [f for f in fips if f]
        if fips:
            area_where = f"AND area_type = 2 AND area_code IN ({','.join('?' * len(fips))})"
            area_params = fips
    elif geo_key == "metro" and geo_values:
        bls = ["00" + str(c).zfill(5) for c in geo_values]
        area_where = f"AND area_type = 4 AND area_code IN ({','.join('?' * len(bls))})"
        area_params = bls

    latest_emp = pd.read_sql_query(
        f"""SELECT occ_code, SUM(tot_emp) AS tot_emp
            FROM oes_employment
            WHERE year = (SELECT MAX(year) FROM oes_employment)
              AND occ_code IN ({soc_ph}) {area_where}
            GROUP BY occ_code""",
        conn,
        params=list(soc_codes) + area_params,
    )
    conn.close()

    if latest_emp.empty:
        return None

    # Employment projections
    df_proj = get_employment_projections(
        soc_codes=soc_codes, geo_key=geo_key, geo_values=tuple(geo_values),
    )
    if df_proj.empty or "cagr" not in df_proj.columns:
        return None

    merged = df_proj.merge(latest_emp, on="occ_code", how="inner")
    merged = merged.dropna(subset=["cagr", "tot_emp"])
    if merged.empty or merged["tot_emp"].sum() <= 0:
        return None

    return float(
        (merged["cagr"] * merged["tot_emp"]).sum() / merged["tot_emp"].sum()
    )


@st.cache_data(show_spinner=False, ttl=600)
def run_google_trends_query(
    cip_patterns: tuple,
    geo_key: str,
    geo_values: tuple,
):
    """Query Google Trends interest data for selected CIP codes and geography.

    Returns dict with:
      - 'time_series': DataFrame(date, interest) — national monthly averages
      - 'geo_interest': float or None — interest index for selected geography
      - 'search_terms': list[str] — search terms used
      - 'state_data': DataFrame(state_abbr, interest) — all states
      - 'top_metros': DataFrame(cbsa_code, cbsa_name, interest) — top 15 metros
      - 'per_cip_time': DataFrame(date, cipcode, search_term, interest) — per-CIP
    Or None if no data is available.
    """
    conn = get_conn()
    try:
        conn.execute("SELECT 1 FROM google_trends_time LIMIT 1")
    except Exception:
        conn.close()
        return None

    if not cip_patterns:
        conn.close()
        return None

    # Build CIP filter
    cip_clauses, cip_params = [], []
    for p in cip_patterns:
        cip_clauses.append("cipcode LIKE ?" if "%" in p else "cipcode = ?")
        cip_params.append(p)
    cip_where = " OR ".join(cip_clauses)

    # 1. National time series (aggregate to monthly, average across matched CIPs)
    time_sql = f"""
        SELECT SUBSTR(date, 1, 7) AS month, AVG(interest) AS interest
        FROM google_trends_time
        WHERE ({cip_where})
        GROUP BY month
        ORDER BY month
    """
    df_time = pd.read_sql_query(time_sql, conn, params=cip_params)
    if not df_time.empty:
        df_time["date"] = pd.to_datetime(df_time["month"] + "-01")
        df_time = df_time[["date", "interest"]]

    # 2. Geographic interest
    geo_interest = None
    if geo_key == "national":
        if not df_time.empty:
            geo_interest = df_time.tail(12)["interest"].mean()
    elif geo_key == "state" and geo_values:
        st_ph = ",".join("?" * len(geo_values))
        row = conn.execute(
            f"SELECT AVG(interest) FROM google_trends_state "
            f"WHERE ({cip_where}) AND state_abbr IN ({st_ph})",
            cip_params + list(geo_values),
        ).fetchone()
        if row and row[0] is not None:
            geo_interest = row[0]
    elif geo_key == "metro" and geo_values:
        cbsa_ph = ",".join("?" * len(geo_values))
        try:
            row = conn.execute(
                f"""SELECT AVG(cbsa_interest) FROM (
                    SELECT gt.cipcode,
                        SUM(gt.interest * w.weight) / SUM(w.weight) AS cbsa_interest
                    FROM google_trends_dma gt
                    JOIN dma_cbsa_weights w ON gt.dma_code = w.dma_code
                    WHERE ({cip_where}) AND w.cbsa_code IN ({cbsa_ph})
                    GROUP BY gt.cipcode
                )""",
                cip_params + list(geo_values),
            ).fetchone()
            if row and row[0] is not None:
                geo_interest = row[0]
        except Exception:
            pass  # dma_cbsa_weights table may not exist

    # 3. Search terms used
    terms = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT search_term FROM google_trends_time WHERE ({cip_where})",
            cip_params,
        ).fetchall()
    ]

    # 4. State-level interest (all states, for choropleth map)
    state_sql = f"""
        SELECT state_abbr, AVG(interest) AS interest
        FROM google_trends_state
        WHERE ({cip_where})
        GROUP BY state_abbr
        ORDER BY interest DESC
    """
    df_states = pd.read_sql_query(state_sql, conn, params=cip_params)

    # 5. Top metro markets (DMA interest weighted into CBSAs, top 15)
    try:
        metro_sql = f"""
            SELECT w.cbsa_code, w.cbsa_name,
                   SUM(gt.interest * w.weight) / SUM(w.weight) AS interest
            FROM google_trends_dma gt
            JOIN dma_cbsa_weights w ON gt.dma_code = w.dma_code
            WHERE ({cip_where}) AND gt.interest > 0
            GROUP BY w.cbsa_code
            ORDER BY interest DESC
            LIMIT 15
        """
        df_metros = pd.read_sql_query(metro_sql, conn, params=cip_params)
    except Exception:
        df_metros = pd.DataFrame(columns=["cbsa_code", "cbsa_name", "interest"])

    # 6. Per-CIP time series (for multi-program comparison)
    per_cip_sql = f"""
        SELECT SUBSTR(date, 1, 7) AS month, cipcode, search_term,
               AVG(interest) AS interest
        FROM google_trends_time
        WHERE ({cip_where})
        GROUP BY cipcode, month
        ORDER BY cipcode, month
    """
    df_per_cip = pd.read_sql_query(per_cip_sql, conn, params=cip_params)
    if not df_per_cip.empty:
        df_per_cip["date"] = pd.to_datetime(df_per_cip["month"] + "-01")
        df_per_cip = df_per_cip[["date", "cipcode", "search_term", "interest"]]

    if df_time.empty:
        conn.close()
        return None

    # ── Volume calibration ────────────────────────────────────────────────
    # Check if search_volume_calibration table exists and load ratios
    has_volume = False
    volume_series = None
    per_cip_volume = None
    geo_volume = None
    est_monthly_vol = None
    state_volume_data = None
    metro_volume_data = None
    try:
        cal_sql = f"""
            SELECT sv.cipcode, sv.est_monthly_vol, sv.anchor_ratio
            FROM search_volume_calibration sv
            WHERE sv.cipcode IN (
                SELECT DISTINCT cipcode FROM google_trends_time WHERE ({cip_where})
            )
        """
        df_cal = pd.read_sql_query(cal_sql, conn, params=cip_params)
        if not df_cal.empty:
            has_volume = True
            # Weighted average est_monthly_vol across selected CIPs
            est_monthly_vol = df_cal["est_monthly_vol"].mean()

            # For time series volume: scale the aggregate interest index
            # Volume = (interest / interest_at_anchor_month) * est_monthly_vol
            anchor_month_interest = None
            march_row = df_time[df_time["date"] == pd.Timestamp("2025-03-01")]
            if not march_row.empty:
                anchor_month_interest = march_row["interest"].iloc[0]
            if anchor_month_interest and anchor_month_interest > 0:
                volume_series = df_time.copy()
                volume_series["volume"] = (
                    volume_series["interest"] / anchor_month_interest * est_monthly_vol
                ).round(0).astype(int)
            else:
                # Fallback: use the series max as reference
                max_interest = df_time["interest"].max()
                if max_interest > 0:
                    volume_series = df_time.copy()
                    # Scale so that peak = est_monthly_vol * (100 / avg_peak_ratio)
                    volume_series["volume"] = (
                        volume_series["interest"] / max_interest
                        * est_monthly_vol * (100 / df_time["interest"].mean())
                    ).round(0).astype(int)

            # Per-CIP volume series
            if not df_per_cip.empty:
                per_cip_volume = df_per_cip.merge(
                    df_cal[["cipcode", "est_monthly_vol"]],
                    on="cipcode", how="left",
                )
                # For each CIP, find its March 2025 interest as anchor
                per_cip_volume["volume"] = 0
                for cip in per_cip_volume["cipcode"].unique():
                    mask = per_cip_volume["cipcode"] == cip
                    cip_data = per_cip_volume.loc[mask]
                    march = cip_data[cip_data["date"] == pd.Timestamp("2025-03-01")]
                    cip_vol = cip_data["est_monthly_vol"].iloc[0]
                    if not march.empty and march["interest"].iloc[0] > 0:
                        anchor_int = march["interest"].iloc[0]
                    else:
                        anchor_int = cip_data["interest"].max()
                    if anchor_int > 0 and pd.notna(cip_vol):
                        per_cip_volume.loc[mask, "volume"] = (
                            cip_data["interest"] / anchor_int * cip_vol
                        ).round(0).astype(int)

            # Geographic volume: scale geo_interest by volume
            if geo_interest is not None and anchor_month_interest and anchor_month_interest > 0:
                geo_volume = round(geo_interest / anchor_month_interest * est_monthly_vol)

            # State-level volume: weight interest by state population
            # volume_share = (interest × population) / Σ(interest × population)
            # state_volume = volume_share × national_monthly_vol
            state_volume_data = None
            if not df_states.empty:
                try:
                    df_state_pop = pd.read_sql_query(
                        "SELECT state_abbr, population FROM state_populations",
                        conn,
                    )
                    sv = df_states.merge(df_state_pop, on="state_abbr", how="inner")
                    sv["weighted"] = sv["interest"] * sv["population"]
                    total_weighted = sv["weighted"].sum()
                    if total_weighted > 0:
                        sv["volume"] = (
                            sv["weighted"] / total_weighted * est_monthly_vol
                        ).round(0).astype(int)
                        state_volume_data = sv[["state_abbr", "interest", "volume"]]
                except Exception:
                    pass

            # Metro-level volume: compute for ALL metros (not just top 15)
            # so volume shares are accurate, then take top 15 by volume
            metro_volume_data = None
            try:
                all_metro_sql = f"""
                    SELECT w.cbsa_code, w.cbsa_name,
                           SUM(gt.interest * w.weight) / SUM(w.weight) AS interest
                    FROM google_trends_dma gt
                    JOIN dma_cbsa_weights w ON gt.dma_code = w.dma_code
                    WHERE ({cip_where}) AND gt.interest > 0
                    GROUP BY w.cbsa_code
                """
                df_all_metros = pd.read_sql_query(
                    all_metro_sql, conn, params=cip_params,
                )
                if not df_all_metros.empty:
                    df_cbsa_pop = pd.read_sql_query(
                        "SELECT cbsa_code, population FROM cbsa_populations",
                        conn,
                    )
                    mv = df_all_metros.merge(
                        df_cbsa_pop, on="cbsa_code", how="inner",
                    )
                    mv["weighted"] = mv["interest"] * mv["population"]
                    total_weighted = mv["weighted"].sum()
                    if total_weighted > 0:
                        mv["volume"] = (
                            mv["weighted"] / total_weighted * est_monthly_vol
                        ).round(0).astype(int)
                        # Top 15 by volume for display
                        metro_volume_data = (
                            mv[["cbsa_code", "cbsa_name", "interest", "volume"]]
                            .sort_values("volume", ascending=False)
                            .head(15)
                        )
            except Exception:
                pass
    except Exception as _vol_err:
        import traceback as _tb
        _vol_debug = _tb.format_exc()
        # Store debug info so the UI can display it
        has_volume = False
        _vol_error_msg = f"{type(_vol_err).__name__}: {_vol_err}\n{_vol_debug}"
    else:
        _vol_error_msg = None

    conn.close()

    return {
        "time_series": df_time,
        "geo_interest": round(geo_interest, 1) if geo_interest is not None else None,
        "search_terms": terms,
        "state_data": df_states,
        "top_metros": df_metros,
        "per_cip_time": df_per_cip,
        "has_volume": has_volume,
        "volume_series": volume_series,
        "per_cip_volume": per_cip_volume,
        "geo_volume": geo_volume,
        "est_monthly_vol": est_monthly_vol,
        "state_volume_data": state_volume_data if has_volume else None,
        "metro_volume_data": metro_volume_data if has_volume else None,
        "_vol_error_msg": _vol_error_msg,
    }



# ── Excel export helper ──────────────────────────────────────────────────────

_VI_ORANGE = "F26822"
_VI_DARK = "333333"
_HEADER_FILL = PatternFill(start_color=_VI_ORANGE, end_color=_VI_ORANGE, fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10, color=_VI_DARK)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)
_PCT_FMT = '0.0%'
_MONEY_FMT = '$#,##0'
_NUM_FMT = '#,##0'


def _style_sheet(ws, df, pct_cols=None, money_cols=None, num_cols=None):
    """Apply VI-branded formatting to a worksheet built from a DataFrame."""
    pct_cols = set(pct_cols or [])
    money_cols = set(money_cols or [])
    num_cols = set(num_cols or [])

    # Header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            col_name = df.columns[col_idx - 1]
            if col_name in pct_cols:
                cell.number_format = _PCT_FMT
            elif col_name in money_cols:
                cell.number_format = _MONEY_FMT
            elif col_name in num_cols:
                cell.number_format = _NUM_FMT

    # Auto-width columns (capped at 40)
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(ws.max_row + 1, 102)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Freeze header row
    ws.freeze_panes = "A2"


def build_export_workbook(sheets_data):
    """Build an openpyxl Workbook from a list of (sheet_name, df, fmt_opts) tuples.

    fmt_opts is a dict with optional keys: pct_cols, money_cols, num_cols.
    Returns bytes of the .xlsx file.
    """
    from openpyxl import Workbook

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    for sheet_name, df, fmt_opts in sheets_data:
        if df is None or df.empty:
            continue
        # Truncate sheet name to 31 chars (Excel limit)
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(title=safe_name)

        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Convert numpy types to native Python
                if isinstance(value, (np.integer,)):
                    value = int(value)
                elif isinstance(value, (np.floating,)):
                    value = float(value) if pd.notna(value) else None
                elif pd.isna(value) if isinstance(value, float) else False:
                    value = None
                cell.value = value

        _style_sheet(
            ws, df,
            pct_cols=fmt_opts.get("pct_cols"),
            money_cols=fmt_opts.get("money_cols"),
            num_cols=fmt_opts.get("num_cols"),
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── CSV (zip) export ──────────────────────────────────────────────────────────
def build_csv_zip(sheets_data) -> bytes:
    """Bundle each section into a single CSV and zip them together.

    sheets_data: list of (name, df, fmt_opts) tuples (fmt_opts unused here).
    """
    import zipfile

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, df, _ in sheets_data:
            if df is None or getattr(df, "empty", True):
                continue
            safe = (
                name.lower()
                .replace(" / ", "_")
                .replace(" ", "_")
                .replace("/", "_")
            )
            zf.writestr(f"{safe}.csv", df.to_csv(index=False))
    buf.seek(0)
    return buf.getvalue()


# ── PDF report ────────────────────────────────────────────────────────────────
def build_pdf_report(sheets_data, *, report_meta: dict) -> bytes:
    """Build a stylized, VI-branded PDF report.

    sheets_data: list of (name, df, fmt_opts) tuples.
    report_meta: dict with keys 'title', 'subtitle' (optional),
                 'geo_label', 'cip_display', 'level_str'.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageBreak, PageTemplate, Paragraph, Spacer,
        Table, TableStyle,
    )

    VI_ORANGE = colors.HexColor("#f26822")
    VI_BLUE = colors.HexColor("#0f86c1")
    INK = colors.HexColor("#1F2937")
    MUTED = colors.HexColor("#6B7280")
    HAIRLINE = colors.HexColor("#E5E7EB")
    SOFT_BG = colors.HexColor("#F9FAFB")
    SOFT_ACCENT = colors.HexColor("#FFF5EE")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "vi_title", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=22, leading=28,
        textColor=INK, spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "vi_subtitle", parent=styles["Normal"],
        fontName="Helvetica", fontSize=11, leading=15,
        textColor=MUTED, spaceAfter=18,
    )
    section_style = ParagraphStyle(
        "vi_section", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=14, leading=18,
        textColor=VI_ORANGE, spaceBefore=14, spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "vi_body", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9, leading=12,
        textColor=INK,
    )
    meta_label_style = ParagraphStyle(
        "vi_meta_label", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8, leading=11,
        textColor=MUTED,
    )
    meta_value_style = ParagraphStyle(
        "vi_meta_value", parent=styles["Normal"],
        fontName="Helvetica", fontSize=10, leading=13,
        textColor=INK, spaceAfter=4,
    )
    footer_style = ParagraphStyle(
        "vi_footer", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7, leading=10,
        textColor=MUTED, alignment=1,
    )

    def _fmt(value, col_name, fmt_opts):
        if value is None:
            return "—"
        try:
            if pd.isna(value):
                return "—"
        except (TypeError, ValueError):
            pass
        money_cols = set(fmt_opts.get("money_cols") or [])
        pct_cols = set(fmt_opts.get("pct_cols") or [])
        num_cols = set(fmt_opts.get("num_cols") or [])
        if col_name in money_cols and isinstance(value, (int, float, np.integer, np.floating)):
            return f"${value:,.0f}"
        if col_name in pct_cols and isinstance(value, (int, float, np.integer, np.floating)):
            v = value if abs(value) > 1 else value * 100
            return f"{v:+.1f}%"
        if col_name in num_cols and isinstance(value, (int, float, np.integer, np.floating)):
            return f"{value:,.0f}"
        if isinstance(value, (np.integer,)):
            return f"{int(value):,}"
        if isinstance(value, (np.floating, float)):
            return f"{float(value):,.2f}"
        return str(value)

    page_w, page_h = landscape(LETTER)
    margin_l = margin_r = 0.5 * inch
    margin_t = 1.0 * inch
    margin_b = 0.7 * inch
    frame = Frame(
        margin_l, margin_b,
        page_w - margin_l - margin_r,
        page_h - margin_t - margin_b,
        id="content",
    )

    def _draw_header_footer(canvas, doc):
        canvas.saveState()
        # Top brand bar
        canvas.setFillColor(VI_ORANGE)
        canvas.rect(0, page_h - 0.4 * inch, page_w, 0.4 * inch, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 11)
        canvas.drawString(margin_l, page_h - 0.27 * inch, "VALIDATED INSIGHTS")
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(
            page_w - margin_r, page_h - 0.27 * inch,
            "IPEDS Completions Explorer",
        )
        # Footer
        canvas.setFillColor(MUTED)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawString(
            margin_l, 0.4 * inch,
            f"Generated {pd.Timestamp.now().strftime('%B %d, %Y')}",
        )
        canvas.drawRightString(
            page_w - margin_r, 0.4 * inch,
            f"Page {doc.page}",
        )
        canvas.setStrokeColor(HAIRLINE)
        canvas.setLineWidth(0.5)
        canvas.line(margin_l, 0.55 * inch, page_w - margin_r, 0.55 * inch)
        canvas.restoreState()

    buf = BytesIO()
    doc = BaseDocTemplate(
        buf,
        pagesize=landscape(LETTER),
        leftMargin=margin_l, rightMargin=margin_r,
        topMargin=margin_t, bottomMargin=margin_b,
        title=report_meta.get("title", "IPEDS Report"),
        author="Validated Insights",
    )
    doc.addPageTemplates([
        PageTemplate(id="main", frames=[frame], onPage=_draw_header_footer),
    ])

    story = []

    # ── Cover/header block ──
    story.append(Paragraph(report_meta.get("title", "IPEDS Report"), title_style))
    if report_meta.get("subtitle"):
        story.append(Paragraph(report_meta["subtitle"], subtitle_style))

    # Filter summary card
    meta_rows = [
        [Paragraph("GEOGRAPHY", meta_label_style), Paragraph("PROGRAM", meta_label_style), Paragraph("AWARD LEVEL", meta_label_style)],
        [
            Paragraph(report_meta.get("geo_label", "—"), meta_value_style),
            Paragraph(report_meta.get("cip_display", "—"), meta_value_style),
            Paragraph(report_meta.get("level_str", "—"), meta_value_style),
        ],
    ]
    avail_w = page_w - margin_l - margin_r
    meta_tbl = Table(meta_rows, colWidths=[avail_w / 3.0] * 3)
    meta_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), SOFT_ACCENT),
        ("BOX", (0, 0), (-1, -1), 0.6, HAIRLINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LINEBEFORE", (1, 0), (1, -1), 0.5, HAIRLINE),
        ("LINEBEFORE", (2, 0), (2, -1), 0.5, HAIRLINE),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 12))

    # ── Per-section content ──
    for idx, (name, df, fmt_opts) in enumerate(sheets_data):
        if df is None or getattr(df, "empty", True):
            continue
        if idx > 0:
            story.append(PageBreak())
        story.append(Paragraph(name, section_style))

        # Cap rows per section to keep PDF size sane
        ROW_CAP = 60
        truncated = len(df) > ROW_CAP
        view = df.head(ROW_CAP).copy()

        # Build header + body
        cols = list(view.columns)
        header = [Paragraph(f"<b>{c}</b>", body_style) for c in cols]
        body = []
        for _, row in view.iterrows():
            body.append([
                Paragraph(_fmt(row[c], c, fmt_opts), body_style)
                for c in cols
            ])

        n_cols = len(cols)
        col_widths = [avail_w / n_cols] * n_cols
        # Give the leading text column more weight if it looks like a label column
        if cols and cols[0] in ("Institution", "Occupation", "City", "Field", "State"):
            extra = avail_w * 0.18
            col_widths = [avail_w * 0.30] + [(avail_w - avail_w * 0.30) / max(1, n_cols - 1)] * (n_cols - 1) if n_cols > 1 else col_widths

        tbl = Table([header] + body, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), VI_ORANGE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, SOFT_BG]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, VI_ORANGE),
            ("GRID", (0, 1), (-1, -1), 0.25, HAIRLINE),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)
        if truncated:
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                f"<i>Showing first {ROW_CAP:,} of {len(df):,} rows. "
                f"Export as Excel or CSV for the full dataset.</i>",
                ParagraphStyle("trunc", parent=body_style, textColor=MUTED, fontSize=8),
            ))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ── Rankings page ─────────────────────────────────────────────────────────────


@st.cache_data(show_spinner=False, ttl=600)
def _rank_programs_cached(
    geo_key: str,
    geo_values: tuple,
    awlevels: tuple,
    cip_family: str | None,
    min_completions: int,
):
    conn = get_conn()
    try:
        return _rankings.score_programs_for_geo(
            conn=conn,
            geo_key=geo_key,
            geo_values=geo_values,
            awlevels=awlevels,
            stabbr_to_fips=STABBR_TO_FIPS,
            cip_family=cip_family,
            min_completions=min_completions,
        )
    finally:
        conn.close()


@st.cache_data(show_spinner=False, ttl=600)
def _rank_markets_cached(
    cipcode: str,
    awlevels: tuple,
    market_grain: str,
    top_n: int | None,
    min_emp: int,
):
    conn = get_conn()
    try:
        return _rankings.score_markets_for_program(
            conn=conn,
            cipcode=cipcode,
            awlevels=awlevels,
            market_grain=market_grain,
            stabbr_to_fips=STABBR_TO_FIPS,
            fips_to_stabbr=FIPS_TO_STABBR,
            excluded_states=EXCLUDED_TERRITORIES,
            top_n=top_n,
            min_emp=min_emp,
        )
    finally:
        conn.close()


def _grade_style(val) -> str:
    """Pandas Styler callback — paint the Grade cell background by letter."""
    if not val or pd.isna(val):
        return ""
    color = _rankings.GRADE_COLORS.get(val, "#666666")
    return f"background-color: {color}; color: #ffffff; font-weight: 600;"


def _render_rankings_page():
    """Demand rankings — Top Programs by Market | Top Markets by Program."""
    st.title("Program/Market Ranker")
    st.caption(
        "Scored, letter-graded leaderboards. The composite score is "
        "anchored on **completions volume + long-term completions trend "
        "(10-year CAGR) + post-COVID completions trend (since AY 2020-21)** "
        "— those three drive the bulk of the rank. Secondary signals: "
        "related-occupation employment (volume / 5-yr growth / BLS "
        "projection), employment-weighted median wage, Scorecard graduate "
        "earnings, Google Trends search interest **and** its multi-year "
        "trend, average program size (completions per reporting "
        "institution), and LMI Institute automation-resilience (11 − risk). "
        "Components are z-scored within the cohort, weighted, and scaled "
        "to a 0-100 composite; letter grades reflect the **percentile "
        "within the cohort**, not an absolute floor. "
        "| Sources: NCES IPEDS Completions, BLS OES, BLS Employment "
        "Projections, U.S. Dept of Education College Scorecard, Google "
        "Trends, LMI Institute Automation Exposure Index."
    )

    rank_mode = st.radio(
        "Ranking type",
        options=["Top Programs by Market", "Top Markets by Program"],
        horizontal=True,
        key="rank_mode",
    )

    # ── Award levels (shared by both modes) ─────────────────────────────────
    level_label_to_code = {v: k for k, v in AWARD_LEVELS.items()}
    level_options = list(AWARD_LEVELS.values())
    if rank_mode == "Top Programs by Market":
        # ── Inputs ──────────────────────────────────────────────────────────
        f1, f2, f3 = st.columns([1.2, 1.0, 1.0])
        with f1:
            geo_type = st.radio(
                "Geography",
                options=["National", "By State", "By Metro Area"],
                key="rank_geo_type",
                horizontal=True,
            )
        with f2:
            chosen_level = st.selectbox(
                "Award level",
                options=level_options,
                index=level_options.index("Bachelor's degree"),
                key="rank_awlevel",
            )
            awlevels = (level_label_to_code[chosen_level],)
        with f3:
            cip_taxonomy_2d = [
                ("(All families)", None),
                ("01 — Agriculture", "01"),
                ("03 — Natural Resources", "03"),
                ("04 — Architecture", "04"),
                ("05 — Area, Ethnic, Cultural Studies", "05"),
                ("09 — Communication, Journalism", "09"),
                ("10 — Communications Technologies", "10"),
                ("11 — Computer & Information Sciences", "11"),
                ("12 — Personal & Culinary Services", "12"),
                ("13 — Education", "13"),
                ("14 — Engineering", "14"),
                ("15 — Engineering Technologies", "15"),
                ("16 — Foreign Languages", "16"),
                ("19 — Family & Consumer Sciences", "19"),
                ("22 — Legal Professions", "22"),
                ("23 — English Language & Literature", "23"),
                ("24 — Liberal Arts & Sciences", "24"),
                ("25 — Library Science", "25"),
                ("26 — Biological & Biomedical Sciences", "26"),
                ("27 — Mathematics & Statistics", "27"),
                ("29 — Military Technologies", "29"),
                ("30 — Multi/Interdisciplinary Studies", "30"),
                ("31 — Parks, Recreation, Fitness", "31"),
                ("38 — Philosophy & Religious Studies", "38"),
                ("39 — Theology & Religious Vocations", "39"),
                ("40 — Physical Sciences", "40"),
                ("41 — Science Technologies", "41"),
                ("42 — Psychology", "42"),
                ("43 — Homeland Security & Law Enforcement", "43"),
                ("44 — Public Administration & Social Service", "44"),
                ("45 — Social Sciences", "45"),
                ("46 — Construction Trades", "46"),
                ("47 — Mechanic & Repair Technologies", "47"),
                ("48 — Precision Production", "48"),
                ("49 — Transportation", "49"),
                ("50 — Visual & Performing Arts", "50"),
                ("51 — Health Professions", "51"),
                ("52 — Business, Management, Marketing", "52"),
                ("54 — History", "54"),
            ]
            fam_label = st.selectbox(
                "CIP family (2-digit)",
                options=[lbl for lbl, _ in cip_taxonomy_2d],
                index=0,
                key="rank_cip_family",
                help=(
                    "Limit the ranking to a single CIP family so the cohort "
                    "is apples-to-apples. Leave on '(All families)' to rank "
                    "every program at this award level together."
                ),
            )
            cip_family = next(code for lbl, code in cip_taxonomy_2d if lbl == fam_label)

        # Geography multi-select
        g_geo_values: tuple = ()
        g_geo_key = "national"
        if geo_type == "By State":
            sts = load_states()
            sel = st.multiselect(
                "State(s)", options=sts, key="rank_state_pick",
                placeholder="Select one or more states…",
            )
            if not sel:
                st.info("Select at least one state to rank programs.")
                return
            g_geo_values = tuple(sel)
            g_geo_key = "state"
        elif geo_type == "By Metro Area":
            cbsas = load_cbsas()
            label_to_code = {label: code for code, label in cbsas}
            sel_labels = st.multiselect(
                "Metro area(s)",
                options=[label for _, label in cbsas],
                key="rank_metro_pick",
                placeholder="Search metro areas…",
            )
            if not sel_labels:
                st.info("Select at least one metro area to rank programs.")
                return
            g_geo_values = tuple(label_to_code[l] for l in sel_labels)
            g_geo_key = "metro"

        min_comp = st.slider(
            "Minimum completions threshold (latest year)",
            min_value=0, max_value=500, value=25, step=5,
            key="rank_min_comp",
            help=(
                "Programs below this threshold of total completions in the "
                "selected geography are excluded from the cohort. Higher = "
                "fewer niche / data-sparse programs."
            ),
        )

        if not st.button("Run rankings", type="primary", key="rank_run_btn"):
            st.caption(
                "Choose your geography, award level, and an optional CIP family, "
                "then hit **Run rankings**."
            )
            return

        with st.spinner("Scoring programs…"):
            df = _rank_programs_cached(
                geo_key=g_geo_key,
                geo_values=g_geo_values,
                awlevels=awlevels,
                cip_family=cip_family,
                min_completions=min_comp,
            )
        if df is None or df.empty:
            st.warning(
                "No programs met the criteria. Try lowering the completions "
                "threshold or selecting a different CIP family."
            )
            return

        _render_rank_table_programs(df)
    else:
        # ── Top Markets by Program ───────────────────────────────────────────
        cip_options = load_cip_options()  # [("51.3801", "51.3801 - Registered Nursing"), ...]
        # Hide catchall residual CIPs ("…, Other") from the picker — see
        # rankings.score_programs_for_geo for the same exclusion. The
        # pattern ", Other" with optional extra whitespace catches both the
        # canonical NCES titles and a handful of double-space variants.
        import re
        _other_re = re.compile(r",\s*Other\s*$", re.IGNORECASE)
        cip_options = [(c, lbl) for c, lbl in cip_options if not _other_re.search(lbl)]
        f1, f2 = st.columns([2.0, 1.0])
        with f1:
            cip_label = st.selectbox(
                "Program (CIP code)",
                options=[label for _, label in cip_options],
                key="rank_cip_pick",
            )
            cipcode = next(code for code, label in cip_options if label == cip_label)
        with f2:
            chosen_level = st.selectbox(
                "Award level",
                options=level_options,
                index=level_options.index("Bachelor's degree"),
                key="rank_market_awlevel",
            )
            awlevels = (level_label_to_code[chosen_level],)
        f3, f4 = st.columns([1.0, 1.0])
        with f3:
            grain = st.radio(
                "Market grain",
                options=["State", "Metro"],
                horizontal=True,
                key="rank_market_grain",
            )
        with f4:
            top_n = st.slider(
                "Show top N markets",
                min_value=10, max_value=200, value=25, step=5,
                key="rank_topn",
            )

        if not st.button("Run rankings", type="primary", key="rank_market_run"):
            st.caption(
                "Pick a CIP, award level, and grain (state vs. metro), then "
                "hit **Run rankings**."
            )
            return

        market_grain = "state" if grain == "State" else "metro"
        with st.spinner("Scoring markets…"):
            df = _rank_markets_cached(
                cipcode=cipcode,
                awlevels=awlevels,
                market_grain=market_grain,
                top_n=top_n,
                min_emp=100 if market_grain == "metro" else 0,
            )
        if df is None or df.empty:
            st.warning(
                "No markets met the criteria. Try the other grain (state vs. "
                "metro) or a different program."
            )
            return

        _render_rank_table_markets(df, market_grain=market_grain, program_label=cip_label)


def _fmt_pct(v):
    if pd.isna(v):
        return "—"
    return f"{v * 100:+.1f}%"


def _fmt_money(v):
    if pd.isna(v):
        return "—"
    return f"${v:,.0f}"


def _fmt_int(v):
    if pd.isna(v):
        return "—"
    return f"{v:,.0f}"


def _fmt_num(v, places=1):
    if pd.isna(v):
        return "—"
    return f"{v:.{places}f}"


def _render_rank_table_programs(df: pd.DataFrame):
    """Format and display the per-program ranking dataframe.

    We pre-format every numeric column to a string because Streamlit's
    st.dataframe ignores pandas Styler.format(). Cell *styling* (the
    coloured Grade pill) still goes through Styler.map(), which IS
    honored.
    """
    display = pd.DataFrame({
        "Rank":            df["rank"],
        "Grade":           df["grade"],
        "Score":           df["composite"].apply(lambda v: _fmt_num(v, 1)),
        "CIP":             df["cipcode"],
        "Program":         df["cipdesc"],
        "Completions":     df["completions"].apply(_fmt_int),
        "Comp 10y CAGR":   df["completions_long_trend"].apply(_fmt_pct)
                            if "completions_long_trend" in df.columns else "—",
        "Comp Post-COVID": df["completions_pc_trend"].apply(_fmt_pct)
                            if "completions_pc_trend" in df.columns else "—",
        "Avg Program Size": df["avg_program_size"].apply(lambda v: _fmt_num(v, 0))
                            if "avg_program_size" in df.columns else "—",
        "Related Emp.":    df["emp_volume"].apply(_fmt_int),
        "Past 5y CAGR":    df["emp_growth"].apply(_fmt_pct),
        "Projected CAGR":  df["emp_projection"].apply(_fmt_pct),
        "Median Wage":     df["wage"].apply(_fmt_money),
        "Grad Earnings":   df["earnings"].apply(_fmt_money)
                            if "earnings" in df.columns else "—",
        "Search Interest": df["search_interest"].apply(lambda v: _fmt_num(v, 1))
                            if "search_interest" in df.columns else "—",
        "Search Trend":    df["search_trend"].apply(_fmt_pct)
                            if "search_trend" in df.columns else "—",
        "Auto Risk":       df["automation_risk"].apply(lambda v: _fmt_num(v, 1))
                            if "automation_risk" in df.columns else "—",
    })

    st.subheader(f"Ranked programs · {len(display):,} in cohort")
    st.caption(
        "Grades are by percentile within this cohort: top 5% = A+, 5-15% = A, "
        "15-25% = A-, 25-35% = B+, 35-45% = B, 45-55% = B-, 55-65% = C+, "
        "65-75% = C, 75-85% = C-, 85-95% = D, bottom 5% = F. The "
        "**Score** column drives the rank."
    )
    styled = display.style.map(_grade_style, subset=["Grade"])
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        height=min(700, 60 + 36 * min(len(display), 18)),
    )

    st.download_button(
        "Download CSV",
        data=df.to_csv(index=False).encode("utf-8"),
        file_name="program_rankings.csv",
        mime="text/csv",
        key="rank_dl_programs",
    )

    # Component-weight footnote
    with st.expander("How the score is calculated"):
        st.markdown(
            "Component weights for the composite score (each component is "
            "z-scored within the cohort and the weighted sum is shifted/scaled "
            "to a roughly 0-100 composite):\n\n"
            + "\n".join(
                f"- **{k}**: {int(v*100)}%"
                for k, v in _rankings.PROGRAM_WEIGHTS.items()
            )
            + "\n\nEmployment volume is log-transformed before z-scoring "
            "because counts span several orders of magnitude. Automation "
            "resilience = 11 − the LMI Institute automation-exposure score "
            "(so a higher value = harder to automate)."
        )


def _render_rank_table_markets(df: pd.DataFrame, market_grain: str, program_label: str):
    """Format and display the per-market ranking dataframe."""
    display = pd.DataFrame({
        "Rank":            df["rank"],
        "Grade":           df["grade"],
        "Score":           df["composite"].apply(lambda v: _fmt_num(v, 1)),
        "Market":          df["area_label"],
        "Local Grads":     df["completions"].apply(_fmt_int)
                            if "completions" in df.columns else "—",
        "Grad 10y CAGR":   df["completions_long_trend"].apply(_fmt_pct)
                            if "completions_long_trend" in df.columns else "—",
        "Grad Post-COVID": df["completions_pc_trend"].apply(_fmt_pct)
                            if "completions_pc_trend" in df.columns else "—",
        "Avg Program Size": df["avg_program_size"].apply(lambda v: _fmt_num(v, 0))
                            if "avg_program_size" in df.columns else "—",
        "Related Emp.":    df["emp_volume"].apply(_fmt_int),
        "LQ":              df["location_quotient"].apply(lambda v: _fmt_num(v, 2))
                            if "location_quotient" in df.columns else "—",
        "Past 5y CAGR":    df["emp_growth"].apply(_fmt_pct),
        "Projected CAGR":  df["emp_projection"].apply(_fmt_pct),
        "Median Wage":     df["wage"].apply(_fmt_money),
        "Search Interest": df["search_interest"].apply(lambda v: _fmt_num(v, 1))
                            if "search_interest" in df.columns else "—",
        "Search Trend":    df["search_trend"].apply(_fmt_pct)
                            if "search_trend" in df.columns else "—",
        "Emp / Grad+1":    df["competition_inv"].apply(lambda v: _fmt_num(v, 1))
                            if "competition_inv" in df.columns else "—",
    })

    st.subheader(
        f"Ranked {market_grain}s · {program_label}"
    )
    st.caption(
        "**LQ** = location quotient (1.0 = same employment share as the "
        "nation; >1.0 = over-indexed). **Emp / Grad+1** is a crude opportunity "
        "score — high values mean lots of jobs relative to local grad supply. "
        "Grades reflect the percentile within this market grain (every state, "
        "or every metro that has at least 100 related jobs)."
    )
    styled = display.style.map(_grade_style, subset=["Grade"])
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        height=min(700, 60 + 36 * min(len(display), 18)),
    )

    st.download_button(
        "Download CSV",
        data=df.to_csv(index=False).encode("utf-8"),
        file_name=f"market_rankings_{market_grain}.csv",
        mime="text/csv",
        key="rank_dl_markets",
    )

    with st.expander("How the score is calculated"):
        st.markdown(
            "Component weights for the market composite score:\n\n"
            + "\n".join(
                f"- **{k}**: {int(v*100)}%"
                for k, v in _rankings.MARKET_WEIGHTS.items()
            )
            + "\n\nEmployment volume is log-transformed before z-scoring. "
            "Location quotient compares the market's share of program-related "
            "employment to the national share. Metro-level Google Trends data "
            "isn't published, so search interest is omitted in metro mode."
        )


# ── UI helpers ────────────────────────────────────────────────────────────────
def vi_kpi_card(
    *,
    label: str,
    value: str,
    icon: str,
    sublabel: str | None = None,
    trend: tuple[str, str] | None = None,
    sentiment: str | None = None,
) -> None:
    """Render a single VI-branded KPI card.

    Designed to be called inside an ``st.columns`` cell. Pure HTML — no
    nested Streamlit containers — so the layout is predictable and the
    card styles live entirely in the global ``.vi-kpi`` stylesheet.

    Args:
        label: Short label under the big number, e.g. ``"Long-Term CAGR"``.
        value: Pre-formatted big-number string, e.g. ``"3,450"`` or ``"+5.2%"``.
        icon: Material Symbols Rounded ligature name, e.g. ``"trending_up"``.
        sublabel: Optional muted line below the label (date range, basis).
        trend: ``(direction, text)`` where direction is ``"up"`` | ``"down"``
            | ``"flat"``. Renders a small colored pill in the top-right.
        sentiment: ``"positive"`` | ``"negative"`` | ``None`` — colors the
            big value semantic green/red (used for signed-% metrics).
    """
    trend_html = ""
    if trend is not None:
        direction, text = trend
        arrow = {"up": "▲", "down": "▼", "flat": "→"}.get(
            direction, ""
        )
        trend_html = (
            f'<span class="vi-trend vi-trend-{direction}">'
            f"{arrow}&nbsp;{text}</span>"
        )
    val_cls = ""
    if sentiment == "positive":
        val_cls = " vi-val-pos"
    elif sentiment == "negative":
        val_cls = " vi-val-neg"
    sub_html = (
        f'<div class="vi-kpi-sub">{sublabel}</div>' if sublabel else ""
    )
    st.markdown(
        f'<div class="vi-kpi">'
        f'<div class="vi-kpi-top">'
        f'<span class="vi-kpi-icon material-symbols-rounded">{icon}</span>'
        f"{trend_html}"
        f"</div>"
        f'<div class="vi-kpi-val{val_cls}">{value}</div>'
        f'<div class="vi-kpi-lbl">{label}</div>'
        f"{sub_html}"
        f"</div>",
        unsafe_allow_html=True,
    )


@contextmanager
def vi_card(
    title: str | None = None,
    *,
    subtitle: str | None = None,
    icon: str | None = None,
):
    """Render a VI-branded card with an optional header (icon + title + sub).

    Use as a context manager and place any Streamlit widgets inside::

        with vi_card("Completions trend", subtitle="Bachelor's · National",
                     icon="show_chart"):
            st.plotly_chart(fig, use_container_width=True)

    The card surface (white background, hairline border, soft shadow,
    rounded corners) is applied via the global ``.vi-card`` CSS rules,
    scoped to ``st.container(border=True)`` via the marker class trick.
    """
    container = st.container(border=True)
    with container:
        st.html(
            '<span class="vi-card-marker" style="display:none"></span>'
        )
        if title or icon:
            icon_html = (
                f'<span class="vi-card-icon material-symbols-rounded">'
                f"{icon}</span>"
                if icon
                else ""
            )
            title_html = (
                f'<div class="vi-card-title">{title}</div>' if title else ""
            )
            sub_html = (
                f'<div class="vi-card-sub">{subtitle}</div>' if subtitle else ""
            )
            st.markdown(
                f'<div class="vi-card-head">'
                f'<div class="vi-card-head-left">{icon_html}'
                f'<div class="vi-card-head-text">{title_html}{sub_html}'
                f"</div></div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        yield container


# ── App ───────────────────────────────────────────────────────────────────────
def main():
    # One-time DB prep
    ensure_cbsa_index()
    ensure_award_levels()

    # ── Global styles ─────────────────────────────────────────────────────────
    # Use st.html (not st.markdown) so the <style> block isn't subject to
    # Streamlit's markdown HTML sanitizer — which in recent versions strips
    # <link> tags and can leak trailing <style> contents as visible text.
    # @import inside <style> handles webfont loading without a separate
    # <link> element.
    st.html(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');

        /* === VI Brand Palette (Pass 1) === */
        :root {
            --vi-orange: #F26822;       /* Primary accent, headlines, CTAs */
            --vi-orange-soft: #FFF5EE;  /* Tint for hovers and badges */
            --vi-orange-deep: #D44E0F;  /* Hover / pressed state */
            --vi-blue: #0F86C1;         /* Links, secondary accents */
            --vi-ink: #333333;          /* VI Gray 1 — dark emphasis */
            --vi-muted: #666666;        /* VI Gray 2 — body / subheads */
            --vi-gray-3: #999999;       /* VI Gray 3 — fine details, disabled */
            --vi-gray-4: #D9D9D9;       /* VI Gray 4 — alt rows, soft borders */
            --vi-hairline: #ECECEC;     /* Thin dividers — lighter than gray-4 */
            --vi-soft-bg: #F9FAFB;      /* Subtle panels */
        }
        html, body, [class*="css"], .stApp, .stMarkdown, .stTextInput,
        .stSelectbox, .stMultiSelect, .stRadio, .stCheckbox, .stMetric,
        .stSidebar, .stButton, .stCaption, .stExpander, .stDataFrame,
        button, input, select, textarea {
            font-family: 'Montserrat', Arial, sans-serif !important;
        }
        .stApp { background-color: #FAFBFC; }
        body { color: var(--vi-muted); }

        /* === Headings (Pass 1) ===
           Per VI brand: H1/H2 are bold 700 VI orange. H3 anchors subsections
           in gray-1 with the orange accent bar (handled below). */
        h1, .stTitle {
            font-family: 'Montserrat', Arial, sans-serif !important;
            color: var(--vi-orange) !important;
            font-weight: 700 !important;
            letter-spacing: -0.015em;
            line-height: 1.15 !important;
        }
        h2 {
            font-family: 'Montserrat', Arial, sans-serif !important;
            color: var(--vi-orange) !important;
            font-weight: 700 !important;
            letter-spacing: -0.01em;
        }
        h3, h4, h5, h6 {
            font-family: 'Montserrat', Arial, sans-serif !important;
            color: var(--vi-ink) !important;
            font-weight: 700 !important;
            letter-spacing: -0.005em;
        }
        /* Sidebar step labels in orange to mirror the brand */
        section[data-testid="stSidebar"] h1,
        section[data-testid="stSidebar"] h2,
        section[data-testid="stSidebar"] h3 {
            font-family: 'Montserrat', Arial, sans-serif !important;
            color: var(--vi-orange) !important;
            font-weight: 700 !important;
        }

        /* === Page header treatment (Pass 2) ===
           Make the main page H1 confident; treat the caption directly
           beneath it as a subtitle with tighter spacing. */
        [data-testid="stMain"] [data-testid="stHeading"] h1 {
            font-size: 2.4rem !important;
            margin-bottom: 0.35rem !important;
        }
        [data-testid="stMain"] [data-testid="stHeading"]:has(h1) +
            [data-testid="stCaptionContainer"],
        [data-testid="stMain"] [data-testid="stHeading"]:has(h1) + .stCaption {
            font-size: 0.95rem !important;
            line-height: 1.5 !important;
            color: var(--vi-muted) !important;
            margin: 0 0 1.6rem 0 !important;
            display: block;
        }

        /* === Section subheaders with left orange accent (Pass 4) ===
           Excludes the sidebar so step-numbered controls stay clean. */
        section.main [data-testid="stHeading"] h2,
        [data-testid="stMain"] [data-testid="stHeading"] h2,
        section.main [data-testid="stHeading"] h3,
        [data-testid="stMain"] [data-testid="stHeading"] h3 {
            position: relative;
            padding-left: 14px;
            margin-top: 18px;
        }
        section.main [data-testid="stHeading"] h2::before,
        [data-testid="stMain"] [data-testid="stHeading"] h2::before,
        section.main [data-testid="stHeading"] h3::before,
        [data-testid="stMain"] [data-testid="stHeading"] h3::before {
            content: "";
            position: absolute; left: 0; top: 6px; bottom: 6px;
            width: 4px; border-radius: 2px;
            background: var(--vi-orange);
        }

        /* === KPI tiles (Pass 2) ===
           Thin orange accent stripe on top, more padding, larger value,
           clearer label hierarchy. */
        [data-testid="stMetric"] {
            background: white;
            border: 1px solid var(--vi-hairline);
            border-radius: 12px;
            padding: 22px 20px 16px 20px;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
            transition: box-shadow 120ms ease, transform 120ms ease;
            position: relative;
            overflow: hidden;
        }
        [data-testid="stMetric"]::before {
            content: "";
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 3px;
            background: var(--vi-orange);
        }
        [data-testid="stMetric"]:hover {
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.08);
            transform: translateY(-1px);
        }
        [data-testid="stMetricLabel"] {
            font-family: 'Montserrat', Arial, sans-serif !important;
            color: var(--vi-muted) !important;
            font-weight: 600 !important;
            font-size: 0.7rem !important;
            text-transform: uppercase;
            letter-spacing: 0.06em;
        }
        [data-testid="stMetricValue"] {
            font-family: 'Montserrat', Arial, sans-serif !important;
            color: var(--vi-ink) !important;
            font-weight: 700 !important;
            font-size: 1.95rem !important;
            line-height: 1.1 !important;
            margin-top: 4px;
        }
        [data-testid="stMetricDelta"] {
            font-family: 'Montserrat', Arial, sans-serif !important;
            font-weight: 600 !important;
            font-size: 0.78rem !important;
        }

        /* === Dataframes (Pass 3) ===
           Modern Streamlit renders cells to canvas, so the visible header
           row is the main hookpoint. We brand the header row orange and
           round the outer container; alternating row colors aren't
           possible on canvas-rendered grids. */
        [data-testid="stDataFrame"] {
            border-radius: 12px;
            overflow: hidden;
            border: 1px solid var(--vi-hairline);
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
        }
        /* Header row (rendered as DOM elements above the canvas body) */
        [data-testid="stDataFrame"] [role="columnheader"],
        [data-testid="stDataFrame"] [data-testid="stDataFrameHeaderCell"],
        [data-testid="stDataFrame"] .header,
        [data-testid="stDataFrame"] thead th {
            background: var(--vi-orange) !important;
            color: #ffffff !important;
            font-weight: 700 !important;
            border-bottom: 1px solid var(--vi-orange-deep) !important;
        }
        [data-testid="stDataFrame"] [role="columnheader"] *,
        [data-testid="stDataFrame"] [data-testid="stDataFrameHeaderCell"] *,
        [data-testid="stDataFrame"] thead th * {
            color: #ffffff !important;
            fill: #ffffff !important;
        }

        /* Sidebar surface — orange step labels are handled above */
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #FAFBFC 100%);
            border-right: 1px solid var(--vi-hairline);
        }
        section[data-testid="stSidebar"] hr { border-color: var(--vi-hairline); }

        /* Buttons */
        .stButton > button {
            border-radius: 8px !important;
            font-weight: 600 !important;
            transition: all 120ms ease;
        }
        .stButton > button[kind="primary"] {
            background: var(--vi-orange) !important;
            border-color: var(--vi-orange) !important;
            color: white !important;
            box-shadow: 0 1px 2px rgba(242, 104, 34, 0.25);
        }
        .stButton > button[kind="primary"]:hover {
            background: var(--vi-orange-deep) !important;
            border-color: var(--vi-orange-deep) !important;
            box-shadow: 0 3px 10px rgba(242, 104, 34, 0.30);
            transform: translateY(-1px);
        }
        .stDownloadButton > button {
            border-radius: 8px !important;
            font-weight: 600 !important;
            border: 1px solid var(--vi-orange) !important;
            color: var(--vi-orange) !important;
            background: white !important;
        }
        .stDownloadButton > button:hover {
            background: var(--vi-orange-soft) !important;
        }

        /* Top-level Export button — ghost-orange.
           Uses :has() to target only the column that contains our
           vi-export-trigger-anchor marker (placed via st.html before
           the button), so other secondary buttons keep their default style. */
        div[data-testid="stColumn"]:has(.vi-export-trigger-anchor)
            [data-testid="stButton"] > button {
            background: rgba(242, 104, 34, 0.08) !important;
            border: 1.5px solid var(--vi-orange) !important;
            color: var(--vi-orange) !important;
            font-weight: 600 !important;
            box-shadow: none !important;
        }
        div[data-testid="stColumn"]:has(.vi-export-trigger-anchor)
            [data-testid="stButton"] > button:hover {
            background: rgba(242, 104, 34, 0.16) !important;
            border-color: var(--vi-orange-deep) !important;
            color: var(--vi-orange-deep) !important;
            transform: translateY(-1px);
        }
        /* Force the Material icon glyph to inherit the orange color
           (Streamlit hard-codes a default color on icon spans). */
        div[data-testid="stColumn"]:has(.vi-export-trigger-anchor)
            [data-testid="stButton"] [data-testid="stIconMaterial"],
        div[data-testid="stColumn"]:has(.vi-export-trigger-anchor)
            [data-testid="stButton"] [data-testid="stIconMaterial"] * {
            color: var(--vi-orange) !important;
            fill: var(--vi-orange) !important;
        }
        /* Hide the empty marker span (visual only) */
        .vi-export-trigger-anchor { display: none; }

        /* Top export bar */
        .vi-export-bar {
            background: white;
            border: 1px solid var(--vi-hairline);
            border-radius: 12px;
            padding: 14px 16px;
            margin: 4px 0 22px 0;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
        }
        .vi-export-bar [data-baseweb="select"] {
            border-radius: 8px !important;
        }

        /* Map caption above choropleth */
        .vi-map-caption {
            color: var(--vi-muted);
            font-size: 0.78rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin: 18px 0 -4px 0;
        }

        /* Dividers — softer */
        hr { border-color: var(--vi-hairline) !important; opacity: 0.7; }

        /* Caption text */
        .stCaption, [data-testid="stCaptionContainer"] {
            color: var(--vi-muted) !important;
            line-height: 1.55 !important;
        }

        /* Methodology callout (Pass 4) — long st.caption blurbs that
           live directly under a section subheader get a subtle soft-bg
           pill with a left orange accent. Targets the caption that is
           the immediate sibling of an h2 or h3 heading. */
        [data-testid="stMain"] [data-testid="stHeading"]:has(h2) +
            [data-testid="stCaptionContainer"],
        [data-testid="stMain"] [data-testid="stHeading"]:has(h3) +
            [data-testid="stCaptionContainer"] {
            background: var(--vi-soft-bg);
            border-left: 3px solid var(--vi-orange);
            border-radius: 0 8px 8px 0;
            padding: 10px 14px;
            margin: 6px 0 18px 0 !important;
            font-size: 0.86rem !important;
        }

        /* Pass 4 — rhythm: a bit more breathing room between sections */
        [data-testid="stMain"] hr { margin: 28px 0 !important; }

        /* Tighten title spacing */
        .block-container { padding-top: 2.2rem !important; }

        /* Tighten sidebar header — pull logo up under the close button */
        section[data-testid="stSidebar"] .block-container,
        [data-testid="stSidebarUserContent"] {
            padding-top: 0.5rem !important;
        }
        section[data-testid="stSidebar"] [data-testid="stImage"] {
            margin-top: 0 !important;
            margin-bottom: 0 !important;
        }
        section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
            margin-top: 0.1rem !important;
        }

        /* === KPI cards (Explore page) =========================== */
        @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@24,500,0,0&display=block');

        .vi-kpi {
            background: #ffffff;
            border: 1.5px solid var(--vi-hairline);
            border-radius: 16px;
            padding: 1.2rem 1.35rem 1.3rem 1.35rem;
            box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04);
            display: flex;
            flex-direction: column;
            gap: 0.45rem;
            min-height: 168px;
            font-family: 'Montserrat', sans-serif;
            transition: border-color 0.18s ease, box-shadow 0.18s ease,
                        transform 0.18s ease;
        }
        .vi-kpi:hover {
            border-color: var(--vi-orange);
            box-shadow: 0 12px 28px rgba(242, 104, 34, 0.12);
            transform: translateY(-2px);
        }
        .vi-kpi-top {
            display: flex; align-items: center; justify-content: space-between;
            gap: 0.5rem;
        }
        .vi-kpi-icon {
            background: var(--vi-orange);
            color: #ffffff;
            width: 44px; height: 44px;
            border-radius: 50%;
            display: inline-flex; align-items: center; justify-content: center;
            font-size: 24px;
            box-shadow: 0 5px 12px rgba(242, 104, 34, 0.28);
            font-variation-settings: 'FILL' 0, 'wght' 500;
            user-select: none;
        }
        .vi-trend {
            font-family: 'Montserrat', sans-serif;
            font-weight: 600; font-size: 0.78rem;
            padding: 0.24rem 0.6rem;
            border-radius: 999px;
            white-space: nowrap;
            line-height: 1;
        }
        .vi-trend-up   { color: #15803d; background: #DCFCE7; }
        .vi-trend-down { color: #b91c1c; background: #FEE2E2; }
        .vi-trend-flat { color: #4B5563; background: #F3F4F6; }
        .vi-kpi-val {
            font-family: 'Montserrat', sans-serif;
            font-weight: 800;
            font-size: 2.15rem;
            color: var(--vi-ink);
            letter-spacing: -0.02em;
            line-height: 1.05;
            margin-top: 0.4rem;
        }
        .vi-val-pos { color: #15803d; }
        .vi-val-neg { color: #b91c1c; }
        .vi-kpi-lbl {
            font-family: 'Montserrat', sans-serif;
            font-weight: 600;
            font-size: 0.95rem;
            color: var(--vi-muted);
            line-height: 1.25;
            margin-top: 0.15rem;
        }
        .vi-kpi-sub {
            font-family: 'Montserrat', sans-serif;
            font-weight: 400;
            font-size: 0.8rem;
            color: var(--vi-gray-3);
            margin-top: 0.1rem;
        }

        /* === Card surface (charts, maps, ranking) ================ */
        /* Streamlit's st.container(border=True) gives us a
           stVerticalBlockBorderWrapper; the marker scopes the styling
           to containers opened by vi_card(). */
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.vi-card-marker) {
            background: #ffffff !important;
            border: 1.5px solid var(--vi-hairline) !important;
            border-radius: 16px !important;
            box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04) !important;
            margin-bottom: 1.1rem !important;
            transition: border-color 0.18s ease, box-shadow 0.18s ease;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.vi-card-marker):hover {
            border-color: #E5E7EB !important;
            box-shadow: 0 10px 28px rgba(16, 24, 40, 0.06) !important;
        }

        .vi-card-head {
            display: flex; align-items: center; justify-content: space-between;
            gap: 1rem;
            padding-bottom: 0.85rem;
            border-bottom: 1px solid var(--vi-hairline);
            margin-bottom: 0.9rem;
        }
        .vi-card-head-left {
            display: flex; align-items: center; gap: 0.85rem; min-width: 0;
        }
        .vi-card-head-text { min-width: 0; }
        .vi-card-icon {
            background: var(--vi-orange-soft);
            color: var(--vi-orange);
            width: 40px; height: 40px;
            border-radius: 10px;
            display: inline-flex; align-items: center; justify-content: center;
            font-size: 22px;
            font-variation-settings: 'FILL' 0, 'wght' 500;
            user-select: none; flex: 0 0 40px;
        }
        .vi-card-title {
            font-family: 'Montserrat', sans-serif;
            font-weight: 700;
            font-size: 1.05rem;
            color: var(--vi-ink);
            line-height: 1.2;
            white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
        }
        .vi-card-sub {
            font-family: 'Montserrat', sans-serif;
            font-weight: 400;
            font-size: 0.82rem;
            color: var(--vi-gray-3);
            margin-top: 0.15rem;
            white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
        }
        </style>
        """
    )

    # ── Landing page (post-login picker) ──────────────────────────────────────
    if not st.session_state.get("nav_choice"):
        st.html(
            """
            <style>
            @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');

            section[data-testid="stSidebar"] { display: none !important; }
            div[data-testid="collapsedControl"] { display: none !important; }
            button[data-testid="baseButton-headerNoPadding"] { display: none !important; }

            .vi-landing-title {
                font-family: 'Montserrat', sans-serif;
                font-size: 2.9rem;
                font-weight: 800;
                color: #F26822;
                text-align: center;
                line-height: 1.1;
                letter-spacing: -0.02em;
                margin: 0 0 0.5rem 0;
            }
            .vi-landing-sub {
                font-family: 'Montserrat', sans-serif;
                font-size: 1.05rem;
                font-weight: 400;
                color: #666666;
                text-align: center;
                margin: 0 0 2.75rem 0;
            }

            /* Modern card buttons: white surface, icon badge on top, label
               below. Vertical stack guarantees the icon and text never
               overlap, regardless of widths. */
            div[data-testid="stButton"] button {
                height: 196px;
                display: flex !important;
                flex-direction: column !important;
                align-items: center !important;
                justify-content: center !important;
                gap: 1.15rem !important;
                background: #ffffff !important;
                color: #333333 !important;
                border: 1.5px solid #ECECEC !important;
                border-radius: 20px !important;
                font-family: 'Montserrat', sans-serif !important;
                font-size: 1.55rem !important;
                font-weight: 700 !important;
                letter-spacing: -0.01em !important;
                box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04) !important;
                transition: border-color 0.18s ease, box-shadow 0.18s ease,
                            transform 0.18s ease, color 0.18s ease !important;
            }
            div[data-testid="stButton"] button p { margin: 0 !important; }

            /* Stack icon above label: the inner span that holds both the
               icon wrapper and the text markdown is flex-row by default. */
            div[data-testid="stButton"] button span:has(> div[data-testid="stMarkdownContainer"]) {
                flex-direction: column !important;
                align-items: center !important;
                gap: 1.15rem !important;
            }

            /* The span wrapping the material glyph becomes the orange circle */
            div[data-testid="stButton"] button span:has(> [data-testid="stIconMaterial"]) {
                width: 74px !important;
                height: 74px !important;
                flex: 0 0 74px !important;
                border-radius: 50% !important;
                background: #F26822 !important;
                display: inline-flex !important;
                align-items: center !important;
                justify-content: center !important;
                box-shadow: 0 8px 18px rgba(242, 104, 34, 0.28) !important;
                transition: transform 0.18s ease, box-shadow 0.18s ease !important;
            }
            div[data-testid="stButton"] button [data-testid="stIconMaterial"],
            div[data-testid="stButton"] button span[class*="icon"] {
                color: #ffffff !important;
                fill: #ffffff !important;
                background: transparent !important;
                font-size: 2.4rem !important;
                width: auto !important;
                height: auto !important;
                line-height: 1 !important;
                margin: 0 !important;
            }

            div[data-testid="stButton"] button:hover {
                border-color: #F26822 !important;
                color: #F26822 !important;
                box-shadow: 0 14px 34px rgba(242, 104, 34, 0.18) !important;
                transform: translateY(-4px) !important;
                background: #ffffff !important;
            }
            div[data-testid="stButton"] button:hover span:has(> [data-testid="stIconMaterial"]) {
                transform: scale(1.06) !important;
                box-shadow: 0 10px 22px rgba(242, 104, 34, 0.38) !important;
            }
            div[data-testid="stButton"] button:active { transform: translateY(-1px) !important; }
            div[data-testid="stButton"] button:focus {
                outline: none !important;
                box-shadow: 0 0 0 3px rgba(242, 104, 34, 0.22) !important;
            }

            .vi-card-cap {
                font-family: 'Montserrat', sans-serif;
                font-size: 0.92rem;
                font-weight: 400;
                color: #999999;
                text-align: center;
                margin: 0.85rem 0 0 0;
            }
            </style>
            """
        )
        for _ in range(3):
            st.write("")
        st.markdown(
            '<h1 class="vi-landing-title">VI Data Explorer</h1>'
            '<p class="vi-landing-sub">Choose a workspace to get started</p>',
            unsafe_allow_html=True,
        )
        _, c1, _gap, c2, _ = st.columns([1, 3, 0.4, 3, 1])
        with c1:
            if st.button(
                "Explore",
                icon=":material/explore:",
                use_container_width=True,
                key="land_explore",
            ):
                st.session_state["view_mode"] = "Explore"
                st.session_state["nav_choice"] = "Explore"
                st.rerun()
            st.markdown(
                '<p class="vi-card-cap">Browse completions, trends &amp; geography</p>',
                unsafe_allow_html=True,
            )
        with c2:
            if st.button(
                "Rank",
                icon=":material/leaderboard:",
                use_container_width=True,
                key="land_rank",
            ):
                st.session_state["view_mode"] = "Rank"
                st.session_state["nav_choice"] = "Rank"
                st.rerun()
            st.markdown(
                '<p class="vi-card-cap">Rank programs &amp; compare competitors</p>',
                unsafe_allow_html=True,
            )
        return

    # ── Sidebar ───────────────────────────────────────────────────────────────
    # ── Preset program definitions ──────────────────────────────────────────
    # Each preset: display name -> dict with "cips" (list of 6-digit codes)
    # and "level" (label matching AWARD_LEVELS values or AGGREGATE_LEVELS keys)
    PROGRAM_PRESETS = {
        "MBA": {
            "cips": ["52.0101", "52.0201", "52.1301"],
            "level": "Master's degree",
        },
    }

    _windows = get_data_windows()
    with st.sidebar:
        st.image("vi-logo.png", width=80)

        # DB version diagnostic (small caption at top of sidebar)
        try:
            _diag_conn = get_conn()
            # Detect DB version by feature presence (best-effort, fail-soft)
            try:
                _has_auto = _diag_conn.execute(
                    "SELECT 1 FROM occ_automation_risk LIMIT 1"
                ).fetchone() is not None
            except Exception:
                _has_auto = False
            _diag_conn.close()
            _db_ver = "v1.7" if _has_auto else "v1.6"
            st.caption(f"DB: {_db_ver}")
        except Exception:
            st.caption("DB: unknown")

        st.divider()
        view_mode = st.radio(
            "View",
            options=["Explore", "Rank"],
            index=0,
            key="view_mode",
            help=(
                "Explore: deep-dive trend charts for one program + geography. "
                "Rank: scored, letter-graded leaderboards — either top "
                "programs for a chosen market, or top markets for a chosen "
                "program."
            ),
        )

        # Quick-select presets
        preset_names = ["— Select a program —"] + list(PROGRAM_PRESETS.keys())
        chosen_preset = st.selectbox(
            "Quick Select",
            options=preset_names,
            index=0,
            key="preset_select",
        )

        if chosen_preset != "— Select a program —":
            preset = PROGRAM_PRESETS[chosen_preset]
            st.session_state["_preset_cips"] = preset["cips"]
            st.session_state["_preset_level"] = preset["level"]
        else:
            st.session_state.pop("_preset_cips", None)
            st.session_state.pop("_preset_level", None)

        st.divider()

        # 1. Geography
        st.markdown("### 1 · Geography")
        geo_type = st.radio(
            "scope",
            ["National", "By State", "By Metro Area"],
            label_visibility="collapsed",
        )

        geo_values = []
        selected_metro_labels = []
        all_states = False
        all_metros = False

        if geo_type == "By State":
            states = load_states()
            all_states = st.checkbox("All states", value=False, key="all_states")
            if all_states:
                geo_values = list(states)
            else:
                selected_states = st.multiselect(
                    "State(s):",
                    options=states,
                    placeholder="Select states…",
                )
                geo_values = selected_states

        elif geo_type == "By Metro Area":
            cbsa_list = load_cbsas()
            cbsa_display_to_code = {label: code for code, label in cbsa_list}
            all_metros = st.checkbox("All metro areas", value=False, key="all_metros")
            if all_metros:
                selected_metro_labels = [label for _, label in cbsa_list]
                geo_values = [code for code, _ in cbsa_list]
            else:
                selected_metro_labels = st.multiselect(
                    "Metro area(s):",
                    options=[label for _, label in cbsa_list],
                    placeholder="Search metro areas…",
                )
                geo_values = [cbsa_display_to_code[l] for l in selected_metro_labels]

        st.divider()

        # 2. Subject (6-digit CIP)
        st.markdown("### 2 · Subject")

        all_cips = st.checkbox("All CIP codes", value=False, key="all_cips")

        cip_options = load_cip_options()
        cip_label_to_code = {label: code for code, label in cip_options}

        if all_cips:
            selected_cip_labels = []
            cip_patterns = ()  # empty = no filter = all
        else:
            # Use preset CIP codes if a quick-select is active, else default
            if "_preset_cips" in st.session_state:
                _pcips = st.session_state["_preset_cips"]
                default_cip_labels = [
                    l for _, l in cip_options
                    if any(l.startswith(c) for c in _pcips)
                ]
            else:
                default_cip_labels = [l for _, l in cip_options if l.startswith("51.3801")]
            selected_cip_labels = st.multiselect(
                "CIP code(s):",
                options=[label for _, label in cip_options],
                default=default_cip_labels,
                placeholder="Search by code or name…",
                label_visibility="collapsed",
            )
            cip_patterns = tuple(cip_label_to_code[l] for l in selected_cip_labels)

        st.caption(
            "🔍 [Look up CIP codes](https://nces.ed.gov/ipeds/cipcode/default.aspx?y=56)"
        )

        st.divider()

        # 3. Award Level
        st.markdown("### 3 · Program Level")

        all_levels = st.checkbox("All award levels", value=False, key="all_levels")

        # Build option list: individual levels + aggregate groups
        # Note: awlevels 20 and 21 replaced awlevel 1 starting in IPEDS C2021
        # (2020-21 collection). "Undergraduate Certificate" bundles all three
        # so the chart stays continuous across the taxonomy change.
        AGGREGATE_LEVELS = {
            "Undergraduate Certificate": (1, 2, 4, 20, 21),
            "Graduate Certificate": (6, 8),
            "Doctoral Degree": (17, 18, 19),
        }
        level_option_labels = list(AGGREGATE_LEVELS.keys()) + [v for v in AWARD_LEVELS.values()]
        level_label_to_code = {v: k for k, v in AWARD_LEVELS.items()}

        if all_levels:
            selected_level_labels = list(AWARD_LEVELS.values())
            selected_awlevels = tuple(AWARD_LEVELS.keys())
        else:
            # Use preset level if a quick-select is active, else default
            if "_preset_level" in st.session_state:
                default_levels = [st.session_state["_preset_level"]]
            else:
                default_levels = ["Bachelor's degree"]
            selected_level_labels = st.multiselect(
                "Award level(s):",
                options=level_option_labels,
                default=default_levels,
                placeholder="Choose levels…",
                label_visibility="collapsed",
            )
            # Expand aggregate groups into individual awlevel codes
            awlevel_set = set()
            for lbl in selected_level_labels:
                if lbl in AGGREGATE_LEVELS:
                    awlevel_set.update(AGGREGATE_LEVELS[lbl])
                else:
                    awlevel_set.add(level_label_to_code[lbl])
            selected_awlevels = tuple(sorted(awlevel_set))

    # ── Main area ─────────────────────────────────────────────────────────────
    # When the user picked the Rank view, render that page and short-circuit
    # everything below. The Explore page sees its existing flow unchanged.
    if view_mode == "Rank":
        _render_rankings_page()
        return

    st.title("Program/Market Explorer")
    if _windows.get("completions"):
        _w_min, _w_max = _windows["completions"]
        _comp_window_str = f"AY {_ay_label(_w_min)} – AY {_ay_label(_w_max)}"
    else:
        _comp_window_str = "available academic years"
    st.caption(
        "Total degrees and certificates awarded by Title-IV-eligible, "
        "IPEDS-reporting U.S. postsecondary institutions, "
        "counted by 6-digit CIP code and award level. "
        f"Coverage: {_comp_window_str}. "
        "| Source: NCES IPEDS Completions Survey (file C{year}_A, "
        "first-major awards only)."
    )

    # Determine geo_key for query — "All states" is functionally national
    if geo_type == "By State" and all_states:
        geo_key = "national"
    else:
        geo_key = {"National": "national", "By State": "state", "By Metro Area": "metro"}[geo_type]

    # Validate — show landing if incomplete
    geo_ready = (geo_type == "National") or bool(geo_values)
    cip_ready = all_cips or bool(cip_patterns)
    level_ready = all_levels or bool(selected_awlevels)

    if not (geo_ready and cip_ready and level_ready):
        c1, c2, c3 = st.columns(3)
        status = lambda ok: "✅" if ok else "⬜"
        c1.info(f"{status(geo_ready)} **Step 1:** Select a geography")
        c2.info(f"{status(cip_ready)} **Step 2:** Select CIP code(s)")
        c3.info(f"{status(level_ready)} **Step 3:** Select program level(s)")
        st.divider()
        st.markdown(
            "Use the sidebar to build your query. This tool searches "
            f"**{3_000_000:,}+** completions records across 10 academic years "
            "from ~7,000 U.S. postsecondary institutions."
        )
        return

    # ── Query ─────────────────────────────────────────────────────────────────
    with st.spinner("Querying…"):
        df = run_query(
            cip_patterns=cip_patterns,
            awlevels=selected_awlevels,
            geo_key=geo_key,
            geo_values=tuple(geo_values),
            split_by_level=True,
        )
        df_inst = run_institution_query(
            cip_patterns=cip_patterns,
            awlevels=selected_awlevels,
            geo_key=geo_key,
            geo_values=tuple(geo_values),
        )
        program_counts = run_program_count_query(
            cip_patterns=cip_patterns,
            awlevels=selected_awlevels,
            geo_key=geo_key,
            geo_values=tuple(geo_values),
        )

    if df.empty:
        st.warning(
            "No completions found for these filters. "
            "Try selecting a broader CIP series, more award levels, or a larger geography."
        )
        return

    # ── Build labels ──────────────────────────────────────────────────────────
    if geo_type == "National":
        geo_label = "United States"
    elif geo_type == "By State":
        geo_label = "All States" if all_states else ", ".join(geo_values)
    else:
        if all_metros:
            geo_label = "All BLS Metro Areas"
        elif selected_metro_labels:
            geo_label = ", ".join(selected_metro_labels)
        else:
            geo_label = "Selected Metro Areas"

    if all_cips:
        cip_display = "All Programs"
    elif len(selected_cip_labels) == 1:
        cip_display = selected_cip_labels[0].split(" \u2013 ", 1)[-1]
    elif len(selected_cip_labels) <= 3:
        cip_display = " / ".join(l.split(" \u2013 ", 1)[-1] for l in selected_cip_labels)
    else:
        cip_display = f"{len(selected_cip_labels)} CIP codes"

    if all_levels:
        level_str = "All Award Levels"
    elif len(selected_level_labels) <= 2:
        level_str = " & ".join(selected_level_labels)
    else:
        level_str = f"{len(selected_level_labels)} award levels"

    # IPEDS file C{YYYY}_A reports awards conferred July YYYY-1 through
    # June YYYY (AY (YYYY-1)-YYYY), and the loader stores YYYY in `year`.
    # So DB year=Y means AY (Y-1)-Y.
    def yr_label(y):
        return f"{y - 1}–{str(y)[-2:]}"

    def yr_label_short(y):
        return f"'{str(y - 1)[-2:]}–'{str(y)[-2:]}"

    all_years = sorted(df["year"].unique())
    year_tick_labels = [yr_label(y) for y in all_years]

    # ── Compute unified projection ───────────────────────────────────────────
    df_totals = df.groupby("year")["completions"].sum()
    sel_dict = df_totals.to_dict()

    # Get employment CAGR if specific CIPs selected
    emp_cagr_for_completions = None
    if not all_cips:
        emp_cagr_for_completions = get_emp_proj_cagr(
            cip_patterns=cip_patterns,
            awlevels=selected_awlevels,
            geo_key=geo_key,
            geo_values=tuple(geo_values),
        )

    projection, proj_components = compute_unified_projection(
        sel_dict,
        emp_cagr=emp_cagr_for_completions,
        program_counts=program_counts,
    )

    # Extract capacity CAGR for the projected program count bars
    capacity_cagr = proj_components.get("capacity")

    # ── Unified export (Excel / CSV / PDF) ────────────────────────────────────
    def _collect_export_sheets():
        """Assemble (name, df, fmt_opts) tuples for every section that has data.

        Returns a list. Each format (Excel / CSV-zip / PDF) consumes the same
        list, optionally filtered by user-selected section names. The Summary
        sheet is always inserted first.
        """
        sheets = []

        # 1. Completions Trend
        if "award_level_name" in df.columns:
            trend_df = df[["year", "award_level_name", "completions"]].copy()
            trend_df["year"] = trend_df["year"].apply(yr_label)
            trend_df.columns = ["Year", "Award Level", "Completions"]
        else:
            trend_df = df.groupby("year")["completions"].sum().reset_index()
            trend_df["year"] = trend_df["year"].apply(yr_label)
            trend_df.columns = ["Year", "Completions"]

        # Add YoY %
        _totals = df.groupby("year")["completions"].sum().reset_index()
        _totals = _totals.sort_values("year")
        _totals["YoY % Change"] = _totals["completions"].pct_change()
        _totals["year"] = _totals["year"].apply(yr_label)
        _yoy_map = dict(zip(_totals["year"], _totals["YoY % Change"]))
        trend_df["YoY % Change"] = trend_df["Year"].map(_yoy_map)

        sheets.append(("Completions Trend", trend_df, {
            "num_cols": ["Completions"],
            "pct_cols": ["YoY % Change"],
        }))

        # 2. By Institution
        if not df_inst.empty:
            meta = (
                df_inst.sort_values("year")
                .groupby("unitid")[["instnm", "city", "stabbr", "control_name"]]
                .last()
                .reset_index()
            )
            pivot = df_inst.pivot_table(
                index="unitid", columns="year", values="completions",
                aggfunc="sum", fill_value=0,
            ).reset_index()
            pivot = pivot.merge(meta, on="unitid", how="left")
            pivot.columns.name = None
            yr_cols = sorted([c for c in pivot.columns if isinstance(c, int)])
            first_col, last_col = yr_cols[0], yr_cols[-1]
            n_years = last_col - first_col
            col_3ago = last_col - 3

            def _inst_cagr(row, start_col, n):
                fv, lv = row[start_col], row[last_col]
                if fv > 0 and lv > 0 and n > 0:
                    return ((lv / fv) ** (1 / n) - 1)
                return None

            if col_3ago in yr_cols:
                pivot["Post-COVID CAGR"] = pivot.apply(lambda r: _inst_cagr(r, col_3ago, 3), axis=1)
            pivot["Long-Term CAGR"] = pivot.apply(lambda r: _inst_cagr(r, first_col, n_years), axis=1)
            pivot = pivot.rename(columns={y: yr_label(y) for y in yr_cols})
            control_map = {"Public": "Public", "Private nonprofit": "Private", "Private for-profit": "For-Profit"}
            pivot["control_name"] = pivot["control_name"].map(control_map).fillna(pivot["control_name"])
            pivot["city"] = pivot["city"] + ", " + pivot["stabbr"]
            pivot = pivot.drop(columns=["unitid", "stabbr"])
            pivot = pivot.rename(columns={"instnm": "Institution", "city": "City", "control_name": "Control"})
            yr_labels = [yr_label(y) for y in yr_cols]
            cagr_cols = [c for c in ["Post-COVID CAGR", "Long-Term CAGR"] if c in pivot.columns]
            pivot = pivot[["Institution", "City", "Control"] + yr_labels + cagr_cols]
            last_yr_lbl = yr_label(last_col)
            pivot = pivot.sort_values(last_yr_lbl, ascending=False, na_position="last").reset_index(drop=True)

            sheets.append(("By Institution", pivot, {
                "num_cols": yr_labels,
                "pct_cols": cagr_cols,
            }))

        # 3. Distance Education Programs
        if not all_cips:
            try:
                _dep_conn = get_conn()
                _dep_conn.execute("SELECT 1 FROM completions_dep LIMIT 1")
                _dep_conn.close()
                _dep_df = run_dep_query(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )
                if _dep_df is not None and not _dep_df.empty:
                    dep_export = _dep_df.copy()
                    dep_export["year"] = dep_export["year"].apply(yr_label)
                    dep_export = dep_export.rename(columns={
                        "year": "Year",
                        "programs": "Total Programs",
                        "programs_de_any": "DE Programs (Any)",
                        "pct_de_any": "DE Share %",
                    })
                    cols_to_keep = [c for c in ["Year", "Total Programs", "DE Programs (Any)", "DE Share %"] if c in dep_export.columns]
                    dep_export = dep_export[cols_to_keep]
                    sheets.append(("Distance Education", dep_export, {
                        "num_cols": ["Total Programs", "DE Programs (Any)"],
                    }))
            except Exception:
                pass

        # 4. Graduate Outcomes (Scorecard)
        if not all_cips:
            try:
                _sc_conn = get_conn()
                _sc_conn.execute("SELECT 1 FROM college_scorecard LIMIT 1")
                _sc_conn.close()
                _sc_df = run_scorecard_query(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )
                if not _sc_df.empty:
                    _sc_df = (
                        _sc_df.sort_values("earn_mdn_4yr", ascending=False)
                        .drop_duplicates(subset=["unitid"], keep="first")
                    )
                    sc_export = _sc_df.rename(columns={
                        "instnm": "Institution",
                        "city": "City",
                        "control_name": "Control",
                        "earn_mdn_4yr": "Median Earnings (4yr)",
                        "debt_all_stgp_eval_mdn": "Median Debt",
                        "debt_to_earnings": "Debt/Earnings",
                    })
                    sc_export = sc_export.sort_values(
                        "Debt/Earnings", ascending=True, na_position="last"
                    )
                    sc_export = sc_export[
                        ["Institution", "City", "Control",
                         "Median Earnings (4yr)", "Median Debt", "Debt/Earnings"]
                    ].reset_index(drop=True)
                    sheets.append(("Graduate Outcomes", sc_export, {
                        "money_cols": ["Median Earnings (4yr)", "Median Debt"],
                    }))
            except Exception:
                pass

        # 5. Employment by Occupation
        if not all_cips:
            try:
                _conn = get_conn()
                _conn.execute("SELECT 1 FROM oes_employment LIMIT 1")
                _conn.execute("SELECT 1 FROM cip_soc_crosswalk LIMIT 1")
                _conn.close()
                _emp_df = run_employment_query(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )
                if not _emp_df.empty:
                    _latest_yr = int(_emp_df["year"].max())
                    _hist_years = sorted(int(y) for y in _emp_df["year"].unique())
                    _emp_latest = _emp_df[_emp_df["year"] == _latest_yr].copy()

                    # Per-occupation projections (CAGR is geo-best-match)
                    _emp_proj = get_employment_projections(
                        soc_codes=tuple(_emp_latest["occ_code"].unique()),
                        geo_key=geo_key,
                        geo_values=tuple(geo_values),
                    )
                    _cagr_map = (
                        {} if _emp_proj.empty or "cagr" not in _emp_proj.columns
                        else dict(zip(_emp_proj["occ_code"], _emp_proj["cagr"]))
                    )

                    # Bolt on LMII automation risk score per occupation
                    _emp_risk = get_automation_risk(
                        tuple(_emp_latest["occ_code"].unique())
                    )

                    # Pivot historical employment so each year is a column
                    _emp_pivot = _emp_df.pivot_table(
                        index="occ_code",
                        columns="year",
                        values="tot_emp",
                        aggfunc="sum",
                    ).reset_index()
                    _emp_pivot.columns.name = None
                    _hist_col_map = {y: f"Emp {int(y)}" for y in _hist_years}
                    _emp_pivot = _emp_pivot.rename(columns=_hist_col_map)

                    # Project forward through 2029 using each SOC's own CAGR
                    _proj_target_year = 2029
                    _proj_years = list(range(_latest_yr + 1, _proj_target_year + 1))
                    _latest_emp_map = dict(zip(
                        _emp_latest["occ_code"], _emp_latest["tot_emp"]
                    ))

                    def _proj_val(soc, year):
                        cagr = _cagr_map.get(soc)
                        base = _latest_emp_map.get(soc)
                        if cagr is None or base is None:
                            return None
                        return int(round(base * (1 + cagr) ** (year - _latest_yr)))

                    for _y in _proj_years:
                        _emp_pivot[f"Emp {_y} (proj.)"] = _emp_pivot["occ_code"].apply(
                            lambda s, y=_y: _proj_val(s, y)
                        )

                    # Latest-year metadata: title, wage, risk, CAGR
                    _meta = _emp_latest[["occ_code", "occ_title", "a_median"]].copy()
                    if not _emp_risk.empty:
                        _meta = _meta.merge(
                            _emp_risk[["occ_code", "risk_score"]],
                            on="occ_code",
                            how="left",
                        )
                    _meta["Projected CAGR"] = _meta["occ_code"].map(_cagr_map)

                    emp_export = _meta.merge(_emp_pivot, on="occ_code", how="left")
                    emp_export = emp_export.rename(columns={
                        "occ_code": "SOC Code",
                        "occ_title": "Occupation",
                        "a_median": "Median Annual Wage",
                        "risk_score": "Automation Risk (1-10)",
                    })

                    _hist_cols = [
                        _hist_col_map[y] for y in _hist_years
                        if _hist_col_map[y] in emp_export.columns
                    ]
                    _proj_cols = [
                        f"Emp {y} (proj.)" for y in _proj_years
                        if f"Emp {y} (proj.)" in emp_export.columns
                    ]
                    cols = (
                        ["SOC Code", "Occupation"]
                        + _hist_cols
                        + _proj_cols
                        + [c for c in [
                            "Median Annual Wage",
                            "Automation Risk (1-10)",
                            "Projected CAGR",
                        ] if c in emp_export.columns]
                    )
                    emp_export = emp_export[cols]

                    _sort_col = (
                        _hist_col_map[_latest_yr]
                        if _hist_col_map[_latest_yr] in emp_export.columns
                        else (_hist_cols[-1] if _hist_cols else None)
                    )
                    if _sort_col:
                        emp_export = emp_export.sort_values(
                            _sort_col, ascending=False, na_position="last"
                        ).reset_index(drop=True)

                    sheets.append(("Employment", emp_export, {
                        "num_cols": _hist_cols + _proj_cols,
                        "money_cols": ["Median Annual Wage"],
                        "pct_cols": ["Projected CAGR"],
                    }))
            except Exception:
                pass

        # 6. Summary / metadata tab
        summary_rows = [
            {"Field": "Report Generated", "Value": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")},
            {"Field": "Geography", "Value": geo_label},
            {"Field": "CIP Code(s)", "Value": "All Programs" if all_cips else "; ".join(selected_cip_labels)},
            {"Field": "Award Level(s)", "Value": "All Award Levels" if all_levels else "; ".join(selected_level_labels)},
        ]
        summary_df = pd.DataFrame(summary_rows)
        sheets.insert(0, ("Summary", summary_df, {}))

        return sheets

    cip_slug = "all_programs" if all_cips else (
        "_".join(cip_label_to_code[l] for l in selected_cip_labels) or "completions"
    )
    geo_slug = geo_label.replace(", ", "_").replace(" ", "_")
    base_fname = f"ipeds_{cip_slug}_{geo_slug}"

    # Candidate section list — stable, doesn't require running the queries.
    # Some sections only fill in for specific filter combinations (e.g. DEP
    # and Scorecard need a CIP selected); we surface them in the picker
    # regardless and the underlying collector silently drops empty ones.
    selectable_sections = [
        "Completions Trend",
        "By Institution",
        "Distance Education",
        "Graduate Outcomes",
        "Employment",
    ]

    # Reset stored exports if the underlying filters changed
    _filters_signature = (cip_slug, geo_slug, tuple(selected_awlevels))
    if st.session_state.get("_export_filters_sig") != _filters_signature:
        for k in (
            "_export_payload", "_export_fname", "_export_mime",
            "_export_filters_sig", "_export_format_label",
        ):
            st.session_state.pop(k, None)
        st.session_state["_export_filters_sig"] = _filters_signature

    # ── Export modal ──────────────────────────────────────────────────────────
    @st.dialog("Export report", width="large")
    def _export_dialog():
        st.caption(
            f"Building from current filters · "
            f"{cip_display} · {level_str} · {geo_label}"
        )
        chosen_sections = st.multiselect(
            "Sections to include",
            options=selectable_sections,
            default=[],
            placeholder="All sections (default)",
            key="export_sections",
            help="Leave empty to include every section, or pick specific ones.",
        )
        export_format = st.selectbox(
            "Format",
            options=["PDF report", "Excel (.xlsx)", "CSV (.zip)"],
            key="export_format",
            help=(
                "PDF: stylized branded report. Excel: multi-tab workbook. "
                "CSV: zip archive with one CSV per section."
            ),
        )

        st.markdown(
            "<div style='height:6px'></div>", unsafe_allow_html=True,
        )

        _go = st.button(
            "Generate", type="primary", use_container_width=True,
            key="export_generate",
        )

        # Build the report when Generate is pressed. We render the download
        # button below in the SAME dialog turn — calling st.rerun() here
        # would close the dialog and the user would never see the download.
        if _go:
            keep = set(chosen_sections) if chosen_sections else set(selectable_sections)
            with st.spinner("Collecting section data…"):
                all_sheets = _collect_export_sheets()
            filtered = [s for s in all_sheets if s[0] == "Summary" or s[0] in keep]

            with st.spinner(f"Building {export_format}…"):
                try:
                    if export_format == "Excel (.xlsx)":
                        payload = build_export_workbook(filtered)
                        fname = f"{base_fname}.xlsx"
                        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        fmt_label = "Excel"
                    elif export_format == "CSV (.zip)":
                        payload = build_csv_zip(filtered)
                        fname = f"{base_fname}.zip"
                        mime = "application/zip"
                        fmt_label = "CSV"
                    else:  # PDF report
                        payload = build_pdf_report(
                            [s for s in filtered if s[0] != "Summary"],
                            report_meta={
                                "title": "IPEDS Completions Report",
                                "subtitle": f"{cip_display} · {level_str}",
                                "geo_label": geo_label,
                                "cip_display": cip_display,
                                "level_str": level_str,
                            },
                        )
                        fname = f"{base_fname}.pdf"
                        mime = "application/pdf"
                        fmt_label = "PDF"
                except Exception as e:
                    st.error(f"Couldn't build {export_format}: {e}")
                    payload = None

            if payload:
                st.session_state["_export_payload"] = payload
                st.session_state["_export_fname"] = fname
                st.session_state["_export_mime"] = mime
                st.session_state["_export_format_label"] = fmt_label

        # Render the download button inline whenever a payload is available.
        # This runs on the same dialog turn as Generate (so the user sees the
        # button immediately) AND on subsequent reopens of the dialog.
        if "_export_payload" in st.session_state:
            st.download_button(
                f"⬇  Download {st.session_state.get('_export_format_label', '')}"
                f" — {st.session_state.get('_export_fname', '')}",
                data=st.session_state["_export_payload"],
                file_name=st.session_state["_export_fname"],
                mime=st.session_state["_export_mime"],
                use_container_width=True,
                key="export_download",
                type="primary",
            )

    # Single trigger button — pinned to the right.
    # Styled in the global stylesheet via the .vi-export-trigger marker class
    # (see the .vi-export-trigger rule near the bottom of the global <style>).
    _trig_l, _trig_r = st.columns([5, 1])
    with _trig_r:
        st.html('<span class="vi-export-trigger-anchor"></span>')
        if st.button(
            "Export",
            icon=":material/file_download:",
            use_container_width=True,
            key="export_open",
        ):
            _export_dialog()

    # ── Summary metrics ───────────────────────────────────────────────────────
    agg = df_totals
    first_yr, last_yr = agg.index.min(), agg.index.max()
    last_val = int(agg[last_yr])

    # 10-year CAGR
    n10 = last_yr - first_yr
    first_val = int(agg[first_yr])
    cagr_10 = (last_val / first_val) ** (1 / n10) - 1 if first_val and n10 > 0 else None

    # 3-year CAGR
    yr_3ago = last_yr - 3
    val_3ago = int(agg[yr_3ago]) if yr_3ago in agg.index else None
    cagr_3 = (last_val / val_3ago) ** (1 / 3) - 1 if val_3ago else None

    # Projected CAGR – use blended rate directly when available so the
    # header metric matches the footnote (avoids rounding drift from
    # back-computing CAGR out of integer-rounded projection values).
    if proj_components and "blended_rate" in proj_components:
        cagr_proj = proj_components["blended_rate"]
        proj_last_yr = last_yr + (len(projection) if projection else 5)
    elif projection and last_val > 0:
        proj_last_yr, proj_last_val = projection[-1]
        n_proj = proj_last_yr - last_yr
        cagr_proj = (proj_last_val / last_val) ** (1 / n_proj) - 1 if n_proj > 0 else None
    else:
        proj_last_yr, cagr_proj = last_yr + 5, None

    # Institution count
    n_inst = df_inst["unitid"].nunique()

    # YoY trend chip for the Completions card — compare the latest year
    # to the immediately prior year (if present).
    prior_yr = last_yr - 1
    if prior_yr in agg.index and int(agg[prior_yr]) > 0:
        _yoy = last_val / int(agg[prior_yr]) - 1
        _yoy_trend = (
            "up" if _yoy > 0 else "down" if _yoy < 0 else "flat",
            f"{_yoy:+.1%} YoY",
        )
    else:
        _yoy_trend = None

    def _sent(rate: float | None) -> str | None:
        if rate is None:
            return None
        return "positive" if rate >= 0 else "negative"

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
        vi_kpi_card(
            label="Latest-Year Completions",
            value=f"{last_val:,}",
            icon="school",
            sublabel=f"AY {yr_label(last_yr)}",
            trend=_yoy_trend,
        )
    with m2:
        vi_kpi_card(
            label="Long-Term CAGR",
            value=f"{cagr_10:+.1%}" if cagr_10 is not None else "N/A",
            icon="trending_up",
            sublabel=f"{yr_label(first_yr)} → {yr_label(last_yr)}",
            sentiment=_sent(cagr_10),
        )
    with m3:
        vi_kpi_card(
            label="Post-COVID CAGR",
            value=f"{cagr_3:+.1%}" if cagr_3 is not None else "N/A",
            icon="speed",
            sublabel=f"{yr_label(yr_3ago)} → {yr_label(last_yr)}",
            sentiment=_sent(cagr_3),
        )
    with m4:
        vi_kpi_card(
            label="Projected CAGR",
            value=f"{cagr_proj:+.1%}" if cagr_proj is not None else "N/A",
            icon="insights",
            sublabel=f"{yr_label(last_yr)} → {yr_label(proj_last_yr)}",
            sentiment=_sent(cagr_proj),
        )
    with m5:
        vi_kpi_card(
            label="Reporting Institutions",
            value=f"{n_inst:,}",
            icon="account_balance",
            sublabel="in this cohort",
        )

    # ── Chart (dual-axis: completions line + programs offered bars) ──────────
    # Title now lives in the surrounding vi_card header (see below); the
    # Plotly fig no longer carries one, so its top margin can be tightened.
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # ── Programs Offered bars (secondary y-axis, behind everything) ───────
    _pc_years = sorted(program_counts.keys())
    _pc_vals = [program_counts[y] for y in _pc_years]
    if _pc_years:
        fig.add_trace(
            go.Bar(
                x=_pc_years,
                y=_pc_vals,
                name="Programs Offered",
                marker=dict(color="rgba(15, 134, 193, 0.18)"),
                hovertemplate="%{y:,} programs<extra></extra>",
                showlegend=True,
            ),
            secondary_y=True,
        )

    # ── Completions line (primary y-axis) ─────────────────────────────────
    # Always sum across selected award levels so groupings (e.g.
    # "Undergraduate Certificate" → awlevels 1/2/4/20/21) render as a
    # single trend line rather than one line per underlying level.
    df_agg = df.groupby("year")["completions"].sum().reset_index()
    fig.add_trace(
        go.Scatter(
            x=df_agg["year"],
            y=df_agg["completions"],
            mode="lines+markers+text",
            name="Completions",
            line=dict(width=2.5, color="#f26822"),
            marker=dict(size=9, color="#f26822"),
            text=[f"{v:,}" for v in df_agg["completions"]],
            textposition="top center",
            textfont=dict(size=11),
            hovertemplate="%{y:,.0f} completions<extra></extra>",
            showlegend=True,
        ),
        secondary_y=False,
    )

    # ── Projection (single unified line) ─────────────────────────────────────
    chart_years = list(all_years)
    last_actual_val = int(df_totals[all_years[-1]]) if projection else 0

    if projection:
        proj_years_list = [p[0] for p in projection]
        proj_vals = [p[1] for p in projection]
        chart_years = list(all_years) + proj_years_list

        # Gray shaded projection region
        fig.add_vrect(
            x0=all_years[-1] + 0.5,
            x1=proj_years_list[-1] + 0.5,
            fillcolor="#E5E7EB", opacity=0.3, layer="below", line_width=0,
        )

        # Projected program count bars (semi-transparent)
        if program_counts and capacity_cagr is not None:
            _last_pc = program_counts.get(all_years[-1], _pc_vals[-1] if _pc_vals else 0)
            _proj_pc_yrs = []
            _proj_pc_vals = []
            for i, y in enumerate(proj_years_list):
                _proj_pc_yrs.append(y)
                _proj_pc_vals.append(
                    max(int(round(_last_pc * (1 + capacity_cagr) ** (i + 1))), 0)
                )
            fig.add_trace(
                go.Bar(
                    x=_proj_pc_yrs,
                    y=_proj_pc_vals,
                    name="Programs (projected)",
                    marker=dict(
                        color="rgba(15, 134, 193, 0.08)",
                        line=dict(color="rgba(15, 134, 193, 0.25)", width=1),
                    ),
                    hovertemplate="%{y:,} programs (projected)<extra></extra>",
                    showlegend=False,
                ),
                secondary_y=True,
            )

        # Single projection line
        fig.add_trace(
            go.Scatter(
                x=[all_years[-1]] + proj_years_list,
                y=[last_actual_val] + proj_vals,
                mode="lines+markers+text",
                name="Projection",
                line=dict(color="rgba(242, 104, 34, 0.6)", width=3, dash="dash"),
                marker=dict(size=7, symbol="diamond", color="rgba(242, 104, 34, 0.6)"),
                text=[""] + [f"{v:,}" for v in proj_vals],
                textposition="top center",
                textfont=dict(size=10, color="rgba(242, 104, 34, 0.75)"),
                hovertemplate="%{y:,.0f} (projected)<extra></extra>",
                showlegend=False,
            ),
            secondary_y=False,
        )

    chart_tick_labels = [yr_label(y) for y in chart_years]

    fig.update_layout(
        xaxis=dict(
            tickmode="array",
            tickvals=chart_years,
            ticktext=chart_tick_labels,
            tickangle=-30,
            showgrid=True,
            gridcolor="#F3F4F6",
            gridwidth=1,
        ),
        hovermode="x unified",
        showlegend=False,
        height=500,
        margin=dict(t=20, b=60, l=70, r=70),
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Montserrat, Arial, sans-serif", size=13, color="#333333"),
        bargap=0.35,
    )
    fig.update_yaxes(
        title_text="Total Completions",
        tickformat=",",
        showgrid=True,
        gridcolor="#F3F4F6",
        gridwidth=1,
        zeroline=False,
        rangemode="tozero",
        secondary_y=False,
    )
    fig.update_yaxes(
        title_text="Programs Offered",
        tickformat=",",
        showgrid=False,
        zeroline=False,
        rangemode="tozero",
        secondary_y=True,
    )

    with vi_card(
        title=cip_display,
        subtitle=f"{level_str} · {geo_label}",
        icon="show_chart",
    ):
        st.plotly_chart(fig, use_container_width=True)

    # Projection methodology note
    if projection and proj_components:
        _parts = []
        _w = proj_components.get("weights", {})
        if "trend" in _w:
            _trend_rate = proj_components.get("trend")
            _trend_str = f" {_trend_rate:+.1%}/yr" if _trend_rate is not None else ""
            _parts.append(f"completions trend{_trend_str}")
        if "employment" in _w:
            _parts.append(
                f"BLS employment growth {emp_cagr_for_completions:+.1%}/yr"
            )
        if "capacity" in _w:
            _cap = proj_components["capacity"]
            _parts.append(
                f"program capacity {_cap:+.1%}/yr"
            )
        _blend_rate = proj_components.get("blended_rate")
        _rate_str = f" Blended rate: {_blend_rate:+.1%}/yr." if _blend_rate is not None else ""
        st.caption(
            f"**Projection** blends {', '.join(_parts)}.{_rate_str}"
        )

    # ── YoY change bar chart ───────────────────────────────────────────────────
    df_yoy = df.groupby("year")["completions"].sum().reset_index().sort_values("year")
    df_yoy["yoy"] = df_yoy["completions"].pct_change() * 100
    df_yoy = df_yoy.dropna(subset=["yoy"])
    df_yoy["color"] = df_yoy["yoy"].apply(lambda v: "#16a34a" if v >= 0 else "#dc2626")
    df_yoy["text"] = df_yoy["yoy"].apply(lambda v: f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%")

    # Actuals bars
    fig_yoy = go.Figure(go.Bar(
        x=df_yoy["year"],
        y=df_yoy["yoy"],
        marker_color=df_yoy["color"],
        text=df_yoy["text"],
        textposition="outside",
        textfont=dict(size=10, family="Montserrat, Arial, sans-serif", color="#333333"),
        hovertemplate="%{text}<extra></extra>",
        name="Actual",
        showlegend=False,
    ))

    # Projected YoY bars (semi-transparent)
    _yoy_proj = projection
    if _yoy_proj:
        last_actual = int(df_totals[all_years[-1]])
        proj_chain = [last_actual] + [p[1] for p in _yoy_proj]
        proj_yoy_years = [p[0] for p in _yoy_proj]
        proj_yoy_vals = [
            ((proj_chain[i + 1] - proj_chain[i]) / proj_chain[i] * 100)
            if proj_chain[i] > 0 else 0
            for i in range(len(_yoy_proj))
        ]
        proj_yoy_colors = [
            "rgba(22, 163, 74, 0.35)" if v >= 0 else "rgba(220, 38, 38, 0.35)"
            for v in proj_yoy_vals
        ]
        proj_yoy_text = [
            f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%" for v in proj_yoy_vals
        ]
        proj_yoy_textcolors = [
            "rgba(22, 163, 74, 0.6)" if v >= 0 else "rgba(220, 38, 38, 0.6)"
            for v in proj_yoy_vals
        ]
        fig_yoy.add_trace(go.Bar(
            x=proj_yoy_years,
            y=proj_yoy_vals,
            marker_color=proj_yoy_colors,
            text=proj_yoy_text,
            textposition="outside",
            textfont=dict(
                size=10,
                family="Montserrat, Arial, sans-serif",
                color=proj_yoy_textcolors,
            ),
            hovertemplate="%{text} (projected)<extra></extra>",
            name="Projected",
            showlegend=False,
        ))

        # Faint gray shaded region matching the main chart
        fig_yoy.add_vrect(
            x0=all_years[-1] + 0.5,
            x1=proj_yoy_years[-1] + 0.5,
            fillcolor="#E5E7EB",
            opacity=0.3,
            layer="below",
            line_width=0,
        )

    fig_yoy.update_layout(
        xaxis=dict(
            tickmode="array",
            tickvals=chart_years,
            ticktext=chart_tick_labels,
            tickangle=-30,
            showgrid=True,
            gridcolor="#F3F4F6",
            gridwidth=1,
            range=[chart_years[0] - 0.5, chart_years[-1] + 0.5],
        ),
        yaxis=dict(
            ticksuffix="%",
            tickformat=".1f",
            showgrid=True,
            gridcolor="#F3F4F6",
            gridwidth=1,
            zeroline=True,
            zerolinecolor="#999999",
            zerolinewidth=1,
        ),
        height=220,
        margin=dict(t=10, b=60, l=70, r=20),
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Montserrat, Arial, sans-serif", size=12, color="#333333"),
        showlegend=False,
    )
    with vi_card(
        title="Year-over-Year % Change",
        subtitle="Actual and projected annual change in completions",
        icon="bar_chart",
    ):
        st.plotly_chart(fig_yoy, use_container_width=True)

    # ── Completions geographic distribution (map + ranking) ──────────────────
    if not all_cips:
        _latest_year = int(max(df["year"].unique()))
        _base_year_growth = _latest_year - 3  # post-COVID 3-yr CAGR window

        # Metric toggle (Volume / Growth / Projected) — drives BOTH the map
        # and the ranking. State/Metro toggle is a sub-control inside the
        # ranking column.
        st.markdown(
            f"<div class='vi-map-caption'>Geographic distribution · "
            f"{yr_label(_latest_year)} completions</div>",
            unsafe_allow_html=True,
        )
        _comp_metric = st.radio(
            "Metric",
            ["Volume", "Growth", "Projected"],
            index=0,
            horizontal=True,
            key="comp_metric",
            label_visibility="collapsed",
        )

        # Resolve SOC codes once — needed only for Projected.
        _comp_soc_codes = (
            resolve_soc_codes_for_cips(cip_patterns, selected_awlevels)
            if _comp_metric == "Projected" else tuple()
        )

        # ── Build the per-state dataframe driving the choropleth ─────────
        if _comp_metric == "Volume":
            _state_df = run_completions_by_state_query(
                cip_patterns=cip_patterns,
                awlevels=selected_awlevels,
                year=_latest_year,
            )
            _state_value_col = "completions"
            _value_label = "Completions"
            _value_format = "{:,.0f}"
            _map_title = f"Completions by State — {yr_label(_latest_year)}"
            _rank_title = f"Top States — {yr_label(_latest_year)}"
        elif _comp_metric == "Growth":
            _state_df = run_completions_state_cagr(
                cip_patterns=cip_patterns,
                awlevels=selected_awlevels,
                base_year=_base_year_growth,
                end_year=_latest_year,
            )
            _state_value_col = "cagr"
            _value_label = "Post-COVID CAGR"
            _value_format = "{:+.1%}"
            _map_title = (
                f"Post-COVID Completions CAGR — "
                f"{yr_label(_base_year_growth)} → {yr_label(_latest_year)}"
            )
            _rank_title = "Fastest-Growing States"
        else:  # Projected
            if not _comp_soc_codes:
                st.info(
                    "Projected change requires CIP codes that map to BLS "
                    "occupations. Try a different program selection."
                )
                _state_df = pd.DataFrame()
            else:
                _state_df = run_employment_projection_state(_comp_soc_codes)
            _state_value_col = "cagr"
            _value_label = "Projected CAGR (related occs)"
            _value_format = "{:+.1%}"
            _map_title = "Projected Annual Change — Related Occupations"
            _rank_title = "Top States by Projected Growth"

        # Need at least 3 states with the chosen metric to render the map.
        _has_state_data = (
            not _state_df.empty
            and "stabbr" in _state_df.columns
            and _state_df[_state_value_col].notna().sum() >= 3
        )

        if _has_state_data:
            _state_df = _state_df.dropna(subset=[_state_value_col])
            _map_col, _rank_col = st.columns([3, 2])
            with _map_col:
                _fig_map = vi_choropleth(
                    _state_df["stabbr"],
                    _state_df[_state_value_col],
                    title="",
                    colorbar_title=_value_label,
                    hover_format=_value_format,
                    hover_label=_value_label,
                )
                with vi_card(
                    title=_map_title,
                    subtitle=f"{_value_label} by state",
                    icon="public",
                ):
                    st.plotly_chart(_fig_map, use_container_width=True)

            with _rank_col:
                _rank_grain = st.radio(
                    "Show ranking by:",
                    ["State", "Metro"],
                    index=0,
                    horizontal=True,
                    key="comp_rank_grain",
                    label_visibility="collapsed",
                )

                if _rank_grain == "State":
                    _rank_df = _state_df.sort_values(
                        _state_value_col, ascending=False, na_position="last"
                    ).head(15)
                    _rank_fig = vi_ranking_bar(
                        _rank_df["stabbr"],
                        _rank_df[_state_value_col],
                        title="",
                        value_label=_value_label,
                        value_format=_value_format,
                    )
                    _rank_header = _rank_title
                else:  # Metro
                    if _comp_metric == "Volume":
                        _metro_df = run_completions_by_metro_query(
                            cip_patterns=cip_patterns,
                            awlevels=selected_awlevels,
                            year=_latest_year, top_n=15,
                        )
                        _metro_value_col = "completions"
                    elif _comp_metric == "Growth":
                        _metro_df = run_completions_metro_cagr(
                            cip_patterns=cip_patterns,
                            awlevels=selected_awlevels,
                            base_year=_base_year_growth,
                            end_year=_latest_year, top_n=15,
                        )
                        _metro_value_col = "cagr"
                    else:  # Projected
                        _metro_df = (
                            run_employment_projection_metro(
                                _comp_soc_codes, top_n=15,
                            )
                            if _comp_soc_codes else pd.DataFrame()
                        )
                        _metro_value_col = "cagr"

                    if _metro_df is None or _metro_df.empty:
                        st.info("No metro-level data for this metric.")
                        _rank_fig = None
                        _rank_header = _rank_title.replace("States", "Metros")
                    else:
                        _rank_fig = vi_ranking_bar(
                            _metro_df["cbsa_name"],
                            _metro_df[_metro_value_col],
                            title="",
                            value_label=_value_label,
                            value_format=_value_format,
                        )
                        _rank_header = _rank_title.replace("States", "Metros")

                if _rank_fig is not None:
                    with vi_card(
                        title=_rank_header,
                        subtitle=f"Top by {_value_label.lower()}",
                        icon="leaderboard",
                    ):
                        st.plotly_chart(_rank_fig, use_container_width=True)

    # Footnote listing CIP codes and award levels included
    if all_cips:
        _cip_note = "All CIP codes"
    else:
        _cip_bullets = "  \n".join(f"- **{l}**" for l in selected_cip_labels)
        _cip_note = f"**{len(selected_cip_labels)}** CIP code(s):  \n{_cip_bullets}"
    _level_note = "All award levels" if all_levels else ", ".join(selected_level_labels)
    st.caption(f"Includes {_cip_note}  \nAward level(s): {_level_note}")

    # ── Search Interest Trends ────────────────────────────────────────────────
    st.divider()
    st.subheader("Search Interest Trends")

    _trends_ok = False
    try:
        _t_conn = get_conn()
        _t_conn.execute("SELECT 1 FROM google_trends_time LIMIT 1")
        _t_conn.close()
        _trends_ok = True
    except Exception:
        pass

    if not _trends_ok:
        st.info(
            "Google Trends data not loaded. Run `python load_google_trends.py` "
            "to download search interest data."
        )
    elif all_cips:
        st.info(
            "Search interest data is shown when specific CIP code(s) are selected. "
            "Deselect **All CIP codes** and choose program(s) to see search trends."
        )
    else:
        with st.spinner("Querying search interest data..."):
            trends_data = run_google_trends_query(
                cip_patterns=cip_patterns,
                geo_key=geo_key,
                geo_values=tuple(geo_values),
            )

        if trends_data is None:
            st.info("No search interest data available for the selected program(s).")
        else:
            df_trend = trends_data["time_series"]
            _geo_interest = trends_data["geo_interest"]
            _search_terms = trends_data["search_terms"]
            _has_volume = trends_data.get("has_volume", False)
            _volume_series = trends_data.get("volume_series")
            _per_cip_volume = trends_data.get("per_cip_volume")
            _geo_volume = trends_data.get("geo_volume")
            _est_monthly_vol = trends_data.get("est_monthly_vol")

            # Volume is default when calibration data exists
            _show_volume = (
                _has_volume
                and _volume_series is not None
                and not _volume_series.empty
            )

            # Helper to format volume numbers
            def _fmt_vol(v):
                if v is None:
                    return "N/A"
                if v >= 1_000_000:
                    return f"{v / 1_000_000:.1f}M"
                if v >= 1_000:
                    return f"{v / 1_000:.1f}K"
                return f"{v:,.0f}"

            # ── Rolling comparison metrics (all based on volume) ─────────
            _vol_src = _volume_series if _show_volume else None
            _int_src = df_trend

            # Monthly volume = rolling 12-month average
            _avg_monthly_vol = None
            if _vol_src is not None and len(_vol_src) >= 12:
                _avg_monthly_vol = _vol_src.tail(12)["volume"].mean()

            # MoM: most recent month vs previous month
            _mom_change = None
            if _vol_src is not None and len(_vol_src) >= 2:
                _cur_m = _vol_src["volume"].iloc[-1]
                _prev_m = _vol_src["volume"].iloc[-2]
                if _prev_m > 0:
                    _mom_change = (_cur_m - _prev_m) / _prev_m

            # QoQ: most recent 3 months vs prior 3 months
            _qoq_change = None
            if _vol_src is not None and len(_vol_src) >= 6:
                _cur_q = _vol_src.tail(3)["volume"].mean()
                _prev_q = _vol_src.iloc[-6:-3]["volume"].mean()
                if _prev_q > 0:
                    _qoq_change = (_cur_q - _prev_q) / _prev_q

            # YoY: most recent 12 months vs prior 12 months
            _yoy_change = None
            if _vol_src is not None and len(_vol_src) >= 24:
                _cur_y = _vol_src.tail(12)["volume"].mean()
                _prev_y = _vol_src.iloc[-24:-12]["volume"].mean()
                if _prev_y > 0:
                    _yoy_change = (_cur_y - _prev_y) / _prev_y
            elif len(_int_src) >= 24:
                _cur_y = _int_src.tail(12)["interest"].mean()
                _prev_y = _int_src.iloc[-24:-12]["interest"].mean()
                if _prev_y > 0:
                    _yoy_change = (_cur_y - _prev_y) / _prev_y

            # Metrics — 4 columns when volume is available
            if _show_volume and _avg_monthly_vol is not None:
                m1, m2, m3, m4 = st.columns(4)
                m1.metric(
                    "Monthly Search Volume",
                    _fmt_vol(round(_avg_monthly_vol)),
                    help="Average estimated monthly searches over the "
                         "most recent 12 months.",
                )
                m2.metric(
                    "MoM Change",
                    f"{_mom_change:+.1%}" if _mom_change is not None else "N/A",
                    help="Most recent month vs. previous month.",
                )
                m3.metric(
                    "QoQ Change",
                    f"{_qoq_change:+.1%}" if _qoq_change is not None else "N/A",
                    help="Most recent 3 months vs. prior 3 months.",
                )
                m4.metric(
                    "YoY Change",
                    f"{_yoy_change:+.1%}" if _yoy_change is not None else "N/A",
                    help="Most recent 12 months vs. prior 12 months.",
                )
            else:
                # Fallback: interest-only metrics
                _peak_idx = df_trend["interest"].idxmax()
                _peak_date = df_trend.loc[_peak_idx, "date"]
                t1, t2, t3 = st.columns(3)
                t1.metric(
                    f"Search Interest ({geo_label})",
                    f"{_geo_interest:.0f}/100"
                    if _geo_interest is not None else "N/A",
                    help="Google Trends relative interest (0=none, 100=peak).",
                )
                t2.metric(
                    "YoY Change",
                    f"{_yoy_change:+.1%}" if _yoy_change is not None else "N/A",
                )
                t3.metric(
                    "Peak Interest",
                    f"{_peak_date.strftime('%b %Y')}"
                    if pd.notna(_peak_date) else "N/A",
                )

            # ── Chart: Dual-axis (volume left, index right) or index-only
            df_per_cip = trends_data["per_cip_time"]
            _multi_cip = (
                not df_per_cip.empty
                and df_per_cip["cipcode"].nunique() > 1
            )

            _trend_colors = [
                "#8B5CF6", "#0f86c1", "#e87537", "#10B981", "#EF4444",
                "#F59E0B", "#EC4899", "#14B8A6", "#6366F1", "#F97316",
            ]

            if _show_volume:
                # ── Dual-axis chart: volume (left) + interest index (right)
                fig_trend = make_subplots(specs=[[{"secondary_y": True}]])

                if _multi_cip:
                    _use_vol = (
                        _per_cip_volume is not None
                        and "volume" in _per_cip_volume.columns
                    )
                    _vdf = _per_cip_volume if _use_vol else df_per_cip
                    for idx, (cip, grp) in enumerate(
                        _vdf.groupby("cipcode", sort=False)
                    ):
                        _color = _trend_colors[idx % len(_trend_colors)]
                        _label = grp["search_term"].iloc[0]
                        # Volume line (left axis)
                        if _use_vol:
                            fig_trend.add_trace(go.Scatter(
                                x=grp["date"], y=grp["volume"],
                                mode="lines", name=_label,
                                line=dict(width=2, color=_color),
                                hovertemplate=(
                                    f"<b>{_label}</b><br>"
                                    "%{x|%b %Y}<br>"
                                    "Volume: %{y:,.0f}<extra></extra>"
                                ),
                            ), secondary_y=False)

                    # Interest index as dashed line on right axis
                    fig_trend.add_trace(go.Scatter(
                        x=df_trend["date"], y=df_trend["interest"],
                        mode="lines", name="Interest Index",
                        line=dict(width=1.5, color="#9CA3AF", dash="dot"),
                        hovertemplate=(
                            "<b>Interest Index</b><br>"
                            "%{x|%b %Y}<br>"
                            "Index: %{y:.0f}/100<extra></extra>"
                        ),
                    ), secondary_y=True)

                    _show_legend = True
                    _chart_title = (
                        "<b>Estimated Monthly Search Volume by Program</b>"
                    )
                else:
                    # Single CIP: volume area + interest dashed line
                    fig_trend.add_trace(go.Scatter(
                        x=_volume_series["date"],
                        y=_volume_series["volume"],
                        mode="lines", name="Est. Volume",
                        line=dict(width=2, color="#8B5CF6"),
                        fill="tozeroy",
                        fillcolor="rgba(139, 92, 246, 0.1)",
                        hovertemplate=(
                            "<b>%{x|%b %Y}</b><br>"
                            "Volume: %{y:,.0f}<extra></extra>"
                        ),
                    ), secondary_y=False)

                    fig_trend.add_trace(go.Scatter(
                        x=df_trend["date"], y=df_trend["interest"],
                        mode="lines", name="Interest Index",
                        line=dict(width=1.5, color="#9CA3AF", dash="dot"),
                        hovertemplate=(
                            "<b>%{x|%b %Y}</b><br>"
                            "Index: %{y:.0f}/100<extra></extra>"
                        ),
                    ), secondary_y=True)

                    _show_legend = True
                    _chart_title = (
                        "<b>Estimated Monthly Search Volume Over Time</b>"
                    )

                fig_trend.update_yaxes(
                    title_text="Est. Monthly Searches",
                    showgrid=True, gridcolor="#F3F4F6", gridwidth=1,
                    rangemode="tozero",
                    secondary_y=False,
                )
                fig_trend.update_yaxes(
                    title_text="Interest Index (0-100)",
                    showgrid=False,
                    rangemode="tozero", range=[0, 105],
                    secondary_y=True,
                )

            else:
                # ── Interest-only chart (no volume calibration) ──────────
                fig_trend = go.Figure()

                if _multi_cip:
                    for idx, (cip, grp) in enumerate(
                        df_per_cip.groupby("cipcode", sort=False)
                    ):
                        _color = _trend_colors[idx % len(_trend_colors)]
                        _label = grp["search_term"].iloc[0]
                        fig_trend.add_trace(go.Scatter(
                            x=grp["date"], y=grp["interest"],
                            mode="lines", name=_label,
                            line=dict(width=2, color=_color),
                            hovertemplate=(
                                f"<b>{_label}</b><br>"
                                "%{x|%b %Y}<br>"
                                "Interest: %{y:.0f}<extra></extra>"
                            ),
                        ))
                    fig_trend.add_trace(go.Scatter(
                        x=df_trend["date"], y=df_trend["interest"],
                        mode="lines", name="Average (all selected)",
                        line=dict(width=2.5, color="#333333", dash="dash"),
                        hovertemplate=(
                            "<b>Average</b><br>"
                            "%{x|%b %Y}<br>"
                            "Interest: %{y:.0f}<extra></extra>"
                        ),
                    ))
                    _show_legend = True
                    _chart_title = "<b>National Search Interest by Program</b>"
                else:
                    fig_trend.add_trace(go.Scatter(
                        x=df_trend["date"], y=df_trend["interest"],
                        mode="lines", name="Search Interest",
                        line=dict(width=2, color="#8B5CF6"),
                        fill="tozeroy",
                        fillcolor="rgba(139, 92, 246, 0.1)",
                        hovertemplate=(
                            "<b>%{x|%b %Y}</b><br>"
                            "Interest: %{y:.0f}<extra></extra>"
                        ),
                    ))
                    _show_legend = False
                    _chart_title = "<b>National Search Interest Over Time</b>"

                fig_trend.update_yaxes(
                    title="Interest (0-100)",
                    showgrid=True, gridcolor="#F3F4F6", gridwidth=1,
                    rangemode="tozero", range=[0, 105],
                )

            fig_trend.update_layout(
                title=dict(
                    text=_chart_title,
                    font=dict(size=15), x=0, xanchor="left",
                ),
                xaxis=dict(
                    title="", showgrid=True,
                    gridcolor="#F3F4F6", gridwidth=1,
                ),
                height=400,
                margin=dict(t=60, b=40, l=60, r=60),
                plot_bgcolor="white",
                paper_bgcolor="white",
                font=dict(
                    family="Montserrat, Arial, sans-serif",
                    size=12, color="#333333",
                ),
                showlegend=_show_legend,
                legend=dict(
                    orientation="h", yanchor="bottom", y=1.02,
                    xanchor="left", x=0, font=dict(size=11),
                ),
                hovermode="x unified",
            )
            st.plotly_chart(fig_trend, use_container_width=True)

            # ── Month-over-Month % change bar chart ──────────────────────
            # Mirrors the YoY bar chart in the Completions section, but at
            # monthly resolution. Trimmed to the most recent 24 months so
            # the bars stay readable. Falls back to the interest index when
            # no volume calibration is present.
            _mom_src = (
                _volume_series.copy()
                if (_show_volume and _volume_series is not None
                    and not _volume_series.empty)
                else df_trend.copy()
            )
            _mom_value_col = "volume" if "volume" in _mom_src.columns else "interest"
            if len(_mom_src) >= 2:
                _mom_src = _mom_src.sort_values("date").tail(25).copy()
                _mom_src["mom"] = _mom_src[_mom_value_col].pct_change() * 100
                _mom_src = _mom_src.dropna(subset=["mom"])
                if not _mom_src.empty:
                    _mom_src["color"] = _mom_src["mom"].apply(
                        lambda v: "#16a34a" if v >= 0 else "#dc2626"
                    )
                    _mom_src["text"] = _mom_src["mom"].apply(
                        lambda v: f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%"
                    )
                    fig_mom = go.Figure(go.Bar(
                        x=_mom_src["date"],
                        y=_mom_src["mom"],
                        marker_color=_mom_src["color"],
                        text=_mom_src["text"],
                        textposition="outside",
                        textfont=dict(
                            size=9,
                            family="Montserrat, Arial, sans-serif",
                            color="#333333",
                        ),
                        hovertemplate="%{x|%b %Y}<br>%{text}<extra></extra>",
                    ))
                    fig_mom.update_layout(
                        title=dict(
                            text="Month-over-Month % Change",
                            font=dict(size=13), x=0, xanchor="left",
                        ),
                        xaxis=dict(
                            title="", showgrid=True, gridcolor="#F3F4F6",
                            tickformat="%b %Y", tickangle=-45,
                        ),
                        yaxis=dict(
                            ticksuffix="%", tickformat=".1f",
                            showgrid=True, gridcolor="#F3F4F6",
                            zeroline=True, zerolinecolor="#999999",
                            zerolinewidth=1,
                        ),
                        height=240,
                        margin=dict(t=40, b=70, l=70, r=20),
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        font=dict(
                            family="Montserrat, Arial, sans-serif",
                            size=12, color="#333333",
                        ),
                        showlegend=False,
                    )
                    st.plotly_chart(fig_mom, use_container_width=True)

            # ── Chart 2 & 3: State Map + Top Metro Markets (side by side)
            df_states = trends_data["state_data"]
            df_metros = trends_data["top_metros"]
            _state_vol = trends_data.get("state_volume_data")
            _metro_vol = trends_data.get("metro_volume_data")

            _has_state_chart = not df_states.empty
            _has_metro_chart = not df_metros.empty and len(df_metros) > 0

            if _has_state_chart or _has_metro_chart:
                # ── Volume / Growth metric toggle ─────────────────────
                # Growth is only available when state-level monthly data
                # has been loaded (run load_google_trends_state_time.py).
                # We probe the table once and offer the option only if at
                # least one (CIP, state) pair has time-series data — for
                # the user's currently-selected CIPs.
                _state_time_loaded = (
                    search_traffic_state_time_coverage(cip_patterns) > 0
                )
                _metric_opts = ["Volume"]
                if _state_time_loaded:
                    _metric_opts.append("Growth")

                if len(_metric_opts) > 1:
                    _search_metric = st.radio(
                        "Metric",
                        _metric_opts,
                        index=0,
                        horizontal=True,
                        key="search_metric",
                        label_visibility="collapsed",
                    )
                else:
                    _search_metric = "Volume"
                    st.caption(
                        "Tip: load per-state historical Trends "
                        "(`python load_google_trends_state_time.py "
                        f"--cips {','.join(cip_patterns) if cip_patterns else '...'}`) "
                        "to enable a Growth view that compares the most "
                        "recent 12 months against the prior 12 months by state."
                    )

                # Layout — same 3:2 in Volume mode; state-only in Growth.
                if _has_state_chart:
                    col_map, col_bar = st.columns([3, 2])
                else:
                    col_map = None
                    col_bar = st.container()

                if _search_metric == "Growth":
                    _growth_df = run_search_traffic_state_growth(cip_patterns)
                else:
                    _growth_df = None

                # ── Choropleth map ────────────────────────────────────
                if _has_state_chart:
                    with col_map:
                        if _search_metric == "Growth":
                            if _growth_df is None or _growth_df.empty:
                                st.info(
                                    "No state-level historical data is "
                                    "loaded yet for the selected CIP(s). "
                                    "Run `load_google_trends_state_time.py`."
                                )
                            else:
                                fig_growth_map = vi_choropleth(
                                    _growth_df["stabbr"],
                                    _growth_df["pct_change"],
                                    title="Search Interest — 12-mo Rolling Change by State",
                                    colorbar_title="12-mo % change",
                                    hover_format="{:+.1%}",
                                    hover_label="12-mo change",
                                )
                                st.plotly_chart(
                                    fig_growth_map, use_container_width=True,
                                )
                        else:
                            _use_state_vol = (
                                _show_volume
                                and _state_vol is not None
                                and not _state_vol.empty
                            )
                            _map_z = (
                                _state_vol["volume"] if _use_state_vol
                                else df_states["interest"]
                            )
                            _map_locs = (
                                _state_vol["state_abbr"] if _use_state_vol
                                else df_states["state_abbr"]
                            )
                            _map_cbar_title = (
                                "Est. Searches" if _use_state_vol else "Interest"
                            )
                            _map_hover = (
                                "<b>%{location}</b><br>"
                                "Est. Searches: %{z:,.0f}"
                                "<extra></extra>"
                            ) if _use_state_vol else (
                                "<b>%{location}</b><br>"
                                "Interest: %{z:.0f}/100"
                                "<extra></extra>"
                            )
                            _map_title = (
                                "<b>Estimated Search Volume by State</b>"
                                if _use_state_vol
                                else "<b>Search Interest by State</b>"
                            )

                            fig_map = go.Figure(go.Choropleth(
                                locations=_map_locs,
                                z=_map_z,
                                locationmode="USA-states",
                                colorscale=VI_CHOROPLETH_SCALE,
                                marker=dict(line=dict(color="white", width=0.6)),
                                colorbar=dict(
                                    title=_map_cbar_title,
                                    thickness=12,
                                    len=0.6,
                                    tickfont=dict(size=10),
                                ),
                                hovertemplate=_map_hover,
                            ))
                            fig_map.update_layout(
                                title=dict(
                                    text=_map_title,
                                    font=dict(size=14), x=0, xanchor="left",
                                ),
                                geo=dict(
                                    scope="usa",
                                    bgcolor="white",
                                    lakecolor="white",
                                    showlakes=True,
                                    landcolor="#FAFAFA",
                                    projection_type="albers usa",
                                ),
                                height=340,
                                margin=dict(t=50, b=10, l=10, r=10),
                                paper_bgcolor="white",
                                font=dict(
                                    family="Montserrat, Arial, sans-serif",
                                    size=12, color="#333333",
                                ),
                            )
                            st.plotly_chart(fig_map, use_container_width=True)

                # ── Ranking (State / Metro toggle) ────────────────────
                # Same VI-orange ranking helper as Completions/Employment.
                # In Growth mode we skip the State/Metro toggle since
                # metro-level historical Trends data isn't loaded.
                with col_bar:
                    if _search_metric == "Growth":
                        if _growth_df is None or _growth_df.empty:
                            _rank_search_fig = None
                        else:
                            _top = _growth_df.sort_values(
                                "pct_change", ascending=False,
                            ).head(15)
                            _rank_search_fig = vi_ranking_bar(
                                _top["stabbr"],
                                _top["pct_change"],
                                title="Fastest-Growing States (12-mo Δ)",
                                value_label="12-mo change",
                                value_format="{:+.1%}",
                            )
                    else:
                        # Volume — preserve the existing State/Metro toggle.
                        _grain_opts = []
                        if _has_state_chart:
                            _grain_opts.append("State")
                        if _has_metro_chart:
                            _grain_opts.append("Metro")

                        if len(_grain_opts) > 1:
                            _search_grain = st.radio(
                                "Show ranking by:",
                                _grain_opts,
                                index=0,
                                horizontal=True,
                                key="search_rank_grain",
                                label_visibility="collapsed",
                            )
                        else:
                            _search_grain = _grain_opts[0]

                        _use_vol = _show_volume

                        if _search_grain == "State":
                            _vol_ok = (
                                _use_vol and _state_vol is not None
                                and not _state_vol.empty
                            )
                            _src = _state_vol if _vol_ok else df_states
                            _label_col = "state_abbr"
                            _val_col = "volume" if _vol_ok else "interest"
                            _value_label = "Est. Searches" if _vol_ok else "Interest"
                            _value_format = (
                                "{:,.0f}" if _vol_ok else "{:.0f}/100"
                            )
                            _title = (
                                "Top States (by Est. Volume)" if _vol_ok
                                else "Top States (by Interest)"
                            )
                        else:  # Metro
                            _vol_ok = (
                                _use_vol and _metro_vol is not None
                                and not _metro_vol.empty
                            )
                            _src = _metro_vol if _vol_ok else df_metros
                            _label_col = "cbsa_name"
                            _val_col = "volume" if _vol_ok else "interest"
                            _value_label = "Est. Searches" if _vol_ok else "Interest"
                            _value_format = (
                                "{:,.0f}" if _vol_ok else "{:.0f}/100"
                            )
                            _title = (
                                "Top Metros (by Est. Volume)" if _vol_ok
                                else "Top Metros (by Interest)"
                            )

                        _top = (
                            _src.sort_values(_val_col, ascending=False)
                                .head(15)
                        )
                        _rank_search_fig = vi_ranking_bar(
                            _top[_label_col],
                            _top[_val_col],
                            title=_title,
                            value_label=_value_label,
                            value_format=_value_format,
                        )

                    if _rank_search_fig is not None:
                        st.plotly_chart(
                            _rank_search_fig, use_container_width=True,
                        )

            _terms_display = ", ".join(
                f"**{t}**" for t in _search_terms[:5]
            )
            _more_terms = (
                f" (+{len(_search_terms) - 5} more)"
                if len(_search_terms) > 5 else ""
            )
            if _show_volume:
                st.caption(
                    f"Estimated monthly search volumes are calibrated from "
                    f"Google Trends indices using known keyword volumes "
                    f"(anchor: 'nursing degree' = 146K searches, March 2025). "
                    f"Search term(s): {_terms_display}{_more_terms} "
                    f"| Source: Google Trends + keyword volume calibration"
                )
            else:
                st.caption(
                    f"Search interest reflects relative Google search volume "
                    f"(0 = no interest, 100 = peak over period). "
                    f"Search term(s): {_terms_display}{_more_terms} "
                    f"| Source: Google Trends"
                )

    # ── Related Employment by Occupation ─────────────────────────────────────
    st.divider()
    st.subheader("Related Employment by Occupation")
    if _windows.get("oes"):
        _oes_min, _oes_max = _windows["oes"]
        _oes_window = f"{_oes_min} – {_oes_max}"
    else:
        _oes_window = "available years"
    st.caption(
        "Total employment and median annual wages for occupations linked to "
        "the selected CIP code(s) via the NCES CIP-SOC crosswalk. "
        "Pre-2019 BLS data uses SOC 2010 codes, which are bridged to SOC 2018 "
        "via the BLS crossover table; combined codes (e.g. 15-1256) are "
        "remapped to their primary detail code. **Projected CAGR** is the "
        "weighted-average BLS Employment Projections growth rate (current → "
        "+10 yr) across the matched occupations, weighted by latest-year "
        f"employment. **Automation Risk** is the LMI Institute Automation "
        "Exposure Index (2019 OES vintage), a 1–10 score derived from O*NET "
        "ability, work-activity, and work-context attributes; 1 = least "
        "exposed, 10 = most exposed. The aggregate shown is weighted by "
        f"latest-year employment. Coverage: {_oes_window}. "
        "| Sources: BLS Occupational Employment & Wage Statistics (OEWS), "
        "BLS Employment Projections, NCES CIP-to-SOC Crosswalk, "
        "LMI Institute Automation Exposure Index (2019 OES)."
    )

    if all_cips:
        st.info(
            "Employment data is shown when specific CIP code(s) are selected. "
            "Deselect 'All CIP codes' and choose program(s) to see related occupations."
        )
    else:
        # Check if OES tables exist
        _oes_ok = False
        try:
            _conn = get_conn()
            _conn.execute("SELECT 1 FROM oes_employment LIMIT 1")
            _conn.execute("SELECT 1 FROM cip_soc_crosswalk LIMIT 1")
            _conn.close()
            _oes_ok = True
        except Exception:
            pass

        if not _oes_ok:
            st.warning(
                "Employment data not loaded. Run `python load_oes_data.py` to download "
                "BLS OES data and the CIP-SOC crosswalk."
            )
        else:
            with st.spinner("Querying employment data..."):
                df_emp = run_employment_query(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )

            if df_emp.empty:
                st.info("No employment data found for the selected program(s) and geography.")
            else:
                # Fetch projected growth for the same occupations
                soc_codes_for_proj = tuple(df_emp["occ_code"].unique())
                df_proj = get_employment_projections(
                    soc_codes=soc_codes_for_proj,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )

                # Aggregate total employment across all occupations by year
                emp_by_year = df_emp.groupby("year")["tot_emp"].sum().reset_index()
                emp_by_year = emp_by_year.sort_values("year")

                latest_emp_year = df_emp["year"].max()

                # Compute weighted average projected CAGR across all related occupations
                proj_avg_cagr = None
                if not df_proj.empty and "cagr" in df_proj.columns:
                    # Weight by latest-year employment
                    latest_emp = df_emp[df_emp["year"] == latest_emp_year][["occ_code", "tot_emp"]]
                    proj_merged = df_proj.merge(latest_emp, on="occ_code", how="inner")
                    proj_merged = proj_merged.dropna(subset=["cagr", "tot_emp"])
                    if not proj_merged.empty and proj_merged["tot_emp"].sum() > 0:
                        proj_avg_cagr = (
                            (proj_merged["cagr"] * proj_merged["tot_emp"]).sum()
                            / proj_merged["tot_emp"].sum()
                        )

                # Employment metrics
                emp_years = sorted(emp_by_year["year"].unique())
                if len(emp_years) >= 2:
                    emp_latest = int(emp_by_year[emp_by_year["year"] == emp_years[-1]]["tot_emp"].iloc[0])
                    emp_first = int(emp_by_year[emp_by_year["year"] == emp_years[0]]["tot_emp"].iloc[0])
                    emp_n = emp_years[-1] - emp_years[0]
                    emp_cagr = ((emp_latest / emp_first) ** (1 / emp_n) - 1) if emp_first > 0 and emp_n > 0 else None

                    # Median wage
                    latest_wages = df_emp[df_emp["year"] == latest_emp_year]
                    wage_weighted = latest_wages.dropna(subset=["a_median", "tot_emp"])
                    if not wage_weighted.empty:
                        avg_median_wage = int(
                            (wage_weighted["a_median"] * wage_weighted["tot_emp"]).sum()
                            / wage_weighted["tot_emp"].sum()
                        )
                    else:
                        avg_median_wage = None

                    n_occs = df_emp["occ_code"].nunique()

                    # 3-year CAGR (mirrors completions section)
                    emp_3yr_cagr = None
                    emp_yr_3ago = latest_emp_year - 3
                    if emp_yr_3ago in emp_by_year["year"].values:
                        emp_3ago_val = int(emp_by_year[emp_by_year["year"] == emp_yr_3ago]["tot_emp"].iloc[0])
                        if emp_3ago_val > 0:
                            emp_3yr_cagr = (emp_latest / emp_3ago_val) ** (1 / 3) - 1

                    # Employment-weighted automation risk (LMII 2019 OES index, 1-10)
                    df_risk = get_automation_risk(
                        tuple(df_emp["occ_code"].unique())
                    )
                    emp_weighted_risk = None
                    risk_coverage = 0.0
                    if not df_risk.empty:
                        latest_with_risk = (
                            latest_wages[["occ_code", "tot_emp"]]
                            .merge(df_risk, on="occ_code", how="left")
                            .dropna(subset=["risk_score"])
                        )
                        if not latest_with_risk.empty and latest_with_risk["tot_emp"].sum() > 0:
                            emp_weighted_risk = (
                                (latest_with_risk["risk_score"] * latest_with_risk["tot_emp"]).sum()
                                / latest_with_risk["tot_emp"].sum()
                            )
                            risk_coverage = (
                                latest_with_risk["tot_emp"].sum()
                                / latest_wages["tot_emp"].sum()
                            )

                    em1, em2, em3, em4, em5, em6 = st.columns(6)
                    em1.metric(
                        f"{latest_emp_year} Related Employment",
                        f"{emp_latest:,}",
                    )
                    em2.metric(
                        f"Long-Term CAGR ({emp_years[0]} → {emp_years[-1]})",
                        f"{emp_cagr:+.1%}" if emp_cagr is not None else "N/A",
                    )
                    em3.metric(
                        f"Post-COVID CAGR ({emp_yr_3ago} → {latest_emp_year})",
                        f"{emp_3yr_cagr:+.1%}" if emp_3yr_cagr is not None else "N/A",
                    )
                    em4.metric(
                        f"Projected CAGR ({latest_emp_year} → 2029)",
                        f"{proj_avg_cagr:+.1%}" if proj_avg_cagr is not None else "N/A",
                    )
                    em5.metric(
                        f"Wtd. Median Wage ({latest_emp_year})",
                        f"${avg_median_wage:,}" if avg_median_wage else "N/A",
                    )
                    if emp_weighted_risk is not None:
                        em6.metric(
                            "Automation Risk (1–10)",
                            f"{emp_weighted_risk:.1f}",
                            help=(
                                "Employment-weighted LMI Institute Automation "
                                "Exposure Index across the related occupations. "
                                "1 = least exposed, 10 = most exposed. The score "
                                "ranks each occupation by its mix of O*NET "
                                "abilities, work activities, and work contexts "
                                "(routine/manual tasks score higher; abstract, "
                                "cognitive, and interpersonal tasks score lower). "
                                "**Not predictive** — a high score means the "
                                "task content is theoretically automatable, not "
                                "that workers are at imminent risk; cost, "
                                "policy, public acceptance, and workforce "
                                "factors all moderate real-world adoption. "
                                f"Covers {risk_coverage:.0%} of latest-year "
                                "related employment. Source: LMI Institute, "
                                "Automation Exposure Index (2019 OES vintage, "
                                "lmiontheweb.org)."
                            ),
                        )
                    else:
                        em6.metric("Automation Risk (1–10)", "N/A")

                # Line chart: aggregated employment across all related occupations
                if not emp_by_year.empty:
                    fig_emp = go.Figure()
                    # Historical line (solid) with data labels
                    fig_emp.add_trace(go.Scatter(
                        x=emp_by_year["year"],
                        y=emp_by_year["tot_emp"],
                        mode="lines+markers+text",
                        name="Total Related Employment",
                        line=dict(width=2.5, color=EMPLOYMENT_COLORS[0]),
                        marker=dict(size=7),
                        textposition="top center",
                        texttemplate="%{y:,.0f}",
                        textfont=dict(size=10),
                        hovertemplate="<b>%{x}</b><br>%{y:,.0f} employed<extra></extra>",
                    ))

                    # Add dotted projection line using weighted avg CAGR
                    emp_tick_years = sorted(emp_by_year["year"].unique())
                    if proj_avg_cagr is not None:
                        proj_target_year = 2029
                        base_val = emp_by_year[
                            emp_by_year["year"] == latest_emp_year
                        ]["tot_emp"].iloc[0]
                        proj_years = list(range(latest_emp_year, proj_target_year + 1))
                        proj_vals = [
                            base_val * (1 + proj_avg_cagr) ** (y - latest_emp_year)
                            for y in proj_years
                        ]

                        # Faint gray shaded region over the projection area
                        fig_emp.add_vrect(
                            x0=latest_emp_year + 0.5,
                            x1=proj_target_year + 0.5,
                            fillcolor="#E5E7EB",
                            opacity=0.3,
                            layer="below",
                            line_width=0,
                        )

                        # Projection line with diamond markers and data labels
                        fig_emp.add_trace(go.Scatter(
                            x=[latest_emp_year] + proj_years[1:],
                            y=[base_val] + proj_vals[1:],
                            mode="lines+markers+text",
                            name="Projected",
                            line=dict(dash="dash", width=2.5, color="rgba(15, 134, 193, 0.45)"),
                            marker=dict(size=7, symbol="diamond", color="rgba(15, 134, 193, 0.45)"),
                            text=[""] + [f"{int(v):,}" for v in proj_vals[1:]],
                            textposition="top center",
                            textfont=dict(size=10, color="rgba(15, 134, 193, 0.6)"),
                            hovertemplate="<b>%{x} (projected)</b><br>%{y:,.0f} employed<extra></extra>",
                        ))
                        emp_tick_years = sorted(set(emp_tick_years) | set(proj_years[1:]))

                    fig_emp.update_layout(
                        title="<b>Employment Trend: All Related Occupations</b>",
                        xaxis=dict(
                            title="",
                            tickmode="array",
                            tickvals=emp_tick_years,
                            ticktext=[str(y) for y in emp_tick_years],
                            tickangle=-30,
                            showgrid=True,
                            gridcolor="#F3F4F6",
                        ),
                        yaxis=dict(
                            title="Total Employment",
                            tickformat=",",
                            showgrid=True,
                            gridcolor="#F3F4F6",
                            rangemode="tozero",
                        ),
                        showlegend=False,
                        height=480,
                        margin=dict(t=80, b=60, l=70, r=20),
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        font=dict(family="Montserrat, Arial, sans-serif", size=12, color="#333333"),
                    )
                    st.plotly_chart(fig_emp, use_container_width=True)

                    # ── Employment YoY change bar chart ───────────────────────
                    emp_yoy = emp_by_year.copy().sort_values("year")
                    emp_yoy["yoy"] = emp_yoy["tot_emp"].pct_change() * 100
                    emp_yoy = emp_yoy.dropna(subset=["yoy"])
                    emp_yoy["color"] = emp_yoy["yoy"].apply(
                        lambda v: "#16a34a" if v >= 0 else "#dc2626"
                    )
                    emp_yoy["text"] = emp_yoy["yoy"].apply(
                        lambda v: f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%"
                    )

                    fig_emp_yoy = go.Figure(go.Bar(
                        x=emp_yoy["year"],
                        y=emp_yoy["yoy"],
                        marker_color=emp_yoy["color"],
                        text=emp_yoy["text"],
                        textposition="outside",
                        textfont=dict(size=10, family="Montserrat, Arial, sans-serif", color="#333333"),
                        hovertemplate="%{text}<extra></extra>",
                        name="Actual",
                        showlegend=False,
                    ))

                    # Projected YoY bars
                    if proj_avg_cagr is not None:
                        proj_chain = [base_val] + proj_vals[1:]
                        proj_yoy_years = proj_years[1:]
                        proj_yoy_vals = [
                            ((proj_chain[i + 1] - proj_chain[i]) / proj_chain[i] * 100)
                            if proj_chain[i] > 0 else 0
                            for i in range(len(proj_chain) - 1)
                        ]
                        proj_yoy_colors = [
                            "rgba(22, 163, 74, 0.35)" if v >= 0 else "rgba(220, 38, 38, 0.35)"
                            for v in proj_yoy_vals
                        ]
                        proj_yoy_text = [
                            f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%" for v in proj_yoy_vals
                        ]
                        proj_yoy_textcolors = [
                            "rgba(22, 163, 74, 0.6)" if v >= 0 else "rgba(220, 38, 38, 0.6)"
                            for v in proj_yoy_vals
                        ]
                        fig_emp_yoy.add_trace(go.Bar(
                            x=proj_yoy_years,
                            y=proj_yoy_vals,
                            marker_color=proj_yoy_colors,
                            text=proj_yoy_text,
                            textposition="outside",
                            textfont=dict(
                                size=10,
                                family="Montserrat, Arial, sans-serif",
                                color=proj_yoy_textcolors,
                            ),
                            hovertemplate="%{text} (projected)<extra></extra>",
                            name="Projected",
                            showlegend=False,
                        ))

                        fig_emp_yoy.add_vrect(
                            x0=latest_emp_year + 0.5,
                            x1=proj_yoy_years[-1] + 0.5,
                            fillcolor="#E5E7EB",
                            opacity=0.3,
                            layer="below",
                            line_width=0,
                        )

                    emp_yoy_tick_years = sorted(set(emp_tick_years) | set(emp_yoy["year"].unique()))
                    if proj_avg_cagr is not None:
                        emp_yoy_tick_years = sorted(set(emp_yoy_tick_years) | set(proj_yoy_years))

                    fig_emp_yoy.update_layout(
                        xaxis=dict(
                            tickmode="array",
                            tickvals=emp_yoy_tick_years,
                            ticktext=[str(y) for y in emp_yoy_tick_years],
                            tickangle=-30,
                            showgrid=True,
                            gridcolor="#F3F4F6",
                            gridwidth=1,
                            range=[emp_yoy_tick_years[0] - 0.5, emp_yoy_tick_years[-1] + 0.5],
                        ),
                        yaxis=dict(
                            ticksuffix="%",
                            tickformat=".1f",
                            showgrid=True,
                            gridcolor="#F3F4F6",
                            gridwidth=1,
                            zeroline=True,
                            zerolinecolor="#999999",
                            zerolinewidth=1,
                        ),
                        title=dict(text="Year-over-Year % Change", font=dict(size=13), x=0, xanchor="left"),
                        height=220,
                        margin=dict(t=40, b=60, l=70, r=20),
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        font=dict(family="Montserrat, Arial, sans-serif", size=12, color="#333333"),
                        showlegend=False,
                    )
                    st.plotly_chart(fig_emp_yoy, use_container_width=True)

                    # ── Related Employment geographic distribution ────────────
                    _emp_socs = tuple(df_emp["occ_code"].unique())
                    _emp_base_year_growth = int(latest_emp_year) - 3
                    st.markdown(
                        f"<div class='vi-map-caption'>Geographic distribution · "
                        f"{int(latest_emp_year)} related employment</div>",
                        unsafe_allow_html=True,
                    )
                    _emp_metric = st.radio(
                        "Metric",
                        ["Volume", "Growth", "Projected"],
                        index=0,
                        horizontal=True,
                        key="emp_metric",
                        label_visibility="collapsed",
                    )

                    # ── Per-state dataframe driving the choropleth ──────
                    if _emp_metric == "Volume":
                        _emp_state_df = run_employment_by_state_query(
                            soc_codes=_emp_socs,
                            year=int(latest_emp_year),
                        )
                        _emp_state_value_col = "tot_emp"
                        _emp_value_label = "Employed"
                        _emp_value_format = "{:,.0f}"
                        _emp_map_title = (
                            f"Related Employment by State — {int(latest_emp_year)}"
                        )
                        _emp_rank_title = f"Top States — {int(latest_emp_year)}"
                    elif _emp_metric == "Growth":
                        _emp_state_df = run_employment_state_cagr(
                            soc_codes=_emp_socs,
                            base_year=_emp_base_year_growth,
                            end_year=int(latest_emp_year),
                        )
                        _emp_state_value_col = "cagr"
                        _emp_value_label = "Post-COVID CAGR"
                        _emp_value_format = "{:+.1%}"
                        _emp_map_title = (
                            f"Post-COVID Employment CAGR — "
                            f"{_emp_base_year_growth} → {int(latest_emp_year)}"
                        )
                        _emp_rank_title = "Fastest-Growing States"
                    else:  # Projected
                        _emp_state_df = run_employment_projection_state(_emp_socs)
                        _emp_state_value_col = "cagr"
                        _emp_value_label = "Projected CAGR"
                        _emp_value_format = "{:+.1%}"
                        _emp_map_title = "Projected Annual Change — Related Occupations"
                        _emp_rank_title = "Top States by Projected Growth"

                    _emp_has_state = (
                        not _emp_state_df.empty
                        and "stabbr" in _emp_state_df.columns
                        and _emp_state_df[_emp_state_value_col].notna().sum() >= 5
                    )

                    if _emp_has_state:
                        _emp_state_df = _emp_state_df.dropna(
                            subset=[_emp_state_value_col]
                        )
                        _emp_map_col, _emp_rank_col = st.columns([3, 2])
                        with _emp_map_col:
                            fig_emp_map = vi_choropleth(
                                _emp_state_df["stabbr"],
                                _emp_state_df[_emp_state_value_col],
                                title=_emp_map_title,
                                colorbar_title=_emp_value_label,
                                hover_format=_emp_value_format,
                                hover_label=_emp_value_label,
                            )
                            st.plotly_chart(fig_emp_map, use_container_width=True)

                        with _emp_rank_col:
                            _emp_rank_grain = st.radio(
                                "Show ranking by:",
                                ["State", "Metro"],
                                index=0,
                                horizontal=True,
                                key="emp_rank_grain",
                                label_visibility="collapsed",
                            )
                            if _emp_rank_grain == "State":
                                _emp_rank_df = _emp_state_df.sort_values(
                                    _emp_state_value_col,
                                    ascending=False,
                                    na_position="last",
                                ).head(15)
                                _emp_rank_fig = vi_ranking_bar(
                                    _emp_rank_df["stabbr"],
                                    _emp_rank_df[_emp_state_value_col],
                                    title=_emp_rank_title,
                                    value_label=_emp_value_label,
                                    value_format=_emp_value_format,
                                )
                            else:  # Metro
                                if _emp_metric == "Volume":
                                    _df_metros_emp = run_employment_by_metro_query(
                                        soc_codes=_emp_socs,
                                        year=int(latest_emp_year),
                                        top_n=15,
                                    )
                                    _emp_metro_value_col = "tot_emp"
                                elif _emp_metric == "Growth":
                                    _df_metros_emp = run_employment_metro_cagr(
                                        soc_codes=_emp_socs,
                                        base_year=_emp_base_year_growth,
                                        end_year=int(latest_emp_year),
                                        top_n=15,
                                    )
                                    _emp_metro_value_col = "cagr"
                                else:  # Projected
                                    _df_metros_emp = run_employment_projection_metro(
                                        _emp_socs, top_n=15,
                                    )
                                    _emp_metro_value_col = "cagr"

                                if _df_metros_emp is None or _df_metros_emp.empty:
                                    st.info("No metro-level data for this metric.")
                                    _emp_rank_fig = None
                                else:
                                    _emp_rank_fig = vi_ranking_bar(
                                        _df_metros_emp["cbsa_name"],
                                        _df_metros_emp[_emp_metro_value_col],
                                        title=_emp_rank_title.replace(
                                            "States", "Metros"
                                        ),
                                        value_label=_emp_value_label,
                                        value_format=_emp_value_format,
                                    )
                            if _emp_rank_fig is not None:
                                st.plotly_chart(_emp_rank_fig, use_container_width=True)

                # Footnote listing the occupations included in the aggregate
                occ_list = (
                    df_emp[["occ_code", "occ_title"]]
                    .drop_duplicates("occ_code")
                    .sort_values("occ_code")
                )
                # Attach automation risk (1-10) per occupation when available
                occ_list = occ_list.merge(
                    get_automation_risk(tuple(occ_list["occ_code"])),
                    on="occ_code",
                    how="left",
                )

                def _risk_tag(score):
                    if pd.isna(score):
                        return ""
                    s = int(score)
                    return f" · automation risk **{s}/10**"

                occ_bullets = "  \n".join(
                    f"- **{row.occ_code}** {row.occ_title}{_risk_tag(row.risk_score)}"
                    for row in occ_list.itertuples()
                )
                st.caption(
                    f"Aggregate includes **{len(occ_list)}** related occupations "
                    f"(SOC codes mapped via CIP-SOC crosswalk · automation risk "
                    f"per LMI Institute 2019 OES Automation Exposure Index, "
                    f"1 = low / 10 = high):  \n{occ_bullets}"
                )
                # Automation-risk methodology footnote
                _risk_n = occ_list["risk_score"].notna().sum()
                if _risk_n > 0:
                    st.caption(
                        ":information_source: **About the automation-risk score.** "
                        "The LMI Institute Automation Exposure Index ranks each "
                        "occupation on a 1–10 scale (1 = least exposed, 10 = "
                        "most exposed) using O*NET data on abilities, work "
                        "activities, and work contexts. Tasks that are routine "
                        "or manual score higher; tasks that are abstract, "
                        "cognitive, or interpersonal (e.g. empathy, creativity, "
                        "judgment) score lower. The index is built on the 2019 "
                        "OES occupational structure, so a small number of newer "
                        "BLS combined or split SOC codes have no score and are "
                        "omitted from the bullet list above. "
                        "**The score is not predictive.** A high score means "
                        "the task content is theoretically automatable, not "
                        "that workers are in imminent danger of displacement — "
                        "cost and complexity, policy and regulation, public "
                        "acceptance, and workforce dynamics all moderate "
                        "real-world adoption, and in many occupations "
                        "automation increases productivity rather than "
                        "eliminating jobs. "
                        "| Source: LMI Institute, Automation Exposure Index "
                        "(2019 OES vintage), [lmiontheweb.org]"
                        "(https://www.lmiontheweb.org)."
                    )


    # ── By Institution ────────────────────────────────────────────────────────
    st.divider()
    st.subheader("Completions by Institution")
    st.caption(
        "Annual completions per institution for the selected filters, "
        "sorted by latest-year volume. "
        "**Long-Term CAGR** = compound annual growth rate across the full "
        "available data window (≈10 one-year intervals between bookend years). "
        "**Post-COVID CAGR** = the same formula applied only to the three "
        "most recent years (starting AY 2020-21, the first full pandemic "
        "academic year), isolating current momentum from pre-pandemic trend. "
        "Both use `(end / start)^(1/n) − 1`. Institutions with zero "
        "completions in either bookend year show no CAGR. "
        "| Source: NCES IPEDS Completions Survey."
    )

    if df_inst.empty:
        st.info("No institution-level data for these filters.")
    else:
        # Get latest metadata per unitid (name may change across years)
        meta = (
            df_inst.sort_values("year")
            .groupby("unitid")[["instnm", "city", "stabbr", "control_name"]]
            .last()
            .reset_index()
        )

        # Pivot on unitid only so name changes don't split rows
        pivot = df_inst.pivot_table(
            index="unitid",
            columns="year",
            values="completions",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
        pivot = pivot.merge(meta, on="unitid", how="left")
        pivot.columns.name = None
        yr_cols = sorted([c for c in pivot.columns if isinstance(c, int)])

        # CAGR per institution (stored as %, e.g. 2.3 means 2.3%)
        first_col, last_col = yr_cols[0], yr_cols[-1]
        n_years = last_col - first_col
        col_3ago = last_col - 3

        def inst_cagr(row, start_col, n):
            fv, lv = row[start_col], row[last_col]
            if fv > 0 and lv > 0 and n > 0:
                return ((lv / fv) ** (1 / n) - 1) * 100
            return None

        if col_3ago in yr_cols:
            pivot["Post-COVID CAGR"] = pivot.apply(lambda r: inst_cagr(r, col_3ago, 3), axis=1)
        pivot["Long-Term CAGR"] = pivot.apply(lambda r: inst_cagr(r, first_col, n_years), axis=1)
        pivot = pivot.rename(columns={y: yr_label_short(y) for y in yr_cols})
        last_yr_short = yr_label_short(last_col)
        pivot = pivot.sort_values(last_yr_short, ascending=False, na_position="last").reset_index(drop=True)
        control_map = {"Public": "Public", "Private nonprofit": "Private", "Private for-profit": "For-Profit"}
        pivot["control_name"] = pivot["control_name"].map(control_map).fillna(pivot["control_name"])
        pivot["city"] = pivot["city"] + ", " + pivot["stabbr"]
        pivot = pivot.drop(columns=["unitid", "stabbr"])
        pivot = pivot.rename(columns={"instnm": "Institution", "city": "City", "control_name": "Control"})
        cagr_cols = [c for c in ["Post-COVID CAGR", "Long-Term CAGR"] if c in pivot.columns]
        yr_short_labels = [yr_label_short(y) for y in yr_cols]
        pivot = pivot[["Institution", "City", "Control"] + yr_short_labels + cagr_cols]

        n_institutions = len(pivot)
        st.caption(f"{n_institutions:,} institutions reported completions for these filters")

        # Smaller font for the institution table
        st.markdown(
            "<style>div[data-testid='stDataFrame'] table {font-size: 0.78rem;}</style>",
            unsafe_allow_html=True,
        )

        # Compute column widths so the table fits without horizontal scroll.
        n_yr = len(yr_short_labels)
        n_cagr = len(cagr_cols)
        yr_col_w = 62
        cagr_col_w = 72
        control_col_w = 68
        fixed_w = n_yr * yr_col_w + n_cagr * cagr_col_w + control_col_w
        remaining = max(400, 1100 - fixed_w)
        inst_w = int(remaining * 0.6)
        city_w = remaining - inst_w

        col_cfg = {
            "Institution": st.column_config.TextColumn("Institution", width=inst_w),
            "City": st.column_config.TextColumn("City", width=city_w),
            "Control": st.column_config.TextColumn("Control", width=control_col_w),
            **{
                yr_label_short(y): st.column_config.NumberColumn(
                    yr_label_short(y), format="%,d", width=yr_col_w,
                )
                for y in yr_cols
            },
        }
        if "Post-COVID CAGR" in cagr_cols:
            col_cfg["Post-COVID CAGR"] = st.column_config.NumberColumn(
                f"Post-COVID CAGR ({yr_label_short(col_3ago)} → {yr_label_short(last_col)})",
                format="%.1f%%",
                width=cagr_col_w,
            )
        if "Long-Term CAGR" in cagr_cols:
            col_cfg["Long-Term CAGR"] = st.column_config.NumberColumn(
                f"Long-Term CAGR ({yr_label_short(first_col)} → {yr_label_short(last_col)})",
                format="%.1f%%",
                width=cagr_col_w,
            )

        st.dataframe(
            pivot,
            use_container_width=True,
            hide_index=True,
            column_config=col_cfg,
        )


    # ── Graduate Outcomes (College Scorecard) ────────────────────────────────
    st.divider()
    st.subheader("Graduate Outcomes")
    st.caption(
        "**Median Earnings (4yr)** = median annual earnings of graduates "
        "measured ~4 years after program completion (most-recent pooled cohort, "
        "Title-IV-aided completers only, IRS earnings). "
        "**Median Debt** = median total federal student-loan debt at "
        "completion. **Debt/Earnings** = the ratio of those two — lower is "
        "better. Outcomes are reported at the **4-digit CIP** level, so a "
        "row may pool several closely-related 6-digit programs. "
        "| Source: U.S. Department of Education College Scorecard, "
        "Field-of-Study (Most-Recent-Cohorts) file."
    )

    _scorecard_ok = False
    try:
        _sc_conn = get_conn()
        _sc_conn.execute("SELECT 1 FROM college_scorecard LIMIT 1")
        _sc_conn.close()
        _scorecard_ok = True
    except Exception:
        pass

    if not _scorecard_ok:
        st.info("Scorecard outcomes data not available.")
    elif all_cips:
        st.info(
            "Graduate outcomes data is shown when specific CIP code(s) are selected. "
            "Deselect **All CIP codes** and choose program(s) to see outcomes."
        )
    else:
        with st.spinner("Querying graduate outcomes..."):
            df_sc = run_scorecard_query(
                cip_patterns=cip_patterns,
                awlevels=selected_awlevels,
                geo_key=geo_key,
                geo_values=tuple(geo_values),
            )

        if df_sc.empty:
            st.info(
                "No graduate outcomes data available for these filters. "
                "Scorecard data is reported at the 4-digit CIP level and may not "
                "cover all institution / program combinations."
            )
        else:
            # Deduplicate across distance modes (keep highest earnings per institution)
            df_sc = (
                df_sc
                .sort_values("earn_mdn_4yr", ascending=False)
                .drop_duplicates(subset=["unitid"], keep="first")
            )

            # ── Metrics row ───────────────────────────────────────────────
            n_inst = df_sc["unitid"].nunique()
            med_earn = df_sc["earn_mdn_4yr"].median()

            _sc_debt = df_sc.dropna(subset=["debt_all_stgp_eval_mdn"])
            med_debt = _sc_debt["debt_all_stgp_eval_mdn"].median() if not _sc_debt.empty else None

            # Compute D/E from the displayed medians so the ratio is consistent
            # with the earnings and debt metrics shown (ratio-of-medians, not
            # median-of-ratios which can diverge due to Simpson's-paradox effects).
            if med_earn and med_earn > 0 and med_debt is not None:
                med_dte = med_debt / med_earn
            else:
                med_dte = None

            sc1, sc2, sc3, sc4 = st.columns(4)
            sc1.metric(
                "Median Earnings (4yr Post-Grad)",
                f"${med_earn:,.0f}" if med_earn else "N/A",
            )
            sc2.metric(
                "Median Debt at Completion",
                f"${med_debt:,.0f}" if med_debt else "N/A",
            )
            sc3.metric(
                "Median Debt-to-Earnings",
                f"{med_dte:.2f}x" if med_dte else "N/A",
            )
            sc4.metric("Institutions with Data", f"{n_inst:,}")

            # ── Detail table ──────────────────────────────────────────────
            sc_display = df_sc.rename(columns={
                "instnm": "Institution",
                "city": "City",
                "control_name": "Control",
                "earn_mdn_4yr": "Median Earnings (4yr)",
                "debt_all_stgp_eval_mdn": "Median Debt",
                "debt_to_earnings": "Debt/Earnings",
            })
            sc_display = sc_display.sort_values(
                "Debt/Earnings", ascending=True, na_position="last"
            )
            sc_display = sc_display[
                ["Institution", "City", "Control",
                 "Median Earnings (4yr)", "Median Debt", "Debt/Earnings"]
            ].reset_index(drop=True)

            # Check if any selected 6-digit CIPs share the same 4-digit prefix
            # (Scorecard data is at 4-digit granularity)
            if cip_patterns:
                _sc_4digit = {p[:5] for p in cip_patterns if "%" not in p}
                _sc_6digit = {p for p in cip_patterns if "%" not in p and len(p) > 5}
                if len(_sc_6digit) > 0 and len(_sc_4digit) < len(_sc_6digit):
                    st.caption(
                        f":information_source: College Scorecard reports outcomes at the 4-digit CIP level "
                        f"(e.g. {sorted(_sc_4digit)[0]}), so results may include related programs "
                        f"that share the same prefix."
                    )

            st.caption(f"{len(sc_display):,} program–institution combinations with earnings data")

            sc_col_cfg = {
                "Institution": st.column_config.TextColumn("Institution", width=250),
                "City": st.column_config.TextColumn("City", width=160),
                "Control": st.column_config.TextColumn("Control", width=85),
                "Median Earnings (4yr)": st.column_config.NumberColumn(
                    "Median Earnings (4yr)", format="$%,.0f", width=155,
                ),
                "Median Debt": st.column_config.NumberColumn(
                    "Median Debt", format="$%,.0f", width=120,
                ),
                "Debt/Earnings": st.column_config.NumberColumn(
                    "Debt/Earnings", format="%.2fx", width=115,
                ),
            }

            st.dataframe(
                sc_display,
                use_container_width=True,
                hide_index=True,
                column_config=sc_col_cfg,
                height=min(len(sc_display) * 35 + 40, 600),
            )

    # ── Distance Education Programs ──────────────────────────────────────────
    # Hidden from UI via SHOW_DISTANCE_EDUCATION_UI; backend retained.
    if SHOW_DISTANCE_EDUCATION_UI:
        st.divider()
        st.subheader("Distance Education Programs")
        if _windows.get("dep"):
            _dep_min, _dep_max = _windows["dep"]
            _dep_window = f"AY {_ay_label(_dep_min)} – AY {_ay_label(_dep_max)}"
        else:
            _dep_window = "available years"
        st.caption(
            "Counts of distinct programs offered (institution × CIP × award level) "
            "and the share available via distance education. "
            "**DE Programs (Any)** = programs that can be completed entirely or "
            "partially via distance education (combines IPEDS DEP fields *DE* "
            "and *DE_SOME*). 6-digit CIP summaries only — 2-digit rollup rows are "
            "excluded to prevent double-counting. "
            f"Coverage: {_dep_window}. "
            "| Source: IPEDS Completions Distance-Education Programs (DEP) survey."
        )

        _dep_ok = False
        try:
            _dep_conn = get_conn()
            _dep_conn.execute("SELECT 1 FROM completions_dep LIMIT 1")
            _dep_conn.close()
            _dep_ok = True
        except Exception:
            pass

        if not _dep_ok:
            st.info("Distance education program data not loaded.")
        elif all_cips:
            st.info(
                "Distance education data is shown when specific CIP code(s) are "
                "selected. Deselect **All CIP codes** and choose program(s)."
            )
        else:
            with st.spinner("Querying distance education data..."):
                df_dep = run_dep_query(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )

            if df_dep is None or df_dep.empty:
                st.info("No distance education program data for these filters.")
            else:
                # ── Metrics ──────────────────────────────────────────────
                _dep_latest = df_dep.iloc[-1]
                _dep_earliest = df_dep.iloc[0]
                _latest_yr = int(_dep_latest["year"])
                _earliest_yr = int(_dep_earliest["year"])

                _total_progs = int(_dep_latest["programs"])
                _de_any = int(_dep_latest["programs_de_any"])
                _pct_de = _dep_latest["pct_de_any"]

                # Change in DE % over full period
                _pct_de_first = _dep_earliest["pct_de_any"]
                _pct_change = (
                    _pct_de - _pct_de_first
                    if pd.notna(_pct_de) and pd.notna(_pct_de_first) else None
                )

                d1, d2, d3, d4 = st.columns(4)
                d1.metric(
                    f"Programs Offered ({yr_label(_latest_yr)})",
                    f"{_total_progs:,}",
                )
                d2.metric(
                    "DE Programs (Any)",
                    f"{_de_any:,}",
                    help="Programs completable entirely or partially "
                         "via distance education.",
                )
                d3.metric(
                    "DE Share",
                    f"{_pct_de:.1f}%",
                )
                d4.metric(
                    f"DE Share Change ({yr_label(_earliest_yr)}-{yr_label(_latest_yr)})",
                    f"{_pct_change:+.1f} pp" if _pct_change is not None else "N/A",
                    help="Percentage point change in DE share over the period.",
                )

                # ── Dual-axis chart: programs (bars) + DE % (line) ───────
                fig_dep = make_subplots(specs=[[{"secondary_y": True}]])

                fig_dep.add_trace(go.Bar(
                    x=df_dep["year"].apply(yr_label),
                    y=df_dep["programs"],
                    name="Total Programs",
                    marker=dict(color="rgba(15, 134, 193, 0.25)"),
                    hovertemplate="%{y:,} programs<extra></extra>",
                ), secondary_y=False)

                fig_dep.add_trace(go.Bar(
                    x=df_dep["year"].apply(yr_label),
                    y=df_dep["programs_de_any"],
                    name="DE Programs (Any)",
                    marker=dict(color="rgba(139, 92, 246, 0.6)"),
                    hovertemplate="%{y:,} DE programs<extra></extra>",
                ), secondary_y=False)

                fig_dep.add_trace(go.Scatter(
                    x=df_dep["year"].apply(yr_label),
                    y=df_dep["pct_de_any"],
                    name="DE Share %",
                    mode="lines+markers",
                    line=dict(width=2.5, color="#f26822"),
                    marker=dict(size=6, color="#f26822"),
                    hovertemplate="%{y:.1f}%<extra></extra>",
                ), secondary_y=True)

                fig_dep.update_yaxes(
                    title_text="Programs Offered",
                    showgrid=True, gridcolor="#F3F4F6", gridwidth=1,
                    rangemode="tozero",
                    secondary_y=False,
                )
                fig_dep.update_yaxes(
                    title_text="DE Share (%)",
                    showgrid=False,
                    rangemode="tozero",
                    ticksuffix="%",
                    secondary_y=True,
                )
                fig_dep.update_layout(
                    title=dict(
                        text="<b>Distance Education Programs Over Time</b>",
                        font=dict(size=15), x=0, xanchor="left",
                    ),
                    xaxis=dict(title="", showgrid=False),
                    height=400,
                    margin=dict(t=60, b=40, l=60, r=60),
                    plot_bgcolor="white",
                    paper_bgcolor="white",
                    font=dict(
                        family="Montserrat, Arial, sans-serif",
                        size=12, color="#333333",
                    ),
                    barmode="overlay",
                    showlegend=True,
                    legend=dict(
                        orientation="h", yanchor="bottom", y=1.02,
                        xanchor="left", x=0, font=dict(size=11),
                    ),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_dep, use_container_width=True)

                # ── DE Share by State (latest year) ──────────────────────
                _dep_latest_yr = int(_dep_latest["year"])
                _df_dep_states = run_dep_by_state_query(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    year=_dep_latest_yr,
                )
                # Only show if we have at least 5 states reporting and >= 1 DE program
                if (
                    not _df_dep_states.empty
                    and len(_df_dep_states) >= 5
                    and _df_dep_states["programs_de_any"].sum() > 0
                ):
                    fig_dep_map = vi_choropleth(
                        _df_dep_states["stabbr"],
                        _df_dep_states["pct_de_any"],
                        title=f"Distance-Education Share by State — {yr_label(_dep_latest_yr)}",
                        colorbar_title="DE %",
                        hover_format="{:.1f}%",
                        hover_label="DE share",
                    )
                    st.plotly_chart(fig_dep_map, use_container_width=True)

    # ── Job Posting Trends (Coresignal) ── hidden from UI; backend retained ──
    # The Coresignal pipeline (queries, title resolution, geography mapping)
    # stays loaded so it can be exercised by other code paths or re-enabled
    # by flipping SHOW_JOB_POSTINGS_UI at the top of this file. None of the
    # rendering below executes while the flag is False.
    try:
        _cs_api_key = st.secrets["coresignal"]["api_key"]
    except (KeyError, FileNotFoundError):
        _cs_api_key = ""

    _is_admin = True

    if SHOW_JOB_POSTINGS_UI and _cs_api_key and _is_admin:
        st.divider()
        st.subheader("Job Posting Trends")
        st.caption(
            "Monthly counts of **newly created** job postings (not the running "
            "total of active postings) for occupations linked to the selected "
            "program(s). Postings are matched by occupation title — only "
            "specific titles from the CIP-SOC crosswalk are queried; broad "
            "catch-all titles (e.g. *Managers, All Other*) are excluded to "
            "reduce noise. Live API call: data refreshes hourly. "
            "| Source: Coresignal Base Jobs API."
        )

    if SHOW_JOB_POSTINGS_UI and _cs_api_key and _is_admin and all_cips:
        st.info(
            "Job posting data is shown when specific CIP code(s) are selected. "
            "Deselect **All CIP codes** and choose program(s) to see posting trends."
        )
    elif SHOW_JOB_POSTINGS_UI and _cs_api_key and _is_admin and not all_cips:
        _cs_key = f"cs_{cip_patterns}_{selected_awlevels}_{geo_key}_{geo_values}"
        if st.button("Pull Job Posting Trends", key="cs_pull"):
            with st.spinner("Querying job posting trends (this may take a moment)..."):
                st.session_state["_cs_data"] = run_coresignal_trend(
                    cip_patterns=cip_patterns,
                    awlevels=selected_awlevels,
                    geo_key=geo_key,
                    geo_values=tuple(geo_values),
                )

        cs_data = st.session_state.get("_cs_data")
        if cs_data is None:
            pass  # no data yet or no results
        elif cs_data == "empty":
            st.info("No job posting data found for the selected program(s) and geography.")
        else:
            trend_df = cs_data["trend_df"]
            current_active = cs_data["current_active"]
            cs_titles = cs_data["search_titles"]
            # ── Metrics row ──────────────────────────────────────────────
            mc1, mc2, mc3 = st.columns(3)
            mc1.metric("Currently Active Postings", f"{current_active:,}")

            latest_month = trend_df["postings"].iloc[-1]
            mc2.metric("New Postings (latest month)", f"{latest_month:,}")

            # 3-month vs prior 3-month change
            if len(trend_df) >= 6:
                recent_3 = trend_df["postings"].iloc[-3:].sum()
                prior_3 = trend_df["postings"].iloc[-6:-3].sum()
                if prior_3 > 0:
                    qoq_change = (recent_3 - prior_3) / prior_3
                    mc3.metric(
                        "3-Month Change",
                        f"{qoq_change:+.1%}",
                        delta=f"{qoq_change:+.1%}",
                    )

            # ── Trend line chart ─────────────────────────────────────────
            fig_trend = go.Figure()
            fig_trend.add_trace(go.Scatter(
                x=trend_df["month"],
                y=trend_df["postings"],
                mode="lines+markers",
                line=dict(color="#f26822", width=3),
                marker=dict(size=7, color="#f26822"),
                hovertemplate="%{x}<br>%{y:,.0f} postings<extra></extra>",
            ))
            fig_trend.update_layout(
                title=dict(
                    text="New Job Postings per Month",
                    font=dict(size=13),
                    x=0,
                    xanchor="left",
                ),
                height=350,
                margin=dict(t=40, b=60, l=70, r=20),
                plot_bgcolor="white",
                paper_bgcolor="white",
                font=dict(family="Montserrat, Arial, sans-serif", size=12, color="#333333"),
                xaxis=dict(
                    title="Month",
                    tickangle=-45,
                    showgrid=True,
                    gridcolor="#F3F4F6",
                    gridwidth=1,
                ),
                yaxis=dict(
                    title="Postings",
                    showgrid=True,
                    gridcolor="#F3F4F6",
                    gridwidth=1,
                    rangemode="tozero",
                ),
                showlegend=False,
            )
            st.plotly_chart(fig_trend, use_container_width=True)

            # ── YoY % change bar chart (if we have 12+ months) ──────────
            if len(trend_df) >= 2:
                trend_df_yoy = trend_df.copy()
                trend_df_yoy["pct_change"] = trend_df_yoy["postings"].pct_change() * 100
                trend_df_yoy = trend_df_yoy.dropna(subset=["pct_change"])

                if not trend_df_yoy.empty:
                    colors = [
                        "#0f86c1" if v >= 0 else "#E74C3C"
                        for v in trend_df_yoy["pct_change"]
                    ]
                    fig_yoy = go.Figure(go.Bar(
                        x=trend_df_yoy["month"],
                        y=trend_df_yoy["pct_change"],
                        marker_color=colors,
                        hovertemplate="%{x}<br>%{y:+.1f}%<extra></extra>",
                    ))
                    fig_yoy.update_layout(
                        title=dict(
                            text="Month-over-Month % Change",
                            font=dict(size=13),
                            x=0,
                            xanchor="left",
                        ),
                        height=220,
                        margin=dict(t=40, b=60, l=70, r=20),
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        font=dict(family="Montserrat, Arial, sans-serif", size=12, color="#333333"),
                        xaxis=dict(tickangle=-45, showgrid=True, gridcolor="#F3F4F6"),
                        yaxis=dict(
                            ticksuffix="%",
                            tickformat=".0f",
                            showgrid=True,
                            gridcolor="#F3F4F6",
                            zeroline=True,
                            zerolinecolor="#999999",
                            zerolinewidth=1,
                        ),
                        showlegend=False,
                    )
                    st.plotly_chart(fig_yoy, use_container_width=True)

            # ── Caption ──────────────────────────────────────────────────
            titles_str = ", ".join(f"**{t}**" for t in cs_titles)
            st.caption(
                f"Occupation titles searched: {titles_str}. "
                f"Monthly counts reflect new postings created in each month (not cumulative). "
                f"Data refreshes hourly."
            )




if __name__ == "__main__":
    main()
