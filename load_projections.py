"""
Load employment projections into the IPEDS SQLite database.

Usage:
    python load_projections.py

Sources:
  - Stage 1: BLS national occupation projections (2024-2034)
  - Stage 2: ProjectionsCentral state projections (2022-2032)
  - Stage 3: State LMI metro scrapers (varies by state)
  - Fallback: State-level CAGR applied to metros in states without scrapers

Creates table:
  - employment_projections (occ_code, geo_level, geo_code, geo_name, base_year,
                            proj_year, base_emp, proj_emp, pct_change, cagr, source)
"""

import io
import json
import re
import sqlite3
import sys
import time
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
import requests

DB_PATH = Path(__file__).parent / "ipeds.db"
RAW_DIR = Path(__file__).parent / "raw"
RAW_DIR.mkdir(exist_ok=True)

BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# FIPS code -> state name mapping
FIPS_TO_STATE = {
    "01": "Alabama", "02": "Alaska", "04": "Arizona", "05": "Arkansas",
    "06": "California", "08": "Colorado", "09": "Connecticut", "10": "Delaware",
    "11": "District of Columbia", "12": "Florida", "13": "Georgia", "15": "Hawaii",
    "16": "Idaho", "17": "Illinois", "18": "Indiana", "19": "Iowa",
    "20": "Kansas", "21": "Kentucky", "22": "Louisiana", "23": "Maine",
    "24": "Maryland", "25": "Massachusetts", "26": "Michigan", "27": "Minnesota",
    "28": "Mississippi", "29": "Missouri", "30": "Montana", "31": "Nebraska",
    "32": "Nevada", "33": "New Hampshire", "34": "New Jersey", "35": "New Mexico",
    "36": "New York", "37": "North Carolina", "38": "North Dakota", "39": "Ohio",
    "40": "Oklahoma", "41": "Oregon", "42": "Pennsylvania", "44": "Rhode Island",
    "45": "South Carolina", "46": "South Dakota", "47": "Tennessee", "48": "Texas",
    "49": "Utah", "50": "Vermont", "51": "Virginia", "53": "Washington",
    "54": "West Virginia", "55": "Wisconsin", "56": "Wyoming",
}


def create_table(conn):
    """Create employment_projections table (drop if exists)."""
    conn.execute("DROP TABLE IF EXISTS employment_projections")
    conn.execute("""
        CREATE TABLE employment_projections (
            occ_code    TEXT    NOT NULL,
            geo_level   TEXT    NOT NULL,
            geo_code    TEXT    NOT NULL,
            geo_name    TEXT,
            base_year   INTEGER NOT NULL,
            proj_year   INTEGER NOT NULL,
            base_emp    INTEGER,
            proj_emp    INTEGER,
            pct_change  REAL,
            cagr        REAL,
            source      TEXT,
            PRIMARY KEY (occ_code, geo_level, geo_code)
        )
    """)
    conn.commit()
    print("Created employment_projections table")


def create_coverage_table(conn):
    """Create metro_projection_coverage table to track scraper vs fallback status."""
    conn.execute("DROP TABLE IF EXISTS metro_projection_coverage")
    conn.execute("""
        CREATE TABLE metro_projection_coverage (
            cbsa_code       TEXT PRIMARY KEY,
            cbsa_name       TEXT,
            state_abbr      TEXT,
            state_fips      TEXT,
            source          TEXT NOT NULL,   -- 'state_lmi' or 'state_fallback'
            scraper_state   TEXT,            -- which state scraper provided data (e.g. 'CA', 'PA')
            occ_count       INTEGER,         -- number of occupations with projection data
            avg_cagr        REAL,            -- average CAGR across occupations
            last_updated    TEXT             -- timestamp of last load
        )
    """)
    conn.commit()


def populate_coverage_table(conn):
    """Populate metro_projection_coverage from employment_projections data."""
    import re
    from datetime import datetime

    ABBR_TO_FIPS = {
        "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06",
        "CO": "08", "CT": "09", "DE": "10", "DC": "11", "FL": "12",
        "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18",
        "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23",
        "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28",
        "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
        "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38",
        "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44",
        "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49",
        "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55",
        "WY": "56",
    }
    FIPS_TO_ABBR = {v: k for k, v in ABBR_TO_FIPS.items()}

    # Build CBSA -> scraper state mapping from scraper config dicts
    CBSA_TO_SCRAPER = {}
    for cbsa in CA_FILE_TO_CBSA.values():
        CBSA_TO_SCRAPER[cbsa] = "CA"
    for _fname, (cbsa, _name) in PA_MSA_FILES.items():
        CBSA_TO_SCRAPER[cbsa] = "PA"
    CBSA_TO_SCRAPER[VT_BURLINGTON_CBSA] = "VT"
    for cbsa_list in NC_REGION_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "NC"
    for cbsa_list in TX_WDA_TO_CBSA.values():
        for cbsa in cbsa_list:
            CBSA_TO_SCRAPER[cbsa] = "TX"
    for cbsa_list in FL_CS_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "FL"
    for _rnum, (_, cbsa_list) in MI_REGION_TO_CBSA.items():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "MI"
    for _area, (cbsa_code, _name) in OK_AREA_TO_CBSA.items():
        CBSA_TO_SCRAPER[cbsa_code] = "OK"
    for cbsa_list in GA_LWDA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "GA"
    for cbsa_list in WI_WDA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "WI"
    for _area, (cbsa_code, _name) in OH_AREA_TO_CBSA.items():
        CBSA_TO_SCRAPER[cbsa_code] = "OH"
    for cbsa_list in NY_REGION_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "NY"
    for _region, (_p1, _p2, cbsa_list) in AL_REGION_TO_CBSA.items():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "AL"
    for cbsa_list in TN_REGION_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "TN"
    for cbsa_list in WA_WDA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "WA"
    for cbsa_list in OR_REGION_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "OR"
    for cbsa_list in WY_REGION_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "WY"
    for cbsa_list in IN_EGR_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "IN"
    for cbsa_list in LA_RLMA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "LA"
    for cbsa_list in MD_WIA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "MD"
    for cbsa_list in VA_LWDA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "VA"
    for cbsa_list in MA_WDA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "MA"
    for cbsa_list in SC_WDA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "SC"
    for cbsa_list in MO_FILES.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "MO"
    for cbsa_list in IA_FILES.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "IA"
    for cbsa_list in IL_EDR_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "IL"
    for cbsa_code in CO_MSA_TO_CBSA:
        CBSA_TO_SCRAPER[cbsa_code] = "CO"
    for cbsa_list in AZ_AREA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "AZ"
    for cbsa_list in WV_WIA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "WV"
    for cbsa_list in KY_LWA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "KY"
    for cbsa_list in ID_FILES.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "ID"
    for cbsa_list in UT_FILES.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "UT"
    for cbsa_list in MT_REGION_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "MT"
    for cbsa_list in AR_FILES.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "AR"
    for cbsa_list in MN_AREA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "MN"
    for cbsa_list in CT_FILES.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "CT"
    for cbsa_list in KS_AREA_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "KS"
    for cbsa_code, _name in NM_MSA_TO_CBSA.values():
        CBSA_TO_SCRAPER[cbsa_code] = "NM"
    for cbsa_code, _name in NJ_COUNTY_TO_CBSA.values():
        CBSA_TO_SCRAPER[cbsa_code] = "NJ"
    for cbsa_code, _name in MS_FILES.values():
        CBSA_TO_SCRAPER[cbsa_code] = "MS"
    for cbsa_code, _name in NV_SHEET_TO_CBSA.values():
        CBSA_TO_SCRAPER[cbsa_code] = "NV"
    for cbsa_code, _name in SD_SHEET_TO_CBSA.values():
        CBSA_TO_SCRAPER[cbsa_code] = "SD"
    for cbsa_code, _name in DE_COUNTY_TO_CBSA.values():
        CBSA_TO_SCRAPER[cbsa_code] = "DE"
    for cbsa_list in NH_SHEET_TO_CBSA.values():
        for cbsa_code, _name in cbsa_list:
            CBSA_TO_SCRAPER[cbsa_code] = "NH"
    for cbsa_code, _name in HI_FILES.values():
        CBSA_TO_SCRAPER[cbsa_code] = "HI"

    print("\n=== Populating Metro Coverage Table ===")
    conn.execute("DELETE FROM metro_projection_coverage")

    rows = conn.execute("""
        SELECT geo_code, geo_name, source, COUNT(*) as occ_count, AVG(cagr) as avg_cagr
        FROM employment_projections
        WHERE geo_level = 'metro'
        GROUP BY geo_code, geo_name, source
        ORDER BY geo_name
    """).fetchall()

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inserts = []
    for geo_code, geo_name, source, occ_count, avg_cagr in rows:
        # Extract state abbreviation from geo_name (e.g., "Dallas-Fort Worth, TX")
        match = re.search(r',\s*([A-Z]{2})', geo_name or "")
        state_abbr = match.group(1) if match else None
        state_fips = ABBR_TO_FIPS.get(state_abbr) if state_abbr else None

        # For scraper rows, derive state from CBSA mapping
        scraper_state = None
        if source == "state_lmi":
            scraper_state = CBSA_TO_SCRAPER.get(geo_code)
            if scraper_state and not state_abbr:
                state_abbr = scraper_state
                state_fips = ABBR_TO_FIPS.get(state_abbr)

        inserts.append((
            geo_code, geo_name, state_abbr, state_fips, source,
            scraper_state, occ_count,
            round(avg_cagr, 6) if avg_cagr else None, now,
        ))

    conn.executemany("""
        INSERT OR REPLACE INTO metro_projection_coverage
        (cbsa_code, cbsa_name, state_abbr, state_fips, source,
         scraper_state, occ_count, avg_cagr, last_updated)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, inserts)
    conn.commit()

    # Print summary
    scraper_count = sum(1 for r in inserts if r[4] == "state_lmi")
    adjusted_count = sum(1 for r in inserts if r[4] == "adjusted_estimate")
    print(f"  {scraper_count} CBSAs with scraper data (state_lmi)")
    print(f"  {adjusted_count} CBSAs with adjusted estimates")

    # Per-state summary for adjusted
    from collections import Counter
    adjusted_states = Counter(r[2] for r in inserts if r[4] == "adjusted_estimate" and r[2])
    if adjusted_states:
        print(f"\n  Adjusted estimate metros by state:")
        for st, cnt in adjusted_states.most_common(15):
            print(f"    {st}: {cnt} CBSAs")

    scraper_sources = Counter(r[5] for r in inserts if r[5])
    if scraper_sources:
        print(f"\n  Scraper metros by source:")
        for src, cnt in scraper_sources.most_common():
            print(f"    {src}: {cnt} CBSAs")

    return len(inserts)


def calc_cagr(base, proj, years):
    """Calculate compound annual growth rate."""
    if base and proj and base > 0 and years > 0:
        return (proj / base) ** (1 / years) - 1
    return None


# ---------------------------------------------------------------------------
# Stage 1: BLS National Projections (2024-2034)
# ---------------------------------------------------------------------------

def load_bls_national(conn):
    """Download and parse BLS occupation projections xlsx."""
    print("\n=== Stage 1: BLS National Projections ===")

    url = "https://www.bls.gov/emp/ind-occ-matrix/occupation.xlsx"
    local_path = RAW_DIR / "bls_occupation_projections.xlsx"

    # Download if not cached
    if not local_path.exists():
        print(f"  Downloading {url} ...")
        resp = requests.get(url, headers={
            "User-Agent": BROWSER_UA,
            "Referer": "https://www.bls.gov/emp/data/occupational-data.htm",
        }, timeout=60)
        resp.raise_for_status()
        local_path.write_bytes(resp.content)
        print(f"  Saved {len(resp.content):,} bytes to {local_path.name}")
    else:
        print(f"  Using cached {local_path.name}")

    # Parse "Table 1.2" sheet
    df = pd.read_excel(local_path, sheet_name="Table 1.2")

    # Identify columns by position (header names vary):
    # Col 0: Title, Col 1: 2018 SOC code, Col 2: Occupation type,
    # Col 3: Employment 2024 (thousands), Col 4: Employment 2034 (thousands)
    cols = df.columns.tolist()
    df.columns = [f"col_{i}" for i in range(len(cols))]

    # Filter to "Line item" rows (detailed occupations, not summaries)
    df = df[df["col_2"].astype(str).str.strip() == "Line item"].copy()

    # Parse SOC code
    df["occ_code"] = df["col_1"].astype(str).str.strip()
    df = df[df["occ_code"].str.match(r"^\d{2}-\d{4}$")].copy()

    # Parse employment (in thousands)
    df["base_emp"] = pd.to_numeric(df["col_3"], errors="coerce") * 1000
    df["proj_emp"] = pd.to_numeric(df["col_4"], errors="coerce") * 1000

    # Drop rows with missing employment
    df = df.dropna(subset=["base_emp", "proj_emp"])
    df["base_emp"] = df["base_emp"].astype(int)
    df["proj_emp"] = df["proj_emp"].astype(int)

    # Calculate CAGR (10-year projection)
    df["cagr"] = df.apply(
        lambda r: calc_cagr(r["base_emp"], r["proj_emp"], 10), axis=1
    )
    df["pct_change"] = (df["proj_emp"] - df["base_emp"]) / df["base_emp"] * 100

    # Build insertion rows
    rows = []
    for _, r in df.iterrows():
        rows.append((
            r["occ_code"], "national", "99", "National",
            2024, 2034, int(r["base_emp"]), int(r["proj_emp"]),
            round(r["pct_change"], 2) if pd.notna(r["pct_change"]) else None,
            round(r["cagr"], 6) if pd.notna(r["cagr"]) else None,
            "bls",
        ))

    conn.executemany("""
        INSERT OR REPLACE INTO employment_projections
        (occ_code, geo_level, geo_code, geo_name, base_year, proj_year,
         base_emp, proj_emp, pct_change, cagr, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, rows)
    conn.commit()
    print(f"  Inserted {len(rows):,} national projections")
    return len(rows)


# ---------------------------------------------------------------------------
# Stage 2: ProjectionsCentral State Projections (2022-2032)
# ---------------------------------------------------------------------------

def load_projections_central(conn):
    """Paginate through ProjectionsCentral API for state-level projections."""
    print("\n=== Stage 2: ProjectionsCentral State Projections ===")

    base_url = "https://public.projectionscentral.org/Projections/LongTermRestJson/all"
    headers = {
        "User-Agent": BROWSER_UA,
        "Origin": "https://projectionscentral.org",
        "Accept": "application/json",
    }

    all_rows = []
    page = 0
    items_per_page = 1000

    while True:
        url = f"{base_url}?items_per_page={items_per_page}&page={page}"
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        data = resp.json()

        rows = data.get("rows", [])
        if not rows:
            break

        for r in rows:
            stfips = r.get("STFIPS")
            if stfips is None or int(stfips) == 0:
                continue  # Skip national rows, we use BLS
            if int(stfips) >= 57:
                continue  # Skip territories (PR=72, GU=66, VI=78, etc.)

            occ_code = r.get("OccCode", "").strip()
            if not occ_code or not occ_code[0].isdigit():
                continue

            base_emp = r.get("Base")
            proj_emp = r.get("Projected")
            pct_change = r.get("PercentChange")
            area_name = r.get("Area", "")

            # Parse numeric values
            try:
                base_emp = int(str(base_emp).replace(",", "")) if base_emp else None
                proj_emp = int(str(proj_emp).replace(",", "")) if proj_emp else None
                pct_change = float(pct_change) if pct_change else None
            except (ValueError, TypeError):
                continue

            geo_code = str(int(stfips)).zfill(2)
            geo_name = FIPS_TO_STATE.get(geo_code, area_name)

            cagr = calc_cagr(base_emp, proj_emp, 10) if base_emp and proj_emp else None

            all_rows.append((
                occ_code, "state", geo_code, geo_name,
                2022, 2032,
                base_emp, proj_emp,
                round(pct_change, 2) if pct_change is not None else None,
                round(cagr, 6) if cagr is not None else None,
                "projections_central",
            ))

        pager = data.get("pager", {})
        total_items = pager.get("total_items", 0)
        current_page = pager.get("current_page", 0)
        total_pages = pager.get("total_pages", 0)

        page += 1
        if page % 5 == 0 or page >= total_pages:
            print(f"  Page {page}/{total_pages} — {len(all_rows):,} state rows so far")

        if page >= total_pages:
            break

        # Be polite to the API
        time.sleep(0.2)

    conn.executemany("""
        INSERT OR REPLACE INTO employment_projections
        (occ_code, geo_level, geo_code, geo_name, base_year, proj_year,
         base_emp, proj_emp, pct_change, cagr, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, all_rows)
    conn.commit()
    print(f"  Inserted {len(all_rows):,} state projections")
    return len(all_rows)


# ---------------------------------------------------------------------------
# Stage 3: Metro Projections — State LMI Scrapers
# ---------------------------------------------------------------------------
# Each state function returns a list of tuples matching the insertion format.
# We'll add scrapers incrementally.

def load_metro_scrapers(conn):
    """Run all available state metro scrapers."""
    print("\n=== Stage 3: Metro Projections (State LMI Scrapers) ===")

    # (name, func, needs_conn) — scrapers that need DB access for SOC mapping
    scrapers = [
        # Phase A
        ("California", load_california_metros, False),
        ("New York", load_new_york_metros, False),
        ("Pennsylvania", load_pennsylvania_metros, False),
        ("Washington", load_washington_metros, False),
        ("Vermont", load_vermont_metros, False),
        ("North Carolina", load_north_carolina_metros, False),
        # Phase B
        ("Texas", load_texas_metros, True),
        ("Florida", load_florida_metros, False),
        ("Michigan", load_michigan_metros, False),
        ("Oklahoma", load_oklahoma_metros, False),
        ("Georgia", load_georgia_metros, False),
        ("Wisconsin", load_wisconsin_metros, False),
        ("Ohio", load_ohio_metros, False),
        ("Alabama", load_alabama_metros, False),
        ("Tennessee", load_tennessee_metros, False),
        ("Oregon", load_oregon_metros, False),
        ("Wyoming", load_wyoming_metros, False),
        ("Indiana", load_indiana_metros, False),
        ("Louisiana", load_louisiana_metros, False),
        ("Maryland", load_maryland_metros, False),
        ("Virginia", load_virginia_metros, False),
        ("Massachusetts", load_massachusetts_metros, False),
        ("South Carolina", load_south_carolina_metros, False),
        ("Missouri", load_missouri_metros, False),
        ("Iowa", load_iowa_metros, False),
        ("Illinois", load_illinois_metros, False),
        ("Colorado", load_colorado_metros, False),
        ("Arizona", load_arizona_metros, False),
        ("West Virginia", load_west_virginia_metros, False),
        ("Kentucky", load_kentucky_metros, False),
        ("Idaho", load_idaho_metros, False),
        ("Utah", load_utah_metros, False),
        ("Montana", load_montana_metros, False),
        ("Arkansas", load_arkansas_metros, False),
        ("Minnesota", load_minnesota_metros, False),
        ("Connecticut", load_connecticut_metros, False),
        ("Kansas", load_kansas_metros, False),
        ("New Mexico", load_new_mexico_metros, False),
        ("New Jersey", load_new_jersey_metros, False),
        ("Mississippi", load_mississippi_metros, False),
        ("Nevada", load_nevada_metros, False),
        ("South Dakota", load_south_dakota_metros, False),
        ("Delaware", load_delaware_metros, False),
        ("New Hampshire", load_new_hampshire_metros, False),
        ("Hawaii", load_hawaii_metros, False),
    ]

    total = 0
    for state_name, func, needs_conn in scrapers:
        try:
            rows = func(conn) if needs_conn else func()
            if rows:
                conn.executemany("""
                    INSERT OR REPLACE INTO employment_projections
                    (occ_code, geo_level, geo_code, geo_name, base_year, proj_year,
                     base_emp, proj_emp, pct_change, cagr, source)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, rows)
                conn.commit()
                total += len(rows)
                print(f"  {state_name}: {len(rows):,} metro rows")
            else:
                print(f"  {state_name}: no data (scraper returned empty)")
        except Exception as e:
            print(f"  {state_name}: FAILED -- {e}")

    print(f"  Total metro rows from scrapers: {total:,}")
    return total


# --- State Scrapers (Phase A) ---

# California: file prefix -> CBSA code mapping
CA_FILE_TO_CBSA = {
    "bake$OccProj.xlsx": "12540",   # Bakersfield
    "chic$OccProj.xlsx": "17020",   # Chico
    "ecen$OccProj.xlsx": "20940",   # El Centro
    "frsn$OccProj.xlsx": "23420",   # Fresno
    "hanf$OccProj.xlsx": "25260",   # Hanford-Corcoran
    "la$OccProj.xlsx": "31080",     # Los Angeles (MD -> LA-LB-Anaheim MSA)
    "mad$OccProj.xlsx": "31460",    # Madera
    "merc$OccProj.xlsx": "32900",   # Merced
    "mode$OccProj.xlsx": "33700",   # Modesto
    "napa$OccProj.xlsx": "34900",   # Napa
    "oak$OccProj.xlsx": "41860",    # Oakland (MD -> SF-Oakland-Hayward MSA)
    "oran$OccProj.xlsx": "31080",   # Anaheim (MD -> LA-LB-Anaheim MSA, skip duplicate)
    "redd$OccProj.xlsx": "39820",   # Redding
    "rive$OccProj.xlsx": "40140",   # Riverside-San Bernardino-Ontario
    "sacr$OccProj.xlsx": "40900",   # Sacramento
    "sali$OccProj.xlsx": "41500",   # Salinas
    "sand$OccProj.xlsx": "41740",   # San Diego-Carlsbad
    "sanf$OccProj.xlsx": "41860",   # San Francisco (MD -> SF-Oakland MSA, skip duplicate)
    "sanrf$OccProj.xlsx": "41860",  # San Rafael (MD -> SF-Oakland MSA, skip duplicate)
    "satb$OccProj.xlsx": "42200",   # Santa Maria-Santa Barbara
    "satr$OccProj.xlsx": "42220",   # Santa Rosa
    "scrz$OccProj.xlsx": "42100",   # Santa Cruz-Watsonville
    "sjos$OccProj.xlsx": "41940",   # San Jose-Sunnyvale-Santa Clara
    "slo$OccProj.xlsx": "42020",    # San Luis Obispo
    "stoc$OccProj.xlsx": "44700",   # Stockton-Lodi
    "vall$OccProj.xlsx": "46700",   # Vallejo-Fairfield
    "vent$OccProj.xlsx": "37100",   # Oxnard-Thousand Oaks-Ventura
    "visa$OccProj.xlsx": "47300",   # Visalia-Porterville
    "yuba$OccProj.xlsx": "49700",   # Yuba City
}

# Track which CBSAs we've already inserted (for MD dedup)
_ca_inserted_cbsas = set()


def _parse_ca_occ_file(file_bytes, cbsa_code, area_name):
    """Parse a single California EDD occupational projections XLSX."""
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    # Find header row (contains "SOC" in first or second column)
    header_row = None
    for i in range(min(10, len(df))):
        row_str = " ".join(str(v) for v in df.iloc[i].tolist() if pd.notna(v))
        if "SOC" in row_str and ("Code" in row_str or "Level" in row_str):
            header_row = i
            break
    if header_row is None:
        return []

    # Data starts right after header
    data = df.iloc[header_row + 1:].copy()
    data.columns = range(len(data.columns))

    # CA format: col 0=SOC Level, col 1=SOC Code, col 3=Base Emp, col 4=Proj Emp
    rows = []
    for _, r in data.iterrows():
        soc_level = r.get(0)
        soc_code = str(r.get(1, "")).strip()
        base_emp = r.get(3)
        proj_emp = r.get(4)

        # Only detailed occupations (level 4) with valid SOC codes
        if soc_level != 4:
            continue
        if not soc_code or not soc_code[0].isdigit() or len(soc_code) < 7:
            continue

        try:
            base_emp = int(float(base_emp)) if pd.notna(base_emp) else None
            proj_emp = int(float(proj_emp)) if pd.notna(proj_emp) else None
        except (ValueError, TypeError):
            continue

        if not base_emp or not proj_emp:
            continue

        pct_change = (proj_emp - base_emp) / base_emp * 100 if base_emp > 0 else None
        cagr = calc_cagr(base_emp, proj_emp, 10)

        rows.append((
            soc_code, "metro", cbsa_code, area_name,
            2022, 2032, base_emp, proj_emp,
            round(pct_change, 2) if pct_change is not None else None,
            round(cagr, 6) if cagr is not None else None,
            "state_lmi",
        ))
    return rows


def load_california_metros():
    """California EDD long-term projections by MSA from bulk ZIP."""
    url = "https://labormarketinfo.edd.ca.gov/file/occproj/allOccProj.zip"
    local_path = RAW_DIR / "ca_allOccProj.zip"

    if not local_path.exists():
        print("    CA: Downloading bulk projections ZIP...")
        resp = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=120, verify=False)
        resp.raise_for_status()
        local_path.write_bytes(resp.content)
        print(f"    CA: Saved {len(resp.content):,} bytes")
    else:
        print(f"    CA: Using cached {local_path.name}")

    all_rows = []
    seen_cbsas = set()

    zf = zipfile.ZipFile(local_path)
    for fname in sorted(zf.namelist()):
        if fname not in CA_FILE_TO_CBSA:
            continue
        cbsa = CA_FILE_TO_CBSA[fname]

        # Skip duplicate MDs mapping to same MSA (take first one)
        if cbsa in seen_cbsas:
            continue
        seen_cbsas.add(cbsa)

        with zf.open(fname) as f:
            file_bytes = f.read()

        # Get area name from file
        df_head = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=3)
        area_name = str(df_head.iloc[1][0]) if pd.notna(df_head.iloc[1][0]) else fname

        rows = _parse_ca_occ_file(file_bytes, cbsa, area_name)
        all_rows.extend(rows)

    print(f"    CA: Parsed {len(seen_cbsas)} MSAs, {len(all_rows):,} occupation rows")
    return all_rows


# Pennsylvania MSA file codes -> CBSA mapping
PA_MSA_FILES = {
    "abemsa_ltop.xlsx": ("10900", "Allentown-Bethlehem-Easton"),
    "altoonamsa_ltop.xlsx": ("11020", "Altoona"),
    "bloommsa_ltop.xlsx": ("14100", "Bloomsburg-Berwick"),
    "chambmsa_ltop.xlsx": ("16540", "Chambersburg-Waynesboro"),
    "estrmsa_ltop.xlsx": ("20700", "East Stroudsburg"),
    "eriemsa_ltop.xlsx": ("21500", "Erie"),
    "gettymsa_ltop.xlsx": ("23900", "Gettysburg"),
    "hcmsa_ltop.xlsx": ("25420", "Harrisburg-Carlisle"),
    "johnmsa_ltop.xlsx": ("27780", "Johnstown"),
    "lancmsa_ltop.xlsx": ("29540", "Lancaster"),
    "lebmsa_ltop.xlsx": ("30140", "Lebanon"),
    "philanjmsa_ltop.xlsx": ("37980", "Philadelphia-Camden-Wilmington"),
    "pghmsa_ltop.xlsx": ("38300", "Pittsburgh"),
    "rdgmsa_ltop.xlsx": ("39740", "Reading"),
    "swbmsa_ltop.xlsx": ("42540", "Scranton-Wilkes-Barre"),
    "stcmsa_ltop.xlsx": ("44300", "State College"),
    "wmsptmsa_ltop.xlsx": ("48700", "Williamsport"),
    "yorkmsa_ltop.xlsx": ("49620", "York-Hanover"),
}


def _parse_pa_msa_file(file_bytes, cbsa_code, area_name):
    """Parse a single Pennsylvania MSA projections XLSX."""
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    # PA format: data starts after header rows (around row 7)
    # Col 0: SOC Code, Col 1: Title, Col 3: Base Emp (2022), Col 4: Proj Emp (2032), Col 5: Pct Change
    rows = []
    for i in range(7, len(df)):
        soc_code = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
        if not soc_code or not soc_code[0].isdigit() or len(soc_code) < 7:
            continue

        # Skip summary codes (XX-0000)
        if soc_code.endswith("0000") or soc_code.endswith("000"):
            continue

        base_emp = df.iloc[i, 3]
        proj_emp = df.iloc[i, 4]

        try:
            base_emp = int(float(base_emp)) if pd.notna(base_emp) else None
            proj_emp = int(float(proj_emp)) if pd.notna(proj_emp) else None
        except (ValueError, TypeError):
            continue

        if not base_emp or not proj_emp:
            continue

        pct_change = (proj_emp - base_emp) / base_emp * 100 if base_emp > 0 else None
        cagr = calc_cagr(base_emp, proj_emp, 10)

        rows.append((
            soc_code, "metro", cbsa_code, area_name,
            2022, 2032, base_emp, proj_emp,
            round(pct_change, 2) if pct_change is not None else None,
            round(cagr, 6) if cagr is not None else None,
            "state_lmi",
        ))
    return rows


def load_pennsylvania_metros():
    """Pennsylvania DOLI MSA-level projections (18 MSAs)."""
    base_url = "https://www.pa.gov/content/dam/copapwp-pagov/en/dli/documents/cwia/products/projections/occupational/msa/"

    all_rows = []
    for fname, (cbsa_code, area_name) in PA_MSA_FILES.items():
        local_path = RAW_DIR / f"pa_{fname}"

        if not local_path.exists():
            url = base_url + fname
            resp = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=60)
            resp.raise_for_status()
            local_path.write_bytes(resp.content)

        rows = _parse_pa_msa_file(local_path.read_bytes(), cbsa_code, area_name)
        all_rows.extend(rows)

    print(f"    PA: Parsed {len(PA_MSA_FILES)} MSAs, {len(all_rows):,} occupation rows")
    return all_rows


NY_REGIONAL_FILE = RAW_DIR / "ny_regional_projections.xlsx"

# NY DOL region -> list of (cbsa_code, cbsa_name)
NY_REGION_TO_CBSA = {
    "Capital Region":  [("10580", "Albany-Schenectady-Troy, NY"), ("24020", "Glens Falls, NY")],
    "Central Region":  [("45060", "Syracuse, NY"), ("46540", "Utica-Rome, NY")],
    "Finger Lakes":    [("40380", "Rochester, NY")],
    "Hudson Valley":   [("28740", "Kingston, NY"), ("28880", "Kiryas Joel-Poughkeepsie-Newburgh, NY")],
    "Long Island":     [("35620", "New York-Newark-Jersey City, NY-NJ")],
    "Mohawk Valley":   [("46540", "Utica-Rome, NY")],
    "New York City":   [("35620", "New York-Newark-Jersey City, NY-NJ")],
    "North Country":   [("48060", "Watertown-Fort Drum, NY")],
    "Southern Tier":   [("13780", "Binghamton, NY"), ("21300", "Elmira, NY"), ("27060", "Ithaca, NY")],
    "Western Region":  [("15380", "Buffalo-Cheektowaga, NY")],
}


def load_new_york_metros():
    """Load NY DOL regional occupational projections (2022-2032) from local file."""
    if not NY_REGIONAL_FILE.exists():
        print("    New York: file not found at raw/ny_regional_projections.xlsx, skipping")
        return []

    soc_re = re.compile(r"^\d{2}-\d{4}$")

    try:
        xls = pd.ExcelFile(NY_REGIONAL_FILE, engine="openpyxl")
    except Exception as e:
        print(f"    NY: Error opening file: {e}")
        return []

    rows = []
    cbsa_names = {}
    for region, cbsa_list in NY_REGION_TO_CBSA.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name

        if region not in xls.sheet_names:
            print(f"    NY: Sheet '{region}' not found, skipping")
            continue

        df = pd.read_excel(xls, sheet_name=region, header=None, engine="openpyxl")

        # Format: row 5 = header labels, row 6 = sub-headers, data starts row 8
        # col 0 = Summary Level, col 1 = SOC Code, col 3 = Emp 2022, col 4 = Emp 2032
        for i in range(7, len(df)):
            summary_level = df.iloc[i, 0]
            # Only take detailed occupations (summary level 4 = 6-digit SOC)
            try:
                if int(summary_level) != 4:
                    continue
            except (ValueError, TypeError):
                continue

            soc = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
            if not soc_re.match(soc):
                continue

            try:
                base = int(float(df.iloc[i, 3])) if pd.notna(df.iloc[i, 3]) else 0
                proj = int(float(df.iloc[i, 4])) if pd.notna(df.iloc[i, 4]) else 0
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code, cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # CAGR only — region != CBSA boundaries
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    NY: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# --- Washington: WDA-level projections from ESD (2023-2033) ---

WA_WDA_FILE = RAW_DIR / "wa_long_occ_proj.xlsx"
WA_WDA_URL = "https://esd.wa.gov/media/xlsx/3794/long-occup-proj-alt-2025xlsx"

# Sheet name -> list of (cbsa_code, cbsa_name)
WA_WDA_TO_CBSA = {
    "Benton-Franklin":     [("28420", "Kennewick-Richland, WA")],
    "Eastern":             [("47460", "Walla Walla, WA"), ("30300", "Lewiston, ID-WA")],
    "North Central":       [("48300", "Wenatchee-East Wenatchee, WA")],
    "Northwest":           [("13380", "Bellingham, WA"), ("34580", "Mount Vernon-Anacortes, WA")],
    "Olympic":             [("14740", "Bremerton-Silverdale-Port Orchard, WA")],
    "Pacific Mountain":    [("36500", "Olympia-Lacey-Tumwater, WA"), ("31020", "Longview-Kelso, WA")],
    "Seattle-King County": [("42660", "Seattle-Tacoma-Bellevue, WA")],
    "Snohomish":           [("42660", "Seattle-Tacoma-Bellevue, WA")],
    "South Central":       [("49420", "Yakima, WA")],
    "Spokane":             [("44060", "Spokane-Spokane Valley, WA")],
    "Tacoma-Pierce":       [("42660", "Seattle-Tacoma-Bellevue, WA")],
}


def load_washington_metros():
    """Load Washington WDA-level occupational projections (2023-2033)."""
    cached = WA_WDA_FILE
    if not cached.exists():
        print("    WA: Downloading WDA projections...")
        r = requests.get(WA_WDA_URL, headers=HEADERS)
        r.raise_for_status()
        cached.write_bytes(r.content)
    else:
        print(f"    WA: Using cached {cached.name}")

    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in WA_WDA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    wb = openpyxl.load_workbook(cached, read_only=True, data_only=True)
    for sheet_name, cbsa_list in WA_WDA_TO_CBSA.items():
        if sheet_name not in wb.sheetnames:
            print(f"    WA: Sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        for row_data in ws.iter_rows(min_row=6, values_only=True):
            soc = str(row_data[0]).strip() if row_data[0] else ""
            if not soc_re.match(soc):
                continue
            # Skip summary codes (xx-0000, xx-x000)
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base = int(float(row_data[2]))  # col 2 = Est Employment 2023
                proj = int(float(row_data[4]))  # col 4 = Est Employment 2033
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2023-2033
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2023, 2033,
                        None, None,  # CAGR only — WDA != CBSA
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))
    wb.close()

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    WA: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# --- Vermont: Burlington NECTA from VDOL ---

VT_BURLINGTON_URL = "https://www.vtlmi.info/public/occprjburl.xlsx"
VT_BURLINGTON_CBSA = "15540"
VT_BURLINGTON_NAME = "Burlington-South Burlington, VT"


def load_vermont_metros():
    """Load Burlington NECTA occupational projections from VT DOL."""
    cached = RAW_DIR / "vt_occprjburl.xlsx"
    if cached.exists():
        print(f"    VT: Using cached {cached.name}")
        data = cached.read_bytes()
    else:
        print(f"    VT: Downloading {VT_BURLINGTON_URL}")
        r = requests.get(VT_BURLINGTON_URL, headers={"User-Agent": BROWSER_UA})
        r.raise_for_status()
        data = r.content
        cached.write_bytes(data)

    df = pd.read_excel(io.BytesIO(data), header=None, skiprows=8)
    # Col 0: SOC Code, Col 1: Title, Col 2: Base emp, Col 3: Proj emp,
    # Col 4: Annual Growth Rate (CAGR, already decimal)

    rows = []
    for i in range(len(df)):
        soc = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
        if not soc or not soc[0].isdigit() or len(soc) != 7:
            continue
        # Skip summary codes
        if soc.endswith("0000"):
            continue

        cagr_val = df.iloc[i, 4]
        if pd.isna(cagr_val):
            continue

        try:
            cagr_val = float(cagr_val)
        except (ValueError, TypeError):
            continue

        pct_change = ((1 + cagr_val) ** 10 - 1) * 100  # 10-year total pct change

        rows.append((
            soc, "metro", VT_BURLINGTON_CBSA, VT_BURLINGTON_NAME,
            2022, 2032,
            None,   # base_emp: use projected growth rate only
            None,   # proj_emp: applied to historical actuals in app
            round(pct_change, 2),
            round(cagr_val, 6),
            "state_lmi",
        ))

    print(f"    VT: {len(rows)} occupation rows for Burlington NECTA")
    return rows


# --- North Carolina: Sub-prosperity zone API ---

NC_API_BASE = "https://analytics.nccommerce.com/projections/api/OccProj"
NC_PROJ_ID = 27  # 2024-2034 projection set

# NC sub-prosperity zone code -> list of (CBSA code, CBSA name)
# Zones are larger than MSAs; each zone maps to the primary CBSA(s) it contains.
NC_REGION_TO_CBSA = {
    "3751000016": [("11700", "Asheville, NC")],
    "3751000005": [("25860", "Hickory-Lenoir-Morganton, NC")],
    "3751000014": [("16740", "Charlotte-Concord-Gastonia, NC-SC")],
    "3751000008": [("24660", "Greensboro-High Point, NC"),
                   ("15500", "Burlington, NC")],
    "3751000007": [("49180", "Winston-Salem, NC")],
    "3751000001": [("39580", "Raleigh-Cary, NC"),
                   ("20500", "Durham-Chapel Hill, NC")],
    "3751000002": [("40580", "Rocky Mount, NC")],
    "3751000010": [("22180", "Fayetteville, NC")],
    "3751000009": [("38240", "Pinehurst-Southern Pines, NC")],
    "3751000003": [("24780", "Greenville, NC")],
    "3751000011": [("24140", "Goldsboro, NC")],
    "3751000012": [("27340", "Jacksonville, NC")],
    "3751000013": [("48900", "Wilmington, NC")],
    # Waynesville-Franklin, Boone-Wilkesboro, Elizabeth City: rural, no CBSA match
}


def load_north_carolina_metros():
    """Load NC occupational projections from the NC Commerce API."""
    rows = []
    seen_cbsas = set()

    for region_code, cbsa_list in NC_REGION_TO_CBSA.items():
        url = f"{NC_API_BASE}/{region_code}/6/{NC_PROJ_ID}"
        try:
            r = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            region_name = cbsa_list[0][1] if cbsa_list else region_code
            print(f"      NC {region_name}: FAILED -- {e}")
            continue

        for cbsa_code, cbsa_name in cbsa_list:
            if cbsa_code in seen_cbsas:
                continue
            seen_cbsas.add(cbsa_code)

            for item in data:
                raw_soc = str(item.get("occ_code", ""))
                if len(raw_soc) != 6 or not raw_soc[0].isdigit():
                    continue
                soc = raw_soc[:2] + "-" + raw_soc[2:]  # "111011" -> "11-1011"
                if soc.endswith("0000"):
                    continue

                ann_rate = item.get("ann_grow_rate")
                if ann_rate is None:
                    continue
                try:
                    cagr = float(ann_rate) / 100  # API gives percent, convert to decimal
                except (ValueError, TypeError):
                    continue

                pct_change = ((1 + cagr) ** 10 - 1) * 100  # 10-year total

                rows.append((
                    soc, "metro", cbsa_code, cbsa_name,
                    2024, 2034,
                    None,   # base_emp: growth rate only
                    None,   # proj_emp: applied to historical actuals in app
                    round(pct_change, 2),
                    round(cagr, 6),
                    "state_lmi",
                ))

        time.sleep(0.3)  # Be polite to the API

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    NC: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# --- Texas: WDA-level projections from manually downloaded file ---

# WDA area number -> list of CBSA codes
TX_WDA_TO_CBSA = {
    "1": ["11100"],             # Panhandle -> Amarillo
    "2": ["31180"],             # South Plains -> Lubbock
    "3": ["48660"],             # North Texas -> Wichita Falls
    "4": ["19100"],             # North Central -> DFW
    "5": ["19100"],             # Tarrant County -> DFW
    "6": ["19100"],             # Dallas -> DFW
    "7": ["45500"],             # Northeast Texas -> Texarkana
    "8": ["46340", "30980"],    # East Texas -> Tyler + Longview
    "9": ["10180"],             # West Central -> Abilene
    "10": ["21340"],            # Borderplex -> El Paso
    "11": ["33260", "36220"],   # Permian Basin -> Midland + Odessa
    "12": ["41660"],            # Concho Valley -> San Angelo
    "13": ["47380"],            # Heart of Texas -> Waco
    "14": ["12420"],            # Capital Area -> Austin
    "15": ["12420"],            # Rural Capital -> Austin
    "16": ["17780"],            # Brazos Valley -> College Station-Bryan
    "17": [],                   # Deep East Texas -> rural (no CBSA)
    "18": ["13140"],            # Southeast Texas -> Beaumont-Port Arthur
    "19": [],                   # Golden Crescent -> Victoria (not in DB)
    "20": ["41700"],            # Alamo -> San Antonio
    "21": ["29700"],            # South Texas -> Laredo
    "22": ["18580"],            # Coastal Bend -> Corpus Christi
    "23": ["32580"],            # Lower Rio Grande -> McAllen
    "24": ["15180"],            # Cameron County -> Brownsville-Harlingen
    "25": ["43300"],            # Texoma -> Sherman-Denison
    "26": ["28660"],            # Central Texas -> Killeen-Temple
    "27": [],                   # Middle Rio Grande -> rural (no CBSA)
    "28": ["26420"],            # Gulf Coast -> Houston
}

TX_CBSA_NAMES = {
    "10180": "Abilene, TX", "11100": "Amarillo, TX",
    "12420": "Austin-Round Rock-San Marcos, TX",
    "13140": "Beaumont-Port Arthur, TX", "15180": "Brownsville-Harlingen, TX",
    "17780": "College Station-Bryan, TX", "18580": "Corpus Christi, TX",
    "19100": "Dallas-Fort Worth-Arlington, TX", "21340": "El Paso, TX",
    "26420": "Houston-Pasadena-The Woodlands, TX",
    "28660": "Killeen-Temple, TX", "29700": "Laredo, TX",
    "30980": "Longview, TX", "31180": "Lubbock, TX",
    "32580": "McAllen-Edinburg-Mission, TX", "33260": "Midland, TX",
    "36220": "Odessa, TX", "41660": "San Angelo, TX",
    "41700": "San Antonio-New Braunfels, TX", "43300": "Sherman-Denison, TX",
    "45500": "Texarkana, TX-AR", "46340": "Tyler, TX",
    "47380": "Waco, TX", "48660": "Wichita Falls, TX",
}


def load_texas_metros(conn):
    """
    Load Texas WDA-level projections, mapped to CBSAs.

    TX file has minor-group SOC codes (XX-X000). We map each to its detailed
    SOC codes and apply the same CAGR. For WDAs that map to the same CBSA
    (e.g., DFW = WDAs 4+5+6), we sum base/proj employment across WDAs first,
    then compute the combined CAGR.

    Only the growth rate (CAGR) is meaningful here since WDA boundaries don't
    match metro boundaries. base_emp and proj_emp are stored as NULL.
    """
    tx_path = RAW_DIR / "tx_projections.xlsx"
    if not tx_path.exists():
        print("    Texas: file not found (raw/tx_projections.xlsx)")
        return []

    df = pd.read_excel(tx_path)
    print(f"    Texas: loaded {len(df)} raw rows")

    # Filter: WDA rows only (not statewide), minor groups only (not XX-0000)
    wda_df = df[
        (df["Area"] != "Texas")
        & (~df["SOC Code"].str.endswith("0000"))
    ].copy()
    wda_df["area_num"] = wda_df["Area Number"].astype(str).str.lstrip("0")

    # Get detailed SOC codes from state-level TX projections for mapping
    tx_detailed = pd.read_sql("""
        SELECT DISTINCT occ_code FROM employment_projections
        WHERE geo_level = 'state' AND geo_code = '48'
        AND occ_code NOT LIKE '%-0000'
    """, conn)["occ_code"].tolist()

    # Build minor group -> detailed codes mapping
    minor_to_detailed = {}
    for soc in tx_detailed:
        prefix = soc[:4]  # e.g., '11-2' from '11-2011'
        minor = prefix + "000"
        minor_to_detailed.setdefault(minor, []).append(soc)

    # Aggregate WDA data by CBSA + SOC minor group
    # For multi-WDA CBSAs (DFW, Austin), sum employment then recompute CAGR
    from collections import defaultdict
    cbsa_minor_agg = defaultdict(lambda: {"base": 0, "proj": 0})

    for _, row in wda_df.iterrows():
        area_num = row["area_num"]
        cbsas = TX_WDA_TO_CBSA.get(area_num, [])
        if not cbsas:
            continue

        soc_minor = row["SOC Code"]
        base = row.get("Estimated Employment (2022)", 0) or 0
        proj = row.get("Projected Employment (2032)", 0) or 0

        try:
            base = int(float(base))
            proj = int(float(proj))
        except (ValueError, TypeError):
            continue

        for cbsa in cbsas:
            key = (cbsa, soc_minor)
            cbsa_minor_agg[key]["base"] += base
            cbsa_minor_agg[key]["proj"] += proj

    # Expand minor groups to detailed codes, store only CAGR
    rows = []
    for (cbsa, soc_minor), emp in cbsa_minor_agg.items():
        detailed_codes = minor_to_detailed.get(soc_minor, [])
        if not detailed_codes:
            continue

        base_total = emp["base"]
        proj_total = emp["proj"]
        if base_total <= 0 or proj_total <= 0:
            continue

        pct_change = (proj_total - base_total) / base_total * 100
        cagr = calc_cagr(base_total, proj_total, 10)
        area_name = TX_CBSA_NAMES.get(cbsa, cbsa)

        for det_soc in detailed_codes:
            rows.append((
                det_soc, "metro", cbsa, area_name,
                2022, 2032,
                None,  # base_emp: NULL (WDA doesn't match metro)
                None,  # proj_emp: NULL
                round(pct_change, 2),
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))

    print(f"    Texas: {len(rows):,} metro rows across {len(set(r[2] for r in rows))} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Florida: CareerSource region XLSX files — 2025-2033 projections
# ---------------------------------------------------------------------------

FL_CS_BASE_URL = "https://lmsresources.labormarketinfo.com/library/ep/2025_2033"

# CareerSource region number -> list of (CBSA code, CBSA name)
# Multi-region CBSAs (Tampa, Miami) have multiple regions mapped
FL_CS_TO_CBSA = {
    "01": [("37860", "Pensacola-Ferry Pass-Brent, FL")],
    "02": [("18880", "Crestview-Fort Walton Beach-Destin, FL")],
    # CS 03 (Chipola): rural, no major CBSA
    "04": [("37460", "Panama City, FL")],
    "05": [("45220", "Tallahassee, FL")],
    # CS 06 (North Florida): mostly rural
    "06": [("29380", "Lake City, FL")],
    "08": [("27260", "Jacksonville, FL")],
    "10": [("36100", "Ocala, FL")],
    "12": [("36740", "Orlando-Kissimmee-Sanford, FL")],
    "16": [("45300", "Tampa-St. Petersburg-Clearwater, FL")],   # part of Tampa CBSA
    "17": [("29460", "Lakeland-Winter Haven, FL")],
    "18": [("35840", "North Port-Sarasota-Bradenton, FL")],
    "19": [("42700", "Sebring-Avon Park, FL")],
    "20": [("38940", "Port St. Lucie, FL")],
    "21": [("33100", "Miami-Fort Lauderdale-Pompano Beach, FL")],  # Palm Beach part
    "22": [("33100", "Miami-Fort Lauderdale-Pompano Beach, FL")],  # Broward part
    "23": [("33100", "Miami-Fort Lauderdale-Pompano Beach, FL")],  # Miami-Dade part
    "24": [("15980", "Cape Coral-Fort Myers, FL"),
           ("34940", "Naples-Marco Island, FL")],
    "26": [("23540", "Gainesville, FL")],
    "27": [("37340", "Palm Bay-Melbourne-Titusville, FL"),
           ("19660", "Deltona-Daytona Beach-Ormond Beach, FL")],
    "28": [("45300", "Tampa-St. Petersburg-Clearwater, FL")],   # main Tampa part
}

# CBSAs that need aggregation from multiple CS regions
FL_MULTI_REGION_CBSAS = {
    "45300": ["16", "28"],         # Tampa = Pasco-Hernando + Hillsborough-Pinellas
    "33100": ["21", "22", "23"],   # Miami = Palm Beach + Broward + Miami-Dade
}


def load_florida_metros():
    """Load Florida metro projections from CareerSource region XLSX files."""
    from collections import defaultdict
    import numpy as np

    soc_re = re.compile(r"^\d{2}-\d{4}$")

    # Download and parse each region file
    region_data = {}  # region_num -> DataFrame of (soc_code, base_emp, proj_emp)
    for region_num in FL_CS_TO_CBSA:
        url = f"{FL_CS_BASE_URL}/p33cs{region_num}.xlsx"
        cache_path = RAW_DIR / f"fl_cs{region_num}.xlsx"

        if cache_path.exists():
            print(f"    FL CS{region_num}: Using cached {cache_path.name}")
            content = cache_path.read_bytes()
        else:
            print(f"    FL CS{region_num}: Downloading...")
            r = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=30)
            if r.status_code != 200:
                print(f"    FL CS{region_num}: HTTP {r.status_code}, skipping")
                continue
            content = r.content
            cache_path.write_bytes(content)
            time.sleep(0.3)

        try:
            df = pd.read_excel(
                io.BytesIO(content), sheet_name="Occs", header=None,
                engine="openpyxl"
            )
        except Exception as e:
            print(f"    FL CS{region_num}: Error reading Occs sheet: {e}")
            continue

        # Find header row (contains "SOC Code")
        header_idx = None
        for i in range(min(15, len(df))):
            row_vals = [str(v).strip() for v in df.iloc[i].values if pd.notna(v)]
            if any("SOC Code" in v for v in row_vals):
                header_idx = i
                break

        if header_idx is None:
            print(f"    FL CS{region_num}: Could not find header row, skipping")
            continue

        # Parse data
        parsed = []
        for i in range(header_idx + 1, len(df)):
            soc = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if not soc_re.match(soc):
                continue
            # Skip summary rows (XX-0000)
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base = int(float(df.iloc[i, 2])) if pd.notna(df.iloc[i, 2]) else 0
                proj = int(float(df.iloc[i, 3])) if pd.notna(df.iloc[i, 3]) else 0
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                parsed.append({"soc": soc, "base": base, "proj": proj})

        region_data[region_num] = parsed

    # Build CBSA-level data
    # For multi-region CBSAs, aggregate employment before computing CAGR
    cbsa_soc_agg = defaultdict(lambda: defaultdict(lambda: {"base": 0, "proj": 0}))

    for region_num, data_rows in region_data.items():
        cbsa_list = FL_CS_TO_CBSA.get(region_num, [])
        for cbsa_code, cbsa_name in cbsa_list:
            for item in data_rows:
                cbsa_soc_agg[cbsa_code][item["soc"]]["base"] += item["base"]
                cbsa_soc_agg[cbsa_code][item["soc"]]["proj"] += item["proj"]

    # Build output rows: CAGR only (base/proj = NULL per user directive)
    cbsa_names = {}
    for cbsa_list in FL_CS_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    rows = []
    for cbsa_code, soc_dict in cbsa_soc_agg.items():
        for soc, emp in soc_dict.items():
            if emp["base"] <= 0 or emp["proj"] <= 0:
                continue
            pct_change = (emp["proj"] - emp["base"]) / emp["base"] * 100
            cagr = calc_cagr(emp["base"], emp["proj"], 8)  # 2025-2033 = 8 years
            rows.append((
                soc, "metro", cbsa_code, cbsa_names.get(cbsa_code, cbsa_code),
                2025, 2033,
                None,   # base_emp: NULL (WDA doesn't match metro boundaries)
                None,   # proj_emp: NULL
                round(pct_change, 2),
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    FL: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Michigan: Prosperity Region XLSX files — 2022-2032 projections
# ---------------------------------------------------------------------------

MI_REGION_BASE_URL = "https://milmi.org/_docs/publications/projectionsdata"

# Michigan Prosperity Region -> (URL filename suffix, list of (CBSA, name))
# Regions are prosperity regions, not aligned with CBSA boundaries, so CAGR-only
MI_REGION_TO_CBSA = {
    # Region 1 (Upper Peninsula): no major CBSAs in OES data
    "2": (
        "LongTerm_OccupationProj_2032_Northwest_Prosperity_Region.xlsx",
        [("45900", "Traverse City, MI")],
    ),
    # Region 3 (Northeast): no major CBSAs in OES data
    "4": (
        "LongTerm_OccupationProj_2032_West_Michigan_Prosperity_Alliance.xlsx",
        [
            ("24340", "Grand Rapids-Wyoming-Kentwood, MI"),
            ("34740", "Muskegon-Norton Shores, MI"),
            ("35660", "Niles, MI"),
        ],
    ),
    "5": (
        "LongTerm_OccupationProj_2032_East_Central_Michigan_Prosperity_Region.xlsx",
        [
            ("13020", "Bay City, MI"),
            ("33220", "Midland, MI"),
            ("40980", "Saginaw, MI"),
        ],
    ),
    "6": (
        "LongTerm_OccupationProj_2032_East_Michigan_Prosperity_Region.xlsx",
        [("22420", "Flint, MI")],
    ),
    "7": (
        "LongTerm_OccupationProj_2032_South_Central_Prosperity_Region.xlsx",
        [
            ("12980", "Battle Creek, MI"),
            ("27100", "Jackson, MI"),
            ("28020", "Kalamazoo-Portage, MI"),
            ("29620", "Lansing-East Lansing, MI"),
        ],
    ),
    "8": (
        "LongTerm_OccupationProj_2032_Southwest_Prosperity_Region.xlsx",
        [("43780", "South Bend-Mishawaka, IN-MI")],
    ),
    "9": (
        "LongTerm_OccupationProj_2032_Southeast_Michigan_Prosperity_Region.xlsx",
        [
            ("11460", "Ann Arbor, MI"),
            ("33780", "Monroe, MI"),
        ],
    ),
    "10": (
        "LongTerm_OccupationProj_2032_Detroit_Metro_Prosperity_Region.xlsx",
        [("19820", "Detroit-Warren-Dearborn, MI")],
    ),
}


def load_michigan_metros():
    """Load Michigan metro projections from Prosperity Region XLSX files."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}

    for region_num, (filename, cbsa_list) in MI_REGION_TO_CBSA.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name

        url = f"{MI_REGION_BASE_URL}/{filename}"
        cache_path = RAW_DIR / f"mi_region{region_num}.xlsx"

        if cache_path.exists():
            print(f"    MI Region {region_num}: Using cached {cache_path.name}")
            content = cache_path.read_bytes()
        else:
            print(f"    MI Region {region_num}: Downloading...")
            r = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=30)
            if r.status_code != 200:
                print(f"    MI Region {region_num}: HTTP {r.status_code}, skipping")
                continue
            content = r.content
            cache_path.write_bytes(content)
            time.sleep(0.3)

        try:
            df = pd.read_excel(
                io.BytesIO(content), sheet_name=0, header=None, engine="openpyxl"
            )
        except Exception as e:
            print(f"    MI Region {region_num}: Error reading file: {e}")
            continue

        # Find header row (contains "SOC Code")
        header_idx = None
        for i in range(min(10, len(df))):
            row_vals = [str(v).strip() for v in df.iloc[i].values if pd.notna(v)]
            if any("SOC Code" in v for v in row_vals):
                header_idx = i
                break

        if header_idx is None:
            print(f"    MI Region {region_num}: Could not find header row, skipping")
            continue

        # Parse data: col 0 = SOC Code, col 2 = Base employment, col 3 = Projected
        for i in range(header_idx + 1, len(df)):
            soc = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if not soc_re.match(soc):
                continue
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base = int(float(df.iloc[i, 2])) if pd.notna(df.iloc[i, 2]) else 0
                proj = int(float(df.iloc[i, 3])) if pd.notna(df.iloc[i, 3]) else 0
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032 = 10 years
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code, cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # CAGR only — region != CBSA boundaries
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MI: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Oklahoma: Single XLSX with MSA-level occupational projections (2022-2032)
# ---------------------------------------------------------------------------

OK_MSA_URL = (
    "https://oklahoma.gov/content/dam/ok/en/oesc/documents/labor-market/"
    "wage-occupation-and-industry-reports/industry-and-occupational-employment-projections/"
    "long-term-employment-projections/2022-2032/"
    "ok-msa-industry-and-occupation-projections-2022-2032.xlsx"
)

# Oklahoma area names in the file -> CBSA code/name
OK_AREA_TO_CBSA = {
    "Enid, OK":           ("21420", "Enid, OK"),
    "Lawton, OK":         ("30020", "Lawton, OK"),
    "Oklahoma City, OK":  ("36420", "Oklahoma City, OK"),
    "Tulsa, OK":          ("46140", "Tulsa, OK"),
}


def load_oklahoma_metros():
    """Load Oklahoma metro projections from MSA-level XLSX (2022-2032)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")

    cache_path = RAW_DIR / "ok_msa_projections.xlsx"
    if cache_path.exists():
        print("    OK: Using cached ok_msa_projections.xlsx")
        content = cache_path.read_bytes()
    else:
        print("    OK: Downloading MSA projections...")
        r = requests.get(OK_MSA_URL, headers={"User-Agent": BROWSER_UA}, timeout=60)
        if r.status_code != 200:
            print(f"    OK: HTTP {r.status_code}")
            return []
        content = r.content
        cache_path.write_bytes(content)

    try:
        df = pd.read_excel(
            io.BytesIO(content), sheet_name="Occupational Projections",
            header=None, engine="openpyxl"
        )
    except Exception as e:
        print(f"    OK: Error reading sheet: {e}")
        return []

    # Find header row (contains "Area Name" and "SOC Code")
    header_idx = None
    for i in range(min(10, len(df))):
        row_vals = [str(v).strip() for v in df.iloc[i].values if pd.notna(v)]
        if any("Area Name" in v for v in row_vals) and any("SOC Code" in v for v in row_vals):
            header_idx = i
            break

    if header_idx is None:
        print("    OK: Could not find header row")
        return []

    rows = []
    # Parse: col 0 = Area Name, col 1 = SOC Code, col 3 = Emp 2022, col 4 = Emp 2032
    for i in range(header_idx + 1, len(df)):
        area = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
        soc = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""

        if area not in OK_AREA_TO_CBSA:
            continue
        if not soc_re.match(soc):
            continue
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        cbsa_code, cbsa_name = OK_AREA_TO_CBSA[area]

        try:
            base = int(float(df.iloc[i, 3])) if pd.notna(df.iloc[i, 3]) else 0
            proj = int(float(df.iloc[i, 4])) if pd.notna(df.iloc[i, 4]) else 0
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2022-2032 = 10 years
            rows.append((
                soc, "metro", cbsa_code, cbsa_name,
                2022, 2032,
                base, proj,  # OK data is MSA-level, store actuals
                round(pct_change, 2),
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    OK: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Ohio — MSA projections from OhioLMI dashboard export (CSV)
# ---------------------------------------------------------------------------

OH_MSA_CSV = RAW_DIR / "oh_msa_projections.csv"

OH_AREA_TO_CBSA = {
    "Akron, OH Metropolitan Statistical Area":            ("10420", "Akron, OH"),
    "Canton-Massillon, OH Metropolitan Statistical Area": ("15940", "Canton-Massillon, OH"),
    "Cincinnati, OH-KY-IN Metropolitan Statistical Area": ("17140", "Cincinnati, OH-KY-IN"),
    "Cleveland-Elyria, OH Metropolitan Statistical Area": ("17410", "Cleveland, OH"),
    "Columbus, OH Metropolitan Statistical Area":         ("18140", "Columbus, OH"),
    "Dayton, OH Metropolitan Statistical Area":           ("19430", "Dayton-Kettering-Beavercreek, OH"),
    "Toledo, OH Metropolitan Statistical Area":           ("45780", "Toledo, OH"),
    "Youngstown-Warren, OH Metropolitan Statistical Area":("49660", "Youngstown-Warren, OH"),
}


def load_ohio_metros():
    """Load Ohio metro projections from MSA-level CSV (2022-2032)."""
    if not OH_MSA_CSV.exists():
        print("    OH: CSV not found at raw/oh_msa_projections.csv, skipping")
        return []

    df = pd.read_csv(OH_MSA_CSV)

    # Pivot: each occupation+area has 8 rows (one per measure). We need
    # "Base Year Employment" and "Projected Employment".
    base_df = df[df["Measure Names"] == "Base Year Employment"][
        ["Area", "Occupation Code", "Measure Values"]
    ].rename(columns={"Measure Values": "base_emp"})

    proj_df = df[df["Measure Names"] == "Projected Employment"][
        ["Area", "Occupation Code", "Measure Values"]
    ].rename(columns={"Measure Values": "proj_emp"})

    merged = base_df.merge(proj_df, on=["Area", "Occupation Code"], how="inner")

    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []

    for _, row in merged.iterrows():
        area = row["Area"]
        if area not in OH_AREA_TO_CBSA:
            continue

        # Convert OH SOC format (no hyphen, e.g., "111011") to standard "11-1011"
        raw_soc = str(row["Occupation Code"]).strip()
        if len(raw_soc) == 6 and raw_soc.isdigit():
            soc = f"{raw_soc[:2]}-{raw_soc[2:]}"
        else:
            continue

        if not soc_re.match(soc):
            continue
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        cbsa_code, cbsa_name = OH_AREA_TO_CBSA[area]

        try:
            base = int(float(row["base_emp"]))
            proj = int(float(row["proj_emp"]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2022-2032
            rows.append((
                soc, "metro", cbsa_code, cbsa_name,
                2022, 2032,
                base, proj,  # MSA-level actuals
                round(pct_change, 2),
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    OH: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Alabama — Workforce Region projections from PDF
# ---------------------------------------------------------------------------

AL_PDF_URL = "https://www2.labor.alabama.gov/projections/OPLTAL.pdf"

# Region -> (detail page range 0-indexed start, end exclusive), list of (cbsa, name)
AL_REGION_TO_CBSA = {
    "Region 1": (52, 79, [
        ("26620", "Huntsville, AL"),
        ("22520", "Florence-Muscle Shoals, AL"),
        ("19460", "Decatur, AL"),
    ]),
    "Region 2": (81, 111, [
        ("11500", "Anniston-Oxford, AL"),
        ("23460", "Gadsden, AL"),
    ]),
    "Region 3": (113, 145, [
        ("46220", "Tuscaloosa, AL"),
    ]),
    "Region 4": (147, 181, [
        ("13820", "Birmingham, AL"),
    ]),
    "Region 5": (183, 216, [
        ("33860", "Montgomery, AL"),
        ("12220", "Auburn-Opelika, AL"),
    ]),
    "Region 6": (219, 248, [
        ("20020", "Dothan, AL"),
    ]),
    "Region 7": (250, 283, [
        ("33660", "Mobile, AL"),
        ("19300", "Daphne-Fairhope-Foley, AL"),
    ]),
}


def load_alabama_metros():
    """Load Alabama metro projections from PDF (2022-2032), parsed by region."""
    import pdfplumber

    cache_path = RAW_DIR / "al_projections.pdf"
    if cache_path.exists():
        print("    AL: Using cached al_projections.pdf")
    else:
        print("    AL: Downloading PDF...")
        r = requests.get(AL_PDF_URL, headers={"User-Agent": BROWSER_UA}, timeout=120)
        if r.status_code != 200:
            print(f"    AL: HTTP {r.status_code}")
            return []
        cache_path.write_bytes(r.content)

    soc_re = re.compile(r"^(\d{2}-\d{4})\s")
    num_re = re.compile(r"(-?[\d,]+\.?\d*)")

    pdf = pdfplumber.open(str(cache_path))
    rows = []
    cbsa_names = {}

    for region_name, (pg_start, pg_end, cbsa_list) in AL_REGION_TO_CBSA.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name

        region_occs = 0
        for pg_idx in range(pg_start, pg_end):
            if pg_idx >= len(pdf.pages):
                break
            page = pdf.pages[pg_idx]
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                line = line.strip()
                m = soc_re.match(line)
                if not m:
                    continue
                soc = m.group(1)
                if soc.endswith("0000") or soc.endswith("000"):
                    continue

                # Extract numbers; last 7 are: emp2022 emp2032 ann_growth total exit transfer growth
                nums = num_re.findall(line)
                if len(nums) < 7:
                    continue

                data_nums = nums[-7:]
                try:
                    base = int(data_nums[0].replace(",", ""))
                    proj = int(data_nums[1].replace(",", ""))
                except (ValueError, TypeError):
                    continue

                if base > 0 and proj > 0:
                    pct_change = (proj - base) / base * 100
                    cagr = calc_cagr(base, proj, 10)  # 2022-2032
                    for cbsa_code, _name in cbsa_list:
                        rows.append((
                            soc, "metro", cbsa_code,
                            cbsa_names.get(cbsa_code, cbsa_code),
                            2022, 2032,
                            None, None,  # CAGR only — region != CBSA
                            round(pct_change, 2),
                            round(cagr, 6) if cagr is not None else None,
                            "state_lmi",
                        ))
                    region_occs += 1

        print(f"    AL {region_name}: {region_occs} occupations")

    pdf.close()

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    AL: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Tennessee — Regional projections from TN DLWD (HTML table as .xls)
# ---------------------------------------------------------------------------

TN_FILE = RAW_DIR / "tn_msa_projections.xls"

TN_REGION_TO_CBSA = {
    "East TN":            [("28940", "Knoxville, TN")],
    "Greater Memphis":    [("32820", "Memphis, TN-MS-AR")],
    "Northeast TN":       [("27740", "Johnson City, TN"),
                           ("28700", "Kingsport-Bristol, TN-VA")],
    "Northern Middle TN": [("34980", "Nashville-Davidson--Murfreesboro--Franklin, TN"),
                           ("17300", "Clarksville, TN-KY")],
    "Southeast TN":       [("16860", "Chattanooga, TN-GA"),
                           ("17420", "Cleveland, TN")],
    "Southern Middle TN": [("34980", "Nashville-Davidson--Murfreesboro--Franklin, TN")],
    "Southwest TN":       [("27180", "Jackson, TN")],
    "Upper Cumberland":   [("34100", "Morristown, TN")],
    # "Northwest TN" has no major OES CBSA
}


def load_tennessee_metros():
    """Load Tennessee regional projections from HTML .xls file (2022-2032)."""
    if not TN_FILE.exists():
        print("    TN: File not found at raw/tn_msa_projections.xls, skipping")
        return []

    try:
        dfs = pd.read_html(str(TN_FILE))
        df = dfs[0]
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    except Exception as e:
        print(f"    TN: Error reading file: {e}")
        return []

    # Filter to regional data (areatype 17), exclude statewide (01)
    df = df[df["areatype"].astype(str) == "17"].copy()

    rows = []
    cbsa_names = {}
    for cbsa_list in TN_REGION_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for _, row in df.iterrows():
        area = str(row["areaname"]).strip()
        if area not in TN_REGION_TO_CBSA:
            continue

        cbsa_list = TN_REGION_TO_CBSA[area]

        # SOC code in matoccode — format "000000" (no hyphen)
        raw_soc = str(row["matoccode"]).strip()
        if len(raw_soc) == 6 and raw_soc.isdigit():
            soc = f"{raw_soc[:2]}-{raw_soc[2:]}"
        else:
            continue

        if not re.match(r"^\d{2}-\d{4}$", soc):
            continue
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        try:
            base = int(float(row["estoccprj"]))
            proj = int(float(row["projoccprj"]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2022-2032
            for cbsa_code, _name in cbsa_list:
                rows.append((
                    soc, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2022, 2032,
                    None, None,  # CAGR only — region != CBSA
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    TN: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Georgia — LWDA XLSX files from GA DOL
# ---------------------------------------------------------------------------

GA_LWDA_BASE_URL = "https://explorer.gdol.ga.gov/vosnet/mis/occupation/lt"

# LWDA number -> (list of (cbsa_code, cbsa_name))
# LWDAs covering multiple counties may map to multiple CBSAs
# Some LWDAs (e.g., 16) have no OES CBSA and are skipped
GA_LWDA_TO_CBSA = {
    "01": [("19140", "Dalton, GA"), ("40660", "Rome, GA")],           # Northwest Georgia
    "02": [("23580", "Gainesville, GA")],                              # Georgia Mountains
    "03": [("12060", "Atlanta-Sandy Springs-Roswell, GA")],            # City of Atlanta/Fulton
    "04": [("12060", "Atlanta-Sandy Springs-Roswell, GA")],            # Cobb County
    "05": [("12060", "Atlanta-Sandy Springs-Roswell, GA")],            # DeKalb County
    "06": [("12060", "Atlanta-Sandy Springs-Roswell, GA")],            # (combined with 03)
    "07": [("12060", "Atlanta-Sandy Springs-Roswell, GA")],            # Atlanta Regional
    "08": [("17980", "Columbus, GA-AL")],                              # Three Rivers
    "09": [("12020", "Athens-Clarke County, GA")],                     # Northeast Georgia
    "10": [("31420", "Macon-Bibb County, GA")],                       # Macon/Bibb County
    "11": [("31420", "Macon-Bibb County, GA"), ("47580", "Warner Robins, GA")],  # Middle Georgia
    "12": [("12260", "Augusta-Richmond County, GA-SC")],               # Central Savannah River
    "14": [("17980", "Columbus, GA-AL")],                              # Lower Chattahoochee
    "15": [("10500", "Albany, GA")],                                   # Middle Flint
    "17": [("10500", "Albany, GA")],                                   # Southwest Georgia
    "18": [("46660", "Valdosta, GA"), ("15260", "Brunswick-St. Simons, GA")],  # Southern Georgia
    "19": [("42340", "Savannah, GA"), ("25980", "Hinesville, GA")],    # Coastal Georgia
    # LWDA 13 (East Central) and 16 (Heart of GA Altamaha) have no major OES CBSAs
}


def load_georgia_metros():
    """Load Georgia metro projections from LWDA-level XLSX files (2022-2032 or 2020-2030)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}

    for lwda_num, cbsa_list in GA_LWDA_TO_CBSA.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name

        url = f"{GA_LWDA_BASE_URL}/occprj{lwda_num}.xlsx"
        cache_path = RAW_DIR / f"ga_occprj{lwda_num}.xlsx"

        if cache_path.exists():
            print(f"    GA LWDA {lwda_num}: Using cached {cache_path.name}")
            content = cache_path.read_bytes()
        else:
            print(f"    GA LWDA {lwda_num}: Downloading...")
            r = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=30)
            if r.status_code != 200:
                print(f"    GA LWDA {lwda_num}: HTTP {r.status_code}, skipping")
                continue
            content = r.content
            cache_path.write_bytes(content)
            time.sleep(0.3)

        try:
            df = pd.read_excel(
                io.BytesIO(content), sheet_name=0, header=None, engine="openpyxl"
            )
        except Exception as e:
            print(f"    GA LWDA {lwda_num}: Error reading file: {e}")
            continue

        # Detect projection period from header rows (row 3 or row 8)
        # Most files are 2022-2032 but LWDA 10 is 2020-2030
        base_year, proj_year = 2022, 2032
        for i in range(min(10, len(df))):
            cell = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if "2020 to 2030" in cell or "2020" in cell and "2030" in cell:
                base_year, proj_year = 2020, 2030
                break
            if "2022 to 2032" in cell or "2022" in cell and "2032" in cell:
                base_year, proj_year = 2022, 2032
                break

        n_years = proj_year - base_year

        # Find header row: contains "SOC" in col 0 and "Code" in next rows
        # GA format: row 8 = ["", "", "", "2022", "2032", "Total", "Percent", "Annual"]
        #            row 9 = ["SOC", "", "Typical", "Base", "Projected", ...]
        #            row 10 = ["Code", "Occupations", "Education", "Employment", ...]
        # Data starts after row 10 (or after an empty row)
        header_idx = None
        for i in range(min(15, len(df))):
            val = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if val == "Code":
                header_idx = i
                break

        if header_idx is None:
            print(f"    GA LWDA {lwda_num}: Could not find header row, skipping")
            continue

        # Parse data rows: col 0 = SOC Code, col 3 = Base Employment, col 4 = Projected Employment
        for i in range(header_idx + 1, len(df)):
            soc = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if not soc_re.match(soc):
                continue
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base_val = df.iloc[i, 3]
                proj_val = df.iloc[i, 4]
                # Skip suppressed data (marked with '*')
                if str(base_val).strip() == "*" or str(proj_val).strip() == "*":
                    continue
                base = int(float(base_val)) if pd.notna(base_val) else 0
                proj = int(float(proj_val)) if pd.notna(proj_val) else 0
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, n_years)
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code, cbsa_names.get(cbsa_code, cbsa_code),
                        base_year, proj_year,
                        None, None,  # CAGR only — LWDA != CBSA boundaries
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    GA: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Wisconsin — WDA XLSX file from DWD/JobCenter
# ---------------------------------------------------------------------------

WI_WDA_URL = "https://jobcenterofwisconsin.com/wisconomy/wits_info/downloads/projections/occ_lt_wda.xlsx"

# Sheet name -> list of (cbsa_code, cbsa_name) covered by that WDA
WI_WDA_TO_CBSA = {
    "WDA1 SOUTHEAST": [
        ("28450", "Kenosha, WI"),
        ("39540", "Racine-Mount Pleasant, WI"),
    ],
    "WDA2 MILWAUKEE": [
        ("33340", "Milwaukee-Waukesha, WI"),
    ],
    "WDA3 WOW": [
        ("33340", "Milwaukee-Waukesha, WI"),  # Waukesha-Ozaukee-Washington overlap
    ],
    "WDA4 FOX VALLEY": [
        ("11540", "Appleton, WI"),
        ("22540", "Fond du Lac, WI"),
        ("36780", "Oshkosh-Neenah, WI"),
    ],
    "WDA5 BAY AREA": [
        ("24580", "Green Bay, WI"),
        ("43100", "Sheboygan, WI"),
    ],
    "WDA6 NORTH CENTRAL": [
        ("48140", "Wausau, WI"),
    ],
    # WDA7 NORTHWEST — no major OES CBSAs (Eau Claire is in WDA8)
    "WDA8 WEST CENTRAL": [
        ("20740", "Eau Claire, WI"),
    ],
    "WDA9 WESTERN": [
        ("29100", "La Crosse-Onalaska, WI-MN"),
    ],
    "WDA10 SOUTH CENTRAL": [
        ("31540", "Madison, WI"),
        ("27500", "Janesville-Beloit, WI"),
    ],
    # WDA11 SOUTHWEST — no major OES CBSAs (Grant, Iowa, Lafayette counties)
}


def load_wisconsin_metros():
    """Load Wisconsin metro projections from WDA-level XLSX (2022-2032)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")

    cache_path = RAW_DIR / "wi_occ_lt_wda.xlsx"
    if cache_path.exists():
        print("    WI: Using cached wi_occ_lt_wda.xlsx")
        content = cache_path.read_bytes()
    else:
        print("    WI: Downloading WDA projections...")
        r = requests.get(WI_WDA_URL, headers={"User-Agent": BROWSER_UA}, timeout=60)
        if r.status_code != 200:
            print(f"    WI: HTTP {r.status_code}")
            return []
        content = r.content
        cache_path.write_bytes(content)

    try:
        xls = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
    except Exception as e:
        print(f"    WI: Error opening file: {e}")
        return []

    rows = []
    cbsa_names = {}
    for sheet_name, cbsa_list in WI_WDA_TO_CBSA.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name

        if sheet_name not in xls.sheet_names:
            print(f"    WI: Sheet '{sheet_name}' not found, skipping")
            continue

        df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine="openpyxl")

        # Find header row containing "SOC Code"
        header_idx = None
        for i in range(min(10, len(df))):
            val = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if "SOC Code" in val:
                header_idx = i
                break

        if header_idx is None:
            print(f"    WI {sheet_name}: Could not find header row, skipping")
            continue

        # Parse data: col 0 = SOC Code, col 2 = 2022 Employment, col 3 = 2032 Projected
        for i in range(header_idx + 1, len(df)):
            soc = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if not soc_re.match(soc):
                continue
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base = int(float(df.iloc[i, 2])) if pd.notna(df.iloc[i, 2]) else 0
                proj = int(float(df.iloc[i, 3])) if pd.notna(df.iloc[i, 3]) else 0
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code, cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # CAGR only — WDA != CBSA boundaries
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    WI: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Oregon — 8 regional XLSX files from QualityInfo.org (2024-2034)
# ---------------------------------------------------------------------------

OR_REGION_FILES = {
    "east_cascades":     "or_east_cascades_proj.xlsx",
    "eastern_oregon":    "or_eastern_oregon_proj.xlsx",
    "lane":              "or_lane_proj.xlsx",
    "mid-valley":        "or_mid-valley_proj.xlsx",
    "northwest_oregon":  "or_northwest_oregon_proj.xlsx",
    "portland_tri-county": "or_portland_tri-county_proj.xlsx",
    "rogue_valley":      "or_rogue_valley_proj.xlsx",
    "southwestern_oregon": "or_southwestern_oregon_proj.xlsx",
}

OR_REGION_URLS = {
    "east_cascades":       "https://www.qualityinfo.org/documents/20117/102288/East%20Cascades%20Occupational%20Employment%20Projections%202024-2034/d7370b15-5457-825d-b925-e9c1f663ed67?version=2.0",
    "eastern_oregon":      "https://www.qualityinfo.org/documents/20117/102288/Eastern%20Oregon%20Occupational%20Employment%20Projections%202024-2034/40dfa669-0e3f-aadc-a4f9-4cfb526de33e?version=2.0",
    "lane":                "https://www.qualityinfo.org/documents/20117/102288/Lane%20Occupational%20Employment%20Projections%202024-2034/5ee79f6e-dbc4-7bb5-e0d5-c066f5f5bcb8?version=2.0",
    "mid-valley":          "https://www.qualityinfo.org/documents/20117/102288/Mid-Valley%20Occupational%20Employment%20Projections%202024-2034/577bed62-b2a5-d57d-c6ee-58e0c4281e82?version=2.0",
    "northwest_oregon":    "https://www.qualityinfo.org/documents/20117/102288/Northwest%20Oregon%20Occupational%20Employment%20Projections%202024-2034/f4f47f49-f82f-817d-f493-ba9acb06984f?version=2.1",
    "portland_tri-county": "https://www.qualityinfo.org/documents/20117/102288/Portland%20Tri-County%20Occupational%20Projections%202024-2034/d9d47cad-8384-b14e-f61f-7a6f34e45902?version=2.1",
    "rogue_valley":        "https://www.qualityinfo.org/documents/20117/102288/Rogue%20Valley%20Occupational%20Employment%20Projections%202024-2034/cd3576c4-bf12-5581-19cc-6f38123eaf1d?version=2.1",
    "southwestern_oregon": "https://www.qualityinfo.org/documents/20117/102288/Southwestern%20Oregon%20Occupational%20Employment%20Projections%202024-2034/e731d89f-3af2-f245-7d83-03706c8411c0?version=2.1",
}

OR_REGION_TO_CBSA = {
    "east_cascades":       [("13460", "Bend, OR")],
    "lane":                [("21660", "Eugene-Springfield, OR")],
    "mid-valley":          [("41420", "Salem, OR"), ("10540", "Albany, OR"), ("18700", "Corvallis, OR")],
    "northwest_oregon":    [("18700", "Corvallis, OR")],
    "portland_tri-county": [("38900", "Portland-Vancouver-Hillsboro, OR-WA")],
    "rogue_valley":        [("32780", "Medford, OR"), ("24420", "Grants Pass, OR")],
    # eastern_oregon and southwestern_oregon have no OES CBSAs
}


def load_oregon_metros():
    """Load Oregon regional occupational projections (2024-2034)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in OR_REGION_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for region_key, cbsa_list in OR_REGION_TO_CBSA.items():
        fname = OR_REGION_FILES[region_key]
        cached = RAW_DIR / fname
        if not cached.exists():
            url = OR_REGION_URLS[region_key]
            print(f"    OR: Downloading {region_key}...")
            r = requests.get(url, headers=HEADERS)
            r.raise_for_status()
            cached.write_bytes(r.content)
        else:
            print(f"    OR {region_key}: Using cached {fname}")

        df = pd.read_excel(cached, sheet_name=0, header=None)
        # Find header row (contains 'SOC Level' or 'Occupation Code')
        header_row = None
        for i in range(min(10, len(df))):
            vals = [str(v).strip() if pd.notna(v) else "" for v in df.iloc[i]]
            if "Occupation Code" in vals or "SOC Level" in vals:
                header_row = i
                break
        if header_row is None:
            print(f"    OR {region_key}: Could not find header row, skipping")
            continue

        df.columns = [str(v).strip() if pd.notna(v) else f"col{j}" for j, v in enumerate(df.iloc[header_row])]
        df = df.iloc[header_row + 1:].reset_index(drop=True)

        for _, row_data in df.iterrows():
            soc = str(row_data.get("Occupation Code", "")).strip()
            if not soc_re.match(soc):
                continue
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            base_val = row_data.get("Employment  2024", row_data.get("Employment 2024"))
            proj_val = row_data.get("Projected Employment  2034", row_data.get("Projected Employment 2034"))

            if pd.isna(base_val) or pd.isna(proj_val):
                continue
            base_str = str(base_val).strip()
            proj_str = str(proj_val).strip()
            if base_str == "-s-" or proj_str == "-s-":
                continue

            try:
                base = int(float(base_str))
                proj = int(float(proj_str))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2024-2034
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2024, 2034,
                        None, None,  # CAGR only — region != CBSA
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    OR: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Wyoming — Sub-state projections XLSX from DWS (2022-2032)
# ---------------------------------------------------------------------------

WY_FILE = RAW_DIR / "wy_substate_projections.xlsx"
WY_URL = "https://doe.state.wy.us/lmi/projections/2025/LT-Substate/Sub-State_Projections_2022-32_REVISED.xlsx"

# Region name in Master sheet -> (cbsa_code, cbsa_name)
WY_REGION_TO_CBSA = {
    "Casper, WY":   [("16220", "Casper, WY")],
    "Cheyenne, WY": [("16940", "Cheyenne, WY")],
    # Central-Southeast, Northeast, Northwest, Southwest have no OES CBSAs
}


def load_wyoming_metros():
    """Load Wyoming sub-state occupational projections (2022-2032)."""
    cached = WY_FILE
    if not cached.exists():
        print("    WY: Downloading sub-state projections...")
        r = requests.get(WY_URL, headers=HEADERS)
        r.raise_for_status()
        cached.write_bytes(r.content)
    else:
        print(f"    WY: Using cached {cached.name}")

    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in WY_REGION_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    wb = openpyxl.load_workbook(cached, read_only=True, data_only=True)
    ws = wb["Master"]

    for row_data in ws.iter_rows(min_row=6, values_only=True):
        region = str(row_data[0]).strip() if row_data[0] else ""
        if region not in WY_REGION_TO_CBSA:
            continue

        soc = str(row_data[1]).strip() if row_data[1] else ""
        if not soc_re.match(soc):
            continue
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        try:
            base = int(float(row_data[3]))  # col 3 = Employment 2022
            proj = int(float(row_data[4]))  # col 4 = Employment 2032
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)
            cbsa_list = WY_REGION_TO_CBSA[region]
            for cbsa_code, _name in cbsa_list:
                rows.append((
                    soc, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2022, 2032,
                    base, proj,  # MSA-level data — actual employment
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))
    wb.close()

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    WY: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Indiana — EGR-level projections from Hoosiers by the Numbers (2023-2033)
# ---------------------------------------------------------------------------

IN_FILE = RAW_DIR / "in_projections.xls"

# EGR -> list of (cbsa_code, cbsa_name)
# County composition from hoosierdata.in.gov
# EGR 1: Lake, Porter, LaPorte, Jasper, Newton, Pulaski, Starke → Gary/Michigan City
# EGR 2: Elkhart, Kosciusko, Marshall, Fulton, Miami, Wabash, St. Joseph, LaGrange → Elkhart, South Bend
# EGR 3: Allen, Adams, DeKalb, Grant, Huntington, Noble, Steuben, Wells, Whitley → Fort Wayne
# EGR 4: Cass, Clinton, Howard, Tippecanoe, Carroll, Fountain, Montgomery, Warren, White, Benton → Kokomo, Lafayette
# EGR 5: Marion + surrounding donut counties (Hamilton, Hancock, Hendricks, Johnson, Madison, Morgan, Boone, Shelby) → Indianapolis
# EGR 6: Henry, Randolph, Wayne, Rush, Fayette, Union, Delaware, Jay, Blackford → Muncie
# EGR 7: Vigo, Clay, Owen, Putnam, Sullivan, Vermillion, Parke, Greene → Terre Haute
# EGR 8: Monroe, Brown, Lawrence, Orange, Martin, Daviess, Knox, Gibson, Pike → Bloomington
# EGR 9: Bartholomew, Decatur, Jennings, Ripley, Dearborn, Ohio, Franklin, Switzerland → Columbus
# EGR 10: Clark, Floyd, Harrison, Scott, Jackson, Jefferson, Washington → Louisville metro (KY-IN)
# EGR 11: Vanderburgh, Posey, Warrick, Spencer, Perry, Dubois, Crawford → Evansville
IN_EGR_TO_CBSA = {
    "EGR 1 , IN":  [("33140", "Michigan City-La Porte, IN")],
    "EGR 2 , IN":  [("21140", "Elkhart-Goshen, IN"), ("43780", "South Bend-Mishawaka, IN-MI")],
    "EGR 3 , IN":  [("23060", "Fort Wayne, IN")],
    "EGR 4 , IN":  [("29020", "Kokomo, IN"), ("29200", "Lafayette-West Lafayette, IN")],
    "EGR 5 , IN":  [("26900", "Indianapolis-Carmel-Greenwood, IN")],
    "EGR 6 , IN":  [("34620", "Muncie, IN")],
    "EGR 7 , IN":  [("45460", "Terre Haute, IN")],
    "EGR 8 , IN":  [("14020", "Bloomington, IN")],
    "EGR 9 , IN":  [("18020", "Columbus, IN")],
    "EGR 10 , IN": [],  # Clark/Floyd are IN side of Louisville KY-IN metro — skip to avoid double-counting
    "EGR 11 , IN": [("21780", "Evansville, IN")],
}


def load_indiana_metros():
    """Load Indiana EGR-level occupational projections (2023-2033) from HTML .xls."""
    if not IN_FILE.exists():
        print("    IN: File not found at raw/in_projections.xls, skipping")
        return []

    try:
        dfs = pd.read_html(str(IN_FILE))
        df = dfs[1]  # Second table has the data
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    except Exception as e:
        print(f"    IN: Error reading file: {e}")
        return []

    soc_re = re.compile(r"^\d{6}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in IN_EGR_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    # Columns: Geography, State Fips, Geography Code, Year,
    #          Occ. Title/Code (title), Occ. Title/Code (soc), Base Year Emp., Projected Year Emp.
    for _, row_data in df.iterrows():
        geo = str(row_data.iloc[0]).strip()
        if geo not in IN_EGR_TO_CBSA:
            continue
        cbsa_list = IN_EGR_TO_CBSA[geo]
        if not cbsa_list:
            continue

        raw_soc = str(row_data.iloc[5]).strip()  # 2nd 'Occ. Title/Code' = SOC code
        if not soc_re.match(raw_soc):
            continue
        soc = f"{raw_soc[:2]}-{raw_soc[2:]}"

        # Skip summary codes
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        try:
            base = int(float(row_data.iloc[6]))
            proj = int(float(row_data.iloc[7]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2023-2033
            for cbsa_code, _name in cbsa_list:
                rows.append((
                    soc, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2023, 2033,
                    None, None,  # CAGR only — EGR != CBSA
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    IN: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Louisiana — RLMA-level projections from LAWorks (2022-2032)
# ---------------------------------------------------------------------------

LA_RLMA_BASE_URL = "https://www.laworks.net/Downloads/LMI"

# RLMA number -> list of (cbsa_code, cbsa_name)
# RLMA 1: New Orleans region (Jefferson, Orleans, Plaquemines, St. Bernard, St. Charles, St. James, St. John, St. Tammany)
# RLMA 2: Baton Rouge region (E. Baton Rouge, Ascension, E. Feliciana, Iberville, Livingston, Pointe Coupee,
#          St. Helena, Tangipahoa, W. Baton Rouge, W. Feliciana, Washington)
# RLMA 3: Houma region (Assumption, Lafourche, Terrebonne)
# RLMA 4: Lafayette region (Acadia, Evangeline, Iberia, Lafayette, St. Landry, St. Martin, St. Mary, Vermilion)
# RLMA 5: Lake Charles region (Allen, Beauregard, Calcasieu, Cameron, Jeff Davis, Vernon)
# RLMA 6: Alexandria region (Avoyelles, Catahoula, Concordia, Grant, LaSalle, Rapides, Winn)
# RLMA 7: Shreveport region (Bienville, Bossier, Caddo, Claiborne, De Soto, Lincoln, Natchitoches, Red River, Sabine, Webster)
# RLMA 8: Monroe region (Caldwell, E. Carroll, Franklin, Jackson, Madison, Morehouse, Ouachita, Richland, Tensas, Union, W. Carroll)
LA_RLMA_TO_CBSA = {
    1: [("35380", "New Orleans-Metairie, LA"), ("43640", "Slidell-Mandeville-Covington, LA")],
    2: [("12940", "Baton Rouge, LA"), ("25220", "Hammond, LA")],
    3: [("26380", "Houma-Bayou Cane-Thibodaux, LA")],
    4: [("29180", "Lafayette, LA")],
    5: [("29340", "Lake Charles, LA")],
    6: [("10780", "Alexandria, LA")],
    7: [("43340", "Shreveport-Bossier City, LA")],
    8: [("33740", "Monroe, LA")],
}


def load_louisiana_metros():
    """Load Louisiana RLMA-level occupational projections (2022-2032)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in LA_RLMA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for rlma_num, cbsa_list in LA_RLMA_TO_CBSA.items():
        fname = f"la_rlma{rlma_num}.xls"
        cached = RAW_DIR / fname
        if not cached.exists():
            url = f"{LA_RLMA_BASE_URL}/20222032Occupations_AllProjRLMA{rlma_num}.xls"
            print(f"    LA RLMA {rlma_num}: Downloading...")
            r = requests.get(url, headers=HEADERS)
            r.raise_for_status()
            cached.write_bytes(r.content)
        else:
            print(f"    LA RLMA {rlma_num}: Using cached {fname}")

        df = pd.read_excel(cached, header=None)

        # Find header row (contains "Occ." and "Code")
        header_row = None
        for i in range(min(10, len(df))):
            row_str = " ".join([str(v) for v in df.iloc[i] if pd.notna(v)])
            if "Occ" in row_str and "Code" in row_str:
                header_row = i
                break
        if header_row is None:
            print(f"    LA RLMA {rlma_num}: Could not find header, skipping")
            continue

        # Data starts after header row
        # Col 1 = SOC Code, Col 3 = 2022 Estimate, Col 4 = 2032 Projected
        for i in range(header_row + 1, len(df)):
            soc = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
            if not soc_re.match(soc):
                continue
            # Skip summary codes
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base = int(float(df.iloc[i, 3]))
                proj = int(float(df.iloc[i, 4]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # CAGR only — RLMA != CBSA
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    LA: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Maryland – WIA county-level projections (2022-2032)
# ---------------------------------------------------------------------------
MD_FILE = RAW_DIR / "md_wias.xlsx"

# Sheet name -> list of (CBSA code, CBSA name) tuples
MD_WIA_TO_CBSA = {
    "Anne Arundel":   [("12580", "Baltimore-Columbia-Towson, MD")],
    "Baltimore":      [("12580", "Baltimore-Columbia-Towson, MD")],
    "Baltimore City": [("12580", "Baltimore-Columbia-Towson, MD")],
    "Carroll":        [("12580", "Baltimore-Columbia-Towson, MD")],
    "Howard":         [("12580", "Baltimore-Columbia-Towson, MD")],
    "Susquehanna":    [("12580", "Baltimore-Columbia-Towson, MD"),
                       ("37980", "Philadelphia-Camden-Wilmington, PA-NJ-DE-MD")],
    "Upper Shore":    [("12580", "Baltimore-Columbia-Towson, MD")],
    "Montgomery":     [("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV")],
    "Prince George's":[("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV")],
    "Frederick":      [("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV")],
    "Southern MD":    [("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV"),
                       ("30500", "Lexington Park, MD")],
    "Lower Shore":    [("41540", "Salisbury, MD")],
    "Sheet13":        [("25180", "Hagerstown-Martinsburg, MD-WV")],  # Western Maryland
}


def load_maryland_metros():
    """Load Maryland WIA county-level occupational projections (2022-2032)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in MD_WIA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    xls = pd.ExcelFile(MD_FILE)

    for sheet_name, cbsa_list in MD_WIA_TO_CBSA.items():
        if sheet_name not in xls.sheet_names:
            print(f"    MD: Sheet '{sheet_name}' not found, skipping")
            continue

        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)

        # Data layout: Row 0=title, Row 2=header, Row 4=sub-header, data from Row 6
        # Col 0=Occ Code, Col 2=SOCLevel, Col 3=Emp 2022, Col 4=Emp 2032
        for i in range(6, len(df)):
            soc = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
            if not soc_re.match(soc):
                continue

            soc_level = df.iloc[i, 2] if pd.notna(df.iloc[i, 2]) else None
            if soc_level != 4:  # Only detailed occupations
                continue

            try:
                base = int(float(df.iloc[i, 3]))
                proj = int(float(df.iloc[i, 4]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # CAGR only — WIA != CBSA
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MD: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Virginia – LWDA-level projections (2022-2032) from IOMatrix CSV (long format)
# ---------------------------------------------------------------------------
VA_FILE = RAW_DIR / "va_lwda_projections.csv"

VA_LWDA_TO_CBSA = {
    "Southwestern Virginia (LWDA I)":           [("28700", "Kingsport-Bristol, TN-VA")],
    "New River/Mt. Rogers (LWDA II)":           [("13980", "Blacksburg-Christiansburg-Radford, VA")],
    "Western Virginia (LWDA III)":              [("40220", "Roanoke, VA")],
    "Shenandoah Valley (LWDA IV)":              [("25500", "Harrisonburg, VA"),
                                                  ("44420", "Staunton-Stuarts Draft, VA"),
                                                  ("49020", "Winchester, VA-WV")],
    "Piedmont Workforce Network (LWDA VI)":     [("16820", "Charlottesville, VA")],
    "Region 2000/Central VA (LWDA VII)":        [("31340", "Lynchburg, VA")],
    "Capital Region Workforce Partnership (LWDA IX)": [("40060", "Richmond, VA")],
    "Crater Area (LWDA XV)":                    [("40060", "Richmond, VA")],
    "Northern Virginia (LWDA XI)":              [("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV")],
    "Alexandria/Arlington (LWDA XII)":          [("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV")],
    "Hampton Roads (LWDA XVI)":                 [("47260", "Virginia Beach-Chesapeake-Norfolk, VA-NC")],
    # LWDAs with no matching OES CBSAs — skipped:
    # Bay Consortium (LWDA XIII), South Central (LWDA VIII), West Piedmont (LWDA XVII)
}


def load_virginia_metros():
    """Load Virginia LWDA-level occupational projections (2022-2032).

    Source data is in long format with separate rows per measure.
    We pivot to get Base Year Employment and Projected Employment per occupation/area.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in VA_LWDA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    df = pd.read_csv(VA_FILE)

    # Filter to the two measures we need
    base_df = df[df["Measure Names"] == "Base Year Employment"][
        ["Area", "Occupation Code", "Measure Values"]
    ].rename(columns={"Measure Values": "base_emp"})

    proj_df = df[df["Measure Names"] == "Projected Employment"][
        ["Area", "Occupation Code", "Measure Values"]
    ].rename(columns={"Measure Values": "proj_emp"})

    merged = pd.merge(base_df, proj_df, on=["Area", "Occupation Code"], how="inner")

    for _, row in merged.iterrows():
        area = row["Area"]
        if area not in VA_LWDA_TO_CBSA:
            continue

        # Format occ code: 111021 -> "11-1021"
        occ_raw = int(row["Occupation Code"])
        occ_str = f"{occ_raw:06d}"
        soc = f"{occ_str[:2]}-{occ_str[2:]}"

        if not soc_re.match(soc):
            continue
        # Skip summary codes (major/minor groups)
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        try:
            base = int(float(row["base_emp"]))
            proj = int(float(row["proj_emp"]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2022-2032

            for cbsa_code, _name in VA_LWDA_TO_CBSA[area]:
                rows.append((
                    soc, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2022, 2032,
                    None, None,  # CAGR only — LWDA != CBSA
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    VA: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Massachusetts – WDA-level projections (2023-2033), 16 individual CSVs
# ---------------------------------------------------------------------------

# Filename (in raw/) -> list of (CBSA code, CBSA name) tuples
MA_WDA_TO_CBSA = {
    "ma_berkshire.csv":         [("38340", "Pittsfield, MA")],
    "ma_cape island.csv":       [("12700", "Barnstable Town, MA")],
    "ma_FranklinHampshire.csv": [("11200", "Amherst Town-Northampton, MA"),
                                  ("44140", "Springfield, MA")],
    "ma_hampden.csv":           [("44140", "Springfield, MA")],
    "ma_boston.csv":             [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_metro north.csv":       [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_metro south west.csv":  [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_greater lowell.csv":    [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_merrimack valley.csv":  [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_north shore.csv":       [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_south shore.csv":       [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_brockton.csv":          [("14460", "Boston-Cambridge-Newton, MA-NH")],
    "ma_bristol.csv":           [("39300", "Providence-Warwick, RI-MA")],
    "ma_greater new bedford.csv": [("39300", "Providence-Warwick, RI-MA")],
    "ma_central.csv":           [("49340", "Worcester, MA")],
    "ma_north central.csv":     [("49340", "Worcester, MA")],
}


def load_massachusetts_metros():
    """Load Massachusetts WDA-level occupational projections (2023-2033)."""
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in MA_WDA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for filename, cbsa_list in MA_WDA_TO_CBSA.items():
        filepath = RAW_DIR / filename
        if not filepath.exists():
            print(f"    MA: {filename} not found, skipping")
            continue

        # CSV format: row 0 = title, row 1 = blank, row 2 = header, data from row 3
        df = pd.read_csv(filepath, skiprows=2)

        for _, row in df.iterrows():
            # SOC code is 6 digits without hyphen, with trailing spaces
            soc_raw = str(row.iloc[0]).strip()
            if not soc_raw or len(soc_raw) != 6 or not soc_raw.isdigit():
                continue
            soc = f"{soc_raw[:2]}-{soc_raw[2:]}"

            if not soc_re.match(soc):
                continue
            # Skip summary codes
            if soc.endswith("0000") or soc.endswith("000"):
                continue

            try:
                base = int(str(row.iloc[2]).replace(",", "").strip())
                proj = int(str(row.iloc[3]).replace(",", "").strip())
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2023-2033
                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2023, 2033,
                        None, None,  # CAGR only — WDA != CBSA
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MA: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# South Carolina – WDA-level projections (2022-2032) from IOMatrix CSV
# ---------------------------------------------------------------------------
SC_FILE = RAW_DIR / "sc_lwda_projections.csv"

SC_WDA_TO_CBSA = {
    "Catawba":        [("16740", "Charlotte-Concord-Gastonia, NC-SC")],
    "Greenville":     [("24860", "Greenville-Anderson-Greer, SC")],
    "Lowcountry":     [("25940", "Hilton Head Island-Bluffton-Port Royal, SC")],
    "Lower Savannah": [("12260", "Augusta-Richmond County, GA-SC")],
    "Midlands":       [("17900", "Columbia, SC")],
    "Pee Dee":        [("22500", "Florence, SC")],
    "Santee-Lynches": [("44940", "Sumter, SC")],
    "Trident":        [("16700", "Charleston-North Charleston, SC")],
    "Upstate":        [("43900", "Spartanburg, SC")],
    "Waccamaw":       [("34820", "Myrtle Beach-Conway-North Myrtle Beach, SC")],
    "Worklink":       [("24860", "Greenville-Anderson-Greer, SC")],
    # Upper Savannah — no matching OES CBSAs (Greenwood is micropolitan)
}


def load_south_carolina_metros():
    """Load South Carolina WDA-level occupational projections (2022-2032).

    Same IOMatrix long format as Virginia — pivot Base/Projected measures.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in SC_WDA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    df = pd.read_csv(SC_FILE)

    base_df = df[df["Measure Names"] == "Base Year Employment"][
        ["Area", "Occupation Code", "Measure Values"]
    ].rename(columns={"Measure Values": "base_emp"})

    proj_df = df[df["Measure Names"] == "Projected Employment"][
        ["Area", "Occupation Code", "Measure Values"]
    ].rename(columns={"Measure Values": "proj_emp"})

    merged = pd.merge(base_df, proj_df, on=["Area", "Occupation Code"], how="inner")

    for _, row in merged.iterrows():
        area = row["Area"]
        if area not in SC_WDA_TO_CBSA:
            continue

        occ_raw = int(row["Occupation Code"])
        occ_str = f"{occ_raw:06d}"
        soc = f"{occ_str[:2]}-{occ_str[2:]}"

        if not soc_re.match(soc):
            continue
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        try:
            base = int(float(row["base_emp"]))
            proj = int(float(row["proj_emp"]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2022-2032

            for cbsa_code, _name in SC_WDA_TO_CBSA[area]:
                rows.append((
                    soc, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2022, 2032,
                    None, None,  # CAGR only — WDA != CBSA
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    SC: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Missouri — 9 regional XLSX files, "All Occupations" sheet (2022-2032)
# ---------------------------------------------------------------------------

MO_FILES = {
    "raw/mo_kansas_city.xlsx":   [("28140", "Kansas City, MO-KS")],
    "raw/mo_saint_louis.xlsx":   [("41180", "St. Louis, MO-IL")],
    "raw/mo_southeast.xlsx":     [("16020", "Cape Girardeau, MO-IL")],
    "raw/mo_ozark.xlsx":         [("44180", "Springfield, MO")],
    "raw/mo_central.xlsx":       [("17860", "Columbia, MO"),
                                  ("27620", "Jefferson City, MO")],
    "raw/mo_southwest.xlsx":     [("27900", "Joplin, MO")],
    "raw/mo_north.xlsx":         [("41140", "St. Joseph, MO-KS")],
    # West Central and South Central — no OES CBSAs, skip
}


def load_missouri_metros():
    """Load Missouri regional occupational projections (2022-2032).

    Each XLSX has a sheet containing 'All Occupations' in its name.
    Detailed occupation rows have: col 0=Grade, col 2=SOC Code, col 4=Est Emp, col 5=Proj Emp.
    Major group rows have SOC in col 2 as well (XX-0000 pattern) — we skip those.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in MO_FILES.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for fname, cbsa_list in MO_FILES.items():
        fpath = RAW_DIR / Path(fname).name
        if not fpath.exists():
            print(f"    MO: Missing {fpath.name}, skipping")
            continue

        xl = pd.ExcelFile(fpath)
        # Find sheet with "All Occ" in name
        sheet = None
        for s in xl.sheet_names:
            if "All Occ" in s or "All occ" in s:
                sheet = s
                break
        if sheet is None:
            print(f"    MO: No 'All Occupations' sheet in {fpath.name}, skipping")
            continue

        df = pd.read_excel(fpath, sheet_name=sheet, header=None)

        for _, row in df.iterrows():
            # Detailed rows have SOC code in col 2 (e.g., "11-1011")
            soc_val = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
            if not soc_re.match(soc_val):
                continue
            # Skip major groups (XX-0000) and broad groups (XX-X000)
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue

            try:
                base = int(float(row.iloc[4]))
                proj = int(float(row.iloc[5]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MO: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Iowa — 6 LWDA XLSX files, "SOC" sheet (2022-2032)
# ---------------------------------------------------------------------------

IA_FILES = {
    "raw/ia_central.xlsx":           [("19780", "Des Moines-West Des Moines, IA"),
                                      ("11180", "Ames, IA")],
    "raw/ia_east_central.xlsx":      [("16300", "Cedar Rapids, IA"),
                                      ("26980", "Iowa City, IA")],
    "raw/ia_mississippi_valley.xlsx": [("19340", "Davenport-Moline-Rock Island, IA-IL")],
    "raw/ia_northeast.xlsx":         [("47940", "Waterloo-Cedar Falls, IA"),
                                      ("20220", "Dubuque, IA")],
    "raw/ia_iowa_plains.xlsx":       [("43580", "Sioux City, IA-NE-SD"),
                                      ("36540", "Omaha-Council Bluffs, NE-IA")],
    # South Central — no OES CBSAs, skip
}


def load_iowa_metros():
    """Load Iowa LWDA-level occupational projections (2022-2032).

    Each XLSX has a 'SOC' sheet with header at row 2.
    Columns: SOC (hyphenated), ..., 2022 Estimated, 2032 Projected.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in IA_FILES.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for fname, cbsa_list in IA_FILES.items():
        fpath = RAW_DIR / Path(fname).name
        if not fpath.exists():
            print(f"    IA: Missing {fpath.name}, skipping")
            continue

        df = pd.read_excel(fpath, sheet_name="SOC", header=2)

        soc_col = "SOC"
        base_col = "2022 Estimated"
        proj_col = "2032 Projected"

        for _, row in df.iterrows():
            soc_val = str(row.get(soc_col, "")).strip() if pd.notna(row.get(soc_col)) else ""
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue

            try:
                base = int(float(row[base_col]))
                proj = int(float(row[proj_col]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    IA: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Illinois — 10 EDR XLSX files from ZIP (2022-2032)
# ---------------------------------------------------------------------------

IL_EDR_DIR = RAW_DIR / "IDESWebsite Files_OccEDRs_22-32"

IL_EDR_TO_CBSA = {
    1:  [("40420", "Rockford, IL")],
    2:  [("37900", "Peoria, IL"),
         ("19340", "Davenport-Moline-Rock Island, IA-IL")],
    3:  [("16980", "Chicago-Naperville-Elgin, IL-IN-WI"),
         ("28100", "Kankakee, IL")],
    4:  [("14010", "Bloomington, IL"),
         ("16580", "Champaign-Urbana, IL"),
         ("19500", "Decatur, IL")],
    5:  [("44100", "Springfield, IL")],
    6:  [("41180", "St. Louis, MO-IL")],
    # EDRs 7-10: no major OES CBSAs, skip
}


def load_illinois_metros():
    """Load Illinois EDR-level occupational projections (2022-2032).

    Each XLSX: data starts at row 6 (0-indexed), col 0=SOC Code, col 2=Emp 2022, col 4=Emp 2032.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in IL_EDR_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for edr_num, cbsa_list in IL_EDR_TO_CBSA.items():
        fpath = IL_EDR_DIR / f"EDR_{edr_num}_Occ '22-'32.xlsx"
        if not fpath.exists():
            print(f"    IL: Missing EDR_{edr_num}, skipping")
            continue

        df = pd.read_excel(fpath, header=None, skiprows=6)

        for _, row in df.iterrows():
            soc_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue

            try:
                base = int(float(row.iloc[2]))
                proj = int(float(row.iloc[4]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    IL: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Colorado — Single CSV with MSA-level data (2023-2033)
# ---------------------------------------------------------------------------

CO_FILE = RAW_DIR / "co_projections.csv"

CO_MSA_TO_CBSA = {
    "14500": "Boulder, CO",
    "17820": "Colorado Springs, CO",
    "19740": "Denver-Aurora-Lakewood, CO",
    "22660": "Fort Collins, CO",
    "24300": "Grand Junction, CO",
    "24540": "Greeley, CO",
    "39380": "Pueblo, CO",
}


def load_colorado_metros():
    """Load Colorado MSA-level occupational projections (2023-2033).

    Single CSV with all areas. Filter to areatyname='Metropolitan Statistical Area'
    and matincode='0' (total across industries). Area codes and SOC codes have
    comma thousand-separators that need stripping.
    MSA-level data → can store actual employment values.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []

    df = pd.read_csv(CO_FILE, dtype=str)

    # Filter to MSA-level, total industry
    msa = df[(df["areatyname"] == "Metropolitan Statistical Area") & (df["matincode"] == "0")]

    for _, row in msa.iterrows():
        # Clean area code: "14,500" → "14500"
        area_raw = str(row["area"]).replace(",", "").strip()

        if area_raw not in CO_MSA_TO_CBSA:
            continue

        cbsa_code = area_raw
        cbsa_name = CO_MSA_TO_CBSA[cbsa_code]

        # Clean SOC code: "111,021" → "111021" → "11-1021"
        occ_raw = str(row["matoccode"]).replace(",", "").strip()
        if len(occ_raw) < 5:
            continue
        occ_str = occ_raw.zfill(6)
        soc = f"{occ_str[:2]}-{occ_str[2:]}"

        if not soc_re.match(soc):
            continue
        if soc.endswith("0000") or soc.endswith("000"):
            continue

        # Check suppress flag
        if str(row.get("suppress", "0")).strip() == "1":
            continue

        try:
            base = int(float(str(row["estemp"]).replace(",", "")))
            proj = int(float(str(row["projemp"]).replace(",", "")))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2023-2033

            rows.append((
                soc, "metro", cbsa_code, cbsa_name,
                2023, 2033,
                base, proj,  # MSA-level → store actual employment
                round(pct_change, 2),
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    CO: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Arizona — Single XLSX, county-level data (2024-2034)
# ---------------------------------------------------------------------------

AZ_FILE = RAW_DIR / "az_projections.xlsx"

AZ_AREA_TO_CBSA = {
    "Coconino County":     [("22380", "Flagstaff, AZ")],
    "Mohave County":       [("29420", "Lake Havasu City-Kingman, AZ")],
    "Maricopa County":     [("38060", "Phoenix-Mesa-Chandler, AZ")],
    "Yavapai County":      [("39150", "Prescott Valley-Prescott, AZ")],
    "Cochise County":      [("43420", "Sierra Vista-Douglas, AZ")],
    "Pima County":         [("46060", "Tucson, AZ")],
    "Yuma County":         [("49740", "Yuma, AZ")],
}


def load_arizona_metros():
    """Load Arizona county-level occupational projections (2024-2034).

    Single XLSX with all areas in one 'LT OCC' sheet.
    Header at row 2: Area Name, SOC Code, Title, 2024 Estimates, 2034 Projections, ...
    County-level → CAGR only (counties ≠ CBSAs exactly, but close enough for major metros).
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in AZ_AREA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    df = pd.read_excel(AZ_FILE, sheet_name="LT OCC", header=2)
    soc_col = df.columns[1]    # 'SOC Code2 '
    est_col = df.columns[3]    # '2024 Estimates3'
    proj_col = df.columns[4]   # '2034 Projections'

    for _, row in df.iterrows():
        area = str(row["Area Name"]).strip() if pd.notna(row["Area Name"]) else ""
        if area not in AZ_AREA_TO_CBSA:
            continue

        soc_val = str(row[soc_col]).strip() if pd.notna(row[soc_col]) else ""
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue

        try:
            base = int(float(row[est_col]))
            proj = int(float(row[proj_col]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2024-2034

            for cbsa_code, _name in AZ_AREA_TO_CBSA[area]:
                rows.append((
                    soc_val, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2024, 2034,
                    None, None,  # County ≠ CBSA, CAGR only
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    AZ: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# West Virginia — 7 WIA XLSX files (2022-2032)
# ---------------------------------------------------------------------------

WV_WIA_TO_CBSA = {
    "wv_wia1.xlsx": [("13220", "Beckley, WV")],
    "wv_wia2.xlsx": [("26580", "Huntington-Ashland, WV-KY-OH")],
    "wv_wia3.xlsx": [("16620", "Charleston, WV")],
    "wv_wia4.xlsx": [("37620", "Parkersburg-Vienna, WV")],
    "wv_wia5.xlsx": [("48540", "Wheeling, WV-OH"),
                     ("48260", "Weirton-Steubenville, WV-OH")],
    "wv_wia6.xlsx": [("34060", "Morgantown, WV")],
    "wv_wia7.xlsx": [("25180", "Hagerstown-Martinsburg, MD-WV"),
                     ("49020", "Winchester, VA-WV"),
                     ("47900", "Washington-Arlington-Alexandria, DC-VA-MD-WV")],
}


def load_west_virginia_metros():
    """Load West Virginia WIA-level occupational projections (2022-2032).

    Each XLSX has header at row 0: SOC Code, Title, Estimated Employment, Projected Employment, ...
    Data from row 1. All rows are detailed occupations (no major groups).
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in WV_WIA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for fname, cbsa_list in WV_WIA_TO_CBSA.items():
        fpath = RAW_DIR / fname
        if not fpath.exists():
            print(f"    WV: Missing {fname}, skipping")
            continue

        df = pd.read_excel(fpath, header=0)

        for _, row in df.iterrows():
            soc_val = str(row.get("SOC Code", "")).strip() if pd.notna(row.get("SOC Code")) else ""
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue

            try:
                base = int(float(row["Estimated Employment"]))
                proj = int(float(row["Projected Employment"]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # WIA ≠ CBSA, CAGR only
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    WV: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Kentucky — Single XLSX, "Local Workforce Areas" sheet (2023-2033)
# ---------------------------------------------------------------------------

KY_FILE = RAW_DIR / "ky_projections.xlsx"

KY_LWA_TO_CBSA = {
    "Bluegrass":         [("30460", "Lexington-Fayette, KY")],
    "Kentuckiana Works": [("31140", "Louisville/Jefferson County, KY-IN")],
    "Northern Kentucky": [("17140", "Cincinnati, OH-KY-IN")],
    "South Central":     [("14540", "Bowling Green, KY")],
    "Lincoln Trail":     [("21060", "Elizabethtown-Fort Knox, KY"),
                          ("17300", "Clarksville, TN-KY")],
    "Green River":       [("36980", "Owensboro, KY")],
    "West Kentucky":     [("37140", "Paducah, KY-IL")],
    "TENCO":             [("26580", "Huntington-Ashland, WV-KY-OH")],
    # Cumberlands, EKCEP — no major OES CBSAs, skip
}


def load_kentucky_metros():
    """Load Kentucky LWA-level occupational projections (2023-2033).

    Single XLSX with 'Local Workforce Areas' sheet.
    Header at row 0: Area, SOC Title, SOC Code, 2023 Est, 2033 Proj, ...
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in KY_LWA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    df = pd.read_excel(KY_FILE, sheet_name="Local Workforce Areas", header=0)
    soc_col = df.columns[2]   # 'Standard Occupational Classification (SOC) System Code'
    est_col = df.columns[3]   # '2023 Estimated Employment'
    proj_col = df.columns[4]  # '2033 Projected Employment'

    for _, row in df.iterrows():
        area = str(row["Area"]).strip() if pd.notna(row["Area"]) else ""
        if area not in KY_LWA_TO_CBSA:
            continue

        soc_val = str(row[soc_col]).strip() if pd.notna(row[soc_col]) else ""
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue

        try:
            base = int(float(row[est_col]))
            proj = int(float(row[proj_col]))
        except (ValueError, TypeError):
            continue

        if base > 0 and proj > 0:
            pct_change = (proj - base) / base * 100
            cagr = calc_cagr(base, proj, 10)  # 2023-2033

            for cbsa_code, _name in KY_LWA_TO_CBSA[area]:
                rows.append((
                    soc_val, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2023, 2033,
                    None, None,  # LWA ≠ CBSA, CAGR only
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    KY: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Idaho — 6 regional XLSX files, "Table 1.12" sheet (2022-2032)
# ---------------------------------------------------------------------------

ID_FILES = {
    "id_northern_idaho.xlsx":      [("17660", "Coeur d'Alene, ID")],
    "id_north_central_idaho.xlsx": [("30300", "Lewiston, ID-WA")],
    "id_southwest_idaho.xlsx":     [("14260", "Boise City, ID")],
    "id_south_central_idaho.xlsx": [("46300", "Twin Falls, ID")],
    "id_southeast_idaho.xlsx":     [("38540", "Pocatello, ID")],
    "id_eastern_idaho.xlsx":       [("26820", "Idaho Falls, ID")],
}


def load_idaho_metros():
    """Load Idaho regional occupational projections (2022-2032).

    Each XLSX has a 'Table 1.12' sheet with detailed occupations.
    Data from row 5 (0-indexed). Col 0=SOC, Col 1=Title, Col 2=Level (filter 'Occupation'),
    Col 3=Employment 2022, Col 4=Employment 2032.
    Suppressed values shown as '-'.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in ID_FILES.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for fname, cbsa_list in ID_FILES.items():
        fpath = RAW_DIR / fname
        if not fpath.exists():
            print(f"    ID: Missing {fname}, skipping")
            continue

        # Sheet names vary by region: Table 1.12, 1.22, 1.32, etc.
        # The detailed occupation table always ends in '2'
        xl = pd.ExcelFile(fpath)
        detail_sheet = None
        for s in xl.sheet_names:
            if s.startswith("Table") and s.endswith("2"):
                detail_sheet = s
                break
        if detail_sheet is None:
            print(f"    ID: No detailed occupation sheet in {fname}, skipping")
            continue

        df = pd.read_excel(fpath, sheet_name=detail_sheet, header=None, skiprows=5)

        for _, row in df.iterrows():
            soc_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue

            # Level column: filter to 'Occupation' (detailed)
            level = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
            if level != "Occupation":
                continue

            # Suppressed values are '-'
            base_raw = row.iloc[3]
            proj_raw = row.iloc[4]
            if str(base_raw).strip() == "-" or str(proj_raw).strip() == "-":
                continue

            try:
                base = int(float(base_raw))
                proj = int(float(proj_raw))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # Region ≠ CBSA, CAGR only
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    ID: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Utah — 7 regional XLSX files (2022-2032), 5 map to OES CBSAs
# ---------------------------------------------------------------------------

UT_FILES = {
    "ut_salt_lake_city_utah.xlsx":  [("41620", "Salt Lake City, UT")],
    "ut_ogden_utah.xlsx":           [("36260", "Ogden-Clearfield, UT")],
    "ut_provo_orem_lehi_utah.xlsx": [("39340", "Provo-Orem, UT")],
    "ut_logan_utah.xlsx":           [("30860", "Logan, UT-ID")],
    "ut_saint_george_utah.xlsx":    [("41100", "St. George, UT")],
    # High Desert, Wasatch Front Fringe — no OES CBSAs, skip
}


def load_utah_metros():
    """Load Utah regional occupational projections (2022-2032).

    Each XLSX has header at row 0: SOC Code, SOC Title, ..., SOC Level,
    2022 Employment, 2032 Projected Employment, ...
    SOC Level '4' = detailed occupation.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in UT_FILES.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    for fname, cbsa_list in UT_FILES.items():
        fpath = RAW_DIR / fname
        if not fpath.exists():
            print(f"    UT: Missing {fname}, skipping")
            continue

        df = pd.read_excel(fpath, header=0)

        soc_col = "SOC Code"
        est_col = "2022 Employment"
        proj_col = "2032 Projected Employment"
        level_col = "SOC Level (copy)"

        for _, row in df.iterrows():
            # Filter to detailed occupations (level 4)
            level = str(row.get(level_col, "")).strip()
            if level != "4":
                continue

            soc_val = str(row.get(soc_col, "")).strip() if pd.notna(row.get(soc_col)) else ""
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue

            try:
                base = int(float(row[est_col]))
                proj = int(float(row[proj_col]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2022-2032

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2022, 2032,
                        None, None,  # Region ≠ CBSA, CAGR only
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    UT: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Montana — Single XLSX, sub-state regions in columns (2024-2034)
# ---------------------------------------------------------------------------

MT_FILE = RAW_DIR / "mt_projections.xlsx"

MT_REGION_TO_CBSA = {
    1: [("33540", "Missoula, MT")],                          # Northwest
    2: [("14580", "Bozeman, MT"), ("25740", "Helena, MT")],  # Southwest
    3: [("24500", "Great Falls, MT")],                        # North Central
    4: [("13740", "Billings, MT")],                           # South Central
    # Region 5 (Eastern) — no major OES CBSAs, skip
}


def load_montana_metros():
    """Load Montana regional occupational projections (2024-2034).

    Single XLSX 'Disclosable Projections' sheet with all regions in columns.
    Columns r{N}_24 = base emp, r{N}_34 = projected emp, r{N}flag = 0 if disclosable.
    soclevel=4 for detailed occupations.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in MT_REGION_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name

    df = pd.read_excel(MT_FILE, sheet_name="Disclosable Projections", header=0)

    for _, row in df.iterrows():
        # Filter to detailed occupations
        if row.get("soclevel") != 4:
            continue

        soc_val = str(row.get("soccode", "")).strip()
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue

        for region_num, cbsa_list in MT_REGION_TO_CBSA.items():
            flag_col = f"r{region_num}flag"
            base_col = f"r{region_num}_24"
            proj_col = f"r{region_num}_34"

            # Skip suppressed data (flag != 0 or NaN)
            flag = row.get(flag_col)
            if pd.isna(flag) or float(flag) != 0:
                continue

            try:
                base = int(float(row[base_col]))
                proj = int(float(row[proj_col]))
            except (ValueError, TypeError):
                continue

            if base > 0 and proj > 0:
                pct_change = (proj - base) / base * 100
                cagr = calc_cagr(base, proj, 10)  # 2024-2034

                for cbsa_code, _name in cbsa_list:
                    rows.append((
                        soc_val, "metro", cbsa_code,
                        cbsa_names.get(cbsa_code, cbsa_code),
                        2024, 2034,
                        None, None,  # Region ≠ CBSA, CAGR only
                        round(pct_change, 2),
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))

    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MT: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Arkansas  –  10 LWDA XLSX files, 2022-2032
# ---------------------------------------------------------------------------

AR_FILES = {
    "raw/ar_northwest.xlsx":     [("22220", "Fayetteville-Springdale-Rogers, AR")],
    "raw/ar_western.xlsx":       [("22900", "Fort Smith, AR-OK")],
    "raw/ar_west_central.xlsx":  [("26300", "Hot Springs, AR")],
    "raw/ar_southwest.xlsx":     [("45500", "Texarkana, TX-AR")],
    "raw/ar_little_rock.xlsx":   [("30780", "Little Rock-North Little Rock-Conway, AR")],
    "raw/ar_central.xlsx":       [("30780", "Little Rock-North Little Rock-Conway, AR")],
    "raw/ar_northeast.xlsx":     [("27860", "Jonesboro, AR")],
    "raw/ar_eastern.xlsx":       [("32820", "Memphis, TN-MS-AR")],
    # ar_north_central.xlsx  → no OES CBSA
    # ar_southeast.xlsx      → no OES CBSA (Pine Bluff not in OES)
}


def load_arkansas_metros():
    """Load AR LWDA projections from XLSX files.

    Format: Sheet 'Sheet 1', header at row 0.
    Columns: Sort CalcField, Occupation, Soccode, ...,
             Projected Employment at End of Proj. Period (col 8),
             Percent Change in Employment (col 9).
    SOC codes have no dashes (e.g., '353023' → '35-3023').
    No base employment column → derive from projected and pct_change.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for fpath, cbsa_list in AR_FILES.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name
        if not Path(fpath).exists():
            print(f"    [AR] missing {fpath}, skipping")
            continue
        try:
            df = pd.read_excel(fpath, sheet_name="Sheet 1", header=0)
        except Exception as e:
            print(f"    [AR] error reading {fpath}: {e}")
            continue
        for _, r in df.iterrows():
            raw_soc = str(r.get("Soccode", "")).strip()
            if len(raw_soc) == 6:
                soc_val = raw_soc[:2] + "-" + raw_soc[2:]
            else:
                continue
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                proj_emp = int(r.get("Projected Employment at End of Proj. Period", 0))
                pct_change = float(r.get("Percent Change in Employment", 0))
            except (ValueError, TypeError):
                continue
            if proj_emp <= 0:
                continue
            # Derive base employment
            base_emp = round(proj_emp / (1 + pct_change / 100)) if pct_change != -100 else None
            cagr = calc_cagr(base_emp, proj_emp, 10) if base_emp and base_emp > 0 else None
            for cbsa_code, _name in cbsa_list:
                rows.append((
                    soc_val, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2022, 2032,
                    None, None,  # Region ≠ CBSA, CAGR only
                    round(pct_change, 2),
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    AR: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Connecticut  –  5 WDA XLSX files, 2022-2032
# ---------------------------------------------------------------------------

CT_FILES = {
    "raw/ct_eastern.xlsx":       [("35980", "Norwich-New London-Willimantic, CT")],
    "raw/ct_north_central.xlsx": [("25540", "Hartford-West Hartford-East Hartford, CT")],
    "raw/ct_northwest.xlsx":     [("47930", "Waterbury-Shelton, CT")],
    "raw/ct_south_central.xlsx": [("35300", "New Haven, CT")],
    "raw/ct_southwest.xlsx":     [("14860", "Bridgeport-Stamford-Danbury, CT")],
}


def load_connecticut_metros():
    """Load CT WDA projections from XLSX files.

    Each file has an occupation sheet named '<Region> Occ 2022-2032'.
    Header at row 0: SOC Code, Occupation Title, SOC Level,
                     Base Employment 2022, Projected Employment 2032, ...
    SOC Level 4 = detailed.  SOC codes already have dashes.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for fpath, cbsa_list in CT_FILES.items():
        for code, name in cbsa_list:
            cbsa_names[code] = name
        if not Path(fpath).exists():
            print(f"    [CT] missing {fpath}, skipping")
            continue
        try:
            xl = pd.ExcelFile(fpath)
            occ_sheet = [s for s in xl.sheet_names if "Occ" in s]
            if not occ_sheet:
                print(f"    [CT] no Occ sheet in {fpath}, skipping")
                continue
            df = pd.read_excel(xl, sheet_name=occ_sheet[0], header=0)
        except Exception as e:
            print(f"    [CT] error reading {fpath}: {e}")
            continue
        for _, r in df.iterrows():
            soc_level = r.get("SOC Level")
            try:
                if int(soc_level) != 4:
                    continue
            except (ValueError, TypeError):
                continue
            soc_val = str(r.get("SOC Code", "")).strip()
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                base_emp = int(r.get("Base Employment 2022", 0))
                proj_emp = int(r.get("Projected Employment 2032", 0))
            except (ValueError, TypeError):
                continue
            if base_emp <= 0 or proj_emp <= 0:
                continue
            pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
            cagr = calc_cagr(base_emp, proj_emp, 10)
            for cbsa_code, _name in cbsa_list:
                rows.append((
                    soc_val, "metro", cbsa_code,
                    cbsa_names.get(cbsa_code, cbsa_code),
                    2022, 2032,
                    None, None,  # Region != CBSA, CAGR only
                    pct_change,
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    CT: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Kansas  –  single HTML-as-XLS file with 7 LWDA regions, 2022-2032
# ---------------------------------------------------------------------------

KS_FILE = RAW_DIR / "ks_projections.xls"

KS_AREA_TO_CBSA = {
    "Kansas City Area":   [("28140", "Kansas City, MO-KS")],
    "Northeast Area":     [("45820", "Topeka, KS"),
                           ("29940", "Lawrence, KS"),
                           ("31740", "Manhattan, KS")],
    "South Central Area": [("48620", "Wichita, KS")],
    # Northwest Area      → no OES CBSA
    # North Central Area  → no OES CBSA
    # Southeast Area      → no OES CBSA
    # Southwest Area      → no OES CBSA
}


def load_kansas_metros():
    """Load KS LWDA projections from HTML table saved as .xls.

    Parsed with pd.read_html(). Columns include: areaname, occcode,
    codelevel, estoccprj, projoccprj, pchg, estyear, projyear.
    codelevel 6 = detailed.  SOC codes have no dashes.
    Suppressed values shown as 's;' -- skip those rows.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in KS_AREA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name
    if not KS_FILE.exists():
        print(f"    [KS] missing {KS_FILE}, skipping")
        return rows
    dfs = pd.read_html(KS_FILE, header=0)
    df = dfs[0]
    # Filter to sub-state regions (areatype 15) and detailed occupations
    df = df[(df["areatype"] == 15) & (df["codelevel"] == 6)]
    for _, r in df.iterrows():
        area = str(r.get("areaname", "")).strip()
        cbsa_list = KS_AREA_TO_CBSA.get(area)
        if not cbsa_list:
            continue
        raw_soc = str(r.get("occcode", "")).strip()
        if len(raw_soc) == 6:
            soc_val = raw_soc[:2] + "-" + raw_soc[2:]
        else:
            continue
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue
        # Skip suppressed values
        est_raw = r.get("estoccprj")
        proj_raw = r.get("projoccprj")
        try:
            base_emp = int(est_raw)
            proj_emp = int(proj_raw)
        except (ValueError, TypeError):
            continue
        if base_emp <= 0 or proj_emp <= 0:
            continue
        pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
        cagr = calc_cagr(base_emp, proj_emp, 10)
        for cbsa_code, _name in cbsa_list:
            rows.append((
                soc_val, "metro", cbsa_code,
                cbsa_names.get(cbsa_code, cbsa_code),
                2022, 2032,
                None, None,  # Region != CBSA, CAGR only
                pct_change,
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    KS: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# New Mexico  –  single HTML-as-XLS file with MSA-level data, 2023-2033
# ---------------------------------------------------------------------------

NM_FILE = RAW_DIR / "nm_projections.xls"

NM_MSA_TO_CBSA = {
    "Albuquerque MSA": ("10740", "Albuquerque, NM"),
    "Farmington MSA":  ("22140", "Farmington, NM"),
    "Las Cruces MSA":  ("29740", "Las Cruces, NM"),
    "Santa Fe MSA":    ("42140", "Santa Fe, NM"),
}


def load_new_mexico_metros():
    """Load NM MSA-level projections from HTML table saved as .xls.

    Same format as Kansas. areatype 21 = MSA (direct MSA data).
    codelevel 6 = detailed.  SOC codes have no dashes.
    Suppressed values shown as 's;' -- skip those rows.
    MSA-level data -> store actual base_emp and proj_emp.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    if not NM_FILE.exists():
        print(f"    [NM] missing {NM_FILE}, skipping")
        return rows
    dfs = pd.read_html(NM_FILE, header=0)
    df = dfs[0]
    # Filter to MSA-level (areatype 21) and detailed occupations
    df = df[(df["areatype"] == 21) & (df["codelevel"] == 6)]
    for _, r in df.iterrows():
        area = str(r.get("areaname", "")).strip()
        mapping = NM_MSA_TO_CBSA.get(area)
        if not mapping:
            continue
        cbsa_code, cbsa_name = mapping
        raw_soc = str(r.get("occcode", "")).strip()
        if len(raw_soc) == 6:
            soc_val = raw_soc[:2] + "-" + raw_soc[2:]
        else:
            continue
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue
        # Skip suppressed values
        try:
            base_emp = int(r.get("estoccprj"))
            proj_emp = int(r.get("projoccprj"))
        except (ValueError, TypeError):
            continue
        if base_emp <= 0 or proj_emp <= 0:
            continue
        pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
        cagr = calc_cagr(base_emp, proj_emp, 10)
        rows.append((
            soc_val, "metro", cbsa_code, cbsa_name,
            2023, 2033,
            base_emp, proj_emp,  # MSA-level -> actual employment
            pct_change,
            round(cagr, 6) if cagr is not None else None,
            "state_lmi",
        ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    NM: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# New Jersey  –  single XLSX with county-level data, 2022-2032
# ---------------------------------------------------------------------------

NJ_FILE = RAW_DIR / "nj_projections.xlsx"

NJ_COUNTY_TO_CBSA = {
    "Atlantic County":    ("12100", "Atlantic City-Hammonton, NJ"),
    "Mercer County":      ("45940", "Trenton-Princeton, NJ"),
    "Cumberland County":  ("47220", "Vineland, NJ"),
}


def load_new_jersey_metros():
    """Load NJ county-level projections from Master sheet.

    Sheet 'Master', header at row 4.
    Columns: Seq, Area, NAICS + Industry, SOC + Occupation,
             2022 Actual, 2032 Projected, ...
    Filter to NAICS '000000' (Total All Industries).
    SOC code embedded in 'SOC + Occupation' col: first 7 chars = 'XX-XXXX'.
    County = single-county CBSA -> store actual employment.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    if not NJ_FILE.exists():
        print(f"    [NJ] missing {NJ_FILE}, skipping")
        return rows
    df = pd.read_excel(NJ_FILE, sheet_name="Master", header=4)
    # Filter to Total All Industries
    df = df[df["NAICS + Industry"].astype(str).str.startswith("000000")]
    for _, r in df.iterrows():
        area = str(r.get("Area", "")).strip()
        mapping = NJ_COUNTY_TO_CBSA.get(area)
        if not mapping:
            continue
        cbsa_code, cbsa_name = mapping
        soc_occ = str(r.get("SOC + Occupation", ""))
        soc_val = soc_occ[:7].strip()
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue
        try:
            base_emp = int(r.get("2022 Actual", 0))
            proj_emp = int(r.get("2032 Projected", 0))
        except (ValueError, TypeError):
            continue
        if base_emp <= 0 or proj_emp <= 0:
            continue
        pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
        cagr = calc_cagr(base_emp, proj_emp, 10)
        rows.append((
            soc_val, "metro", cbsa_code, cbsa_name,
            2022, 2032,
            base_emp, proj_emp,  # County = CBSA, actual employment
            pct_change,
            round(cagr, 6) if cagr is not None else None,
            "state_lmi",
        ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    NJ: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Mississippi  –  3 MSA PDFs, 2022-2032
# ---------------------------------------------------------------------------

MS_FILES = {
    "raw/ms_jackson.pdf":     ("27140", "Jackson, MS"),
    "raw/ms_gulfport.pdf":    ("25060", "Gulfport-Biloxi, MS"),
    "raw/ms_hattiesburg.pdf": ("25620", "Hattiesburg, MS"),
}


def load_mississippi_metros():
    """Load MS MSA projections from PDF tables via pdfplumber.

    Each PDF is one MSA.  Tables span multiple pages.
    Table rows: [SOC Code, Occupation, 2022 Emp, 2032 Proj Emp, Change, %Change, Openings]
    Employment values have comma separators.  Negative values in parentheses.
    MSA-level data -> store actual employment.
    """
    import pdfplumber
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    for fpath, (cbsa_code, cbsa_name) in MS_FILES.items():
        if not Path(fpath).exists():
            print(f"    [MS] missing {fpath}, skipping")
            continue
        pdf = pdfplumber.open(fpath)
        n = 0
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for trow in table:
                    if not trow or len(trow) < 5:
                        continue
                    soc_val = str(trow[0] or "").strip()
                    if not soc_re.match(soc_val):
                        continue
                    if soc_val.endswith("0000") or soc_val.endswith("000"):
                        continue
                    try:
                        base_str = str(trow[2] or "").replace(",", "").strip()
                        proj_str = str(trow[3] or "").replace(",", "").strip()
                        base_emp = int(base_str)
                        proj_emp = int(proj_str)
                    except (ValueError, TypeError):
                        continue
                    if base_emp <= 0 or proj_emp <= 0:
                        continue
                    pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
                    cagr = calc_cagr(base_emp, proj_emp, 10)
                    rows.append((
                        soc_val, "metro", cbsa_code, cbsa_name,
                        2022, 2032,
                        base_emp, proj_emp,  # MSA-level, actual employment
                        pct_change,
                        round(cagr, 6) if cagr is not None else None,
                        "state_lmi",
                    ))
                    n += 1
        print(f"    MS {cbsa_name}: {n} occupations")
        pdf.close()
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MS: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Nevada  –  single XLSX with MSA-level sheets, 2022-2032
# ---------------------------------------------------------------------------

NV_FILE = RAW_DIR / "nv_projections.xlsx"

NV_SHEET_TO_CBSA = {
    "Las Vegas LT":   ("29820", "Las Vegas-Henderson-North Las Vegas, NV"),
    "Reno LT":        ("39900", "Reno, NV"),
    "Carson City LT": ("16180", "Carson City, NV"),
    # "BOS LT" -> Balance of State, no OES CBSA
}


def load_nevada_metros():
    """Load NV MSA projections from per-sheet XLSX.

    Sheets: 'Las Vegas LT', 'Reno LT', 'Carson City LT'.
    Header at row 3.  Columns: Occupation Type, SOC Code, Occupation Title,
    2022, 2032*, ...
    Filter to Occupation Type == 'Detailed Occupations'.
    SOC codes already have dashes.  MSA-level -> actual employment.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    if not NV_FILE.exists():
        print(f"    [NV] missing {NV_FILE}, skipping")
        return rows
    wb = openpyxl.load_workbook(NV_FILE, read_only=True, data_only=True)
    for sheet_name, (cbsa_code, cbsa_name) in NV_SHEET_TO_CBSA.items():
        if sheet_name not in wb.sheetnames:
            print(f"    [NV] sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=5, values_only=True))  # skip rows 0-3 (header at row 3)
        for r in all_rows:
            if not r or len(r) < 5:
                continue
            occ_type = str(r[0] or "").strip()
            if occ_type != "Detailed Occupations":
                continue
            soc_val = str(r[1] or "").strip()
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                base_emp = int(r[3])
                proj_emp = int(r[4])
            except (ValueError, TypeError):
                continue
            if base_emp <= 0 or proj_emp <= 0:
                continue
            pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
            cagr = calc_cagr(base_emp, proj_emp, 10)
            rows.append((
                soc_val, "metro", cbsa_code, cbsa_name,
                2022, 2032,
                base_emp, proj_emp,  # MSA-level, actual employment
                pct_change,
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))
    wb.close()
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    NV: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# South Dakota  –  single XLSX with MSA-level sheets, 2022-2032
# ---------------------------------------------------------------------------

SD_FILE = RAW_DIR / "sd_projections.xlsx"

SD_SHEET_TO_CBSA = {
    "RCMSAPubLMI22-32": ("39660", "Rapid City, SD"),
    "SFMSAPubLMI22-32": ("43620", "Sioux Falls, SD-MN"),
    # "BOSPubLMI22-32" -> Balance of State, no OES CBSA
}


def load_south_dakota_metros():
    """Load SD MSA projections from per-sheet XLSX.

    Sheets: 'RCMSAPubLMI22-32' (Rapid City), 'SFMSAPubLMI22-32' (Sioux Falls).
    Row 0 = title, Row 1 = column headers (SOC Code, SOC Title, 2022 Employment,
    2032 Employment, ...), Row 2 = sub-headers for openings breakdown.
    Data starts at row 3.  SOC codes already have dashes.
    MSA-level -> actual employment.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    if not SD_FILE.exists():
        print(f"    [SD] missing {SD_FILE}, skipping")
        return rows
    wb = openpyxl.load_workbook(SD_FILE, read_only=True, data_only=True)
    for sheet_name, (cbsa_code, cbsa_name) in SD_SHEET_TO_CBSA.items():
        if sheet_name not in wb.sheetnames:
            print(f"    [SD] sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip rows 0-2 (title + headers)
        for r in all_rows:
            if not r or len(r) < 4:
                continue
            soc_val = str(r[0] or "").strip()
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                base_emp = int(r[2])
                proj_emp = int(r[3])
            except (ValueError, TypeError):
                continue
            if base_emp <= 0 or proj_emp <= 0:
                continue
            pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
            cagr = calc_cagr(base_emp, proj_emp, 10)
            rows.append((
                soc_val, "metro", cbsa_code, cbsa_name,
                2022, 2032,
                base_emp, proj_emp,  # MSA-level, actual employment
                pct_change,
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))
    wb.close()
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    SD: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Delaware  –  long-format CSV with county-level data, 2022-2032
# ---------------------------------------------------------------------------

DE_FILE = RAW_DIR / "de_projections.csv"

# Only Kent County maps to a DE-only OES CBSA.
# New Castle County → part of Philadelphia CBSA (already covered by PA scraper)
# Sussex County → nonmetro
DE_COUNTY_TO_CBSA = {
    "Kent County": ("20100", "Dover, DE"),
}


def load_delaware_metros():
    """Load DE county-level projections from long-format CSV.

    CSV has one row per measure per occupation per area.
    Pivot on Measure Names to get 'Base Year Employment' and 'Projected Employment'.
    Occupation Code is integer without dashes (e.g. 111021 -> 11-1021).
    County = single-county CBSA -> actual employment.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    if not DE_FILE.exists():
        print(f"    [DE] missing {DE_FILE}, skipping")
        return rows
    df = pd.read_csv(DE_FILE)
    for county, (cbsa_code, cbsa_name) in DE_COUNTY_TO_CBSA.items():
        cdf = df[df["Area"] == county]
        measures = cdf[cdf["Measure Names"].isin(["Base Year Employment", "Projected Employment"])]
        pivot = measures.pivot_table(
            index=["Occupation Code", "Occupation Title"],
            columns="Measure Names", values="Measure Values",
        )
        for (occ_code_int, _title), row in pivot.iterrows():
            raw = str(int(occ_code_int)).zfill(6)
            soc_val = raw[:2] + "-" + raw[2:]
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                base_emp = int(row.get("Base Year Employment", 0))
                proj_emp = int(row.get("Projected Employment", 0))
            except (ValueError, TypeError):
                continue
            if base_emp <= 0 or proj_emp <= 0:
                continue
            pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
            cagr = calc_cagr(base_emp, proj_emp, 10)
            rows.append((
                soc_val, "metro", cbsa_code, cbsa_name,
                2022, 2032,
                base_emp, proj_emp,  # County = CBSA, actual employment
                pct_change,
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    DE: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# New Hampshire  –  single XLSX with 9 regional sheets, 2020-2030
# ---------------------------------------------------------------------------

NH_FILE = RAW_DIR / "nh_projections.xlsx"

NH_SHEET_TO_CBSA = {
    "Nashua":   [("31700", "Manchester-Nashua, NH")],
    "Southern": [("31700", "Manchester-Nashua, NH")],
    # Central NH       → Concord area, no OES CBSA
    # Lakes Region     → Laconia area, no OES CBSA
    # North Country    → far north, no OES CBSA
    # Rockingham       → part of Boston CBSA (already covered by MA)
    # Southwest        → Keene area, no OES CBSA
    # Strafford        → Dover-Rochester area, no OES CBSA
    # Upper Valley     → Lebanon area, no OES CBSA
}


def load_new_hampshire_metros():
    """Load NH regional projections from per-sheet XLSX.

    Each sheet = one Regional Planning Commission.
    Row 0 = region name, Row 1-2 = column headers, Row 3+ = data.
    Columns: SOC Code (col 0), Occupation Title (col 1),
    2020 Estimated (col 2), 2030 Projected (col 3), ...
    SOC codes already have dashes.  All SOC levels included, filter to detailed.
    Suppressed values shown as 'n' -> skip.
    2020-2030 timeframe.  Region ≠ CBSA -> CAGR only.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    if not NH_FILE.exists():
        print(f"    [NH] missing {NH_FILE}, skipping")
        return rows
    wb = openpyxl.load_workbook(NH_FILE, read_only=True, data_only=True)
    for sheet_name, cbsa_list in NH_SHEET_TO_CBSA.items():
        if sheet_name not in wb.sheetnames:
            print(f"    [NH] sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip rows 0-2 (title + headers)
        for r in all_rows:
            if not r or len(r) < 4:
                continue
            soc_val = str(r[0] or "").strip()
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                base_emp = int(r[2])
                proj_emp = int(r[3])
            except (ValueError, TypeError):
                continue
            if base_emp <= 0 or proj_emp <= 0:
                continue
            pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
            cagr = calc_cagr(base_emp, proj_emp, 10)
            for cbsa_code, cbsa_name in cbsa_list:
                rows.append((
                    soc_val, "metro", cbsa_code, cbsa_name,
                    2020, 2030,
                    None, None,  # Region ≠ CBSA, CAGR only
                    pct_change,
                    round(cagr, 6) if cagr is not None else None,
                    "state_lmi",
                ))
    wb.close()
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    NH: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Hawaii  –  per-county XLSX files (MSA-level), 2022-2032
# ---------------------------------------------------------------------------

HI_FILES = {
    "raw/hi_honolulu.xlsx": ("46520", "Urban Honolulu, HI"),
    "raw/hi_maui.xlsx":     ("27980", "Kahului-Wailuku, HI"),
    # hi_hawaii.xlsx  → Hawaii County, no OES CBSA
    # hi_kauai.xlsx   → Kauai County, no OES CBSA
}


def load_hawaii_metros():
    """Load HI MSA projections from per-county XLSX files.

    Each file has one sheet.
    Row 0 = title, Row 1-2 = column headers, Row 3+ = data.
    Columns: SOC Code (col 0), Occupation Title (col 1),
    2022 Employment (col 2), 2032 Employment (col 3), ...
    SOC codes already have dashes.  All levels included, filter to detailed.
    Suppressed values shown as '*' -> skip.
    MSA-level -> actual employment.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    for fpath, (cbsa_code, cbsa_name) in HI_FILES.items():
        if not Path(fpath).exists():
            print(f"    [HI] missing {fpath}, skipping")
            continue
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip rows 0-2 (title + headers)
        for r in all_rows:
            if not r or len(r) < 4:
                continue
            soc_val = str(r[0] or "").strip()
            if not soc_re.match(soc_val):
                continue
            if soc_val.endswith("0000") or soc_val.endswith("000"):
                continue
            try:
                base_emp = int(r[2])
                proj_emp = int(r[3])
            except (ValueError, TypeError):
                continue
            if base_emp <= 0 or proj_emp <= 0:
                continue
            pct_change = round((proj_emp - base_emp) / base_emp * 100, 2)
            cagr = calc_cagr(base_emp, proj_emp, 10)
            rows.append((
                soc_val, "metro", cbsa_code, cbsa_name,
                2022, 2032,
                base_emp, proj_emp,  # MSA-level, actual employment
                pct_change,
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))
        wb.close()
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    HI: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Minnesota  –  single CSV with 6 planning regions, 2022-2032
# ---------------------------------------------------------------------------

MN_FILE = RAW_DIR / "mn_projections.csv"

MN_AREA_TO_CBSA = {
    "Central Minnesota":                [("41060", "St. Cloud, MN")],
    "Northeast Minnesota":              [("20260", "Duluth, MN-WI")],
    "Seven County Mpls-St Paul, MN":    [("33460", "Minneapolis-St. Paul-Bloomington, MN-WI")],
    "Southeast Minnesota":              [("40340", "Rochester, MN"),
                                         ("31860", "Mankato, MN")],
    # Northwest Minnesota → no OES CBSA
    # Southwest Minnesota → no OES CBSA
}


def load_minnesota_metros():
    """Load MN planning-region projections from single CSV.

    Columns: Area Name, SOC Code, SOC Level, Title, Estimate Year,
             Estimate Year Employment, Projected Year,
             Projected Year Employment, Percent Change, ...
    SOC Level 6 = detailed.  SOC codes have no dashes.
    """
    soc_re = re.compile(r"^\d{2}-\d{4}$")
    rows = []
    cbsa_names = {}
    for cbsa_list in MN_AREA_TO_CBSA.values():
        for code, name in cbsa_list:
            cbsa_names[code] = name
    if not MN_FILE.exists():
        print(f"    [MN] missing {MN_FILE}, skipping")
        return rows
    df = pd.read_csv(MN_FILE, encoding="utf-8-sig")
    for _, r in df.iterrows():
        area = str(r.get("Area Name", "")).strip()
        cbsa_list = MN_AREA_TO_CBSA.get(area)
        if not cbsa_list:
            continue
        soc_level = str(r.get("SOC Level", "")).strip()
        if soc_level != "6":
            continue
        raw_soc = str(r.get("SOC Code", "")).strip()
        if len(raw_soc) == 6:
            soc_val = raw_soc[:2] + "-" + raw_soc[2:]
        else:
            continue
        if not soc_re.match(soc_val):
            continue
        if soc_val.endswith("0000") or soc_val.endswith("000"):
            continue
        try:
            base_emp = int(r.get("Estimate Year Employment", 0))
            proj_emp = int(r.get("Projected Year Employment", 0))
            pct_change = float(r.get("Percent Change", 0))
        except (ValueError, TypeError):
            continue
        if base_emp <= 0 or proj_emp <= 0:
            continue
        cagr = calc_cagr(base_emp, proj_emp, 10)
        for cbsa_code, _name in cbsa_list:
            rows.append((
                soc_val, "metro", cbsa_code,
                cbsa_names.get(cbsa_code, cbsa_code),
                2022, 2032,
                None, None,  # Region ≠ CBSA, CAGR only
                round(pct_change, 2),
                round(cagr, 6) if cagr is not None else None,
                "state_lmi",
            ))
    n_cbsas = len(set(r[2] for r in rows))
    print(f"    MN: {len(rows):,} occupation rows across {n_cbsas} CBSAs")
    return rows


# ---------------------------------------------------------------------------
# Fallback: Apply state-level CAGR to metros in states without scraper data
# ---------------------------------------------------------------------------

def load_metro_fallback(conn):
    """For metros without scraper data, produce adjusted estimates by blending
    state-level and national-level CAGR with a conservative dampening factor.

    Method (justified & conservative):
      1. For each occupation in a fallback CBSA, get both the state CAGR
         (from ProjectionsCentral) and the national CAGR (from BLS).
      2. Blend:  adjusted = 0.70 * state_cagr + 0.30 * national_cagr
         Rationale: state trends are more relevant than national, but national
         provides grounding for occupations that are outliers at the state level.
      3. Dampen positive growth by 10%:  if adjusted > 0: adjusted *= 0.90
         Rationale: these metros lack their own projections data, typically
         indicating smaller/less-dynamic local economies; tempering growth
         avoids over-optimistic estimates.
      4. Leave negative growth unchanged (already conservative to project decline).
      5. Source tagged 'adjusted_estimate' to distinguish from raw public data.

    If a national CAGR is unavailable for an occupation, fall back to state-only
    with the same 10% dampening on positive values.

    6. Size-based differentiation: Within each state, rank fallback metros by
       their total OES employment.  Apply a small multiplicative adjustment
       (±5%) so larger metros get slightly higher growth and smaller metros
       slightly lower.  This ensures metros in the same state receive distinct
       (non-identical) projections reflecting that larger metros tend to
       capture a greater share of growth.
    """
    print("\n=== Metro Adjusted Estimates (State+National Blend) ===")

    STATE_WEIGHT = 0.70
    NATIONAL_WEIGHT = 0.30
    POSITIVE_DAMPEN = 0.90  # multiply positive CAGR by this
    SIZE_SPREAD = 0.05      # max ±5% relative adjustment based on metro size rank

    import re
    STABBR_TO_FIPS = {
        "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06",
        "CO": "08", "CT": "09", "DE": "10", "DC": "11", "FL": "12",
        "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18",
        "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23",
        "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28",
        "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
        "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38",
        "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44",
        "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49",
        "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55",
        "WY": "56",
    }

    def extract_state_fips(area_name):
        match = re.search(r',\s*([A-Z]{2})', area_name)
        if match:
            return STABBR_TO_FIPS.get(match.group(1))
        return None

    # All OES CBSAs
    cbsa_df = pd.read_sql("""
        SELECT DISTINCT cbsa, area_name
        FROM bls_oes_areas
        WHERE area_type = 'metro' AND cbsa IS NOT NULL AND cbsa != ''
    """, conn)
    if cbsa_df.empty:
        print("  No metro areas found in bls_oes_areas")
        return 0

    cbsa_df["state_fips"] = cbsa_df["area_name"].apply(extract_state_fips)
    cbsa_df = cbsa_df.dropna(subset=["state_fips"])

    # CBSAs already covered by scrapers
    existing = pd.read_sql("""
        SELECT DISTINCT geo_code FROM employment_projections WHERE geo_level = 'metro'
    """, conn)
    existing_cbsas = set(existing["geo_code"].tolist()) if not existing.empty else set()
    cbsa_df = cbsa_df[~cbsa_df["cbsa"].isin(existing_cbsas)]
    if cbsa_df.empty:
        print("  All metros already have scraper data")
        return 0

    # State-level projections (keyed by state_fips + occ_code)
    state_proj = pd.read_sql("""
        SELECT occ_code, geo_code as state_fips, base_year, proj_year, cagr
        FROM employment_projections
        WHERE geo_level = 'state'
    """, conn)
    if state_proj.empty:
        print("  No state projections available for fallback")
        return 0

    # National projections (keyed by occ_code)
    national_proj = pd.read_sql("""
        SELECT occ_code, cagr as nat_cagr
        FROM employment_projections
        WHERE geo_level = 'national'
    """, conn)
    nat_cagr_map = {}
    if not national_proj.empty:
        nat_cagr_map = dict(zip(national_proj["occ_code"], national_proj["nat_cagr"]))

    # ---------- Metro size differentiation ----------
    # Get total employment per CBSA from OES data (most recent year).
    # No "all occupations" summary row exists, so SUM across detailed codes.
    oes_emp = pd.read_sql("""
        SELECT a.cbsa, SUM(o.tot_emp) AS tot_emp
        FROM bls_oes_areas a
        JOIN oes_employment o ON o.area_code = a.bls_area_code
        WHERE a.area_type = 'metro' AND a.cbsa IS NOT NULL AND a.cbsa != ''
          AND o.year = (SELECT MAX(year) FROM oes_employment)
        GROUP BY a.cbsa
        ORDER BY a.cbsa
    """, conn)
    cbsa_emp = {}
    if not oes_emp.empty:
        cbsa_emp = dict(zip(oes_emp["cbsa"], oes_emp["tot_emp"]))

    # Group fallback CBSAs by state and compute size-based adjustment factor.
    # Within each state, rank metros by total employment and assign a factor
    # in the range [-SIZE_SPREAD, +SIZE_SPREAD].  This creates small but
    # meaningful differentiation: e.g., Fargo gets a slightly higher growth
    # estimate than Minot within North Dakota.
    cbsa_size_factor = {}
    state_groups = cbsa_df.groupby("state_fips")
    for _sfips, group in state_groups:
        cbsas = group["cbsa"].tolist()
        if len(cbsas) <= 1:
            # Only one metro -> no differentiation needed
            for c in cbsas:
                cbsa_size_factor[c] = 0.0
            continue
        # Sort by employment (ascending)
        sorted_cbsas = sorted(cbsas, key=lambda c: cbsa_emp.get(c, 0))
        n = len(sorted_cbsas)
        for rank_i, c in enumerate(sorted_cbsas):
            # Linear interpolation: smallest -> -SIZE_SPREAD, largest -> +SIZE_SPREAD
            factor = -SIZE_SPREAD + (2 * SIZE_SPREAD) * (rank_i / (n - 1))
            cbsa_size_factor[c] = factor

    # Log differentiation
    if any(v != 0.0 for v in cbsa_size_factor.values()):
        print("  Size-based differentiation factors:")
        for c in sorted(cbsa_size_factor, key=lambda x: cbsa_size_factor[x]):
            emp = cbsa_emp.get(c, 0)
            factor = cbsa_size_factor[c]
            area = cbsa_df[cbsa_df["cbsa"] == c]["area_name"].iloc[0] if not cbsa_df[cbsa_df["cbsa"] == c].empty else c
            if factor != 0.0:
                print(f"    {c} {area:<50s} emp={emp:>8,}  factor={factor:+.3f}")

    rows = []
    for _, cbsa_row in cbsa_df.iterrows():
        cbsa_code = cbsa_row["cbsa"]
        state_fips = cbsa_row["state_fips"]
        area_name = cbsa_row["area_name"]
        size_factor = cbsa_size_factor.get(cbsa_code, 0.0)

        state_rows = state_proj[state_proj["state_fips"] == state_fips]
        for _, sr in state_rows.iterrows():
            occ = sr["occ_code"]
            st_cagr = sr["cagr"]
            if pd.isna(st_cagr):
                continue

            # Blend with national if available
            nat_cagr = nat_cagr_map.get(occ)
            if nat_cagr is not None and not pd.isna(nat_cagr):
                adjusted = STATE_WEIGHT * st_cagr + NATIONAL_WEIGHT * nat_cagr
            else:
                adjusted = st_cagr

            # Dampen positive growth conservatively
            if adjusted > 0:
                adjusted *= POSITIVE_DAMPEN

            # Apply metro size differentiation (multiplicative on absolute value)
            # Larger metros: factor > 0 -> slightly boost magnitude
            # Smaller metros: factor < 0 -> slightly reduce magnitude
            adjusted *= (1.0 + size_factor)

            # Derive pct_change from CAGR over the state's projection period
            n_years = int(sr["proj_year"]) - int(sr["base_year"])
            if n_years > 0:
                pct_change = round(((1 + adjusted) ** n_years - 1) * 100, 2)
            else:
                pct_change = 0.0

            rows.append((
                occ, "metro", cbsa_code, area_name,
                int(sr["base_year"]), int(sr["proj_year"]),
                None, None,  # no metro-level employment data
                pct_change,
                round(adjusted, 6),
                "adjusted_estimate",
            ))

    if rows:
        conn.executemany("""
            INSERT OR REPLACE INTO employment_projections
            (occ_code, geo_level, geo_code, geo_name, base_year, proj_year,
             base_emp, proj_emp, pct_change, cagr, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()

    covered = len(set(r[2] for r in rows))
    print(f"  Inserted {len(rows):,} adjusted estimate rows covering {covered} CBSAs")
    print(f"  Method: {STATE_WEIGHT:.0%} state + {NATIONAL_WEIGHT:.0%} national CAGR, "
          f"positive growth dampened by {(1 - POSITIVE_DAMPEN):.0%}, "
          f"±{SIZE_SPREAD:.0%} size-based differentiation")
    return len(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("Employment Projections Loader")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)
    create_table(conn)
    create_coverage_table(conn)

    n1 = load_bls_national(conn)
    n2 = load_projections_central(conn)
    n3 = load_metro_scrapers(conn)
    n4 = load_metro_fallback(conn)

    # Build coverage tracking table
    populate_coverage_table(conn)

    # Summary
    print("\n" + "=" * 60)
    print("Summary:")
    cur = conn.execute("""
        SELECT geo_level, source, COUNT(*) as cnt
        FROM employment_projections
        GROUP BY geo_level, source
        ORDER BY geo_level, source
    """)
    for row in cur:
        print(f"  {row[0]:10s} / {row[1]:25s} : {row[2]:>8,} rows")

    total = conn.execute("SELECT COUNT(*) FROM employment_projections").fetchone()[0]
    print(f"\n  TOTAL: {total:,} rows")
    print("=" * 60)

    conn.close()


if __name__ == "__main__":
    main()
